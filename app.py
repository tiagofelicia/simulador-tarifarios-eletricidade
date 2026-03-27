import streamlit as st
import pandas as pd
import datetime
import io
import json
import time
import re
import graficos as gfx
import processamento_dados as proc_dados
import calculos as calc

from st_aggrid import AgGrid, GridOptionsBuilder
from st_aggrid.shared import GridUpdateMode, JsCode
from bs4 import BeautifulSoup # Para processar o resumo HTML
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill # Para formatação Excel
from openpyxl.utils import get_column_letter # Para nomes de colunas Excel
from calendar import monthrange

st.set_page_config(page_title="Simulador de Tarifários Eletricidade 2026: Poupe na Fatura | Tiago Felícia", page_icon="🔌", layout="wide",initial_sidebar_state="collapsed")

# --- Carregar ficheiro Excel do Hugging Face ---

url_excel = "https://huggingface.co/spaces/tiagofelicia/simulador-tarifarios-eletricidade/resolve/main/Tarifarios_%F0%9F%94%8C_Eletricidade_Tiago_Felicia.xlsx"


tarifarios_fixos, tarifarios_indexados, OMIE_PERDAS_CICLOS, CONSTANTES = proc_dados.carregar_dados_excel_elec(url_excel)

potencias_validas = [1.15, 2.3, 3.45, 4.6, 5.75, 6.9, 10.35, 13.8, 17.25, 20.7, 27.6, 34.5, 41.4]
opcoes_horarias_existentes = list(tarifarios_fixos['opcao_horaria_e_ciclo'].dropna().unique())

# --- Mapas para encurtar os parâmetros do URL ---
MAPA_OH_PARA_URL = {
    "Simples": "S",
    "Bi-horário - Ciclo Diário": "BD",
    "Bi-horário - Ciclo Semanal": "BS",
    "Tri-horário - Ciclo Diário": "TD",
    "Tri-horário - Ciclo Semanal": "TS",
    "Tri-horário > 20.7 kVA - Ciclo Diário": "TD-A",
    "Tri-horário > 20.7 kVA - Ciclo Semanal": "TS-A",
}
# Cria o dicionário inverso automaticamente para ler os URLs
MAPA_URL_PARA_OH = {v: k for k, v in MAPA_OH_PARA_URL.items()}

#Funções
def inicializar_estado_e_url():
    """
    Verifica e inicializa o st.session_state, dando prioridade a valores no URL.
    Corre apenas uma vez por sessão, centralizando toda a lógica de arranque.
    """
    if 'estado_inicializado' in st.session_state:
        return

    # 1. Potência
    potencia_no_url = st.query_params.get("p") # Alterado de "potencia_url"
    if potencia_no_url and float(potencia_no_url) in potencias_validas:
        st.session_state.sel_potencia = float(potencia_no_url)
    else:
        st.session_state.sel_potencia = 3.45

    # 2. Opção Horária
    oh_codigo_url = st.query_params.get("oh") # Alterado de "oh_url"
    oh_nome_longo = MAPA_URL_PARA_OH.get(oh_codigo_url) # Traduz o código para o nome completo

    if oh_nome_longo and oh_nome_longo in opcoes_horarias_existentes:
        st.session_state.sel_opcao_horaria = oh_nome_longo
    else:
        st.session_state.sel_opcao_horaria = "Simples"


    # 3. Mês
    if "sel_mes" not in st.session_state:
        meses = ["Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho", "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]
        mes_atual_idx = datetime.datetime.now().month - 1
        st.session_state.sel_mes = meses[mes_atual_idx]

    # 4. Consumos
    st.session_state.exp_consumo_s = st.query_params.get("c_s", "158")
    st.session_state.exp_consumo_v = st.query_params.get("c_v", "63")
    st.session_state.exp_consumo_f = st.query_params.get("c_fv", "95")
    st.session_state.exp_consumo_c = st.query_params.get("c_c", "68")
    st.session_state.exp_consumo_p = st.query_params.get("c_p", "27")

    # 5. Opções Adicionais
    st.session_state.chk_tarifa_social = st.query_params.get("ts") == "1"
    st.session_state.chk_familia_numerosa = st.query_params.get("fn") == "1"
    st.session_state.dgeg_input = float(st.query_params.get("dgeg", 0.07))
    st.session_state.cav_input = float(st.query_params.get("cav", 2.85))
    # Se o parâmetro for "0", a checkbox fica False, senão fica com o seu default (True)
    st.session_state.chk_acp = st.query_params.get("acp") != "0"
    st.session_state.chk_continente = st.query_params.get("cont") != "0"

    # 6. "O Meu Tarifário"
    if st.query_params.get("m_a") == "1":
        st.session_state.chk_meu_tarifario_ativo = True
        
        # Função auxiliar para converter para float de forma segura
        def safe_float_or_none(val):
            return float(val) if val is not None else None

        # Preços (convertidos para float)
        st.session_state.potencia_meu_input_val = safe_float_or_none(st.query_params.get("m_ps"))
        st.session_state.energia_meu_s_input_val = safe_float_or_none(st.query_params.get("m_es"))
        st.session_state.energia_meu_v_input_val = safe_float_or_none(st.query_params.get("m_ev"))
        st.session_state.energia_meu_f_input_val = safe_float_or_none(st.query_params.get("m_efv"))
        st.session_state.energia_meu_c_input_val = safe_float_or_none(st.query_params.get("m_ec"))
        st.session_state.energia_meu_p_input_val = safe_float_or_none(st.query_params.get("m_ep"))
        
        # Flags (lógica original está correta)
        st.session_state.meu_tar_energia_val = st.query_params.get("m_te") != "0"
        st.session_state.meu_tar_potencia_val = st.query_params.get("m_tp") != "0"
        st.session_state.meu_fin_tse_incluido_val = st.query_params.get("m_tse") != "0"
        
        # Descontos/Acréscimos (já estavam a ser convertidos, mantemos para consistência)
        st.session_state.meu_desconto_energia_val = float(st.query_params.get("m_de", 0.0))
        st.session_state.meu_desconto_potencia_val = float(st.query_params.get("m_dp", 0.0))
        st.session_state.meu_desconto_fatura_val = float(st.query_params.get("m_df", 0.0))
        st.session_state.meu_acrescimo_fatura_val = float(st.query_params.get("m_af", 0.0))

    st.session_state.estado_inicializado = True

inicializar_estado_e_url()

def atualizar_url_potencia():
    """
    Callback para atualizar o URL quando a potência muda e para sincronizar
    os consumos se a opção horária for forçada a mudar.
    """
    # Guarda: Se estiver em modo diagrama, não faz nada.
    if 'dados_completos_ficheiro' in st.session_state and st.session_state.get('dados_completos_ficheiro') is not None:
        return
    
    # --- Parte 1: Atualizar o URL da potência ---
    potencia_selecionada = st.session_state.get("sel_potencia")
    if potencia_selecionada:
        # Limpa a chave antiga, se existir
        if "potencia_url" in st.query_params:
            del st.query_params["potencia_url"]
        # Adiciona a nova chave curta
        st.query_params["p"] = str(potencia_selecionada)

    # --- Parte 2: Lógica de Sincronização ---
    opcao_atual = st.session_state.get("sel_opcao_horaria")

    # Replicamos a lógica que filtra as opções horárias válidas
    if potencia_selecionada >= 27.6:
        opcoes_validas_novas = [o for o in opcoes_horarias_existentes if "Tri-horário > 20.7 kVA" in o]
    else:
        opcoes_validas_novas = [o for o in opcoes_horarias_existentes if not "Tri-horário > 20.7 kVA" in o]

    # Se a opção horária atual já não é válida, sabemos que ela vai mudar.
    if opcao_atual not in opcoes_validas_novas and opcoes_validas_novas:
        # A nova opção será a primeira da lista.
        nova_opcao_horaria = opcoes_validas_novas[0]
        
        # Atualizamos o estado da sessão para a nova opção.
        st.session_state.sel_opcao_horaria = nova_opcao_horaria
        
        # Agora, chamamos manualmente o callback da opção horária.
        # Ele já tem toda a lógica correta para repor os consumos e atualizar o URL.
        atualizar_url_opcao_horaria()

def atualizar_url_opcao_horaria():
    """
    Callback para atualizar o URL e o estado dos consumos quando a opção horária muda.
    """
    # Se estiver em modo diagrama, não faz nada.
    if 'dados_completos_ficheiro' in st.session_state and st.session_state.get('dados_completos_ficheiro') is not None:
        return
    
    opcao_selecionada = st.session_state.get("sel_opcao_horaria")

    # --- 1. ATUALIZAR O ESTADO DA SESSÃO PRIMEIRO ---
    # Esta lógica foi movida do corpo principal do script para aqui.
    if opcao_selecionada.lower() == "simples":
        st.session_state.exp_consumo_s = "158"
        st.session_state.exp_consumo_v = "0"
        st.session_state.exp_consumo_f = "0"
        st.session_state.exp_consumo_c = "0"
        st.session_state.exp_consumo_p = "0"
    elif opcao_selecionada.lower().startswith("bi"):
        st.session_state.exp_consumo_s = "0"
        st.session_state.exp_consumo_v = "63"
        st.session_state.exp_consumo_f = "95"
        st.session_state.exp_consumo_c = "0"
        st.session_state.exp_consumo_p = "0"
    elif opcao_selecionada.lower().startswith("tri"):
        st.session_state.exp_consumo_s = "0"
        st.session_state.exp_consumo_v = "63"
        st.session_state.exp_consumo_f = "0"
        st.session_state.exp_consumo_c = "68"
        st.session_state.exp_consumo_p = "27"

    # --- 2. AGORA, ATUALIZAR O URL COM O ESTADO JÁ CORRIGIDO ---
    if opcao_selecionada:
        # Traduz o nome completo para o código curto
        codigo_oh = MAPA_OH_PARA_URL.get(opcao_selecionada)
        if codigo_oh:
            # Limpa a chave antiga, se existir
            if "oh_url" in st.query_params:
                del st.query_params["oh_url"]
            # Adiciona a nova chave curta
            st.query_params["oh"] = codigo_oh
    
    # Esta função agora usará os valores que acabámos de definir.
    atualizar_url_consumos()

    # Também chamar a atualização do "Meu Tarifário" para limpar parâmetros
    # de energia que possam ter ficado do modo anterior.
    atualizar_url_meu_tarifario()

def atualizar_url_opcoes_adicionais():
    """Callback para monitorizar e atualizar o URL com todas as opções adicionais."""
    # Guarda: Se estiver em modo diagrama, não faz nada.
    if 'dados_completos_ficheiro' in st.session_state and st.session_state.get('dados_completos_ficheiro') is not None:
        return

    # Lógica para Tarifa Social e Família Numerosa
    if st.session_state.get("chk_tarifa_social", False):
        st.query_params["ts"] = "1"
    elif "ts" in st.query_params:
        del st.query_params["ts"]

    if st.session_state.get("chk_familia_numerosa", False):
        st.query_params["fn"] = "1"
    elif "fn" in st.query_params:
        del st.query_params["fn"]

    # Lógica para DGEG e CAV
    dgeg_default = 0.07
    if st.session_state.get("dgeg_input", dgeg_default) != dgeg_default:
        st.query_params["dgeg"] = str(st.session_state.dgeg_input)
    elif "dgeg" in st.query_params:
        del st.query_params["dgeg"]

    cav_default = 2.85
    if st.session_state.get("cav_input", cav_default) != cav_default:
        st.query_params["cav"] = str(st.session_state.cav_input)
    elif "cav" in st.query_params:
        del st.query_params["cav"]

    # Lógica para Quota ACP e Desconto Continente
    # Como o padrão é True, só guardamos no URL se o valor for False ("0")
    if not st.session_state.get("chk_acp", True):
        st.query_params["acp"] = "0"
    elif "acp" in st.query_params:
        del st.query_params["acp"]

    if not st.session_state.get("chk_continente", True):
        st.query_params["cont"] = "0"
    elif "cont" in st.query_params:
        del st.query_params["cont"]

def atualizar_url_meu_tarifario():
    """Callback para monitorizar e atualizar o URL com os dados do Meu Tarifário."""
    # Se estiver em modo diagrama, não faz nada.
    if 'dados_completos_ficheiro' in st.session_state and st.session_state.get('dados_completos_ficheiro') is not None:
        return
    
    # Lista de todas as chaves possíveis para o Meu Tarifário no URL
    chaves_meu_tar_url = [
        "m_a", "m_ps", "m_es", "m_ev", "m_efv", "m_ec", "m_ep",
        "m_te", "m_tp", "m_tse", "m_de", "m_dp", "m_df", "m_af"
    ]
    
    # Limpar sempre as chaves antigas
    for chave in chaves_meu_tar_url:
        if chave in st.query_params:
            del st.query_params[chave]

    # Se a secção "Meu Tarifário" não estiver ativa, não fazemos mais nada
    if not st.session_state.get("chk_meu_tarifario_ativo", False):
        return

    # Se estiver ativa, adicionamos a flag e os preços relevantes
    st.query_params["m_a"] = "1" # 'm_a' significa 'meu_ativo'
    
    opcao_selecionada = st.session_state.get("sel_opcao_horaria", "Simples").lower()
    
    # Preços de energia e potência (lógica existente)
    potencia_val = st.session_state.get("potencia_meu_input_val", 0.0)
    if potencia_val: st.query_params["m_ps"] = str(potencia_val)
    
    if "simples" in opcao_selecionada:
        energia_s_val = st.session_state.get("energia_meu_s_input_val", 0.0)
        if energia_s_val: st.query_params["m_es"] = str(energia_s_val)
    elif opcao_selecionada.startswith("bi"):
        energia_v_val = st.session_state.get("energia_meu_v_input_val", 0.0)
        energia_f_val = st.session_state.get("energia_meu_f_input_val", 0.0)
        if energia_v_val: st.query_params["m_ev"] = str(energia_v_val)
        if energia_f_val: st.query_params["m_efv"] = str(energia_f_val)
    elif opcao_selecionada.startswith("tri"):
        energia_v_val = st.session_state.get("energia_meu_v_input_val", 0.0)
        energia_c_val = st.session_state.get("energia_meu_c_input_val", 0.0)
        energia_p_val = st.session_state.get("energia_meu_p_input_val", 0.0)
        if energia_v_val: st.query_params["m_ev"] = str(energia_v_val)
        if energia_c_val: st.query_params["m_ec"] = str(energia_c_val)
        if energia_p_val: st.query_params["m_ep"] = str(energia_p_val)

    # Adicionar as flags e os descontos/acréscimos
    # Para as checkboxes, que por defeito são True, só guardamos no URL se forem False (valor "0")
    if not st.session_state.get("meu_tar_energia_val", True):
        st.query_params["m_te"] = "0"
    if not st.session_state.get("meu_tar_potencia_val", True):
        st.query_params["m_tp"] = "0"
    if not st.session_state.get("meu_fin_tse_incluido_val", True):
        st.query_params["m_tse"] = "0"

    # Para os campos numéricos, que por defeito são 0, só guardamos se tiverem um valor > 0
    desconto_energia_val = st.session_state.get("meu_desconto_energia_val", 0.0)
    if desconto_energia_val: st.query_params["m_de"] = str(desconto_energia_val)
    
    desconto_potencia_val = st.session_state.get("meu_desconto_potencia_val", 0.0)
    if desconto_potencia_val: st.query_params["m_dp"] = str(desconto_potencia_val)

    desconto_fatura_val = st.session_state.get("meu_desconto_fatura_val", 0.0)
    if desconto_fatura_val: st.query_params["m_df"] = str(desconto_fatura_val)
    
    acrescimo_fatura_val = st.session_state.get("meu_acrescimo_fatura_val", 0.0)
    if acrescimo_fatura_val: st.query_params["m_af"] = str(acrescimo_fatura_val)


def atualizar_url_consumos():
    """
    Callback para atualizar o URL apenas com os consumos relevantes
    para a opção horária selecionada, limpando os restantes.
    """
    # Se estiver em modo diagrama, não faz nada.
    if 'dados_completos_ficheiro' in st.session_state and st.session_state.get('dados_completos_ficheiro') is not None:
        return
    
    opcao_selecionada = st.session_state.get("sel_opcao_horaria", "Simples").lower()
    
    # Lista de todas as chaves de consumo possíveis no URL
    chaves_consumo_url = ["c_s", "c_v", "c_fv", "c_c", "c_p"]

    # Limpar todas as chaves de consumo existentes para começar do zero
    for chave in chaves_consumo_url:
        if chave in st.query_params:
            del st.query_params[chave]

    # Adicionar apenas as chaves relevantes para a opção atual
    if "simples" in opcao_selecionada:
        st.query_params["c_s"] = st.session_state.get("exp_consumo_s", "0")
    elif opcao_selecionada.startswith("bi"):
        st.query_params["c_v"] = st.session_state.get("exp_consumo_v", "0")
        st.query_params["c_fv"] = st.session_state.get("exp_consumo_f", "0")
    elif opcao_selecionada.startswith("tri"):
        st.query_params["c_v"] = st.session_state.get("exp_consumo_v", "0")
        st.query_params["c_c"] = st.session_state.get("exp_consumo_c", "0")
        st.query_params["c_p"] = st.session_state.get("exp_consumo_p", "0")

# FUNÇÃO criar_tabela_analise_completa_html
def criar_tabela_analise_completa_html(consumos_agregados, omie_agregados):
    """
    Gera uma tabela HTML detalhada, com cores personalizadas de fundo e texto
    que se adaptam ao tema claro/escuro do Streamlit.
    """
    
    # --- Deteção do Tema Atual do Streamlit ---
    is_dark_theme = st.get_option('theme.base') == 'dark'

    # --- Definição de DUAS paletas de cores ---    
    # Paleta para o Tema Claro (a sua original)
    cores_light = {
        'header': {
            'S':  {'bg': '#A6A6A6'}, 'BD': {'bg': '#A9D08E'}, 'BS': {'bg': '#8EA9DB'},
            'TD': {'bg': '#BF8F00', 'text': '#FFFFFF'}, 'TS': {'bg': '#C65911', 'text': '#FFFFFF'}
        },
        'cell': {
            'S':    {'bg': '#D9D9D9'},
            'BD_V': {'bg': '#C6E0B4'}, 'BD_F': {'bg': '#E2EFDA'},
            'BS_V': {'bg': '#B4C6E7'}, 'BS_F': {'bg': '#D9E1F2'},
            'TD_V': {'bg': '#FFD966'}, 'TD_C': {'bg': '#FFE699'}, 'TD_P': {'bg': '#FFF2CC'},
            'TS_V': {'bg': '#F4B084'}, 'TS_C': {'bg': '#F8CBAD'}, 'TS_P': {'bg': '#FCE4D6'}
        }
    }
    
    # Paleta para o Tema Escuro (cores com melhor contraste)
    cores_dark = {
        'header': {
            'S':  {'bg': '#5A5A5A'}, 'BD': {'bg': '#4B6140'}, 'BS': {'bg': '#3E4C6D'},
            'TD': {'bg': '#8C6600'}, 'TS': {'bg': '#95430D'}
        },
        'cell': {
            'S':    {'bg': '#404040'},
            'BD_V': {'bg': '#384E30'}, 'BD_F': {'bg': '#2E3F27'},
            'BS_V': {'bg': '#2D3850'}, 'BS_F': {'bg': '#242C40'},
            'TD_V': {'bg': '#665000'}, 'TD_C': {'bg': '#594600'}, 'TD_P': {'bg': '#4D3C00'},
            'TS_V': {'bg': '#6F3A1D'}, 'TS_C': {'bg': '#613319'}, 'TS_P': {'bg': '#542C15'}
        }
    }

    # --- Selecionar a paleta e cores de base com base no tema ---
    if is_dark_theme:
        cores = cores_dark
        row_label_bg = '#1E2128'   # Fundo escuro para os rótulos
        row_label_text = '#FFFFFF' # Texto branco
        border_color = '#3E414B'   # Borda mais escura
    else:
        cores = cores_light
        row_label_bg = '#f8f9fa'   # Fundo claro original
        row_label_text = '#212529' # Texto preto
        border_color = '#999'

    # --- Geração do CSS (usa as variáveis de cor dinâmicas) ---
    html = "<style>"
    html += ".analise-table { width: 100%; border-collapse: collapse; margin: 20px 0; font-size: 13px; font-family: sans-serif; text-align: center; }"
    # ---  Usa a cor da borda dinâmica ---
    html += f".analise-table th, .analise-table td {{ padding: 8px 10px; border: 1px solid {border_color}; }}"
    html += ".analise-table thead th { font-weight: bold; }"
    html += ".analise-table .header-main { vertical-align: middle; }"
    # --- Usa as cores de fundo e texto dinâmicas para os rótulos das linhas ---
    html += f".analise-table .row-label {{ text-align: center; font-weight: bold; background-color: {row_label_bg}; color: {row_label_text}; }}"
    
    # Este loop usa a paleta de cores correta (clara ou escura)
    for tipo_estilo, mapa_cores in cores.items():
        for chave, config_cor in mapa_cores.items():
            cor_fundo = config_cor['bg']
            cor_texto = config_cor.get('text')
            
            if not cor_texto:
                try:
                    r, g, b = int(cor_fundo[1:3], 16), int(cor_fundo[3:5], 16), int(cor_fundo[5:7], 16)
                    cor_texto = '#000000' if (r*0.299 + g*0.587 + b*0.114) > 140 else '#FFFFFF'
                except:
                    cor_texto = '#000000' if is_dark_theme else '#FFFFFF'
            
            html += f".{tipo_estilo}-{chave} {{ background-color: {cor_fundo}; color: {cor_texto}; }}"
    html += "</style>"

    # --- 2. Extração e Cálculo de Todos os Valores ---
    data = {}
    total_kwh_geral = consumos_agregados.get('Simples', 0)
    ciclos_info = {'S': ['S'], 'BD': ['V', 'F'], 'BS': ['V', 'F'], 'TD': ['V', 'C', 'P'], 'TS': ['V', 'C', 'P']}
    for ciclo, periodos in ciclos_info.items():
        total_consumo_ciclo = sum(consumos_agregados.get(ciclo, {}).values()) if ciclo != 'S' else total_kwh_geral
        for periodo in periodos:
            chave_omie = f"{ciclo}_{periodo}" if ciclo != 'S' else 'S'
            chave_kwh = periodo if ciclo != 'S' else 'Simples'
            kwh = consumos_agregados.get(ciclo, {}).get(periodo, 0) if ciclo != 'S' else total_kwh_geral
            data[f"{ciclo}_{periodo}"] = {
                'omie': omie_agregados.get(chave_omie, 0),
                'kwh': kwh,
                'perc': (kwh / total_consumo_ciclo * 100) if total_consumo_ciclo > 0 else (100 if ciclo == 'S' else 0)
            }

    # --- 3. Construção da Tabela HTML ---
    def fnum(n, casas_decimais=0, sufixo=""):
        try:
            return f"{float(n):,.{casas_decimais}f}".replace(",", " ") + sufixo
        except (ValueError, TypeError):
            return "-"

    def criar_celula(valor, classe, casas_decimais=0, sufixo=""):
        return f"<td class='{classe}'>{fnum(valor, casas_decimais, sufixo)}</td>"
    
    html += "<table class='analise-table'>"
    html += "<thead>"
    html += f"<tr><th rowspan='2'></th><th rowspan='2' class='header-S'>Simples</th><th colspan='2' class='header-BD'>Bi-horário Diário</th><th colspan='2' class='header-BS'>Bi-horário Semanal</th><th colspan='3' class='header-TD'>Tri-horário Diário</th><th colspan='3' class='header-TS'>Tri-horário Semanal</th></tr>"
    html += f"<tr class='header-sub'><th class='cell-BD_V'>Vazio</th><th class='cell-BD_F'>Fora Vazio</th><th class='cell-BS_V'>Vazio</th><th class='cell-BS_F'>Fora Vazio</th><th class='cell-TD_V'>Vazio</th><th class='cell-TD_C'>Cheias</th><th class='cell-TD_P'>Ponta</th><th class='cell-TS_V'>Vazio</th><th class='cell-TS_C'>Cheias</th><th class='cell-TS_P'>Ponta</th></tr>"
    html += "</thead><tbody>"
    
    # Linha Média OMIE
    html += '<tr><td class="row-label">Média OMIE (€/MWh)</td>'
    html += f"{criar_celula(data['S_S']['omie'], 'cell-S', 2)}"
    html += f"{criar_celula(data['BD_V']['omie'], 'cell-BD_V', 2)}{criar_celula(data['BD_F']['omie'], 'cell-BD_F', 2)}"
    html += f"{criar_celula(data['BS_V']['omie'], 'cell-BS_V', 2)}{criar_celula(data['BS_F']['omie'], 'cell-BS_F', 2)}"
    html += f"{criar_celula(data['TD_V']['omie'], 'cell-TD_V', 2)}{criar_celula(data['TD_C']['omie'], 'cell-TD_C', 2)}{criar_celula(data['TD_P']['omie'], 'cell-TD_P', 2)}"
    html += f"{criar_celula(data['TS_V']['omie'], 'cell-TS_V', 2)}{criar_celula(data['TS_C']['omie'], 'cell-TS_C', 2)}{criar_celula(data['TS_P']['omie'], 'cell-TS_P', 2)}</tr>"
    
    # Linha Consumo Real (kWh)
    html += '<tr><td class="row-label">Consumo Real (kWh)</td>'
    html += f"{criar_celula(data['S_S']['kwh'], 'cell-S', 0)}"
    html += f"{criar_celula(data['BD_V']['kwh'], 'cell-BD_V', 0)}{criar_celula(data['BD_F']['kwh'], 'cell-BD_F', 0)}"
    html += f"{criar_celula(data['BS_V']['kwh'], 'cell-BS_V', 0)}{criar_celula(data['BS_F']['kwh'], 'cell-BS_F', 0)}"
    html += f"{criar_celula(data['TD_V']['kwh'], 'cell-TD_V', 0)}{criar_celula(data['TD_C']['kwh'], 'cell-TD_C', 0)}{criar_celula(data['TD_P']['kwh'], 'cell-TD_P', 0)}"
    html += f"{criar_celula(data['TS_V']['kwh'], 'cell-TS_V', 0)}{criar_celula(data['TS_C']['kwh'], 'cell-TS_C', 0)}{criar_celula(data['TS_P']['kwh'], 'cell-TS_P', 0)}</tr>"

    # Linha Consumo %
    html += '<tr><td class="row-label">Consumo %</td>'
    html += f"{criar_celula(data['S_S']['perc'], 'cell-S', 1, '%')}"
    html += f"{criar_celula(data['BD_V']['perc'], 'cell-BD_V', 1, '%')}{criar_celula(data['BD_F']['perc'], 'cell-BD_F', 1, '%')}"
    html += f"{criar_celula(data['BS_V']['perc'], 'cell-BS_V', 1, '%')}{criar_celula(data['BS_F']['perc'], 'cell-BS_F', 1, '%')}"
    html += f"{criar_celula(data['TD_V']['perc'], 'cell-TD_V', 1, '%')}{criar_celula(data['TD_C']['perc'], 'cell-TD_C', 1, '%')}{criar_celula(data['TD_P']['perc'], 'cell-TD_P', 1, '%')}"
    html += f"{criar_celula(data['TS_V']['perc'], 'cell-TS_V', 1, '%')}{criar_celula(data['TS_C']['perc'], 'cell-TS_C', 1, '%')}{criar_celula(data['TS_P']['perc'], 'cell-TS_P', 1, '%')}</tr>"
    
    html += "</tbody></table>"
    return html

# --- Função para REINICIAR o simulador para os valores padrão ---
def reiniciar_simulador():
    """
    Repõe explicitamente todos os valores de session_state para os seus defaults.
    """
    
    # Lista de todas as chaves a serem completamente removidas do session_state
    chaves_a_apagar = [
        # Chaves de ficheiro, datas e estado relacionado
        'dados_completos_ficheiro', 'nome_ficheiro_processado',
        'min_date_ficheiro', 'max_date_ficheiro', 
        'data_inicio_val', 'data_fim_val', 'dias_manual_val', 
        'session_initialized_dates', 'previous_mes_for_dates',
        'data_inicio_anterior_val', 'data_fim_anterior_val',
        'dados_consumo_processados', 'dias_manual_input_key',
        
        # Chaves de consumo manual e seus valores válidos anteriores
        'exp_consumo_s', 'exp_consumo_v', 'exp_consumo_f', 'exp_consumo_c', 'exp_consumo_p',
        'exp_consumo_s_anterior_valido', 'exp_consumo_v_anterior_valido', 
        'exp_consumo_f_anterior_valido', 'exp_consumo_tc_anterior_valido', 
        'exp_consumo_tp_anterior_valido',

        # Chaves de OMIE manual e estado de edição
        'omie_s_input_field', 'omie_v_input_field', 'omie_f_input_field', 
        'omie_c_input_field', 'omie_p_input_field', 'omie_foi_editado_manualmente',
        'last_omie_dependency_key_for_inputs',

        # Chaves do "Meu Tarifário"
        "energia_meu_s_input_val", "potencia_meu_input_val", "energia_meu_v_input_val",
        "energia_meu_f_input_val", "energia_meu_c_input_val", "energia_meu_p_input_val",
        "meu_tar_energia_val", "meu_tar_potencia_val", "meu_fin_tse_incluido_val",
        "meu_desconto_energia_val", "meu_desconto_potencia_val", "meu_desconto_fatura_val",
        "meu_acrescimo_fatura_val", 'meu_tarifario_calculado'
    ]

    for key in chaves_a_apagar:
        if key in st.session_state:
            del st.session_state[key]

    # 1. Repor os inputs principais
    st.session_state.sel_potencia = 3.45  # Default da potência
    st.session_state.sel_opcao_horaria = "Simples" # Default da opção horária
    meses = ["Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho", "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]
    mes_atual_idx = datetime.datetime.now().month - 1
    st.session_state.sel_mes = meses[mes_atual_idx] # Default do mês

    # 2. Repor os inputs de consumo para os valores padrão
    st.session_state.exp_consumo_s = "158"
    st.session_state.exp_consumo_v = "63"
    st.session_state.exp_consumo_f = "95"
    st.session_state.exp_consumo_c = "68"
    st.session_state.exp_consumo_p = "27"

    # 4. Repor as checkboxes de opções adicionais
    st.session_state.chk_tarifa_social = False
    st.session_state.chk_familia_numerosa = False
    st.session_state.chk_acp = True
    st.session_state.chk_continente = True
    st.session_state.chk_modo_comparacao_opcoes = False
            
    # 6. Repor os filtros da tabela
    st.session_state.filter_segmento_selectbox = "Residencial"
    st.session_state.filter_tipos_multi = []
    st.session_state.filter_faturacao_selectbox = "Todas"
    st.session_state.filter_pagamento_selectbox = "Todos"

    st.query_params.clear()  # Limpa todos os parâmetros do URL

    # Definir a flag de reinício
    st.session_state.app_just_reset = True

    st.success("Simulador reiniciado para os valores padrão.")


# --- Obter valor constante do Financiamento TSE ---
FINANCIAMENTO_TSE_VAL = calc.obter_constante("Financiamento_TSE", CONSTANTES)

# --- Obter valor constante da Quota ACP ---
VALOR_QUOTA_ACP_MENSAL = calc.obter_constante("Quota_ACP", CONSTANTES)

def preparar_dados_para_graficos(df_consumos_filtrado, df_omie_filtrado, opcao_horaria_selecionada, dias_periodo):
    """
    Prepara os dados agregados para os gráficos Highcharts no MODO DIAGRAMA.
    - Gráfico Horário: Consumo TOTAL vs. Média OMIE. Tooltip mostra TOTAL e MÉDIA.
    - Gráfico Diário: Consumo empilhado vs. Média OMIE por período.
    """
    if df_consumos_filtrado.empty or df_omie_filtrado.empty:
        return None, None

    df_merged = pd.merge(df_consumos_filtrado, df_omie_filtrado, on='DataHora', how='inner')
    if df_merged.empty:
        st.warning("Não foi possível alinhar dados de consumo e OMIE para os gráficos.")
        return None, None

    # --- Lógica do Título do Ciclo e Períodos ---
    oh_lower = opcao_horaria_selecionada.lower()
    titulo_ciclo = "Simples"
    ciclo_a_usar = None
    periodos_ciclo = []
    
    nomes_periodos = {'V': 'Vazio', 'F': 'Fora Vazio', 'C': 'Cheias', 'P': 'Ponta'}
    
    cores_consumo_diario = {'V_bi': '#A9D18E', 'V_tri': '#BF9000', 'F': '#E2F0D9', 'C': '#FFD966', 'P': '#FFF2CC'}
    cores_consumo_semanal = {'V_bi': '#8FAADC', 'V_tri': '#C55A11', 'F': '#DAE3F3', 'C': '#F4B183', 'P': '#FBE5D6'}
    
    cores_omie = {'S': '#FF0000','V': '#000000', 'F': '#FFC000', 'C': '#2F5597', 'P': '#00B050'}
    
    cores_consumo_a_usar = cores_consumo_diario if "diário" in oh_lower else cores_consumo_semanal

    if oh_lower.startswith("bi"):
        periodos_ciclo = ['V', 'F']
        if "diário" in oh_lower:
            ciclo_a_usar = 'BD'; titulo_ciclo = "Bi-Horário - Ciclo Diário"
        else:
            ciclo_a_usar = 'BS'; titulo_ciclo = "Bi-Horário - Ciclo Semanal"
            
    elif oh_lower.startswith("tri"):
        periodos_ciclo = ['V', 'C', 'P']
        if "diário" in oh_lower:
            ciclo_a_usar = 'TD'; titulo_ciclo = "Tri-Horário - Ciclo Diário"
        else:
            ciclo_a_usar = 'TS'; titulo_ciclo = "Tri-Horário - Ciclo Semanal"
            
    # --- 1. Gráfico Horário (com barras empilhadas e OMIE por período) ---
    df_horario = df_merged.copy()
    df_horario['HoraParaAgrupar'] = (df_horario['DataHora'] - pd.Timedelta(seconds=1)).dt.hour
    
    num_dias = dias_periodo if dias_periodo > 0 else 1
    
    series_horario = []
    
    if not ciclo_a_usar:
        agg_horario = df_horario.groupby('HoraParaAgrupar').agg(
            Consumo_kWh_Total=('Consumo (kWh)', 'sum')
        ).reindex(range(24), fill_value=0)
        
        agg_horario['Consumo_kWh_Medio'] = agg_horario['Consumo_kWh_Total'] / num_dias
        
        data_points_horario = [
            {'y': row['Consumo_kWh_Total'], 'media': row['Consumo_kWh_Medio']}
            for _, row in agg_horario.iterrows()
        ]
        series_horario.append({"name": "Consumo por hora (kWh)", "type": "column", "data": data_points_horario, "yAxis": 0, "color": "#BFBFBF"})
    else:
        agg_total_horario_periodo = df_horario.groupby(['HoraParaAgrupar', ciclo_a_usar])['Consumo (kWh)'].sum().unstack(fill_value=0)
        agg_media_horario_periodo = agg_total_horario_periodo / num_dias
        
        for p in reversed(periodos_ciclo):
            if p in agg_total_horario_periodo.columns:
                
                dados_t_p = agg_total_horario_periodo[p].reindex(range(24), fill_value=0)
                dados_m_p = agg_media_horario_periodo[p].reindex(range(24), fill_value=0)
                
                data_points_periodo = [
                    {'y': total, 'media': media} 
                    for total, media in zip(dados_t_p, dados_m_p)
                ]
                
                cor_key = 'V_tri' if p == 'V' and oh_lower.startswith("tri") else ('V_bi' if p == 'V' else p)
                series_horario.append({
                    "name": f"Consumo {nomes_periodos.get(p, p)} (kWh)",
                    "type": "column",
                    "data": data_points_periodo,
                    "yAxis": 0,
                    "color": cores_consumo_a_usar.get(cor_key)
                })

    # Adicionar as linhas OMIE (visível e ocultas)
    agg_omie_horario_simples = df_horario.groupby('HoraParaAgrupar')['OMIE'].mean().reindex(range(24))
    dados_omie_simples_final = agg_omie_horario_simples.round(2).where(pd.notna(agg_omie_horario_simples), None).tolist() # Converte NaN para None
    series_horario.append({
        "name": "Média horária OMIE (€/MWh)", "type": "line", 
        "data": dados_omie_simples_final, "yAxis": 1, "color": cores_omie.get('S')
    })
    
    if ciclo_a_usar:
        agg_omie_horario_periodos = df_horario.groupby(['HoraParaAgrupar', ciclo_a_usar])['OMIE'].mean().unstack()
        for p in periodos_ciclo:
            if p in agg_omie_horario_periodos.columns:
                dados_omie_p = agg_omie_horario_periodos[p].reindex(range(24))
                dados_omie_p_final = dados_omie_p.round(2).where(pd.notna(dados_omie_p), None).tolist() # Converte NaN para None
                series_horario.append({
                    "name": f"Média OMIE {nomes_periodos.get(p, p)} (€/MWh)", "type": "line",
                    "data": dados_omie_p_final, "yAxis": 1, 
                    "color": cores_omie.get(p), "visible": False
                })

    dados_grafico_horario = {
        'titulo': f'Consumo por Hora vs. Preço OMIE Horário ({titulo_ciclo})',
        'titulo_eixo_y1': 'Consumo por hora (kWh)',
        'titulo_eixo_y2': 'Média horária OMIE (€/MWh)',
        'categorias': [f"{h}h-24h" if h == 23 else f"{h}h-{h + 1}h" for h in range(24)],
        'series': series_horario
    }

    # --- 2. Gráfico Diário ---
    df_diario = df_merged.copy()
    df_diario['data_dia'] = pd.to_datetime(df_diario['DataHora'].dt.date)
    
    agg_diario_base = df_diario.groupby('data_dia').agg(
        Consumo_kWh=('Consumo (kWh)', 'sum'),
        Media_OMIE_Simples=('OMIE', 'mean')
    ).sort_index()
    
    series_diario = []
    
    if not ciclo_a_usar:
        series_diario.insert(0, {"name": "Consumo por dia (kWh)", "type": "column", "data": agg_diario_base['Consumo_kWh'].round(2).where(pd.notna, None).tolist(), "yAxis": 0, "color": "#BFBFBF"})
    else:
        agg_consumo_periodos = df_diario.groupby(['data_dia', ciclo_a_usar])['Consumo (kWh)'].sum().unstack(fill_value=0)
        agg_consumo_periodos = agg_consumo_periodos.reindex(agg_diario_base.index)
        for p in periodos_ciclo:
            if p in agg_consumo_periodos.columns:
                cor_a_usar = cores_consumo_a_usar.get(p)
                if p == 'V':
                    chave_cor = 'V_tri' if oh_lower.startswith("tri") else 'V_bi'
                    cor_a_usar = cores_consumo_a_usar.get(chave_cor)

                series_diario.insert(0, {
                    "name": f"Consumo {nomes_periodos.get(p, p)} (kWh)", "type": "column",
                    "data": agg_consumo_periodos[p].round(2).where(pd.notna, None).tolist(),
                    "yAxis": 0, "color": cor_a_usar
                })

    # A lógica de conversão para NaN -> None é aplicada aqui
    dados_omie_diario_simples_final = agg_diario_base['Media_OMIE_Simples'].round(2).where(pd.notna(agg_diario_base['Media_OMIE_Simples']), None).tolist()
    series_diario.append({"name": "Média diária OMIE (€/MWh)", "type": "line", "data": dados_omie_diario_simples_final, "yAxis": 1, "color": cores_omie.get('S')})
    
    if ciclo_a_usar:
        agg_omie_periodos = df_diario.groupby(['data_dia', ciclo_a_usar])['OMIE'].mean().unstack()
        agg_omie_periodos = agg_omie_periodos.reindex(agg_diario_base.index)
        for p in periodos_ciclo:
            if p in agg_omie_periodos.columns:
                dados_omie_p_diario_final = agg_omie_periodos[p].round(2).where(pd.notna(agg_omie_periodos[p]), None).tolist()
                series_diario.append({
                    "name": f"Média OMIE {nomes_periodos.get(p, p)} (€/MWh)", "type": "line",
                    "data": dados_omie_p_diario_final,
                    "yAxis": 1, "color": cores_omie.get(p), "visible": False
                })

    dados_grafico_diario = {
        'titulo': f'Consumo Diário vs. Preço Médio OMIE ({titulo_ciclo})',
        'titulo_eixo_y1': 'Consumo (kWh)',
        'titulo_eixo_y2': 'Média diária OMIE (€/MWh)',
        'categorias': agg_diario_base.index.strftime('%d/%m/%Y').tolist(),
        'series': series_diario
    }
    
    return dados_grafico_horario, dados_grafico_diario

### Função para preparar os dados para os gráficos do Modo Manual ###
def preparar_dados_grafico_manual(df_omie, data_inicio_periodo, data_fim_periodo, data_split_spot_futuros, opcao_horaria_selecionada):
    """
    Prepara os dados para um único gráfico dinâmico no Modo Manual, mostrando
    múltiplas séries por período horário e distinguindo entre Spot e Futuros.
    """
    if df_omie.empty:
        return None

    # 1. Filtrar os dados OMIE para o período selecionado
    df_periodo = df_omie[
        (df_omie['DataHora'] >= pd.to_datetime(data_inicio_periodo)) &
        (df_omie['DataHora'] <= pd.to_datetime(data_fim_periodo) + pd.Timedelta(hours=23, minutes=59))
    ].copy()

    if df_periodo.empty:
        st.warning("Não foram encontrados dados de mercado para o período selecionado.")
        return None

    df_periodo['data_dia'] = (df_periodo['DataHora']).dt.date
    
    # 2. Determinar os períodos e ciclo relevantes
    periodos_a_processar = []
    ciclo_a_usar = None
    cores_periodos = {}
    oh_lower = opcao_horaria_selecionada.lower()

    if oh_lower == "simples":
        periodos_a_processar = ['S']
        cores_periodos = {'S': ('#FF0000', '#ED7D31')} # Cor Spot, Cor Futuros
    elif oh_lower.startswith("bi"):
        ciclo_a_usar = 'BD' if "diário" in oh_lower else 'BS'
        periodos_a_processar = ['V', 'F']
        cores_periodos = {'V': ('#00B050', '#7CFC00'), 'F': ('#FFC000', '#FFD700')}
    elif oh_lower.startswith("tri"):
        ciclo_a_usar = 'TD' if "diário" in oh_lower else 'TS'
        periodos_a_processar = ['V', 'C', 'P']
        cores_periodos = {'V': ('#00B050', '#7CFC00'), 'C': ('#C00000', '#FF6347'), 'P': ('#7030A0', '#9370DB')}

    if ciclo_a_usar and ciclo_a_usar not in df_periodo.columns:
        st.warning(f"Não foi possível gerar gráficos detalhados. A coluna de ciclo '{ciclo_a_usar}' não foi encontrada nos dados.")
        return None
        
    # 3. Preparar a base do gráfico
    todas_as_datas = pd.to_datetime(sorted(df_periodo['data_dia'].unique()))
    categorias_eixo_x = todas_as_datas.strftime('%d/%m').tolist()
    
    series_grafico = []

    # 4. Construir as séries para cada período
    for periodo in periodos_a_processar:
        if periodo == 'S':
            df_agg_periodo = df_periodo.groupby('data_dia')['OMIE'].mean()
        else:
            df_agg_periodo = df_periodo[df_periodo[ciclo_a_usar] == periodo].groupby('data_dia')['OMIE'].mean()
        
        # Alinhar com todas as datas para lidar com dias em falta
        df_agg_periodo = df_agg_periodo.reindex(todas_as_datas.date)
        
        dados_spot, dados_futuros = [], []
        for data, valor in df_agg_periodo.items():
            valor_arredondado = round(valor, 2) if pd.notna(valor) else None
            if data <= data_split_spot_futuros:
                dados_spot.append(valor_arredondado)
                dados_futuros.append(None)
            else:
                dados_spot.append(None)
                dados_futuros.append(valor_arredondado)
        
        nome_periodo_legenda = {'S':'Simples', 'V':'Vazio', 'F':'Fora Vazio', 'C':'Cheias', 'P':'Ponta'}.get(periodo, periodo)
        cor_spot, cor_futuros = cores_periodos[periodo]

        if any(v is not None for v in dados_spot):
            series_grafico.append({"name": f"{nome_periodo_legenda} Spot", "data": dados_spot, "color": cor_spot})
        if any(v is not None for v in dados_futuros):
            series_grafico.append({"name": f"{nome_periodo_legenda} Futuros", "data": dados_futuros, "color": cor_futuros, "dashStyle": "shortdot"})

    if not series_grafico:
        return None

    return [{
        'id': 'grafico_evolucao_omie',
        'titulo': f'Evolução Diária OMIE/OMIP - {opcao_horaria_selecionada}',
        'categorias': categorias_eixo_x,
        'series': series_grafico
    }]


def preparar_dados_dia_semana(df_merged):
    """
    Prepara os dados agregados por dia da semana.
    - O gráfico mostra o TOTAL de consumo (empilhado) e a MÉDIA de OMIE.
    - O tooltip do consumo mostra o TOTAL e a MÉDIA diária.
    """
    if df_merged.empty:
        return None

    df_semana = df_merged.copy()
    df_semana['dia_da_semana'] = df_semana['DataHora'].dt.dayofweek
    
    day_counts = df_semana.groupby(df_semana['DataHora'].dt.date)['dia_da_semana'].first().value_counts().reindex(range(7), fill_value=0)
    
    series_grafico = []
    
    oh_lower = st.session_state.get('sel_opcao_horaria', 'simples').lower()
    
    titulo_ciclo = "Simples"
    ciclo_a_usar = None
    periodos_ciclo = []
    
    if oh_lower.startswith("bi"):
        ciclo_a_usar = 'BD' if "diário" in oh_lower else 'BS'
        periodos_ciclo = ['V', 'F']
        titulo_ciclo = "Bi-Horário - Ciclo Diário" if "diário" in oh_lower else "Bi-Horário - Ciclo Semanal"
    elif oh_lower.startswith("tri"):
        ciclo_a_usar = 'TD' if "diário" in oh_lower else 'TS'
        periodos_ciclo = ['V', 'C', 'P']
        titulo_ciclo = "Tri-Horário - Ciclo Diário" if "diário" in oh_lower else "Tri-Horário - Ciclo Semanal"
    
    nomes_periodos = {'V': 'Vazio', 'F': 'Fora Vazio', 'C': 'Cheias', 'P': 'Ponta'}
    cores_consumo_diario = {'V_bi': '#A9D18E', 'V_tri': '#BF9000', 'F': '#E2EFDA', 'C': '#FFE699', 'P': '#FFF2CC'}
    cores_consumo_semanal = {'V_bi': '#8FAADC', 'V_tri': '#C55A11', 'F': '#DAE3F3', 'C': '#F4B183', 'P': '#FBE5D6'}
    cores_consumo_a_usar = cores_consumo_diario if "diário" in oh_lower else cores_consumo_semanal
    cores_omie = {'S': '#FF0000', 'V': '#000000', 'F': '#FFC000', 'C': '#2F5597', 'P': '#00B050'}

    if ciclo_a_usar and ciclo_a_usar in df_semana.columns:
        consumo_total_periodo = df_semana.groupby(['dia_da_semana', ciclo_a_usar])['Consumo (kWh)'].sum().unstack(fill_value=0)
        
        for p in reversed(periodos_ciclo):
            if p in consumo_total_periodo.columns:
                media_periodo = (consumo_total_periodo[p] / day_counts).fillna(0)
                data_points = [{'y': consumo_total_periodo[p].get(i, 0), 'media': media_periodo.get(i, 0)} for i in range(7)]

                cor_key = 'V_tri' if p == 'V' and oh_lower.startswith("tri") else ('V_bi' if p == 'V' else p)
                series_grafico.append({
                    "name": f"Consumo {nomes_periodos.get(p, p)} (kWh)", "type": "column",
                    "data": data_points,
                    "yAxis": 0, "color": cores_consumo_a_usar.get(cor_key)
                })
    else:
        agg_total_consumo = df_semana.groupby('dia_da_semana')['Consumo (kWh)'].sum()
        agg_media_consumo = (agg_total_consumo / day_counts).fillna(0)
        data_points = [{'y': agg_total_consumo.get(i, 0), 'media': agg_media_consumo.get(i, 0)} for i in range(7)]
        series_grafico.append({"name": "Consumo Total (kWh)", "type": "column", "data": data_points, "yAxis": 0, "color": "#BFBFBF"})
    
    # Adicionar as linhas OMIE (visível e ocultas)
    agg_media_omie_simples = df_semana.groupby('dia_da_semana')['OMIE'].mean().reindex(range(7))
    dados_omie_simples_final = agg_media_omie_simples.round(2).where(pd.notna(agg_media_omie_simples), None).tolist() # Converte NaN para None
    series_grafico.append({
        "name": "Média OMIE (€/MWh)", "type": "line", 
        "data": dados_omie_simples_final, "yAxis": 1, "color": cores_omie.get('S')
    })
    
    if ciclo_a_usar and ciclo_a_usar in df_semana.columns:
        agg_omie_semana_periodos = df_semana.groupby(['dia_da_semana', ciclo_a_usar])['OMIE'].mean().unstack()
        for p in periodos_ciclo:
            if p in agg_omie_semana_periodos.columns:
                dados_omie_p = agg_omie_semana_periodos[p].reindex(range(7))
                dados_omie_p_final = dados_omie_p.round(2).where(pd.notna(dados_omie_p), None).tolist() # Converte NaN para None
                series_grafico.append({
                    "name": f"Média OMIE {nomes_periodos.get(p, p)} (€/MWh)", "type": "line",
                    "data": dados_omie_p_final, "yAxis": 1, 
                    "color": cores_omie.get(p), "visible": False
                })

    return {
        'titulo': f'Consumo e Preço Médio OMIE por Dia da Semana ({titulo_ciclo})',
        'titulo_eixo_y1': 'Consumo Total (kWh)',
        'titulo_eixo_y2': 'Média OMIE (€/MWh)',
        'categorias': ['Segunda', 'Terça', 'Quarta', 'Quinta', 'Sexta', 'Sábado', 'Domingo'],
        'series': series_grafico
    }

def extrair_nome_base_tarifario(nome_completo):
    """
    Extrai o nome base de um tarifário, removendo sufixos e textos em parênteses.
    Ex: "EDP - Tarifa X (Desconto Y) - Diagrama" -> "EDP - Tarifa X"
    """
    if not isinstance(nome_completo, str):
        return ""
    
    # Remove os sufixos de modo e perfil
    nome_sem_sufixo = nome_completo.replace(" - Diagrama", "").replace(" - Perfil", "")
    
    # Remove qualquer texto a partir do primeiro parêntese
    nome_base = re.split(r'\s*\(', nome_sem_sufixo)[0]
    
    return nome_base.strip()

# --- Inicializar lista de resultados ---
resultados_list = []

# --- Título e Botão de Limpeza Geral ---

# Linha 1: Logo e Título
col_logo, col_titulo = st.columns([1, 5])

with col_logo:
    st.image("https://huggingface.co/spaces/tiagofelicia/simulador-tarifarios-eletricidade/resolve/main/Logo_Tiago_Felicia.png", width=180)

with col_titulo:
    st.title("🔌 Tiago Felícia - Simulador de Tarifários de Eletricidade 2026")

st.button(
    "🧹 Limpar e Reiniciar Simulador",
    on_click=reiniciar_simulador,
    help="Repõe todos os campos do simulador para os valores iniciais.",
    use_container_width=True
)

# ##################################################################
# INÍCIO DO BLOCO - GUIA RÁPIDO
# ##################################################################

with st.expander("❓ Como Usar o Simulador de Tarifários de Eletricidade (Guia Rápido)", expanded=False):
    st.markdown("""
    Bem-vindo! Esta ferramenta ajuda-o a descobrir o tarifário de eletricidade mais económico para si. Siga os passos abaixo para começar a poupar.

    #### **Passo 1: Escolha o seu Ponto de Partida**
    Primeiro, defina a sua **Potência Contratada** e **Opção Horária**. Depois, escolha como quer fornecer os seus dados de consumo.

    * **Opção A: 📊 Análise Detalhada (com Ficheiro E-Redes) - *Recomendado***
        1.  **Carregue o Ficheiro:** Na secção "📂 Carregar Diagrama de Carga", envie o seu ficheiro `.xlsx` da E-Redes. Pode obtê-lo em [balcaodigital.e-redes.pt](https://balcaodigital.e-redes.pt/consumptions/history).
        2.  **Selecione o Período:** Escolha as datas que pretende analisar.
        3.  **Vantagem:** Esta é a forma mais precisa. O simulador usa os seus consumos a cada 15 minutos para uma análise exata, incluindo **gráficos detalhados** e uma análise da sua potência contratada.

    * **Opção B: ✍️ Estimativa Rápida (Modo Manual)**
        1.  **Defina o Período:** Escolha o **mês** ou as **datas** para a simulação.
        2.  **Insira os Consumos:** Preencha os seus consumos mensais estimados (kWh) nos campos que aparecem (ex: Vazio, Fora de Vazio).
        3.  **Vantagem:** Ideal para simulações rápidas e previsões de custos sem precisar do ficheiro da E-Redes.

    #### **Passo 2: ⚙️ Configure a Simulação**
    Depois de inserir os seus consumos, pode refinar a simulação.

    * **☀️ Simular Autoconsumo (Opcional - Modo Diagrama):** Se ativou a **Opção A**, pode abrir o *expander* de autoconsumo para simular o impacto de painéis solares fotovoltaicos e ver como o seu consumo da rede diminui.
        1.  **Configure o Sistema:** Defina a potência dos painéis (kWp), a localização (distrito), inclinação e orientação.
        2.  Para uma simulação de painéis fotovoltaicos mais detalhada e com baterias https://www.tiagofelicia.pt/autoconsumo-tiagofelicia.html
        3.  **Selecione o Perfil:** Após a simulação, na secção **"⚙️ Selecione os consumos a usar..."**, escolha se os cálculos finais devem usar o seu consumo original ou o novo consumo já com o abate da produção solar.

    * **➕ Opções Adicionais:** No *expander* de "Opções Adicionais", pode ativar benefícios como a **Tarifa Social** ou incluir descontos específicos (ACP, Continente).

    #### **Passo 3: 🏆 Encontre a Melhor Tarifa**
    A tabela de resultados no final da página é a sua ferramenta principal.

    * **Ordenar por Custo:** Clique no cabeçalho da coluna **"Total (€)"** para ordenar os tarifários do mais barato para o mais caro.
    * **Explorar Detalhes:** Passe o rato sobre os preços ou sobre o custo total para ver um **resumo detalhado** dos cálculos, incluindo todas as taxas e impostos.
    * **Filtrar Resultados:** Use os filtros no topo da tabela para refinar a sua pesquisa por tipo de tarifário (Fixo, Indexado), segmento, etc.
    * **O Seu Pódio:** No final, a secção **"🏆 O Seu Pódio da Poupança"** destaca as 3 opções mais económicas para si.
    * **Comparar Opções Horárias:** Ative a opção "🔬 **Comparar custos entre diferentes Opções Horárias**" para ver uma tabela especial que mostra quanto pagaria em cada tarifário se tivesse um ciclo diferente (Simples, Bi-Horário, etc.).

    > **Dica Pro:** Use a secção **"🧾 O Meu Tarifário"** para introduzir os preços da sua fatura atual e compará-la diretamente com todas as ofertas do mercado. Assim, saberá exatamente quanto pode poupar!
    """)

# ##################################################################
# FIM DO BLOCO - GUIA RÁPIDO
# ##################################################################

# ##################################################################
# INÍCIO DO BLOCO - FAQ (Perguntas Frequentes)
# ##################################################################

with st.expander("❔ Perguntas Frequentes (FAQ)", expanded=False):
    st.markdown("""
    ### Perguntas Gerais

    **P: De onde vêm os dados dos tarifários e dos preços de mercado (OMIE)?**
    
    **R:** Todos os dados dos tarifários são recolhidos a partir das informações públicas disponibilizadas pelos comercializadores nos seus websites. Os preços do mercado ibérico (OMIE) e os perfis de consumo da ERSE são obtidos de fontes oficiais para garantir a máxima precisão nos cálculos dos tarifários indexados. Os dados são atualizados regularmente para refletir as condições atuais do mercado.

    **P: Alguns tarifários têm custo na energia diferentes do que têm no seu site institucional, porquê?**
    
    **R:** Alguns comercializadores não incluem o valor do financiamento da Tarifa Social de Eletricidade (TSE) no custo base da energia. Para que se possa comparar entre todos de forma igual, nesses casos junto o custo de financiamento da TSE ao custo base. Para confirmar, passe o rato por cima do valor da energia na tabela para ver a decomposição detalhada do preço.

    **P: Tenho um tarifário com o mesmo nome que o do simulador, mas tem valores diferentes na energia e/ou potência, porquê?**
    
    **R:** Muitos comercializadores alteram os valores nos seus tarifários mas mantêm as denominações dos mesmos. Este simulador só tem os valores para a última versão do tarifário.

    **P: O simulador é 100% preciso?**
    
    **R:** O objetivo é ser o mais preciso possível. Para tarifários **fixos**, a precisão é muito elevada. Para tarifários **indexados**, o custo final é uma estimativa baseada em médias e perfis de consumo. A forma mais rigorosa de simulação é sempre através do carregamento do seu **diagrama de carga da E-Redes**, pois utiliza o seu perfil de consumo real. (neste caso, apenas dados históricos)
    
    **P: Porque é que o modo "Análise Detalhada" com o ficheiro da E-Redes é recomendado?**
    
    **R:** Este modo utiliza o seu consumo real registado a cada 15 minutos. Isto permite um cálculo exato do custo para qualquer tipo de tarifário, incluindo os **indexados quarto-horários (dinâmicos)**. Além disso, só neste modo é possível fazer uma análise precisa da sua **potência contratada** e simular o impacto do **autoconsumo** com painéis solares fotovoltaicos.
    
    ---
    
    ### Dados e Períodos
    **P: Porque é que a simulação só está disponível a partir de 01/01/2025?**
    
    **R:** A simulação está focada no ano de 2026 para garantir que os cálculos utilizam as Tarifas de Acesso às Redes (TAR) e outras taxas e impostos que estão em vigor. Utilizar dados de consumo de anos anteriores com as tarifas atuais poderia levar a resultados imprecisos, uma vez que as condições do mercado e a regulação mudam anualmente. O simulador ignora automaticamente quaisquer dados anteriores a esta data para garantir a relevância dos resultados.
                
    **P:  Qual a diferença entre escolher um Mês e um Período de Datas manual?**
    
    **R:** 
    * **Escolher um Mês**: É a forma mais simples e rápida. O simulador seleciona automaticamente o primeiro e o último dia desse mês e usa o número de dias correto (ex: 30 para Abril, 31 para Maio).

    * **Selecionar Datas**: Oferece total flexibilidade para analisar um período específico (ex: uma semana, uma quinzena, ou um período que abranja partes de dois meses). O número de dias é calculado com base no intervalo que definir.
                
    **P: Como são tratados os fins de semana e feriados nos ciclos Bi e Tri-Horário?**
    
    **R:** O simulador utiliza os ciclos horários oficiais definidos pela ERSE. Nos **Ciclos Semanais** (ex: Bi-Horário Semanal), os fins de semana têm um tratamento específico, geralmente com mais horas a contar como "Vazio". Nos **Ciclos Diários**, a contagem das horas é a mesma para todos os dias da semana. Os feriados não têm regras especificas em BTN, sendo tratados como dias "normais". O simulador aplica estas regras automaticamente com base nos dados oficiais.           
                
    ---
    
    ### Funcionalidades do Simulador

    **P: O que significa a opção "Comparar custos entre diferentes Opções Horárias"?**
    
    **R:** Ao ativar esta opção, o simulador recalcula o custo de cada tarifário para diferentes ciclos horários (Simples, Bi-Horário, etc.), assumindo os mesmos consumos totais que inseriu. Isto é útil para perceber se, com o seu padrão de consumo atual, compensaria mudar de opção horária.

    **P: Como funciona a secção "O Meu Tarifário"?**
    
    **R:** Esta secção permite-lhe introduzir os preços da sua fatura atual (energia em €/kWh e potência em €/dia) para comparar diretamente com todas as outras ofertas do mercado. É fundamental verificar na sua fatura se os preços que insere já incluem as **TAR (Tarifas de Acesso às Redes)** para que a comparação seja correta.

    **P: Para que serve a secção "Comparar outro Tarifário Personalizado?"**
    
    **R:** Esta funcionalidade é uma ferramenta poderosa para criar cenários hipotéticos. Permite-lhe construir até três estruturas tarifárias (Simples, Bi-Horária e Tri-Horária) com os seus próprios preços. É ideal para:

    * Simular uma oferta que recebeu e que ainda não está na lista.

    * Perceber qual seria o custo se o seu comercializador atual oferecesse uma opção horária diferente.

    * Testar o impacto de futuras subidas ou descidas de preços nos seus custos.
                
    **P: Como funciona a simulação de Autoconsumo com painéis solares fotovoltaicos?**
    
    **R:** Ao ativar esta opção (disponível no modo de diagrama de carga), o simulador estima a produção de energia de um sistema fotovoltaico com base na potência, localização e orientação que definir. Essa produção é depois subtraída do seu consumo a cada 15 minutos. O resultado é um novo perfil de "consumo da rede", que mostra quanta energia realmente precisou de comprar. Os cálculos dos tarifários podem então ser feitos com base neste novo consumo, mostrando a poupança real que o autoconsumo pode gerar.

    **P: A simulação de autoconsumo tem em conta os dias de chuva ou de sol??**
    
    **R:** A simulação utiliza **perfis de produção solar médios mensais para cada distrito**, baseados em dados históricos do PVGIS (Photovoltaic Geographical Information System). Isto significa que a produção estimada para um dia de Julho, por exemplo, representa um "dia médio" de Julho, já tendo em conta a média de horas de sol e de nebulosidade para esse mês nessa região. Não simula a produção para um dia específico com as suas condições meteorológicas reais, mas oferece uma estimativa muito realista para um período de análise mais longo.

    **P: Os valores OMIE para datas futuras são reais?**
    
    **R:** Não. Os valores OMIE apresentados para datas futuras baseiam-se nos preços do mercado de futuros (OMIP). Estes valores representam a expectativa do mercado para os preços da eletricidade, mas não são uma garantia. Servem como a melhor estimativa disponível para simular custos em tarifários indexados para períodos que ainda não ocorreram. A data de atualização destes valores está indicada no final da página.

    **P: Como são tratados descontos especiais como os do ACP ou Continente?**
    
    **R:** O simulador tenta replicar as condições comerciais o mais fielmente possível. Para a parceria **Goldenergy/ACP**, pode optar por incluir o valor da quota mensal no custo final. Para os tarifários **Galp/Continente**, o simulador calcula o valor do desconto em Cartão Continente e subtrai-o ao custo total, refletindo a poupança real na sua carteira. Pode ativar ou desativar estas opções na secção "Opções Adicionais de Simulação".                

    **P: O que são as colunas "Custo com o seu Perfil Real (€)" e "Custo com Perfil Padrão ERSE (€)"?**
    
    **R:** Esta análise, disponível no modo de diagrama de carga, compara duas coisas: o custo exato do seu consumo (Perfil Real) e o custo que teria se o seu consumo seguisse um perfil médio definido pela ERSE (Perfil Padrão). Uma diferença negativa (a verde) significa que o seu padrão de consumo pessoal é mais económico que a média, o que é ótimo!

    **P: O que está incluído no "Total (€)"? É o valor final da fatura?**
    
    **R:** Sim, o valor "Total (€)" representa a sua fatura final estimada. Ele inclui a soma de várias componentes:

    1. **Custo da Energia**: O consumo (kWh) multiplicado pelo preço da energia (€/kWh) de cada período.

    2. **Custo da Potência**: O preço da potência (€/dia) multiplicado pelo número de dias do período.

    3. **Taxas e Impostos**: A Contribuição Audiovisual (CAV), a Taxa DGEG e o Imposto Especial de Consumo (IEC).

    4. **IVA**: O IVA é aplicado a 6% ou 23% sobre cada componente, de acordo com as regras em vigor (ex: IVA reduzido na potência até 3.45 kVA ou nos primeiros kWh de consumo).

    5. **Descontos/Acréscimos**: Quaisquer descontos de fatura ou acréscimos (como a quota ACP) são aplicados ao valor final.
    Pode ver a decomposição detalhada de todos estes custos passando o rato por cima do valor na coluna "Total (€)".

    **P: A tabela de resultados tem muitas opções. Como posso encontrar rapidamente o que procuro?**
    
    **R:** A tabela é interativa! Pode:

    * **Ordenar**: Clique no cabeçalho de qualquer coluna (como "Total (€)" ou "Potência (€/dia)") para ordenar os resultados do mais baixo para o mais alto, ou vice-versa.

    * **Filtrar**: Use os filtros no topo da tabela para refinar a sua pesquisa por tipo de tarifário (Fixo, Indexado), segmento (Residencial, Empresarial), ou por opções de faturação e pagamento.

    * **Pesquisar**: Na vista detalhada, pode usar a pesquisa dentro das colunas "Comercializador" e "Tarifário" para encontrar uma oferta específica.
                
    ### Termos e Conceitos

    **P: O que é um tarifário indexado? É uma boa opção para mim?**
    
    **R:** Um tarifário indexado tem um preço de energia que varia de acordo com o preço do mercado grossista (OMIE). Este tipo de tarifário pode oferecer uma poupança significativa quando os preços de mercado estão baixos, mas também implica um risco maior se os preços subirem. Pode ser:

       * **Indexado à Média**: O preço baseia-se na média do período de faturação (ex. mensal) dos preços OMIE para os períodos horários (Vazio, Fora de Vazio, etc.).

       * **Indexado Quarto-Horário (ou Dinâmico)**: O preço varia a cada hora (ou 15 minutos), seguindo o preço real do mercado.
    São ideais para quem consegue adaptar o seu consumo às horas mais baratas do dia.

    **P: O que são as TAR (Tarifas de Acesso às Redes)**
    
    **R:** As TAR são tarifas reguladas pela entidade reguladora (ERSE) que pagam pelo uso das infraestruturas elétricas (transporte e distribuição). Todos os consumidores as pagam, independentemente do comercializador. Alguns tarifários apresentam o preço final já com as TAR incluídas, enquanto outros as mostram em separado na fatura. O simulador lida com ambas as situações para garantir uma comparação justa.

    **P: O que significa "Perfil de Consumo" (ex: Perfil A, B, C)?**
    
    **R:** A ERSE agrupa os consumidores em perfis (A, B ou C) com base no seu consumo anual e potência. Estes perfis representam um padrão de consumo médio para cada tipo de cliente. No **Modo Manual**, o simulador usa estes perfis para estimar como o seu consumo se distribui ao longo do dia, o que é crucial para calcular o custo dos tarifários indexados quarto-horários. No **Modo Diagrama**, este perfil não é necessário, pois o simulador usa os seus dados de consumo reais, no entanto para comparação, são apresentados ambos os valores.                

    **P: Porque é que o simulador pergunta se a minha instalação é trifásica?**
    
    **R:** O ficheiro da E-Redes regista a potência total da sua casa em médias de 15 minutos. Num contador monofásico, a potência máxima registada corresponde ao pico da média de consumo da instalação. Num contador trifásico, o consumo total pode distribuir-se pelas três fases. O valor que a E-Redes regista no seu ficheiro é da soma das três fases. É possível que uma única fase tenha um pico de consumo muito elevado num curto espaço de tempo, fazendo o disjuntor disparar, mesmo que a potência total média nesses 15 minutos não ultrapasse o valor contratado.
                
    **P: Como sei se a minha instalação é monofásica ou trifásica?**
    
    **R:** Geralmente, instalações domésticas com potências contratadas até 6.9 kVA são monofásicas. Potências superiores são quase sempre trifásicas. Geralmente esta informação está disponível no seu contador ou no balcão digital da E-Redes. A análise de potência máxima no simulador tem em conta esta diferença.


    """)
# ##################################################################
# FIM DO BLOCO - FAQ (Perguntas Frequentes)
# ##################################################################

# --- 1. INPUTS E DEFINIÇÕES GLOBAIS ---
if "Simples" not in opcoes_horarias_existentes:
    opcoes_horarias = ["Simples"] + sorted(opcoes_horarias_existentes)
else:
    opcoes_horarias = sorted(opcoes_horarias_existentes)

col1, col2 = st.columns(2)

# --- POTÊNCIA ---
with col1:
    potencia = st.selectbox(
        "Potência Contratada (kVA)", 
        potencias_validas, 
        key="sel_potencia",
        help="Potências BTN (1.15 kVA a 41.4 kVA)",
        on_change=atualizar_url_potencia
    )

# --- LÓGICA DE FILTRAGEM E VALIDAÇÃO DA OPÇÃO HORÁRIA ---
if potencia >= 27.6:
    opcoes_validas = [o for o in opcoes_horarias if "Tri-horário > 20.7 kVA" in o]
else:
    opcoes_validas = [o for o in opcoes_horarias if not "Tri-horário > 20.7 kVA" in o]

# --- OPÇÃO HORÁRIA E LÓGICA DE RESET ---
with col2:
    opcao_atual_no_estado = st.session_state.get("sel_opcao_horaria")

    if opcao_atual_no_estado in opcoes_validas:
        default_index_oh = opcoes_validas.index(opcao_atual_no_estado)
    else:
        default_index_oh = 0
        if opcoes_validas:
            st.session_state.sel_opcao_horaria = opcoes_validas[0]
        else:
            st.session_state.sel_opcao_horaria = None

    opcao_horaria = st.selectbox(
        "Opção Horária", 
        opcoes_validas,
        key="sel_opcao_horaria", 
        help="Opção horária contratada",
        on_change=atualizar_url_opcao_horaria
    )



# --- CAMPO DE PARTILHA POR URL ---

st.subheader("⚡ Período e Consumos")
# --- 2. LÓGICA DE UPLOAD E DETERMINAÇÃO DO MODO DE SIMULAÇÃO ---
with st.expander("📂 Carregar Diagrama de Carga da E-Redes (opcional)"):
    # Mensagem informativa para o utilizador
    st.warning("⚠️ O carregamento de ficheiros da E-Redes encontra-se temporariamente indisponível para manutenção.")
    
    # Forçar a variável a None para que o resto do código assuma que não há ficheiro
    uploaded_files = None
    
    # Código original comentado para não ser executado:
    # uploaded_files = st.file_uploader(
    # "Selecione um ou mais ficheiros da E-Redes (apenas do ano atual - outros datas serão ignoradas)", 
    # type=['xlsx'], 
    # key="consumos_uploader",
    # accept_multiple_files=True
    # )
# Lógica para determinar o modo e processar o ficheiro
# Se foram carregados novos ficheiros
if uploaded_files:
    # Criamos uma chave única baseada nos nomes e tamanhos dos ficheiros para saber se mudaram
    chave_ficheiros_atuais = "".join([f.name + str(f.size) for f in uploaded_files])

    if st.session_state.get('chave_ficheiros_processados') != chave_ficheiros_atuais:
        with st.spinner("A processar e validar ficheiros..."):
            
            # Renomear a variável 'erro' para 'mensagem' para maior clareza
            df_combinado, mensagem = proc_dados.validar_e_juntar_ficheiros(uploaded_files)

            # A verificação principal passa a ser sobre o DataFrame
            if df_combinado is None:
                # Se o DataFrame é None, a mensagem é um erro fatal
                st.error(mensagem)
                st.session_state.dados_completos_ficheiro = None
            else:
                # Se o DataFrame existe, o processo foi um sucesso
                st.success("Ficheiros validados e carregados com sucesso!")
                
                # Verificar se há uma mensagem de AVISO para mostrar
                if mensagem:
                    st.warning(mensagem) # Mostra o aviso de que dados antigos foram ignorados

                # Lógica de sucesso
                st.session_state.dados_completos_ficheiro = df_combinado
                st.session_state.chave_ficheiros_processados = chave_ficheiros_atuais
                st.session_state.nomes_ficheiros_processados = ", ".join([f.name for f in uploaded_files])
                
                # Limpar os parâmetros do URL ao entrar em modo diagrama
                st.query_params.clear()

# Se não há ficheiros, mas havia antes, limpar o estado
elif not uploaded_files and 'dados_completos_ficheiro' in st.session_state:
     del st.session_state.dados_completos_ficheiro
     if 'chave_ficheiros_processados' in st.session_state:
         del st.session_state.chave_ficheiros_processados
     if 'nomes_ficheiros_processados' in st.session_state:
         del st.session_state.nomes_ficheiros_processados

# --- 3. DEFINIÇÃO DE DATAS  ---
is_diagram_mode = 'dados_completos_ficheiro' in st.session_state and st.session_state.get('dados_completos_ficheiro') is not None
dias_mes = {"Janeiro":31,"Fevereiro":29,"Março":31,"Abril":30,"Maio":31,"Junho":30,"Julho":31,"Agosto":31,"Setembro":30,"Outubro":31,"Novembro":30,"Dezembro":31}
ano_atual = datetime.datetime.now().year
if ((ano_atual % 4 == 0 and ano_atual % 100 != 0) or (ano_atual % 400 == 0)):
    dias_mes["Fevereiro"] = 29

if is_diagram_mode:
    # ######################
    # --- MODO DIAGRAMA ---
    # ######################
    st.success(f"Modo Diagrama ativo, a utilizar dados de: {st.session_state.get('nomes_ficheiros_processados', 'ficheiro(s) carregado(s)')}")
    
    # --- PASSO 1: INPUTS E FILTRAGEM INICIAL ---
    df_consumos_total = st.session_state.dados_completos_ficheiro
    min_date_ficheiro = df_consumos_total['DataHora'].min().date()
    max_date_ficheiro = df_consumos_total['DataHora'].max().date()
    
    col_f1, col_f2, col_f3 = st.columns(3)
    with col_f1:
        data_inicio = st.date_input("Filtrar Data Inicial", value=min_date_ficheiro, min_value=min_date_ficheiro, max_value=max_date_ficheiro, format="DD/MM/YYYY", key="data_inicio_ficheiro")
    with col_f2:
        data_fim = st.date_input("Filtrar Data Final", value=max_date_ficheiro, min_value=min_date_ficheiro, max_value=max_date_ficheiro, format="DD/MM/YYYY", key="data_fim_ficheiro")

    dias = (data_fim - data_inicio).days + 1 if data_fim >= data_inicio else 0
    dias_default_calculado = dias
    with col_f3:
        gfx.exibir_metrica_personalizada("Nº de Dias", f"{dias} dias")

    mes_num = data_inicio.month
    ano_atual = data_inicio.year
    meses_lista = list(dias_mes.keys())
    mes = meses_lista[mes_num - 1]
    
    df_consumos_bruto_filtrado = df_consumos_total[(df_consumos_total['DataHora'].dt.date >= data_inicio) & (df_consumos_total['DataHora'].dt.date <= data_fim)].copy()
    df_omie_filtrado_para_analise = OMIE_PERDAS_CICLOS[(OMIE_PERDAS_CICLOS['DataHora'] >= pd.to_datetime(data_inicio)) & (OMIE_PERDAS_CICLOS['DataHora'] <= pd.to_datetime(data_fim) + pd.Timedelta(hours=23, minutes=59))].copy()

    # --- PASSO 2: ANÁLISE DO CONSUMO BRUTO (ANTES DO AUTOCONSUMO) ---
    st.markdown("##### Análise de Consumos e Médias OMIE (do(s) ficheiro(s))")
    consumos_agregados_brutos = proc_dados.agregar_consumos_por_periodo(df_consumos_bruto_filtrado, OMIE_PERDAS_CICLOS)
    omie_medios_para_tabela_bruta = proc_dados.calcular_medias_omie_para_todos_ciclos(df_consumos_bruto_filtrado, OMIE_PERDAS_CICLOS)
    tabela_analise_html_bruta = criar_tabela_analise_completa_html(consumos_agregados_brutos, omie_medios_para_tabela_bruta)
    st.markdown(tabela_analise_html_bruta, unsafe_allow_html=True)

    with st.expander("Ver Gráficos de Análise (Consumo vs. OMIE)"):
        df_merged_bruto = pd.merge(df_consumos_bruto_filtrado, df_omie_filtrado_para_analise, on='DataHora', how='inner')
        dados_horario_bruto, dados_diario_bruto = preparar_dados_para_graficos(df_consumos_bruto_filtrado, df_omie_filtrado_para_analise, opcao_horaria, dias)
        dados_semana_bruto = gfx.preparar_dados_dia_semana(df_merged_bruto, st.session_state)
        dados_mensal_bruto = gfx.preparar_dados_mensais(df_merged_bruto, st.session_state)

        if dados_horario_bruto:
            html_grafico_bruto_horario = gfx.gerar_grafico_highcharts('grafico_bruto_horario', dados_horario_bruto)
            st.components.v1.html(html_grafico_bruto_horario, height=620)
        if dados_diario_bruto:
            html_grafico_bruto_diario = gfx.gerar_grafico_highcharts('grafico_bruto_diario', dados_diario_bruto)
            st.components.v1.html(html_grafico_bruto_diario, height=620)
        if dados_semana_bruto:
            st.components.v1.html(gfx.gerar_grafico_highcharts('grafico_bruto_semana', dados_semana_bruto), height=620)
        if dados_mensal_bruto:
            html_grafico_bruto_mensal = gfx.gerar_grafico_highcharts('grafico_bruto_mensal', dados_mensal_bruto)
            st.components.v1.html(html_grafico_bruto_mensal, height=620)

    # --- PASSO 3: LÓGICA DE AUTOCONSUMO (INTERFACE E CÁLCULO) ---
    df_consumos_final_para_calculos = df_consumos_bruto_filtrado.copy()
    
    with st.expander("☀️ Simular Autoconsumo (Painéis Solares Fotovoltaicos)", expanded=False):
        ativar_autoconsumo = st.checkbox("Ativar simulação de autoconsumo - Simplificada", key="chk_autoconsumo_ativo")
        if ativar_autoconsumo:
            distritos_regioes = ['Aveiro', 'Beja', 'Braga', 'Bragança', 'Castelo Branco', 'Coimbra', 'Évora', 'Faro', 'Guarda', 'Leiria', 'Lisboa', 'Portalegre', 'Porto', 'Santarém', 'Setúbal', 'Viana do Castelo', 'Vila Real', 'Viseu', 'Açores (Ponta Delgada)', 'Madeira (Funchal)']
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                potencia_kwp_input = st.number_input("Potência (kWp)", min_value=0.01, value=2.0, step=0.1, format="%.1f", key="solar_potencia")
            with col2:
                distrito_selecionado = st.selectbox("Distrito/Região", distritos_regioes, index=8, key="solar_distrito")
            with col3:
                inclinacao_input = st.number_input("Inclinação (°)", min_value=0, max_value=90, value=35, step=1, key="solar_inclinacao")
            with col4:
                orientacao_selecionada = st.selectbox("Orientação", ["Sul (Ótima)", "Sudeste / Sudoeste", "Este / Oeste"], key="solar_orientacao")

            with st.spinner("A simular produção solar..."):
                df_com_solar = calc.simular_autoconsumo_completo(
                    df_consumos=df_consumos_bruto_filtrado,
                    potencia_kwp=st.session_state.solar_potencia,
                    distrito=st.session_state.solar_distrito,
                    inclinacao=st.session_state.solar_inclinacao,
                    orientacao_str=st.session_state.solar_orientacao
                )
                df_consumos_final_para_calculos = df_com_solar
            
                st.write("##### Resumo da Simulação Solar (para o período selecionado)")
                res_col1, res_col2, res_col3 = st.columns(3)
                with res_col1:
                    gfx.exibir_metrica_personalizada("Produção Solar", f"{df_consumos_final_para_calculos['Producao_Solar_kWh'].sum():.0f} kWh")
                with res_col2:
                    gfx.exibir_metrica_personalizada("Autoconsumo", f"{df_consumos_final_para_calculos['Autoconsumo_kWh'].sum():.0f} kWh")
                with res_col3:
                    gfx.exibir_metrica_personalizada("Excedente", f"{df_consumos_final_para_calculos['Excedente_kWh'].sum():.0f} kWh")

                # <<< GRÁFICO DE AUTOCONSUMO COM HIGHCHARTS >>>
                if not df_consumos_final_para_calculos.empty and df_consumos_final_para_calculos['Producao_Solar_kWh'].sum() > 0:
                    # --- Seletor de Data para o Gráfico ---
                    # Por defeito, sugere o dia de maior produção, mas o utilizador pode alterar.
                    dia_default_grafico = df_consumos_final_para_calculos.groupby(df_consumos_final_para_calculos['DataHora'].dt.date)['Producao_Solar_kWh'].sum().idxmax()
                    
                    dia_selecionado_para_grafico = st.date_input(
                        "Selecione um dia para visualizar no gráfico:",
                        value=dia_default_grafico,
                        min_value=data_inicio, # Usa as datas do filtro principal
                        max_value=data_fim,    # Usa as datas do filtro principal
                        format="DD/MM/YYYY",
                        key="date_input_grafico_solar"
                    )

                    df_dia_exemplo = df_consumos_final_para_calculos[df_consumos_final_para_calculos['DataHora'].dt.date == dia_selecionado_para_grafico].copy()
                    
                    # Preparar os dados para a função do gráfico
                    dados_para_grafico_solar = {
                        'titulo': 'Produção Solar vs. Consumo Horário (no dia selecionado)',
                        'categorias': df_dia_exemplo['DataHora'].dt.strftime('%H:%M').tolist(),
                        'series': [
                            {
                                "name": "Consumo (kWh)",
                                "data": df_dia_exemplo['Consumo (kWh)'].round(3).tolist(),
                                "color": "#2E75B6" # Azul
                            },
                            {
                                "name": "Produção Solar (kWh)",
                                "data": df_dia_exemplo['Producao_Solar_kWh'].round(3).tolist(),
                                "color": "#FFA500" # Laranja
                            }
                        ]
                    }
                    # 1. Gerar o HTML do gráfico usando a função do ficheiro graficos.py
                    html_grafico_solar = gfx.gerar_grafico_solar('grafico_autoconsumo_solar', dados_para_grafico_solar)
                    
                    # 2. Exibir o HTML gerado na página do Streamlit
                    st.components.v1.html(html_grafico_solar, height=420)

                    # Guardar o dataframe com os resultados do autoconsumo no estado da sessão
                    st.session_state.df_resultado_autoconsumo = df_consumos_final_para_calculos.copy()

    # --- PASSO 4: ANÁLISE DO CONSUMO LÍQUIDO (DEPOIS DO AUTOCONSUMO, SE ATIVO) ---
    if st.session_state.get("chk_autoconsumo_ativo"):
        with st.spinner("A simular produção solar..."):
            st.markdown("##### Análise de Consumos e Médias OMIE (Após Autoconsumo)")
            df_para_tabela_liquida = df_consumos_final_para_calculos.copy()
            df_para_tabela_liquida['Consumo (kWh)'] = df_para_tabela_liquida['Consumo_Rede_kWh']
            consumos_agregados_liquidos = proc_dados.agregar_consumos_por_periodo(df_para_tabela_liquida, OMIE_PERDAS_CICLOS)
            omie_medios_para_tabela_liquida = proc_dados.calcular_medias_omie_para_todos_ciclos(df_para_tabela_liquida, OMIE_PERDAS_CICLOS)
            tabela_analise_html_liquida = criar_tabela_analise_completa_html(consumos_agregados_liquidos, omie_medios_para_tabela_liquida)
            st.markdown(tabela_analise_html_liquida, unsafe_allow_html=True)

            with st.expander("Ver Gráficos de Análise (Consumo Após Autoconsumo vs. OMIE)"):
                df_merged_liquido = pd.merge(df_para_tabela_liquida, df_omie_filtrado_para_analise, on='DataHora', how='inner')
                dados_horario_liq, dados_diario_liq = preparar_dados_para_graficos(df_para_tabela_liquida, df_omie_filtrado_para_analise, opcao_horaria, dias)
                dados_semana_liq = gfx.preparar_dados_dia_semana(df_merged_liquido, st.session_state)
                dados_mensal_liq = gfx.preparar_dados_mensais(df_merged_liquido, st.session_state)

                if dados_horario_liq:
                    html_grafico_liq_horario = gfx.gerar_grafico_highcharts('grafico_liq_horario', dados_horario_liq)
                    st.components.v1.html(html_grafico_liq_horario, height=620)
                if dados_diario_liq:
                    html_grafico_liq_diario = gfx.gerar_grafico_highcharts('grafico_liq_diario', dados_diario_liq)
                    st.components.v1.html(html_grafico_liq_diario, height=620)
                if dados_semana_liq:
                    st.components.v1.html(gfx.gerar_grafico_highcharts('grafico_liq_semana', dados_semana_liq), height=620)
                if dados_mensal_liq:
                    html_grafico_liq_mensal = gfx.gerar_grafico_highcharts('grafico_liq_mensal', dados_mensal_liq)
                    st.components.v1.html(html_grafico_liq_mensal, height=620)

    # --- PASSO 5: ESCOLHA E PREPARAÇÃO DAS VARIÁVEIS FINAIS PARA CÁLCULOS ---

    # Por defeito, os consumos a utilizar são os brutos (antes do autoconsumo).
    df_consumos_a_utilizar = df_consumos_bruto_filtrado.copy()

    # Se o autoconsumo foi ativado, damos ao utilizador a escolha.
    if st.session_state.get("chk_autoconsumo_ativo", False):
        
        st.markdown("##### ⚙️ Selecione os consumos a usar nos cálculos dos tarifários:")
        
        # Adicionamos um widget de rádio para a escolha.
        escolha_consumo = st.radio(
            label="Escolha o perfil de consumo para calcular os custos dos tarifários:",
            options=[
                "Consumo Original (sem autoconsumo)",
                "Consumo da Rede (após autoconsumo)"
            ],
            index=1,  # Pré-seleciona a opção "após autoconsumo" por defeito
            key="escolha_consumo_calculo",
            horizontal=True, # Para um layout mais compacto
        )
        
        # Atualizamos o DataFrame a ser usado com base na escolha.
        if escolha_consumo == "Consumo da Rede (após autoconsumo)":
            # 'df_consumos_final_para_calculos' já tem a coluna 'Consumo_Rede_kWh'
            df_consumos_a_utilizar = df_consumos_final_para_calculos.copy()
            # É CRUCIAL renomear a coluna de consumo líquido para o nome padrão 'Consumo (kWh)'
            # que o resto do programa espera.
            df_consumos_a_utilizar['Consumo (kWh)'] = df_consumos_a_utilizar['Consumo_Rede_kWh']
            st.success("Os cálculos na tabela de resultados irão usar os **consumos líquidos (após autoconsumo)**.")
        else:
            st.info("Os cálculos na tabela de resultados irão usar os **consumos originais (sem autoconsumo)**.")

    # AGORA, e só agora, calculamos as variáveis finais de consumo usando o DataFrame escolhido.
    # Este bloco garante que os valores corretos são usados, quer o autoconsumo esteja ativo ou não.
    consumos_para_custos = proc_dados.agregar_consumos_por_periodo(df_consumos_a_utilizar, OMIE_PERDAS_CICLOS)

    consumo_simples = consumos_para_custos.get('Simples', 0)
    consumo_vazio, consumo_fora_vazio, consumo_cheias, consumo_ponta = 0, 0, 0, 0

    # Determina o ciclo Bi-horário com base na opção horária principal
    ciclo_bi = 'BD' if "Diário" in opcao_horaria else 'BS'
    if ciclo_bi in consumos_para_custos:
        consumo_vazio = consumos_para_custos.get(ciclo_bi, {}).get('V', 0)
        consumo_fora_vazio = consumos_para_custos.get(ciclo_bi, {}).get('F', 0)

    # Determina o ciclo Tri-horário com base na opção horária principal
    # Nota: Sobrescreve 'consumo_vazio' se ambos os ciclos existirem, o que é o comportamento esperado.
    ciclo_tri = 'TD' if "Diário" in opcao_horaria else 'TS'
    if ciclo_tri in consumos_para_custos:
        consumo_vazio = consumos_para_custos.get(ciclo_tri, {}).get('V', consumo_vazio)
        consumo_cheias = consumos_para_custos.get(ciclo_tri, {}).get('C', 0)
        consumo_ponta = consumos_para_custos.get(ciclo_tri, {}).get('P', 0)

    # A variável 'consumo' é o total, que corresponde ao valor 'Simples' agregado
    consumo = consumo_simples

    # --- Guardar os consumos FINAIS no estado da sessão ---
    # Isto garante que os valores corretos ficam disponíveis para o resto da aplicação.
    st.session_state['consumos_finais_para_resumo'] = {
        'total': consumo,
        'simples': consumo_simples,
        'vazio': consumo_vazio,
        'fora_vazio': consumo_fora_vazio,
        'cheias': consumo_cheias,
        'ponta': consumo_ponta
    }

else:
    # ####################
    # --- MODO MANUAL ---
    # ####################
    
    # --- Criação das 5 colunas para os widgets ---
    col_mes, col_data_i, col_data_f, col_dias_calc, col_dias_man = st.columns(5)

    with col_mes:
        mes = st.selectbox("Mês", ["Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho", "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"], key="sel_mes", help="Se o mês escolhido já tiver terminado, o valor do OMIE é final, se ainda estiver em curso será com Futuros, que pode consultar no site www.tiagofelicia.pt")

    dias_mes = {"Janeiro": 31, "Fevereiro": 28, "Março": 31, "Abril": 30, "Maio": 31, "Junho": 30, "Julho": 31, "Agosto": 31, "Setembro": 30, "Outubro": 31, "Novembro": 30, "Dezembro": 31}
    ano_atual = datetime.datetime.now().year
    if mes == "Fevereiro" and ((ano_atual % 4 == 0 and ano_atual % 100 != 0) or (ano_atual % 400 == 0)):
        dias_mes["Fevereiro"] = 29
    mes_num = list(dias_mes.keys()).index(mes) + 1

    if 'session_initialized_dates' not in st.session_state:
        hoje = datetime.date.today()
        data_inicial_default_calc = hoje + datetime.timedelta(days=1)
        ano_final_calc, mes_final_calc = (data_inicial_default_calc.year, data_inicial_default_calc.month + 1)
        if mes_final_calc > 12: mes_final_calc, ano_final_calc = 1, ano_final_calc + 1
        dias_no_mes_final = monthrange(ano_final_calc, mes_final_calc)[1]
        dia_final_calc = min(data_inicial_default_calc.day, dias_no_mes_final)
        data_final_bruta = datetime.date(ano_final_calc, mes_final_calc, dia_final_calc)
        data_final_default_calc = data_final_bruta - datetime.timedelta(days=1)
        st.session_state.data_inicio_val = data_inicial_default_calc
        st.session_state.data_fim_val = data_final_default_calc
        st.session_state.previous_mes_for_dates = mes
        st.session_state.session_initialized_dates = True

    if st.session_state.get('previous_mes_for_dates') != mes:
        st.session_state.previous_mes_for_dates = mes
        primeiro_dia_mes_selecionado = datetime.date(ano_atual, mes_num, 1)
        ultimo_dia_mes_selecionado = datetime.date(ano_atual, mes_num, dias_mes[mes])
        st.session_state.data_inicio_val = primeiro_dia_mes_selecionado
        st.session_state.data_fim_val = ultimo_dia_mes_selecionado
        if 'dias_manual_val' in st.session_state: del st.session_state['dias_manual_val']

    data_inicio_anterior = st.session_state.get('data_inicio_val')
    data_fim_anterior = st.session_state.get('data_fim_val')

    # Definir a data mínima permitida
    data_minima_permitida = datetime.date(2025, 1, 1)
    # Definir a data máxima permitida
    data_maxima_permitida = datetime.date(2026, 12, 31)

    with col_data_i:
        data_inicio = st.date_input("Data Inicial", value=st.session_state.data_inicio_val, min_value=data_minima_permitida, max_value=data_maxima_permitida, format="DD/MM/YYYY", key="data_inicio_key_input", help="A partir de 01/01/2025. Se não modificar as datas ou o mês, será calculado a partir do dia seguinte ao atual.")
    # Adicionar uma verificação para garantir que a data inicial não é anterior à mínima
    if data_inicio < data_minima_permitida:
        st.error(f"A Data Inicial selecionada ({data_inicio.strftime('%d/%m/%Y')}) é anterior ao limite de 01/12/2024. Por favor, escolha uma data válida.")
        # Pode optar por parar a execução ou reverter para a data mínima
        data_inicio = data_minima_permitida
        st.session_state.data_inicio_val = data_minima_permitida
        st.rerun()
    with col_data_f:
        data_fim = st.date_input("Data Final", value=st.session_state.data_fim_val, min_value=data_minima_permitida, max_value=data_maxima_permitida, format="DD/MM/YYYY", key="data_fim_key_input", help="De Data Inicial a 31/12/2026. Se não modificar as datas ou o mês, será calculado até um mês após a data inicial.")

    if data_inicio_anterior != data_inicio or data_fim_anterior != data_fim:
        if 'dias_manual_val' in st.session_state: del st.session_state['dias_manual_val']

    # --- Lógica de cálculo e exibição dos dias ---
    dias_calculado = (data_fim - data_inicio).days + 1 if data_fim >= data_inicio else 0

    with col_dias_calc:
        gfx.exibir_metrica_personalizada("Dias (pelas datas)", f"{dias_calculado} dias")

    dias_default_calculado = (data_fim - data_inicio).days + 1 if data_fim >= data_inicio else 0
    with col_dias_man:
        dias_manual_input_val = st.number_input("Nº Dias (manual)", min_value=0, value=st.session_state.get('dias_manual_val', dias_default_calculado), step=1, key="dias_manual_input_key", help="Pode alterar os dias de forma manual, mas dê preferência às datas ou mês, para ter dados mais fidedignos nos tarifários indexados.")
        st.session_state['dias_manual_val'] = dias_manual_input_val

    if pd.isna(dias_manual_input_val) or dias_manual_input_val <= 0:
        dias = dias_default_calculado
    else:
        dias = int(dias_manual_input_val)

    st.write(f"Dias considerados: **{dias} dias**")

# Inicializar variáveis para garantir que existem em todos os caminhos
consumo, consumo_simples, consumo_vazio, consumo_fora_vazio, consumo_cheias, consumo_ponta = 0,0,0,0,0,0
df_consumos_filtrado = pd.DataFrame()

# --- 4. CÁLCULOS UNIFICADOS DE DADOS DE MERCADO E OPÇÕES ---

# Nota sobre OMIE
data_valores_omie_dt = pd.to_datetime(CONSTANTES.loc[CONSTANTES['constante'] == 'Data_Valores_OMIE', 'valor_unitário'].iloc[0]).date()
nota_omie = " (Média Final)" if data_fim <= data_valores_omie_dt else " (Média com Futuros)"

# Cálculos de OMIE e Perdas para o período de simulação ATIVO
df_omie_no_periodo_selecionado = OMIE_PERDAS_CICLOS[
    (OMIE_PERDAS_CICLOS['DataHora'] >= pd.to_datetime(data_inicio)) &
    (OMIE_PERDAS_CICLOS['DataHora'] <= pd.to_datetime(data_fim) + pd.Timedelta(hours=23, minutes=59, seconds=59))
].copy()

if df_omie_no_periodo_selecionado.empty:
    st.error(f"Não foram encontrados dados de mercado OMIE para o período selecionado ({data_inicio.strftime('%d/%m/%Y')} a {data_fim.strftime('%d/%m/%Y')}). Os resultados para tarifários indexados estarão incorretos.")
    omie_medios_calculados_para_todos_ciclos = {}
    perdas_medias = {}
    df_omie_ajustado = pd.DataFrame()
    omie_medio_simples_real_kwh = 0.0
    omie_medios_calculados = {}
else:
    # Mantenha aqui toda a sua lógica detalhada de cálculo de médias, copiada do ficheiro original
    omie_medios_calculados_para_todos_ciclos = {'S': df_omie_no_periodo_selecionado['OMIE'].mean()}
    for ciclo in ['BD', 'BS', 'TD', 'TS']:
        if ciclo in df_omie_no_periodo_selecionado.columns:
            agrupado = df_omie_no_periodo_selecionado.groupby(ciclo)['OMIE'].mean()
            for periodo, media in agrupado.items():
                omie_medios_calculados_para_todos_ciclos[f"{ciclo}_{periodo}"] = media if pd.notna(media) else 0.0
    
    perdas_medias = {}
    perdas_medias['Perdas_M_S'] = df_omie_no_periodo_selecionado['Perdas'].mean()
    for ciclo_base_curto in ['BD', 'BS', 'TD', 'TS']:
        periodos_ciclo = ('V', 'F') if ciclo_base_curto in ['BD', 'BS'] else ('V', 'C', 'P')
        for periodo_perda in periodos_ciclo:
            if ciclo_base_curto in df_omie_no_periodo_selecionado.columns:
                perdas_ciclo_periodo = df_omie_no_periodo_selecionado.groupby(ciclo_base_curto)['Perdas'].mean()
                perdas_medias[f'Perdas_M_{ciclo_base_curto}_{periodo_perda}'] = perdas_ciclo_periodo.get(periodo_perda, 1.0)
            else:
                perdas_medias[f'Perdas_M_{ciclo_base_curto}_{periodo_perda}'] = perdas_medias.get('Perdas_M_S', 1.0)
    
    df_omie_ano_completo_pm = OMIE_PERDAS_CICLOS[OMIE_PERDAS_CICLOS['DataHora'].dt.year == ano_atual].copy()
    if not df_omie_ano_completo_pm.empty:
        perdas_medias['Perdas_Anual_S'] = df_omie_ano_completo_pm['Perdas'].mean()
        for ciclo_anual in ['BD', 'BS', 'TD', 'TS']:
            periodos_ciclo = ('V', 'F') if ciclo_anual in ['BD', 'BS'] else ('V', 'C', 'P')
            for periodo_anual in periodos_ciclo:
                if ciclo_anual in df_omie_ano_completo_pm.columns:
                    perdas_ciclo_anual = df_omie_ano_completo_pm.groupby(ciclo_anual)['Perdas'].mean()
                    perdas_medias[f'Perdas_Anual_{ciclo_anual}_{periodo_anual}'] = perdas_ciclo_anual.get(periodo_anual, 1.0)
                else:
                    perdas_medias[f'Perdas_Anual_{ciclo_anual}_{periodo_anual}'] = perdas_medias.get('Perdas_Anual_S', 1.0)
    
    # Lógica de inputs manuais de OMIE
    omie_medios_calculados = {}
    if 'OMIE' in df_omie_no_periodo_selecionado.columns:
        omie_medios_calculados['S'] = df_omie_no_periodo_selecionado.groupby(pd.Grouper(key='DataHora', freq='D'))['OMIE'].mean().mean()
        if opcao_horaria.lower().startswith("bi"):
            ciclo_col = 'BD' if "diário" in opcao_horaria.lower() else 'BS'
            if ciclo_col in df_omie_no_periodo_selecionado:
                omie_bi = df_omie_no_periodo_selecionado.groupby(ciclo_col)['OMIE'].mean()
                omie_medios_calculados['V'] = omie_bi.get('V', 0.0)
                omie_medios_calculados['F'] = omie_bi.get('F', 0.0)
        elif opcao_horaria.lower().startswith("tri"):
            ciclo_col = 'TD' if "diário" in opcao_horaria.lower() else 'TS'
            if ciclo_col in df_omie_no_periodo_selecionado:
                omie_tri = df_omie_no_periodo_selecionado.groupby(ciclo_col)['OMIE'].mean()
                omie_medios_calculados['V'] = omie_tri.get('V', 0.0)
                omie_medios_calculados['C'] = omie_tri.get('C', 0.0)
                omie_medios_calculados['P'] = omie_tri.get('P', 0.0)

# Ler e Processar a Constante Data_Valores_OMIE
data_valores_omie_dt = None
nota_omie = " (Info OMIE Indisp.)" # Default se a data não puder ser processada

constante_row_data_omie = CONSTANTES[CONSTANTES['constante'] == 'Data_Valores_OMIE']
if not constante_row_data_omie.empty:
    valor_raw = constante_row_data_omie['valor_unitário'].iloc[0]
    if pd.notna(valor_raw):
        try:
            # Pandas geralmente lê datas como datetime64[ns] que se tornam Timestamp
            if isinstance(valor_raw, (datetime.datetime, pd.Timestamp)):
                data_valores_omie_dt = valor_raw.date() # Converter para objeto date
            else:
                # Tentar converter para pd.Timestamp primeiro, depois para date
                timestamp_convertido = pd.to_datetime(valor_raw, errors='coerce')
                if pd.notna(timestamp_convertido):
                    data_valores_omie_dt = timestamp_convertido.date()
            
            if pd.isna(data_valores_omie_dt): # Se a conversão resultou em NaT (Not a Time)
                data_valores_omie_dt = None
                st.warning(f"Não foi possível converter 'Data_Valores_OMIE' para uma data válida: {valor_raw}")
        except Exception as e:
            st.warning(f"Erro ao processar 'Data_Valores_OMIE' ('{valor_raw}'): {e}")
            data_valores_omie_dt = None # Garantir que fica None em caso de erro
    else:
        st.warning("Valor para 'Data_Valores_OMIE' está vazio na folha Constantes.")
else:
    st.warning("Constante 'Data_Valores_OMIE' não encontrada na folha Constantes.")

# Determinar a nota para os inputs OMIE
# 'data_fim' já deve ser um objeto datetime.date
if data_valores_omie_dt and isinstance(data_fim, datetime.date):
    if data_fim <= data_valores_omie_dt:
        nota_omie = " (Média Final)"
    else:
        nota_omie = " (Média com Futuros)"

# --- LÓGICA DE RESET DOS INPUTS OMIE ---
# Gerar uma chave única para os parâmetros que afetam os defaults OMIE
current_omie_dependency_key = f"{data_inicio}-{data_fim}-{opcao_horaria}"

if st.session_state.get('last_omie_dependency_key_for_inputs') != current_omie_dependency_key:
    st.session_state.last_omie_dependency_key_for_inputs = current_omie_dependency_key
    omie_input_keys_to_reset = ['omie_s_input_field', 'omie_v_input_field', 'omie_f_input_field', 'omie_c_input_field', 'omie_p_input_field']
    for key_in_state in omie_input_keys_to_reset:
        if key_in_state in st.session_state:
            del st.session_state[key_in_state]
    st.session_state.omie_foi_editado_manualmente = {}
    # st.write("Debug: OMIE inputs e flags de edição resetados devido à mudança de parâmetros.")

# --- Calcular valores OMIE médios POR PERÍODO (V, F, C, P) e Global (S) ---
# Estes são os valores CALCULADOS da tabela, antes de qualquer input manual
df_omie_no_periodo_selecionado = pd.DataFrame()
if 'DataHora' in OMIE_PERDAS_CICLOS.columns:
    df_omie_no_periodo_selecionado = OMIE_PERDAS_CICLOS[
        (OMIE_PERDAS_CICLOS['DataHora'] >= pd.to_datetime(data_inicio)) & # Usar data_inicio
        (OMIE_PERDAS_CICLOS['DataHora'] <= pd.to_datetime(data_fim) + pd.Timedelta(hours=23, minutes=59, seconds=59)) # Usar data_fim
    ].copy()
else:
    st.warning("Coluna 'DataHora' não encontrada nos dados OMIE. Não é possível calcular médias OMIE.")

omie_medios_calculados = {'S': 0.0, 'V': 0.0, 'F': 0.0, 'C': 0.0, 'P': 0.0}

if not df_omie_no_periodo_selecionado.empty:
    omie_medios_calculados['S'] = df_omie_no_periodo_selecionado['OMIE'].mean()

    if pd.isna(omie_medios_calculados['S']): omie_medios_calculados['S'] = 0.0
    ciclo_bi_col = 'BD' if "Diário" in opcao_horaria else 'BS'
    ciclo_tri_col = 'TD' if "Diário" in opcao_horaria else 'TS'
    if ciclo_bi_col in df_omie_no_periodo_selecionado.columns:
        omie_bi_calculado = df_omie_no_periodo_selecionado.groupby(ciclo_bi_col)['OMIE'].mean()
        omie_medios_calculados['V'] = omie_bi_calculado.get('V', omie_medios_calculados.get('V', 0.0))
        omie_medios_calculados['F'] = omie_bi_calculado.get('F', omie_medios_calculados.get('F', 0.0))
    if ciclo_tri_col in df_omie_no_periodo_selecionado.columns:
        omie_tri_calculado = df_omie_no_periodo_selecionado.groupby(ciclo_tri_col)['OMIE'].mean()
        omie_medios_calculados['V'] = omie_tri_calculado.get('V', omie_medios_calculados.get('V',0.0))
        omie_medios_calculados['C'] = omie_tri_calculado.get('C', omie_medios_calculados.get('C',0.0))
        omie_medios_calculados['P'] = omie_tri_calculado.get('P', omie_medios_calculados.get('P',0.0))

else:
    st.warning("Não existem dados OMIE para o período selecionado. As médias OMIE serão zero.")


# --- Calcular OMIEs médios para TODOS OS CICLOS POSSÍVEIS ---
omie_medios_calculados_para_todos_ciclos = {'S': 0.0} # Inicializar com Simples

if not df_omie_no_periodo_selecionado.empty and 'OMIE' in df_omie_no_periodo_selecionado.columns:
    omie_medios_calculados_para_todos_ciclos['S'] = df_omie_no_periodo_selecionado['OMIE'].mean()
    if pd.isna(omie_medios_calculados_para_todos_ciclos['S']):
        omie_medios_calculados_para_todos_ciclos['S'] = 0.0

    ciclos_a_processar = {
        'BD': ['V', 'F'], 'BS': ['V', 'F'],
        'TD': ['V', 'C', 'P'], 'TS': ['V', 'C', 'P']
    }
    for ciclo_curto, periodos_ciclo in ciclos_a_processar.items():
        if ciclo_curto in df_omie_no_periodo_selecionado.columns:
            omie_ciclo_calculado = df_omie_no_periodo_selecionado.groupby(ciclo_curto)['OMIE'].mean()
            for p_ciclo in periodos_ciclo:
                chave_completa = f"{ciclo_curto}_{p_ciclo}"
                omie_medios_calculados_para_todos_ciclos[chave_completa] = omie_ciclo_calculado.get(p_ciclo, 0.0)
                if pd.isna(omie_medios_calculados_para_todos_ciclos[chave_completa]):
                     omie_medios_calculados_para_todos_ciclos[chave_completa] = 0.0
        else: # Fallback se a coluna do ciclo não existir
            for p_ciclo in periodos_ciclo:
                chave_completa = f"{ciclo_curto}_{p_ciclo}"
                omie_medios_calculados_para_todos_ciclos[chave_completa] = omie_medios_calculados_para_todos_ciclos['S'] # Usa OMIE Simples como fallback


# Garantir que todas as chaves esperadas existem, mesmo que com valor de OMIE Simples como fallback
chaves_omie_esperadas = ['S']
for ciclo_key_esperado in ['BD', 'BS']:
    for periodo_key_esperado in ['V', 'F']:
        chaves_omie_esperadas.append(f"{ciclo_key_esperado}_{periodo_key_esperado}")
for ciclo_key_esperado in ['TD', 'TS']:
    for periodo_key_esperado in ['V', 'C', 'P']:
        chaves_omie_esperadas.append(f"{ciclo_key_esperado}_{periodo_key_esperado}")

for k_omie_calc in chaves_omie_esperadas:
    if k_omie_calc not in omie_medios_calculados_para_todos_ciclos:
        omie_medios_calculados_para_todos_ciclos[k_omie_calc] = omie_medios_calculados_para_todos_ciclos.get('S', 0.0) # Default para OMIE Simples

# --- Inputs Manuais OMIE pelo utilizador e Deteção de Edição ---
st.session_state.omie_foi_editado_manualmente = st.session_state.get('omie_foi_editado_manualmente', {})
omie_medios_calculados = {'S': 0.0, 'V': 0.0, 'F': 0.0, 'C': 0.0, 'P': 0.0} # Recalcular aqui com base em df_omie_no_periodo_selecionado

if not df_omie_no_periodo_selecionado.empty:
    omie_medios_calculados['S'] = df_omie_no_periodo_selecionado['OMIE'].mean()
    if pd.isna(omie_medios_calculados['S']): omie_medios_calculados['S'] = 0.0
    ciclo_bi_col = 'BD' if "Diário" in opcao_horaria else 'BS'
    ciclo_tri_col = 'TD' if "Diário" in opcao_horaria else 'TS'
    if ciclo_bi_col in df_omie_no_periodo_selecionado.columns:
        omie_bi_calculado = df_omie_no_periodo_selecionado.groupby(ciclo_bi_col)['OMIE'].mean()
        omie_medios_calculados['V'] = omie_bi_calculado.get('V', omie_medios_calculados.get('V', 0.0))
        omie_medios_calculados['F'] = omie_bi_calculado.get('F', omie_medios_calculados.get('F', 0.0))
    if ciclo_tri_col in df_omie_no_periodo_selecionado.columns:
        omie_tri_calculado = df_omie_no_periodo_selecionado.groupby(ciclo_tri_col)['OMIE'].mean()
        omie_medios_calculados['V'] = omie_tri_calculado.get('V', omie_medios_calculados.get('V',0.0))
        omie_medios_calculados['C'] = omie_tri_calculado.get('C', omie_medios_calculados.get('C',0.0))
        omie_medios_calculados['P'] = omie_tri_calculado.get('P', omie_medios_calculados.get('P',0.0))
else:
    st.warning("Não existem dados OMIE para o período selecionado. As médias OMIE serão zero.")

# --- 5. LÓGICA DE INPUTS E APRESENTAÇÃO POR MODO ---
consumo_total_final = 0
consumos_repartidos_finais = {}

if is_diagram_mode:
    # ######################
    # --- MODO DIAGRAMA ---
    # ######################    
    df_consumos_filtrado = df_consumos_total[(df_consumos_total['DataHora'].dt.date >= data_inicio) & (df_consumos_total['DataHora'].dt.date <= data_fim)].copy()

    consumos_agregados = proc_dados.agregar_consumos_por_periodo(df_consumos_filtrado, OMIE_PERDAS_CICLOS)
    consumo_simples = consumos_agregados.get('Simples', 0)
    
    ciclo_bi = 'BD' if "Diário" in opcao_horaria else 'BS'
    if ciclo_bi in consumos_agregados:
        consumo_vazio = consumos_agregados.get(ciclo_bi, {}).get('V', 0)
        consumo_fora_vazio = consumos_agregados.get(ciclo_bi, {}).get('F', 0)
    ciclo_tri = 'TD' if "Diário" in opcao_horaria else 'TS'
    if ciclo_tri in consumos_agregados:
        consumo_vazio = consumos_agregados.get(ciclo_tri, {}).get('V', consumo_vazio)
        consumo_cheias = consumos_agregados.get(ciclo_tri, {}).get('C', 0)
        consumo_ponta = consumos_agregados.get(ciclo_tri, {}).get('P', 0)
    
    consumo = consumo_simples

    df_merged_graficos = pd.merge(df_consumos_filtrado, df_omie_no_periodo_selecionado, on='DataHora', how='inner')

    ### INÍCIO SECÇÃO DE ANÁLISE DE POTÊNCIA ###
    st.subheader("⚡ Análise da Potência Contratada")

    is_trifasico = st.checkbox(
        "A minha instalação é Trifásica", 
        key="chk_trifasico",
        help="Selecione esta opção se a sua instalação for trifásica. Neste caso o valor de potência será estimado."
    )

    # A coluna chama-se 'Potencia_kW_Para_Analise'
    coluna_potencia_analise = "Potencia_kW_Para_Analise"

    if not df_consumos_filtrado.empty and coluna_potencia_analise in df_consumos_filtrado.columns:
        pico_potencia_registado = df_consumos_filtrado[coluna_potencia_analise].max()
        potencia_a_comparar = pico_potencia_registado
        nota_trifasico = ""
        nota_trifasico_2 = ""

        if is_trifasico:
            potencia_a_comparar *= 1
            nota_trifasico = "(estimativa para 3 fases)"
            nota_trifasico_2 = "Dado que é uma instalação trifásica, a potência elétrica é distribuida por três fases, sendo o valor da potência máxima tomada a soma das três fases."

        col_p1, col_p2, col_p3 = st.columns(3)
        col_p1.metric("Potência Contratada", f"{potencia} kVA")
        col_p2.metric(f"Potência Máxima Registada (Médias de 15 min) {nota_trifasico}", f"{potencia_a_comparar:.3f} kW", help=nota_trifasico_2)

        percentagem_uso = (potencia_a_comparar / potencia) * 100 if potencia > 0 else 0
        
        recomendacao = ""
        if percentagem_uso > 100:
            recomendacao = f"🔴 **Atenção:** A sua Potência Máxima Registada ({potencia_a_comparar:.2f} kW) ultrapassa a sua potência contratada. Considere aumentar a potência."
        elif percentagem_uso > 85:
            recomendacao = f"✅ **Adequado:** A sua potência contratada parece bem dimensionada."
        elif percentagem_uso > 60:
            recomendacao = f"💡 **Oportunidade de Poupança:** A sua Potência Máxima Registada utiliza entre 60% e 85% da potência contratada. Pode ser possível reduzir a potência."
        else:
            recomendacao = f"💰 **Forte Oportunidade de Poupança:** A sua Potência Máxima Registada utiliza menos de 60% da sua potência contratada. É muito provável que possa reduzir a potência e poupar na fatura."

        col_p3.metric("Utilização da Potência Máxima", f"{percentagem_uso:.1f} %")
        st.markdown(recomendacao)
            
    elif not df_consumos_filtrado.empty:
        st.warning("Não foi possível realizar a análise de potência. Verifique o conteúdo do ficheiro Excel.")

    ### FIM DA SECÇÃO ###

else:
    # ####################
    # --- MODO MANUAL ---
    # ####################

    # --- Alerta sobre o tipo de OMIE ---
    if "Futuros" in nota_omie:
        mensagem_futuros = "ℹ️ Os valores OMIE abaixo são uma média para o período selecionado e incluem dados de futuros (OMIP) - \"Média com Futuros\"."
        gfx.exibir_info_personalizada(mensagem_futuros)
    elif "Final" in nota_omie:
        mensagem_final = "ℹ️ Os valores OMIE abaixo são uma média do OMIE para o período selecionado - \"Média Final\"."
        gfx.exibir_info_personalizada(mensagem_final)
    # Se nota_omie for outra coisa, não mostra nada, o que é um comportamento seguro.

    # --- Lógica de inputs simplificada ---
    if opcao_horaria.lower() == "simples":
        label_s_completo = f"Valor OMIE (€/MWh) - Simples{nota_omie}"
        default_s = round(omie_medios_calculados.get('S', 0.0), 2)
        
        # O widget agora é mais simples
        omie_s_manual = st.number_input(
            label_s_completo,
            value=st.session_state.get('omie_s_input_field', default_s),
            step=1.0, format="%.2f", key="omie_s_input_field"
        )
        # A deteção de edição continua a funcionar
        if omie_s_manual != default_s:
            st.session_state.omie_foi_editado_manualmente['S'] = True

    elif opcao_horaria.lower().startswith("bi"):
        col_omie1, col_omie2 = st.columns(2)
        with col_omie1:
            label_v_completo = f"Valor OMIE (€/MWh) - Vazio{nota_omie}"
            default_v = round(omie_medios_calculados.get('V', 0.0), 2)
            omie_v_manual = st.number_input(
                label_v_completo, value=st.session_state.get('omie_v_input_field', default_v),
                step=1.0, format="%.2f", key="omie_v_input_field"
            )
            if omie_v_manual != default_v:
                st.session_state.omie_foi_editado_manualmente['V'] = True
        with col_omie2:
            label_f_completo = f"Valor OMIE (€/MWh) - Fora Vazio {nota_omie}"
            default_f = round(omie_medios_calculados.get('F', 0.0), 2)
            omie_f_manual = st.number_input(
                label_f_completo, value=st.session_state.get('omie_f_input_field', default_f),
                step=1.0, format="%.2f", key="omie_f_input_field"
            )
            if omie_f_manual != default_f:
                st.session_state.omie_foi_editado_manualmente['F'] = True

    elif opcao_horaria.lower().startswith("tri"):
        col_omie1, col_omie2, col_omie3 = st.columns(3)
        with col_omie1:
            label_v_completo = f"Valor OMIE (€/MWh) - Vazio{nota_omie}"
            default_v = round(omie_medios_calculados.get('V', 0.0), 2)
            omie_v_manual = st.number_input(
                label_v_completo, value=st.session_state.get('omie_v_input_field', default_v),
                step=1.0, format="%.2f", key="omie_v_input_field"
            )
            if omie_v_manual != default_v:
                st.session_state.omie_foi_editado_manualmente['V'] = True
        with col_omie2:
            label_c_completo = f"Valor OMIE (€/MWh) - Cheias{nota_omie}"
            default_c = round(omie_medios_calculados.get('C', 0.0), 2)
            omie_c_manual = st.number_input(
                label_c_completo, value=st.session_state.get('omie_c_input_field', default_c),
                step=1.0, format="%.2f", key="omie_c_input_field"
            )
            if omie_c_manual != default_c:
                st.session_state.omie_foi_editado_manualmente['C'] = True
        with col_omie3:
            label_p_completo = f"Valor OMIE (€/MWh) - Ponta{nota_omie}"
            default_p = round(omie_medios_calculados.get('P', 0.0), 2)
            omie_p_manual = st.number_input(
                label_p_completo, value=st.session_state.get('omie_p_input_field', default_p),
                step=1.0, format="%.2f", key="omie_p_input_field"
            )
            if omie_p_manual != default_p:
                st.session_state.omie_foi_editado_manualmente['P'] = True

    if not is_diagram_mode:
        with st.expander("📊 Ver Gráfico de Evolução dos Preços Médios Diários OMIE PT no Período"):
        # Passamos a data de split para a função
            dados_dos_graficos = preparar_dados_grafico_manual(
                df_omie_no_periodo_selecionado, 
                data_inicio, 
                data_fim, 
                data_valores_omie_dt,
                opcao_horaria
            )

            if dados_dos_graficos:
                st.markdown("---")
                for dados_um_grafico in dados_dos_graficos:
                    html_grafico = gfx.gerar_grafico_highcharts_multi_serie(
                        chart_id=dados_um_grafico['id'],
                        chart_data=dados_um_grafico
                    )
                    st.components.v1.html(html_grafico, height=320)
    
    # --- Alerta para uso de OMIE Manual ---
    alertas_omie_manual = []
    if st.session_state.omie_foi_editado_manualmente.get('S'): alertas_omie_manual.append("Simples")
    if st.session_state.omie_foi_editado_manualmente.get('V'): alertas_omie_manual.append("Vazio")
    if st.session_state.omie_foi_editado_manualmente.get('F'): alertas_omie_manual.append("Fora Vazio")
    if st.session_state.omie_foi_editado_manualmente.get('C'): alertas_omie_manual.append("Cheias")
    if st.session_state.omie_foi_editado_manualmente.get('P'): alertas_omie_manual.append("Ponta")

    if alertas_omie_manual:
        st.info(f"ℹ️ Atenção: Os cálculos estão a utilizar valores OMIE manuais (editados) para o(s) período(s): {', '.join(alertas_omie_manual)}. "
                "Para os tarifários quarto-horários, isto significa que este valor OMIE manual (e não os OMIEs horários) será aplicado a todas as horas desse(s) período(s). "
                "Outros períodos não editados usarão os OMIEs horários (para quarto-horários) ou as médias calculadas (para tarifários de média).")
 
    ### Consumos Manuais
    st.markdown(
        "Introduza o consumo para cada período (kWh). Pode usar somas ou subtrações simples (ex: `100+50+30` ou `100+50-30`). Os valores serão arredondados para o inteiro mais próximo."
    )
    if opcao_horaria.lower() == "simples":
        st.markdown("###### Consumo Simples (kWh)")
        expressao_s = st.text_input(
            "Introduza o consumo total ou um cálculo (ex: 80+20+58 ou 200+58-100)",
            key="exp_consumo_s",
            on_change=atualizar_url_consumos
        )
    
        resultado_s_calc, erro_s = calc.calcular_expressao_matematica_simples(expressao_s, "Simples")
    
        if erro_s:
            st.error(erro_s)
            consumo_simples = int(st.session_state.get('exp_consumo_s_anterior_valido', 0))
        else:
            consumo_simples = resultado_s_calc
            st.session_state.exp_consumo_s_anterior_valido = consumo_simples
            mensagem = f"Consumo Simples considerado: <b>{consumo_simples:.0f} kWh</b>"
            gfx.exibir_info_personalizada(mensagem)
        consumo = consumo_simples

    elif opcao_horaria.lower().startswith("bi"):
        st.markdown("###### Consumos Bi-Horário (kWh)")
        col_bi1, col_bi2 = st.columns(2)
        with col_bi1:
            expressao_v = st.text_input("Vazio (ex: 60+3 ou 70+3-10)", key="exp_consumo_v", on_change=atualizar_url_consumos)
            consumo_vazio, erro_v = calc.calcular_expressao_matematica_simples(expressao_v, "Vazio")
            if erro_v: 
                st.error(erro_v)
                consumo_vazio = int(st.session_state.get('exp_consumo_v_anterior_valido', 0))
            else:
                st.session_state.exp_consumo_v_anterior_valido = consumo_vazio
                mensagem_v = f"Consumo Vazio considerado: <b>{consumo_vazio:.0f} kWh</b>"
                gfx.exibir_info_personalizada(mensagem_v)
        with col_bi2:
            expressao_f = st.text_input("Fora Vazio (ex: 90+5 ou 100+5-10)", key="exp_consumo_f", on_change=atualizar_url_consumos)
            consumo_fora_vazio, erro_f = calc.calcular_expressao_matematica_simples(expressao_f, "Fora Vazio")
            if erro_f: 
                st.error(erro_f)
                consumo_fora_vazio = int(st.session_state.get('exp_consumo_f_anterior_valido', 0))
            else:
                st.session_state.exp_consumo_f_anterior_valido = consumo_fora_vazio
                mensagem_f = f"Consumo Fora Vazio considerado: <b>{consumo_fora_vazio:.0f} kWh</b>"
                gfx.exibir_info_personalizada(mensagem_f)
        consumo = consumo_vazio + consumo_fora_vazio

    elif opcao_horaria.lower().startswith("tri"):
        st.markdown("###### Consumos Tri-Horário (kWh)")
        col_tri1, col_tri2, col_tri3 = st.columns(3)
        with col_tri1:
            expressao_tv = st.text_input("Vazio (ex: 60+3 ou 70+3-10)", key="exp_consumo_v", on_change=atualizar_url_consumos)
            consumo_vazio, erro_tv = calc.calcular_expressao_matematica_simples(expressao_tv, "Vazio (Tri)")
            if erro_tv: 
                st.error(erro_tv)
                consumo_vazio = int(st.session_state.get('exp_consumo_v_anterior_valido', 0))
            else:
                st.session_state.exp_consumo_v_anterior_valido = consumo_vazio
                mensagem_v = f"Consumo Vazio considerado: <b>{consumo_vazio:.0f} kWh</b>"
                gfx.exibir_info_personalizada(mensagem_v)
        with col_tri2:
            expressao_tc = st.text_input("Cheias (ex: 60+8 ou 70+8-10)", key="exp_consumo_c", on_change=atualizar_url_consumos)
            consumo_cheias, erro_tc = calc.calcular_expressao_matematica_simples(expressao_tc, "Cheias (Tri)")
            if erro_tc: 
                st.error(erro_tc)
                consumo_cheias = int(st.session_state.get('exp_consumo_tc_anterior_valido', 0))
            else:
                st.session_state.exp_consumo_tc_anterior_valido = consumo_cheias
                mensagem_c = f"Consumo Cheias considerado: <b>{consumo_cheias:.0f} kWh</b>"
                gfx.exibir_info_personalizada(mensagem_c)
        with col_tri3:
            expressao_tp = st.text_input("Ponta (ex: 20+7 ou 30+7-10)", key="exp_consumo_p", on_change=atualizar_url_consumos)
            consumo_ponta, erro_tp = calc.calcular_expressao_matematica_simples(expressao_tp, "Ponta (Tri)")
            if erro_tp: 
                st.error(erro_tp)
                consumo_ponta = int(st.session_state.get('exp_consumo_tp_anterior_valido', 0))
            else:
                st.session_state.exp_consumo_tp_anterior_valido = consumo_ponta
                mensagem_p = f"Consumo Ponta considerado: <b>{consumo_ponta:.0f} kWh</b>"
                gfx.exibir_info_personalizada(mensagem_p)
        consumo = consumo_vazio + consumo_cheias + consumo_ponta

    st.write(f"Total Consumo a considerar nos cálculos: **{consumo:.0f} kWh**") # Exibir total como inteiro

    consumos_para_custos = {'Simples': consumo} # A chave 'Simples' contém sempre o consumo total

    oh_manual_lower = opcao_horaria.lower()

    if oh_manual_lower.startswith("bi"):
        ciclo_bi_manual = 'BD' if "diário" in oh_manual_lower else 'BS'
        # Inicializa ambos os ciclos para consistência, preenchendo o que for relevante
        consumos_para_custos['BD'] = {}
        consumos_para_custos['BS'] = {}
        consumos_para_custos[ciclo_bi_manual] = {
            'V': consumo_vazio,
            'F': consumo_fora_vazio
        }
    elif oh_manual_lower.startswith("tri"):
        ciclo_tri_manual = 'TD' if "diário" in oh_manual_lower else 'TS'
        # Mesma lógica para o tri-horário
        consumos_para_custos['TD'] = {}
        consumos_para_custos['TS'] = {}
        consumos_para_custos[ciclo_tri_manual] = {
            'V': consumo_vazio,
            'C': consumo_cheias,
            'P': consumo_ponta
        }

# --- Preparar df_omie_ajustado ---
# Começa com os OMIEs HORÁRIOS ORIGINAIS da tabela para o período selecionado
df_omie_ajustado = df_omie_no_periodo_selecionado.copy()
if not df_omie_ajustado.empty:
    if opcao_horaria.lower() == "simples":
        if st.session_state.omie_foi_editado_manualmente.get('S'):
            df_omie_ajustado['OMIE'] = st.session_state.omie_s_input_field

    elif opcao_horaria.lower().startswith("bi"):
        ciclo_col = 'BD' if "Diário" in opcao_horaria else 'BS'
        if ciclo_col in df_omie_ajustado.columns:
            def substituir_bi_se_manual(row):
                omie_original_hora = row['OMIE']
                if row[ciclo_col] == 'V' and st.session_state.omie_foi_editado_manualmente.get('V'):
                    return st.session_state.omie_v_input_field
                elif row[ciclo_col] == 'F' and st.session_state.omie_foi_editado_manualmente.get('F'):
                    return st.session_state.omie_f_input_field
                return omie_original_hora
            if st.session_state.omie_foi_editado_manualmente.get('V') or st.session_state.omie_foi_editado_manualmente.get('F'):
                 df_omie_ajustado['OMIE'] = df_omie_ajustado.apply(substituir_bi_se_manual, axis=1)

    elif opcao_horaria.lower().startswith("tri"):
        ciclo_col = 'TD' if "Diário" in opcao_horaria else 'TS'
        if ciclo_col in df_omie_ajustado.columns:
            def substituir_tri_se_manual(row):
                omie_original_hora = row['OMIE']
                if row[ciclo_col] == 'V' and st.session_state.omie_foi_editado_manualmente.get('V'):
                    return st.session_state.omie_v_input_field
                elif row[ciclo_col] == 'C' and st.session_state.omie_foi_editado_manualmente.get('C'):
                    return st.session_state.omie_c_input_field
                elif row[ciclo_col] == 'P' and st.session_state.omie_foi_editado_manualmente.get('P'):
                    return st.session_state.omie_p_input_field
                return omie_original_hora
            if st.session_state.omie_foi_editado_manualmente.get('V') or \
               st.session_state.omie_foi_editado_manualmente.get('C') or \
               st.session_state.omie_foi_editado_manualmente.get('P'):
                df_omie_ajustado['OMIE'] = df_omie_ajustado.apply(substituir_tri_se_manual, axis=1)
        else:
            st.warning(f"Coluna de ciclo '{ciclo_col}' não encontrada nos dados OMIE. Não é possível aplicar OMIE manual por período horário.")
else: # df_omie_ajustado está vazio porque df_omie_no_periodo_selecionado estava vazio
    st.warning("DataFrame OMIE ajustado está vazio pois não há dados OMIE para o período.")

# --- Recalcular omie_medio_simples_real_kwh com base nos OMIE ajustados ---
# Este valor é usado por alguns tarifários de MÉDIA (ex: LuziGás)
if not df_omie_ajustado.empty and 'OMIE' in df_omie_ajustado.columns:
    omie_medio_simples_real_kwh = df_omie_ajustado['OMIE'].mean() / 1000.0
    if pd.isna(omie_medio_simples_real_kwh): omie_medio_simples_real_kwh = 0.0
else:
    omie_medio_simples_real_kwh = 0.0

# --- Guardar os valores OMIE médios (para tarifários de MÉDIA) ---
# Estes virão dos inputs, que por sua vez são inicializados com as médias calculadas ou com os valores manuais do utilizador.
omie_para_tarifarios_media = {}
if opcao_horaria.lower() == "simples":
    omie_para_tarifarios_media['S'] = st.session_state.get('omie_s_input_field', omie_medios_calculados['S'])
elif opcao_horaria.lower().startswith("bi"):
    omie_para_tarifarios_media['V'] = st.session_state.get('omie_v_input_field', omie_medios_calculados['V'])
    omie_para_tarifarios_media['F'] = st.session_state.get('omie_f_input_field', omie_medios_calculados['F'])
elif opcao_horaria.lower().startswith("tri"):
    omie_para_tarifarios_media['V'] = st.session_state.get('omie_v_input_field', omie_medios_calculados['V'])
    omie_para_tarifarios_media['C'] = st.session_state.get('omie_c_input_field', omie_medios_calculados['C'])
    omie_para_tarifarios_media['P'] = st.session_state.get('omie_p_input_field', omie_medios_calculados['P'])

# --- Calcular Perdas médias para TODOS os ciclos e períodos (GLOBALMENTE) ---
perdas_medias = {}
if not df_omie_no_periodo_selecionado.empty and 'Perdas' in df_omie_no_periodo_selecionado.columns:
    # Médias para o período selecionado
    perdas_medias['Perdas_M_S'] = df_omie_no_periodo_selecionado['Perdas'].mean()
    
    for ciclo_base_curto in ['BD', 'BS']: # Bi-Diário, Bi-Semanal
        if ciclo_base_curto in df_omie_no_periodo_selecionado.columns:
            perdas_ciclo_periodo = df_omie_no_periodo_selecionado.groupby(ciclo_base_curto)['Perdas'].mean()
            perdas_medias[f'Perdas_M_{ciclo_base_curto}_V'] = perdas_ciclo_periodo.get('V', 1.0)
            perdas_medias[f'Perdas_M_{ciclo_base_curto}_F'] = perdas_ciclo_periodo.get('F', 1.0)
        else: # Fallback se coluna de ciclo não existir para o período selecionado
            perdas_medias[f'Perdas_M_{ciclo_base_curto}_V'] = perdas_medias['Perdas_M_S'] # Usa média simples como fallback
            perdas_medias[f'Perdas_M_{ciclo_base_curto}_F'] = perdas_medias['Perdas_M_S']

    for ciclo_base_curto in ['TD', 'TS']: # Tri-Diário, Tri-Semanal
        if ciclo_base_curto in df_omie_no_periodo_selecionado.columns:
            perdas_ciclo_periodo = df_omie_no_periodo_selecionado.groupby(ciclo_base_curto)['Perdas'].mean()
            perdas_medias[f'Perdas_M_{ciclo_base_curto}_V'] = perdas_ciclo_periodo.get('V', 1.0)
            perdas_medias[f'Perdas_M_{ciclo_base_curto}_C'] = perdas_ciclo_periodo.get('C', 1.0)
            perdas_medias[f'Perdas_M_{ciclo_base_curto}_P'] = perdas_ciclo_periodo.get('P', 1.0)
        else: # Fallback
            perdas_medias[f'Perdas_M_{ciclo_base_curto}_V'] = perdas_medias['Perdas_M_S']
            perdas_medias[f'Perdas_M_{ciclo_base_curto}_C'] = perdas_medias['Perdas_M_S']
            perdas_medias[f'Perdas_M_{ciclo_base_curto}_P'] = perdas_medias['Perdas_M_S']

    # Médias para o ano completo
    df_omie_ano_completo_pm = OMIE_PERDAS_CICLOS[OMIE_PERDAS_CICLOS['DataHora'].dt.year == ano_atual].copy()
    if not df_omie_ano_completo_pm.empty and 'Perdas' in df_omie_ano_completo_pm.columns:
        perdas_medias['Perdas_Anual_S'] = df_omie_ano_completo_pm['Perdas'].mean()

        for ciclo_base_curto_anual in ['BD', 'BS']:
            if ciclo_base_curto_anual in df_omie_ano_completo_pm.columns:
                perdas_ciclo_anual = df_omie_ano_completo_pm.groupby(ciclo_base_curto_anual)['Perdas'].mean()
                perdas_medias[f'Perdas_Anual_{ciclo_base_curto_anual}_V'] = perdas_ciclo_anual.get('V', 1.0)
                perdas_medias[f'Perdas_Anual_{ciclo_base_curto_anual}_F'] = perdas_ciclo_anual.get('F', 1.0)
            else: # Fallback
                perdas_medias[f'Perdas_Anual_{ciclo_base_curto_anual}_V'] = perdas_medias.get('Perdas_Anual_S', 1.0)
                perdas_medias[f'Perdas_Anual_{ciclo_base_curto_anual}_F'] = perdas_medias.get('Perdas_Anual_S', 1.0)

        for ciclo_base_curto_anual in ['TD', 'TS']:
            if ciclo_base_curto_anual in df_omie_ano_completo_pm.columns:
                perdas_ciclo_anual = df_omie_ano_completo_pm.groupby(ciclo_base_curto_anual)['Perdas'].mean()
                perdas_medias[f'Perdas_Anual_{ciclo_base_curto_anual}_V'] = perdas_ciclo_anual.get('V', 1.0)
                perdas_medias[f'Perdas_Anual_{ciclo_base_curto_anual}_C'] = perdas_ciclo_anual.get('C', 1.0)
                perdas_medias[f'Perdas_Anual_{ciclo_base_curto_anual}_P'] = perdas_ciclo_anual.get('P', 1.0)
            else: # Fallback
                perdas_medias[f'Perdas_Anual_{ciclo_base_curto_anual}_V'] = perdas_medias.get('Perdas_Anual_S', 1.0)
                perdas_medias[f'Perdas_Anual_{ciclo_base_curto_anual}_C'] = perdas_medias.get('Perdas_Anual_S', 1.0)
                perdas_medias[f'Perdas_Anual_{ciclo_base_curto_anual}_P'] = perdas_medias.get('Perdas_Anual_S', 1.0)
    else:
        st.warning("Não existem dados OMIE para o ano completo. Algumas médias de perdas anuais podem não ser calculadas.")
else:
    st.warning("Não existem dados OMIE ou coluna 'Perdas' para o período selecionado. As médias de perdas podem não ser calculadas corretamente.")
# Garantir que todas as chaves esperadas existem em perdas_medias, mesmo que com default 1.0
chaves_perdas_esperadas = [
    'Perdas_M_S', 'Perdas_Anual_S',
    'Perdas_M_BD_V', 'Perdas_Anual_BD_V', 'Perdas_M_BD_F', 'Perdas_Anual_BD_F',
    'Perdas_M_BS_V', 'Perdas_Anual_BS_V', 'Perdas_M_BS_F', 'Perdas_Anual_BS_F',
    'Perdas_M_TD_V', 'Perdas_Anual_TD_V', 'Perdas_M_TD_C', 'Perdas_Anual_TD_C', 'Perdas_M_TD_P', 'Perdas_Anual_TD_P',
    'Perdas_M_TS_V', 'Perdas_Anual_TS_V', 'Perdas_M_TS_C', 'Perdas_Anual_TS_C', 'Perdas_M_TS_P', 'Perdas_Anual_TS_P',
]
for k_perda in chaves_perdas_esperadas:
    if k_perda not in perdas_medias:
        perdas_medias[k_perda] = 1.0 # Default para perdas

# --- FUNÇÃO AUXILIAR PARA OPÇÕES DE FILTRO ---
def get_filter_options_for_multiselect(df_fixos, df_indexados, column_name):
    options = []
    if df_fixos is not None and column_name in df_fixos.columns:
        options.extend(df_fixos[column_name].astype(str).str.strip().dropna().unique())
    if df_indexados is not None and column_name in df_indexados.columns:
        options.extend(df_indexados[column_name].astype(str).str.strip().dropna().unique())
    
    if column_name == 'Tipo':
        options.append("Pessoal")
    if column_name == 'Segmento':
        options.append("Pessoal")
        
    unique_options_intermediate = set(opt for opt in options if opt and opt.lower() != 'nan')
    
    return sorted(list(unique_options_intermediate))
# --- FIM FUNÇÃO AUXILIAR PARA OPÇÕES DE FILTRO ---

# --- Lógica de Reset dos Inputs de Consumo ao Mudar Opção Horária ---
chave_opcao_horaria_consumo = f"consumo_inputs_para_{opcao_horaria}"

# ... (Restantes inputs: Taxas DGEG/CAV, Consumos, Opções Adicionais, Meu Tarifário) ...
# Expander para as opções que são menos alteradas ou mais específicas
with st.expander("➕ Opções Adicionais de Simulação (Tarifa Social e condicionais)"):
    st.markdown("##### Definição de Taxas Mensais")
    col_taxa1, col_taxa2 = st.columns(2)
    with col_taxa1:
        valor_dgeg_user = st.number_input(
            "Valor DGEG (€/mês)",
            min_value=0.0, step=0.01, value=0.07, # Mantém os teus defaults
            help="Taxa de Exploração da Direção-Geral de Energia e Geologia - Verifique qual o valor cobrado na sua fatura. Em condições normais, para contratos domésticos o valor é de 0,07 €/mês e os não domésticos têm o valor de 0,35 €/mês.",
            key="dgeg_input",
            on_change=atualizar_url_opcoes_adicionais
        )
    with col_taxa2:
        valor_cav_user = st.number_input(
            "Valor Contribuição Audiovisual (€/mês)",
            min_value=0.0, step=0.01, value=2.85, # Mantém os teus defaults
            help="Contribuição Audiovisual (CAV) - Verifique qual o valor cobrado na sua fatura. O valor normal é de 2,85 €/mês. Será 1 €/mês, para alguns casos de Tarifa Social (1º escalão de abono...) Será 0 €/mês, para consumo inferior a 400 kWh/ano.",
            key="cav_input",
            on_change=atualizar_url_opcoes_adicionais
        )
    
    # --- MOSTRAR BENEFÍCIOS APENAS SE POTÊNCIA PERMITE ---
    if potencia <= 6.9:  # <--- CONDIÇÃO AQUI
        st.markdown(r"##### Benefícios e Condições Especiais (para potências $\leq 6.9$ kVA)") # Título condicional
        colx1, colx2 = st.columns(2)
        with colx1:
            tarifa_social = st.checkbox("Tarifa Social",
                                        key="chk_tarifa_social",
                                        help="Só pode ter Tarifa Social para potências até 6.9 kVA.",
                                        on_change=atualizar_url_opcoes_adicionais)
        with colx2:
            familia_numerosa = st.checkbox("Família Numerosa",
                                        key="chk_familia_numerosa",
                                        help="300 kWh com IVA a 6% (em vez dos normais 200 kWh) para potências até 6.9 kVA.",
                                        on_change=atualizar_url_opcoes_adicionais)
    else:
        tarifa_social = False
        familia_numerosa = False
        st.session_state.chk_tarifa_social = False
        st.session_state.chk_familia_numerosa = False

    # --- CONDIÇÕES PARA ACP E CONTINENTE (dentro do expander) ---
    mostrar_widgets_acp_continente = False
    if potencia <= 20.7:
        if opcao_horaria.lower() == "simples" or opcao_horaria.lower().startswith("bi"):
            mostrar_widgets_acp_continente = True

    if mostrar_widgets_acp_continente:
        st.markdown("##### Parcerias e Descontos Específicos")
        colx3, colx4 = st.columns(2)
        with colx3:
            # Padrão simplificado
            incluir_quota_acp = st.checkbox(
                "Incluir Quota ACP",
                key='chk_acp',
                help="Inclui o valor da quota do ACP (4,90 €/mês) no valor do tarifário da parceria GE/ACP.",
                on_change=atualizar_url_opcoes_adicionais
            )
        with colx4:
            # Padrão simplificado
            #desconto_continente = st.checkbox("Desconto Continente",key='chk_continente',help="Comparar o custo total incluindo o desconto do valor do cupão Continente no tarifário Galp&Continente.",on_change=atualizar_url_opcoes_adicionais)
            desconto_continente = False
    else:
        incluir_quota_acp = st.session_state.get('chk_acp', True)
        desconto_continente = st.session_state.get('chk_continente', True)

    # O título mantém-se para organização visual
    #st.markdown("##### Comparação Tarifários Indexados")
    # A variável é agora fixada como True, removendo a checkbox da interface
    comparar_indexados = True

# --- Fim do st.expander "Opções Adicionais de Simulação" ---

# Checkbox para ativar "O Meu Tarifário"
help_O_Meu_Tarifario = """
Para preencher os valores de acordo com o seu tarifário, ou com outro qualquer que queira comparar.

**Atenção às notas sobre as TAR e TSE.**
    """
meu_tarifario_ativo = st.checkbox(
    "**Comparar com O Meu Tarifário?**",
    key="chk_meu_tarifario_ativo",
    help=help_O_Meu_Tarifario,
    on_change=atualizar_url_meu_tarifario
)

# Criação de todos_omie_inputs_utilizador_comp
todos_omie_inputs_utilizador_comp = {
    'S': st.session_state.get('omie_s_input_field', round(omie_medios_calculados.get('S',0), 2)),
    'V': st.session_state.get('omie_v_input_field', round(omie_medios_calculados.get('V',0), 2)),
    'F': st.session_state.get('omie_f_input_field', round(omie_medios_calculados.get('F',0), 2)),
    'C': st.session_state.get('omie_c_input_field', round(omie_medios_calculados.get('C',0), 2)),
    'P': st.session_state.get('omie_p_input_field', round(omie_medios_calculados.get('P',0), 2))
}
# omie_medio_simples_real_kwh já é calculado globalmente (linha 619) e pode ser passado como está
# perdas_medias já é calculado globalmente (linhas 637-678) e pode ser passado como está

ts_global_ativa = tarifa_social # Flag global de TS

# --- "Meu Tarifário" ---
# Exibe o subheader e o conteúdo apenas se a checkbox estiver selecionada
if meu_tarifario_ativo:
    st.subheader("🧾 O Meu Tarifário (para comparação)")

    # Definir chaves para todos os inputs do Meu Tarifário
    # Preços de Energia e Potência
    key_energia_meu_s = "energia_meu_s_input_val"
    key_potencia_meu = "potencia_meu_input_val"
    key_energia_meu_v = "energia_meu_v_input_val"
    key_energia_meu_f = "energia_meu_f_input_val"
    key_energia_meu_c = "energia_meu_c_input_val"
    key_energia_meu_p = "energia_meu_p_input_val"
    # Checkboxes TAR/TSE
    key_meu_tar_energia = "meu_tar_energia_val"
    key_meu_tar_potencia = "meu_tar_potencia_val"
    key_meu_fin_tse_incluido = "meu_fin_tse_incluido_val"
    # Descontos
    key_meu_desconto_energia = "meu_desconto_energia_val"
    key_meu_desconto_potencia = "meu_desconto_potencia_val"
    key_meu_desconto_fatura = "meu_desconto_fatura_val"
    # Acréscimo
    key_meu_acrescimo_fatura = "meu_acrescimo_fatura_val"

    # Lista de todas as chaves do Meu Tarifário para facilitar a limpeza
    chaves_meu_tarifario = [
        key_energia_meu_s, key_potencia_meu, key_energia_meu_v, key_energia_meu_f,
        key_energia_meu_c, key_energia_meu_p, key_meu_tar_energia, key_meu_tar_potencia,
        key_meu_fin_tse_incluido, key_meu_desconto_energia, key_meu_desconto_potencia,
        key_meu_desconto_fatura, key_meu_acrescimo_fatura
    ]

    # Botão para Limpar Dados do Meu Tarifário
    # Colocado antes dos inputs para que a limpeza ocorra antes da renderização dos inputs
    if st.button("🧹 Limpar Dados do Meu Tarifário", key="btn_limpar_meu_tarifario"):
        for k in chaves_meu_tarifario:
            if k in st.session_state:
                del st.session_state[k]
        # Também podemos limpar o resultado calculado, se existir
        if 'meu_tarifario_calculado' in st.session_state:
            del st.session_state['meu_tarifario_calculado']
        st.success("Dados do 'Meu Tarifário' foram repostos.")

    col_user1, col_user2, col_user3, col_user4 = st.columns(4)

    # Usar st.session_state.get(key, default_value_para_o_input)
    # Default None para number_input faz com que apareça vazio (com placeholder se definido no widget)
    # Default True/False para checkboxes
    # Default 0.0 para descontos

    if opcao_horaria.lower() == "simples":
        with col_user1:
            energia_meu = st.number_input("Preço Energia (€/kWh)", min_value=0.0, step=0.001, format="%g",
                                        value=st.session_state.get(key_energia_meu_s, None), key=key_energia_meu_s, on_change=atualizar_url_meu_tarifario)
        with col_user2:
            potencia_meu = st.number_input("Preço Potência (€/dia)", min_value=0.0, step=0.001, format="%g",
                                         value=st.session_state.get(key_potencia_meu, None), key=key_potencia_meu, on_change=atualizar_url_meu_tarifario)
    elif opcao_horaria.lower().startswith("bi"):
        with col_user1:
            energia_vazio_meu = st.number_input("Preço Vazio (€/kWh)", min_value=0.0, step=0.001, format="%g",
                                              value=st.session_state.get(key_energia_meu_v, None), key=key_energia_meu_v, on_change=atualizar_url_meu_tarifario)
        with col_user2:
            energia_fora_vazio_meu = st.number_input("Preço Fora Vazio (€/kWh)", min_value=0.0, step=0.001, format="%g",
                                                   value=st.session_state.get(key_energia_meu_f, None), key=key_energia_meu_f, on_change=atualizar_url_meu_tarifario)
        with col_user3:
            potencia_meu = st.number_input("Preço Potência (€/dia)", min_value=0.0, step=0.001, format="%g",
                                         value=st.session_state.get(key_potencia_meu, None), key=key_potencia_meu, on_change=atualizar_url_meu_tarifario)
    elif opcao_horaria.lower().startswith("tri"):
        with col_user1:
            energia_vazio_meu = st.number_input("Preço Vazio (€/kWh)", min_value=0.0, step=0.001, format="%g",
                                              value=st.session_state.get(key_energia_meu_v, None), key=key_energia_meu_v, on_change=atualizar_url_meu_tarifario)
        with col_user2:
            energia_cheias_meu = st.number_input("Preço Cheias (€/kWh)", min_value=0.0, step=0.001, format="%g",
                                               value=st.session_state.get(key_energia_meu_c, None), key=key_energia_meu_c, on_change=atualizar_url_meu_tarifario)
        with col_user3:
            energia_ponta_meu = st.number_input("Preço Ponta (€/kWh)", min_value=0.0, step=0.001, format="%g",
                                              value=st.session_state.get(key_energia_meu_p, None), key=key_energia_meu_p, on_change=atualizar_url_meu_tarifario)
        with col_user4:
            potencia_meu = st.number_input("Preço Potência (€/dia)", min_value=0.0, step=0.001, format="%g",
                                         value=st.session_state.get(key_potencia_meu, None), key=key_potencia_meu, on_change=atualizar_url_meu_tarifario)

    col_userx1, col_userx2, col_userx3 = st.columns(3)
    with col_userx1:
        tar_incluida_energia_meu = st.checkbox("TAR incluída na Energia?", value=st.session_state.get(key_meu_tar_energia, True), key=key_meu_tar_energia, help="É muito importante saber se os valores têm ou não as TAR (Tarifas de Acesso às Redes). Alguns comercializadores separam na fatura, outros não. Verifique se há alguma referência a Acesso às Redes na fatura em (€/kWh)", on_change=atualizar_url_meu_tarifario)
    with col_userx2:
        tar_incluida_potencia_meu = st.checkbox("TAR incluída na Potência?", value=st.session_state.get(key_meu_tar_potencia, True), key=key_meu_tar_potencia, help="É muito importante saber se os valores têm ou não as TAR (Tarifas de Acesso às Redes). Alguns comercializadores separam na fatura, outros não. Verifique se há alguma referência a Acesso às Redes na fatura em (€/dia)", on_change=atualizar_url_meu_tarifario)
    with col_userx3:
        # A checkbox "Inclui Financiamento TSE?" guarda True se ESTIVER incluído.
        # A variável 'adicionar_financiamento_tse_meu' é o inverso.
        checkbox_tse_incluido_estado = st.checkbox("Inclui Financiamento TSE?", value=st.session_state.get(key_meu_fin_tse_incluido, True), key=key_meu_fin_tse_incluido, help="É importante saber se os valores têm ou não incluido o financiamento da Tarifa Social de Eletricidade (TSE). Alguns comercializadores separam na fatura, outros não. Verifique se há alguma referência a Financiamento Tarifa Social na fatura em (€/kWh)", on_change=atualizar_url_meu_tarifario)
        adicionar_financiamento_tse_meu = not checkbox_tse_incluido_estado


    col_userd1, col_userd2, col_userd3, col_userd4 = st.columns(4)
    with col_userd1:
        desconto_energia = st.number_input("Desconto Energia (%)", min_value=0.0, max_value=100.0, step=0.1,
                                           value=st.session_state.get(key_meu_desconto_energia, 0.0), key=key_meu_desconto_energia, help="O desconto é aplicado a Energia+TAR. Alguns tarifários não aplicam o desconto nas TAR (por exemplo os da Plenitude), pelo que o desconto não pode ser aqui aplicado. Se não tiver, não necessita preencher!", on_change=atualizar_url_meu_tarifario)
    with col_userd2:
        desconto_potencia = st.number_input("Desconto Potência (%)", min_value=0.0, max_value=100.0, step=0.1,
                                            value=st.session_state.get(key_meu_desconto_potencia, 0.0), key=key_meu_desconto_potencia, help="O desconto é aplicado a Potência+TAR. Alguns tarifários não aplicam o desconto nas TAR, pelo que se assim for, o desconto não pode ser aqui aplicado. Se não tiver, não necessita preencher!", on_change=atualizar_url_meu_tarifario)
    with col_userd3:
        desconto_fatura_input_meu = st.number_input("Desconto Fatura (€)", min_value=0.0, step=0.01, format="%.2f",
                                                 value=st.session_state.get(key_meu_desconto_fatura, 0.0), key=key_meu_desconto_fatura, help="Se não tiver, não necessita preencher!", on_change=atualizar_url_meu_tarifario)
    with col_userd4:
        acrescimo_fatura_input_meu = st.number_input("Acréscimo Fatura (€)", min_value=0.0, step=0.01, format="%.2f",
                                                 value=st.session_state.get(key_meu_acrescimo_fatura, 0.0), key=key_meu_acrescimo_fatura, help="Para outros custos fixos na fatura. Se não tiver, não necessita preencher!", on_change=atualizar_url_meu_tarifario)
    
    if st.button("Calcular e Adicionar O Meu Tarifário à Comparação", icon="🧮", type="primary", key="btn_meu_tarifario", use_container_width=True):

        # Esta variável verifica se o número de dias da simulação corresponde a um mês de faturação.
        is_billing_month = 28 <= dias <= 31

        preco_energia_input_meu = {}
        
        # Acesso aos valores via variáveis locais (que o Streamlit preenche a partir dos widgets com keys)
        if opcao_horaria.lower() == "simples":
            preco_energia_input_meu['S'] = float(energia_meu or 0.0) # Usa a variável local 'energia_meu'
            preco_potencia_input_meu = float(potencia_meu or 0.0)
        elif opcao_horaria.lower().startswith("bi"):
            preco_energia_input_meu['V'] = float(energia_vazio_meu or 0.0)
            preco_energia_input_meu['F'] = float(energia_fora_vazio_meu or 0.0)
            preco_potencia_input_meu = float(potencia_meu or 0.0)
        elif opcao_horaria.lower().startswith("tri"):
            preco_energia_input_meu['V'] = float(energia_vazio_meu or 0.0)
            preco_energia_input_meu['C'] = float(energia_cheias_meu or 0.0)
            preco_energia_input_meu['P'] = float(energia_ponta_meu or 0.0)
            preco_potencia_input_meu = float(potencia_meu or 0.0)
        else: # Fallback caso algo corra mal com opcao_horaria
            preco_potencia_input_meu = 0.0

        preco_potencia_input_meu = float(potencia_meu or 0.0) # Input user potencia
        alert_negativo = False
        if preco_potencia_input_meu < 0: alert_negativo = True

        consumos_horarios_para_func = {} # Dicionário consumos para função IVA

        if opcao_horaria.lower() == "simples":
            preco_energia_input_meu['S'] = float(energia_meu or 0.0)
            if preco_energia_input_meu['S'] < 0: alert_negativo = True
            consumos_horarios_para_func = {'S': consumos_para_custos.get('Simples', 0)}
        elif opcao_horaria.lower().startswith("bi"):
            ciclo_a_usar = 'BD' if "Diário" in opcao_horaria else 'BS'

            preco_energia_input_meu['V'] = float(energia_vazio_meu or 0.0)
            preco_energia_input_meu['F'] = float(energia_fora_vazio_meu or 0.0)
            if preco_energia_input_meu['V'] < 0 or preco_energia_input_meu['F'] < 0: alert_negativo = True
            consumos_horarios_para_func = {
                'V': consumos_para_custos.get(ciclo_a_usar, {}).get('V', 0),
                'F': consumos_para_custos.get(ciclo_a_usar, {}).get('F', 0)
            }
        elif opcao_horaria.lower().startswith("tri"):
            ciclo_a_usar = 'TD' if "Diário" in opcao_horaria else 'TS'
            preco_energia_input_meu['V'] = float(energia_vazio_meu or 0.0)
            preco_energia_input_meu['C'] = float(energia_cheias_meu or 0.0)
            preco_energia_input_meu['P'] = float(energia_ponta_meu or 0.0)
            if preco_energia_input_meu['V'] < 0 or preco_energia_input_meu['C'] < 0 or preco_energia_input_meu['P'] < 0: alert_negativo = True
            consumos_horarios_para_func = {
                'V': consumos_para_custos.get(ciclo_a_usar, {}).get('V', 0),
                'C': consumos_para_custos.get(ciclo_a_usar, {}).get('C', 0),
                'P': consumos_para_custos.get(ciclo_a_usar, {}).get('P', 0)
            }

        if alert_negativo:
            st.warning("Atenção: Introduziu um ou mais preços negativos para o seu tarifário.")

    # --- 1. OBTER COMPONENTES BASE (SEM DESCONTOS, SEM TS, SEM IVA) ---

    # ENERGIA (por período p)
        tar_energia_regulada_periodo_meu = {} # TAR da energia por período (€/kWh)
        for p_key in preco_energia_input_meu.keys(): # S, V, F, C, P
            tar_energia_regulada_periodo_meu[p_key] = calc.obter_tar_energia_periodo(opcao_horaria, p_key, potencia, CONSTANTES)

        energia_meu_periodo_comercializador_base = {} # Componente do comercializador para energia (€/kWh)
        for p_key, preco_input_val in preco_energia_input_meu.items():
            preco_input_val_float = float(preco_input_val or 0.0)
            if tar_incluida_energia_meu:
                energia_meu_periodo_comercializador_base[p_key] = preco_input_val_float - tar_energia_regulada_periodo_meu.get(p_key, 0.0)
            else:
                energia_meu_periodo_comercializador_base[p_key] = preco_input_val_float

        # Financiamento TSE (€/kWh, valor único, aplicável a todos os períodos de energia)
        # 'adicionar_financiamento_tse_meu' é (not checkbox_tse_incluido_estado)
        financiamento_tse_a_somar_base = FINANCIAMENTO_TSE_VAL if adicionar_financiamento_tse_meu else 0.0

        # POTÊNCIA (€/dia)
        tar_potencia_regulada_meu_base = calc.obter_tar_dia(potencia, CONSTANTES) # TAR da potência
        preco_potencia_input_meu_float = float(preco_potencia_input_meu or 0.0)
        if tar_incluida_potencia_meu:
            potencia_meu_comercializador_base = preco_potencia_input_meu_float - tar_potencia_regulada_meu_base
        else:
            potencia_meu_comercializador_base = preco_potencia_input_meu_float

    # --- 2. CALCULAR PREÇOS UNITÁRIOS FINAIS (PARA EXIBIR NA TABELA, SEM IVA) ---
    # Estes preços já incluem o desconto percentual do comercializador e o desconto da Tarifa Social.

        preco_energia_final_unitario_sem_iva = {} # Dicionário para {período: preço_final_unitario}
        desconto_monetario_ts_energia = 0.0 # Valor do desconto TS para energia em €/kWh
        if tarifa_social: # Flag global de TS
            desconto_monetario_ts_energia = calc.obter_constante('Desconto TS Energia', CONSTANTES)

        for p_key in energia_meu_periodo_comercializador_base.keys():
            # Base para o desconto percentual da energia (Comercializador + TAR + TSE)
            preco_total_energia_antes_desc_perc = (
                energia_meu_periodo_comercializador_base.get(p_key, 0.0) +
                tar_energia_regulada_periodo_meu.get(p_key, 0.0) +
                financiamento_tse_a_somar_base
            )

            # Aplicar desconto percentual do comercializador à energia
            preco_energia_apos_desc_comerc = preco_total_energia_antes_desc_perc * (1 - (desconto_energia or 0.0) / 100.0)

            # Aplicar desconto da Tarifa Social (se ativo)
            if tarifa_social:
                preco_energia_final_unitario_sem_iva[p_key] = preco_energia_apos_desc_comerc - desconto_monetario_ts_energia
            else:
                preco_energia_final_unitario_sem_iva[p_key] = preco_energia_apos_desc_comerc

        # Preço unitário final da Potência (€/dia, sem IVA)
        desconto_monetario_ts_potencia = 0.0 # Valor do desconto TS para potência em €/dia
        if tarifa_social:
            desconto_monetario_ts_potencia = calc.obter_constante(f'Desconto TS Potencia {potencia}', CONSTANTES)

        # Base para o desconto percentual da potência (Comercializador + TAR)
        preco_total_potencia_antes_desc_perc = potencia_meu_comercializador_base + tar_potencia_regulada_meu_base

        # Aplicar desconto percentual do comercializador à potência
        preco_potencia_apos_desc_comerc = preco_total_potencia_antes_desc_perc * (1 - (desconto_potencia or 0.0) / 100.0)

        # Aplicar desconto da Tarifa Social (se ativo)
        if tarifa_social:
            preco_potencia_final_unitario_sem_iva = max(0.0, preco_potencia_apos_desc_comerc - desconto_monetario_ts_potencia)
        else:
            preco_potencia_final_unitario_sem_iva = preco_potencia_apos_desc_comerc

        desconto_ts_potencia_valor_aplicado_meu = 0.0
        if tarifa_social:
             desconto_ts_potencia_dia_bruto_meu = calc.obter_constante(f'Desconto TS Potencia {potencia}', CONSTANTES)
             # O desconto efetivamente aplicado à TAR para o meu tarifário.
             # tar_potencia_regulada_meu_base é a TAR bruta.
             desconto_ts_potencia_valor_aplicado_meu = min(tar_potencia_regulada_meu_base, desconto_ts_potencia_dia_bruto_meu)

    # --- 3. CALCULAR CUSTOS TOTAIS (COM IVA) E DECOMPOSIÇÃO PARA TOOLTIP ---

        # CUSTO ENERGIA COM IVA
        # Preparar inputs para calcular_custo_energia_com_iva, que usa os preços unitários finais (já com descontos)
        preco_energia_simples_para_iva = None
        precos_energia_horarios_para_iva = {}
        if opcao_horaria.lower() == "simples":
            preco_energia_simples_para_iva = preco_energia_final_unitario_sem_iva.get('S')
        else: # Bi ou Tri
            precos_energia_horarios_para_iva = {
                p: val for p, val in preco_energia_final_unitario_sem_iva.items() if p != 'S'
            }

        decomposicao_custo_energia_meu = calc.calcular_custo_energia_com_iva(
            consumo, # Consumo total global
            preco_energia_simples_para_iva,
            precos_energia_horarios_para_iva,
            dias,
            potencia,
            opcao_horaria,
            consumos_horarios_para_func,
            familia_numerosa
        )

        custo_energia_meu_final_com_iva = decomposicao_custo_energia_meu['custo_com_iva']
        tt_cte_energia_siva_meu = decomposicao_custo_energia_meu['custo_sem_iva']
        tt_cte_energia_iva_6_meu = decomposicao_custo_energia_meu['valor_iva_6']
        tt_cte_energia_iva_23_meu = decomposicao_custo_energia_meu['valor_iva_23']


        # CUSTO POTÊNCIA COM IVA
        # Para aplicar IVA corretamente em potências <= 3.45 kVA, precisamos das componentes "comercializador" e "TAR" APÓS os descontos.

        # Componente do comercializador para potência, após o seu desconto percentual
        comp_comerc_pot_para_iva = potencia_meu_comercializador_base * (1 - (desconto_potencia or 0.0) / 100.0)

        # Componente TAR da potência, após o desconto percentual do comercializador e o desconto TS
        tar_pot_bruta_apos_desc_perc = tar_potencia_regulada_meu_base * (1 - (desconto_potencia or 0.0) / 100.0)

        tar_pot_final_para_iva = 0.0
        if tarifa_social:
            tar_pot_final_para_iva = max(0.0, tar_pot_bruta_apos_desc_perc - desconto_monetario_ts_potencia)
        else:
            tar_pot_final_para_iva = tar_pot_bruta_apos_desc_perc

        decomposicao_custo_potencia_meu = calc.calcular_custo_potencia_com_iva_final(
            comp_comerc_pot_para_iva,
            tar_pot_final_para_iva,
            dias,
            potencia
        )
        custo_potencia_meu_final_com_iva = decomposicao_custo_potencia_meu['custo_com_iva']
        tt_cte_potencia_siva_meu = decomposicao_custo_potencia_meu['custo_sem_iva']
        tt_cte_potencia_iva_6_meu = decomposicao_custo_potencia_meu['valor_iva_6']
        tt_cte_potencia_iva_23 = decomposicao_custo_potencia_meu['valor_iva_23']


        # --- 4. TAXAS ADICIONAIS E CUSTO TOTAL FINAL ---

        # Taxas Adicionais (IEC, DGEG, CAV) - chamada à função não muda
        consumo_total_para_taxas_meu = sum(consumos_horarios_para_func.values())

        decomposicao_taxas_meu = calc.calcular_taxas_adicionais(
            consumo_total_para_taxas_meu,
            dias, tarifa_social,
            valor_dgeg_user, valor_cav_user,
            nome_comercializador_atual="Pessoal",
            aplica_taxa_fixa_mensal=is_billing_month
        )
        taxas_meu_tarifario_com_iva = decomposicao_taxas_meu['custo_com_iva']
        # tt_cte_taxas_siva = decomposicao_taxas_meu['custo_sem_iva'] # Já teremos as taxas s/IVA individuais
        tt_cte_iec_siva = decomposicao_taxas_meu['iec_sem_iva']
        tt_cte_dgeg_siva = decomposicao_taxas_meu['dgeg_sem_iva']
        tt_cte_cav_siva = decomposicao_taxas_meu['cav_sem_iva']
        tt_cte_taxas_iva_6 = decomposicao_taxas_meu['valor_iva_6']
        tt_cte_taxas_iva_23 = decomposicao_taxas_meu['valor_iva_23']


    # Custo Total antes do desconto de fatura em €
        custo_total_antes_desc_fatura = custo_energia_meu_final_com_iva + custo_potencia_meu_final_com_iva + taxas_meu_tarifario_com_iva

        # Aplicar Desconto Fatura (€) - lógica não muda
        # 'desconto_fatura_input_meu' já é o valor do input numérico
        custo_total_meu_tarifario_com_iva = custo_total_antes_desc_fatura - float(desconto_fatura_input_meu or 0.0) + float(acrescimo_fatura_input_meu or 0.0)

        # Calcular totais para o tooltip
        tt_cte_total_siva_meu = tt_cte_energia_siva_meu + tt_cte_potencia_siva_meu + tt_cte_iec_siva + tt_cte_dgeg_siva + tt_cte_cav_siva
        tt_cte_valor_iva_6_total_meu = tt_cte_energia_iva_6_meu + tt_cte_potencia_iva_6_meu + tt_cte_taxas_iva_6
        tt_cte_valor_iva_23_total_meu = tt_cte_energia_iva_23_meu + tt_cte_potencia_iva_23 + tt_cte_taxas_iva_23

        # Calcular Subtotal c/IVA (antes do desconto de fatura)
        tt_cte_subtotal_civa_meu = tt_cte_total_siva_meu + tt_cte_valor_iva_6_total_meu + tt_cte_valor_iva_23_total_meu
        
        # Descontos e Acréscimos Finais
        tt_cte_desc_finais_valor_meu = 0.0
        if 'desconto_fatura_input_meu' in locals() and desconto_fatura_input_meu > 0:
            tt_cte_desc_finais_valor_meu = desconto_fatura_input_meu

        tt_cte_acres_finais_valor_meu = float(acrescimo_fatura_input_meu or 0.0)

        # Adicionar ao resultado_meu_tarifario_dict
        componentes_tooltip_custo_total_dict_meu = {
            'tt_cte_energia_siva': tt_cte_energia_siva_meu,
            'tt_cte_potencia_siva': tt_cte_potencia_siva_meu,
            'tt_cte_iec_siva': tt_cte_iec_siva,
            'tt_cte_dgeg_siva': tt_cte_dgeg_siva,
            'tt_cte_cav_siva': tt_cte_cav_siva,
            'tt_cte_total_siva': tt_cte_total_siva_meu,
            'tt_cte_valor_iva_6_total': tt_cte_valor_iva_6_total_meu,
            'tt_cte_valor_iva_23_total': tt_cte_valor_iva_23_total_meu,
            'tt_cte_subtotal_civa': tt_cte_subtotal_civa_meu,
            'tt_cte_desc_finais_valor': tt_cte_desc_finais_valor_meu,
            'tt_cte_acres_finais_valor': tt_cte_acres_finais_valor_meu
        }

        # --- INÍCIO: CAMPOS PARA TOOLTIPS DE ENERGIA (O MEU TARIFÁRIO) ---
        componentes_tooltip_energia_dict_meu = {}

        # Desconto bruto da Tarifa Social para energia (se TS global estiver ativa)
        desconto_ts_energia_bruto = 0.0
        if tarifa_social: # tarifa_social é a flag global do checkbox
            desconto_ts_energia_bruto = calc.obter_constante('Desconto TS Energia', CONSTANTES)


        for p_key_tooltip in preco_energia_input_meu.keys():
            preco_final_celula_periodo = preco_energia_final_unitario_sem_iva.get(p_key_tooltip, 0.0) # Valor que vai para a célula
    
            # Componentes fixas para o tooltip, conforme as regras:
            tar_bruta_para_tooltip = tar_energia_regulada_periodo_meu.get(p_key_tooltip, 0.0) # Regra 1
    
            # Se checkbox_tse_incluido_estado é True, o TSE está "embutido" e o tooltip só faz uma nota.
            # Se False, o tooltip mostra "Financiamento TSE: VALOR"
            tse_valor_para_soma_tooltip = FINANCIAMENTO_TSE_VAL if not checkbox_tse_incluido_estado else 0.0
    
            desconto_ts_bruto_para_tooltip = desconto_ts_energia_bruto if tarifa_social else 0.0 # Regra 2

            # Calcular a componente "Comercializador (s/TAR)" para o tooltip:
            # É o valor residual para que (Comercializador_Tooltip + TAR_Tooltip + TSE_Adicional_Tooltip - DescontoTS_Tooltip) = PrecoFinalCelula
            comerc_final_para_tooltip = (
                preco_final_celula_periodo -
                tar_bruta_para_tooltip -
                tse_valor_para_soma_tooltip + # Subtrai o valor que o tooltip adicionará
                desconto_ts_bruto_para_tooltip # Adiciona de volta o que o tooltip subtrairá
            )

            # Valores para as flags e campos nominais do tooltip JS:
            tooltip_tse_declarado_incluido = checkbox_tse_incluido_estado # Vem da checkbox do utilizador
            tooltip_tse_valor_nominal = FINANCIAMENTO_TSE_VAL # O JS usa isto se a flag acima for false
            tooltip_ts_aplicada_flag = tarifa_social
            tooltip_ts_desconto_valor_bruto = desconto_ts_energia_bruto if tarifa_social else 0.0 # O JS usa isto se a flag acima for true

            componentes_tooltip_energia_dict_meu[f'tooltip_energia_{p_key_tooltip}_comerc_sem_tar'] = comerc_final_para_tooltip
            componentes_tooltip_energia_dict_meu[f'tooltip_energia_{p_key_tooltip}_tar_bruta'] = tar_bruta_para_tooltip
            componentes_tooltip_energia_dict_meu[f'tooltip_energia_{p_key_tooltip}_tse_declarado_incluido'] = tooltip_tse_declarado_incluido
            componentes_tooltip_energia_dict_meu[f'tooltip_energia_{p_key_tooltip}_tse_valor_nominal'] = tooltip_tse_valor_nominal
            componentes_tooltip_energia_dict_meu[f'tooltip_energia_{p_key_tooltip}_ts_aplicada_flag'] = tooltip_ts_aplicada_flag
            componentes_tooltip_energia_dict_meu[f'tooltip_energia_{p_key_tooltip}_ts_desconto_valor'] = tooltip_ts_desconto_valor_bruto
        # --- FIM: CAMPOS PARA TOOLTIPS DE ENERGIA (O MEU TARIFÁRIO) ---

            # Para o tooltip do Preço Potência (O MEU TARIFÁRIO):
            potencia_comerc_base_meu = potencia_meu_comercializador_base
            tar_potencia_bruta_meu = tar_potencia_regulada_meu_base
    
            # Base para o desconto percentual da potência (Comercializador Base + TAR Bruta)
            base_para_desconto_potencia = potencia_comerc_base_meu + tar_potencia_bruta_meu
    
            # Valor do desconto monetário total para potência
            desconto_monetario_total_potencia = base_para_desconto_potencia * ((desconto_potencia or 0.0) / 100.0)

            # Aplicar desconto primeiro à componente do comercializador
            if desconto_monetario_total_potencia <= potencia_comerc_base_meu:
                pot_comerc_final_tooltip = potencia_comerc_base_meu - desconto_monetario_total_potencia
                pot_tar_final_tooltip = tar_potencia_bruta_meu
            else:
                pot_comerc_final_tooltip = 0.0
                desconto_restante_pot_para_tar = desconto_monetario_total_potencia - potencia_comerc_base_meu
                pot_tar_final_tooltip = max(0.0, tar_potencia_bruta_meu - desconto_restante_pot_para_tar)

    # ... (cálculo de 'desconto_ts_potencia_valor_aplicado_meu' permanece o mesmo) ...
            componentes_tooltip_potencia_dict_meu = {
                'tooltip_pot_comerc_sem_tar': pot_comerc_final_tooltip, # Componente do comercializador s/TAR e s/TS
                'tooltip_pot_tar_bruta': tar_potencia_regulada_meu_base,              # TAR bruta s/TS
                'tooltip_pot_ts_aplicada': tarifa_social,                       # True se TS ativa globalmente
                'tooltip_pot_desconto_ts_valor': desconto_ts_potencia_valor_aplicado_meu if tarifa_social else 0.0, # Valor do desconto TS efetivamente aplicado à TAR
            }

        # --- 5. PREPARAR RESULTADOS PARA EXIBIÇÃO NA TABELA ---

        nome_para_exibir_meu_tarifario = "O Meu Tarifário"
        sufixo_nome = "" # String para o sufixo do nome

        # Converter para float, tratando valores nulos para evitar erros
        desconto = float(desconto_fatura_input_meu or 0.0)
        acrescimo = float(acrescimo_fatura_input_meu or 0.0)

        if desconto > 0 and acrescimo > 0:
            # Cenário 1: Ambos existem, calcular o valor líquido
            valor_liquido = desconto - acrescimo
            
            if valor_liquido > 0:
                # O desconto é maior que o acréscimo -> Resultado final é um Desconto
                custo_sem_ajuste = custo_total_meu_tarifario_com_iva + valor_liquido
                sufixo_nome = f" (Inclui desconto líquido de {valor_liquido:.2f}€ no período, s/ desc.={custo_sem_ajuste:.2f}€)"
            
            elif valor_liquido < 0:
                # O acréscimo é maior que o desconto -> Resultado final é um Acréscimo
                valor_acresc_abs = abs(valor_liquido)
                custo_sem_ajuste = custo_total_meu_tarifario_com_iva - valor_acresc_abs
                sufixo_nome = f" (Inclui acréscimo líquido de {valor_acresc_abs:.2f}€ no período, s/ acrésc.={custo_sem_ajuste:.2f}€)"
            
            # Se valor_liquido for 0, não adicionamos sufixo

        elif desconto > 0:
            # Cenário 2: Apenas desconto existe
            custo_sem_ajuste = custo_total_meu_tarifario_com_iva + desconto
            sufixo_nome = f" (Inclui desconto de {desconto:.2f}€ no período, s/ desc.={custo_sem_ajuste:.2f}€)"

        elif acrescimo > 0:
            # Cenário 3: Apenas acréscimo existe
            custo_sem_ajuste = custo_total_meu_tarifario_com_iva - acrescimo
            sufixo_nome = f" (Inclui acréscimo de {acrescimo:.2f}€ no período, s/ acrésc.={custo_sem_ajuste:.2f}€)"

        # Adicionar o sufixo calculado ao nome final
        nome_para_exibir_meu_tarifario += sufixo_nome

        valores_energia_meu_exibir_dict = {}
        if isinstance(preco_energia_final_unitario_sem_iva, dict):
            for p, v_energia in preco_energia_final_unitario_sem_iva.items():
                periodo_nome = ""
                if p == 'S': periodo_nome = "Simples"
                elif p == 'V': periodo_nome = "Vazio"
                elif p == 'F': periodo_nome = "Fora Vazio"
                elif p == 'C': periodo_nome = "Cheias"
                elif p == 'P': periodo_nome = "Ponta"
                if periodo_nome:
                    valores_energia_meu_exibir_dict[f'{periodo_nome} (€/kWh)'] = round(v_energia, 4)

        resultado_meu_tarifario_dict = {
            'NomeParaExibir': nome_para_exibir_meu_tarifario,
            'LinkAdesao': "-",            
            'Tipo': "Pessoal",
            'Comercializador': "-",
            'Segmento': "Pessoal",
            'Faturação': "-",
            'Pagamento': "-",          
            **valores_energia_meu_exibir_dict,
            'Potência (€/dia)': round(preco_potencia_final_unitario_sem_iva, 4), # Arredondar para exibição
            'Total (€)': round(custo_total_meu_tarifario_com_iva, 2),
            'opcao_horaria_calculada': opcao_horaria,
            # CAMPOS DO TOOLTIP DA POTÊNCIA MEU
            **componentes_tooltip_potencia_dict_meu,
            # CAMPOS DO TOOLTIP DA ENERGIA MEU
            **componentes_tooltip_energia_dict_meu, 
            # CAMPOS DO TOOLTIP DA CUSTO TOTAL MEU
            **componentes_tooltip_custo_total_dict_meu, 
            }
        
        st.session_state['meu_tarifario_calculado'] = resultado_meu_tarifario_dict
        st.success(f"Cálculo para 'O Meu Tarifário' adicionado/atualizado. Custo: {custo_total_meu_tarifario_com_iva:.2f} €")
# --- Fim do if st.button ---

# --- INÍCIO DA SECÇÃO: TARIFÁRIO PERSONALIZADO ---
personalizado_ativo = st.checkbox(
    "**Comparar outro Tarifário Personalizado? (simplificado)**",
    key="chk_pers_ativo",
    help="Crie tarifário personalizado para comparar com os seus consumos. Ideal para comparar diferentes opções tarifárias/horárias. Não permite descontos e acréscimos que existem em 'O Meu Tarifário (para comparação)'."
)

if personalizado_ativo:
    st.info("Introduza os preços para as estruturas tarifárias que pretende simular. Deixe a zero os campos que não quiser calcular.")

    # --- INPUTS DINÂMICOS COM BASE NA OPÇÃO HORÁRIA PRINCIPAL ---
    
    # REGRA 1 E PARTE DAS REGRAS 2 E 3: CAMPOS PARA SIMPLES
    if opcao_horaria.lower() == "simples" or opcao_horaria.lower().startswith("bi") or (opcao_horaria.lower().startswith("tri") and potencia <= 20.7):
        with st.container(border=True):
            st.subheader("Estrutura Simples Personalizada")
            col_s1, col_s2 = st.columns(2)
            with col_s1:
                st.number_input("Preço Energia Simples (€/kWh)", key="pers_energia_s", min_value=0.0, step=0.001, format="%g")
            with col_s2:
                st.number_input("Preço Potência Simples (€/dia)", key="pers_potencia_s", min_value=0.0, step=0.001, format="%g")

    # REGRA 2 E PARTE DA REGRA 3: CAMPOS PARA BI-HORÁRIO
    if opcao_horaria.lower().startswith("bi") or (opcao_horaria.lower().startswith("tri") and potencia <= 20.7):
        with st.container(border=True):
            st.subheader("Estrutura Bi-Horária Personalizada")
            col_b1, col_b2, col_b3 = st.columns(3)
            with col_b1:
                st.number_input("Preço Vazio Bi-Horário (€/kWh)", key="pers_energia_v_bi", min_value=0.0, step=0.001, format="%g")
            with col_b2:
                st.number_input("Preço Fora Vazio Bi-Horário (€/kWh)", key="pers_energia_f_bi", min_value=0.0, step=0.001, format="%g")
            with col_b3:
                st.number_input("Preço Potência Bi-Horário (€/dia)", key="pers_potencia_bi", min_value=0.0, step=0.001, format="%g")

    # REGRA 3 E 4: CAMPOS PARA TRI-HORÁRIO
    if opcao_horaria.lower().startswith("tri"):
        with st.container(border=True):
            st.subheader("Estrutura Tri-Horária Personalizada")
            col_t1, col_t2, col_t3, col_t4 = st.columns(4)
            with col_t1:
                st.number_input("Preço Vazio Tri-Horário (€/kWh)", key="pers_energia_v_tri", min_value=0.0, step=0.001, format="%g")
            with col_t2:
                st.number_input("Preço Cheias Tri-Horário (€/kWh)", key="pers_energia_c_tri", min_value=0.0, step=0.001, format="%g")
            with col_t3:
                st.number_input("Preço Ponta Tri-Horário (€/kWh)", key="pers_energia_p_tri", min_value=0.0, step=0.001, format="%g")
            with col_t4:
                st.number_input("Preço Potência Tri-Horário (€/dia)", key="pers_potencia_tri", min_value=0.0, step=0.001, format="%g")

    # --- Checkboxes comuns para TAR/TSE do tarifário personalizado ---
    with st.container(border=True):
        st.subheader("Opções Comuns ao Tarifário Personalizado")
        col_opt1, col_opt2, col_opt3 = st.columns(3)
        with col_opt1:
            st.checkbox("TAR incluída na Energia?", value=True, key="pers_tar_energia")
        with col_opt2:
            st.checkbox("TAR incluída na Potência?", value=True, key="pers_tar_potencia")
        with col_opt3:
            st.checkbox("Inclui Financiamento TSE?", value=True, key="pers_tse_incluido")

    # --- Botão de cálculo ---
    # ### Lógica do botão "Calcular" ###
    if st.button("Guardar e Configurar Tarifário Personalizado", type="primary", key="btn_calc_pers", use_container_width=True):
        # Apenas guarda os dados no session_state. Não faz cálculos aqui.
        st.session_state['dados_tarifario_personalizado'] = {
            'ativo': True,
            'precos_s': {'energia': st.session_state.get('pers_energia_s', 0.0), 'potencia': st.session_state.get('pers_potencia_s', 0.0)},
            'precos_bi': {'vazio': st.session_state.get('pers_energia_v_bi', 0.0), 'fora_vazio': st.session_state.get('pers_energia_f_bi', 0.0), 'potencia': st.session_state.get('pers_potencia_bi', 0.0)},
            'precos_tri': {'vazio': st.session_state.get('pers_energia_v_tri', 0.0), 'cheias': st.session_state.get('pers_energia_c_tri', 0.0), 'ponta': st.session_state.get('pers_energia_p_tri', 0.0), 'potencia': st.session_state.get('pers_potencia_tri', 0.0)},
            'flags': {
                'tar_energia': st.session_state.get('pers_tar_energia', True),
                'tar_potencia': st.session_state.get('pers_tar_potencia', True),
                'tse_incluido': st.session_state.get('pers_tse_incluido', True)
            }
        }
        st.success("Tarifário Personalizado guardado! Os cálculos serão refletidos nas tabelas.")

    # --- FIM DA SECÇÃO ---        

#Seletor de Modo de Visualização - OPÇÃO HORÁRIA OU DETALHADA
help_modo_de_comparacao = """
    Ative para ver uma tabela que compara o custo de cada tarifário em diferentes opções horárias (Simples, Bi-Horário, Tri-Horário), usando os seus consumos inseridos. Irá desativar a tabela detalhada.

    **Estará ordenada pela opção tarifária escolhida, mas posteriormente pode ordenar por qualquer outra coluna.**
    """
modo_de_comparacao_ativo = st.checkbox(
"🔬 **Comparar custos entre diferentes Opções Horárias para os mesmos consumos?**",
    value=False, # Começa desativado por defeito
    key="chk_modo_comparacao_opcoes",
    help=help_modo_de_comparacao 
)

st.markdown("---") # Separador
st.subheader("🔍 Filtros da Tabela de Resultados")
# --- INÍCIO: FILTROS PARA A TABELA DE RESULTADOS ---
filt_col1, filt_col2, filt_col3, filt_col4 = st.columns(4)

# --- Filtro de Segmento ---
with filt_col1:
    opcoes_filtro_segmento_user = ["Residencial", "Empresarial", "Ambos"]
    # Encontrar o índice da opção default ("Residencial") para o selectbox
    default_index_segmento = opcoes_filtro_segmento_user.index("Residencial")
    
    selected_segmento_user = st.selectbox(
        "Segmento", 
        opcoes_filtro_segmento_user, 
        index=st.session_state.get("filter_segmento_selectbox_index", default_index_segmento), # Guarda o índice para manter o estado
        key="filter_segmento_selectbox",
        help="Escolha o segmento para a simulação."
    )
    # Guarda o índice selecionado no session_state
    st.session_state.filter_segmento_selectbox_index = opcoes_filtro_segmento_user.index(selected_segmento_user)


with filt_col2:
    tipos_options_ms = get_filter_options_for_multiselect(tarifarios_fixos, tarifarios_indexados, 'tipo')
    
    # Criar o texto de ajuda formatado com Markdown
    help_text_formatado = """
    Deixe em branco para mostrar todos os tipos ou selecione um ou mais para filtrar.

    Os tipos de tarifário são:
    * **Fixo**: O preço da energia (€/kWh) é constante durante o contrato.
    * **Indexado Média**: Preço da energia baseado na média do OMIE para os períodos horários.
    * **Indexado quarto-horário**: Preço da energia baseado nos valores OMIE horários/quarto-horários e no perfil de consumo. Também conhecidos como 'Dinâmicos'.
    """
    
    selected_tipos = st.multiselect("Tipo(s) de Tarifário", tipos_options_ms,
                                 default=st.session_state.get("filter_tipos_multi", []),
                                 key="filter_tipos_multi",
                                 help=help_text_formatado)

# --- Filtro de Faturação ---
with filt_col3:
    opcoes_faturacao_user = ["Todas", "Fatura eletrónica", "Fatura em papel"]
    selected_faturacao_user = st.selectbox(
        "Faturação",
        opcoes_faturacao_user,
        index=st.session_state.get("filter_faturacao_selectbox_index", 0), # Default "Todas"
        key="filter_faturacao_selectbox",
        help="Escolha o tipo de faturação pretendido."
    )
    st.session_state.filter_faturacao_selectbox_index = opcoes_faturacao_user.index(selected_faturacao_user)


# --- Filtro de Pagamento ---
with filt_col4:
    opcoes_pagamento_user = ["Todos", "Débito Direto", "Multibanco", "Numerário/Payshop/CTT"]
    selected_pagamento_user = st.selectbox(
        "Pagamento",
        opcoes_pagamento_user,
        index=st.session_state.get("filter_pagamento_selectbox_index", 0), # Default "Todos"
        key="filter_pagamento_selectbox",
        help="Escolha o método de pagamento."
    )
    st.session_state.filter_pagamento_selectbox_index = opcoes_pagamento_user.index(selected_pagamento_user)


# --- Lógica de Aplicação dos Filtros ---

# Começar com os DataFrames completos
tf_processar = tarifarios_fixos.copy()
ti_processar = tarifarios_indexados.copy()

# 1. Lógica para o filtro de Segmento
if selected_segmento_user != "Ambos":
    segmentos_para_filtrar = []
    if selected_segmento_user == "Residencial":
        segmentos_para_filtrar.extend(["Doméstico", "Doméstico e Não Doméstico"])
    elif selected_segmento_user == "Empresarial":
        segmentos_para_filtrar.extend(["Não Doméstico", "Doméstico e Não Doméstico"])
    
    # Aplicar o filtro
    tf_processar = tf_processar[tf_processar['segmento'].astype(str).str.strip().isin(segmentos_para_filtrar)]
    ti_processar = ti_processar[ti_processar['segmento'].astype(str).str.strip().isin(segmentos_para_filtrar)]

# 2. Lógica para o filtro de Tipo
if selected_tipos:
    tf_processar = tf_processar[tf_processar['tipo'].astype(str).str.strip().isin(selected_tipos)]
    ti_processar = ti_processar[ti_processar['tipo'].astype(str).str.strip().isin(selected_tipos)]

# 3. Lógica para o filtro de Faturação
if selected_faturacao_user != "Todas":
    faturacao_para_filtrar = []
    if selected_faturacao_user == "Fatura eletrónica":
        faturacao_para_filtrar.extend(["Fatura eletrónica", "Fatura eletrónica, Fatura em papel"])
    elif selected_faturacao_user == "Fatura em papel":
        faturacao_para_filtrar.extend(["Fatura eletrónica, Fatura em papel"]) 
    
    # Aplicar o filtro
    tf_processar = tf_processar[tf_processar['faturacao'].astype(str).str.strip().isin(faturacao_para_filtrar)]
    ti_processar = ti_processar[ti_processar['faturacao'].astype(str).str.strip().isin(faturacao_para_filtrar)]

# 4. Lógica para o filtro de Pagamento
if selected_pagamento_user != "Todos":
    pagamento_para_filtrar = []
    if selected_pagamento_user == "Débito Direto":
        pagamento_para_filtrar.extend(["Débito Direto", "Débito Direto, Multibanco", "Débito Direto, Multibanco, Numerário/Payshop/CTT"])
    elif selected_pagamento_user == "Multibanco":
        pagamento_para_filtrar.extend(["Débito Direto, Multibanco", "Débito Direto, Multibanco, Numerário/Payshop/CTT"])
    elif selected_pagamento_user == "Numerário/Payshop/CTT":
        pagamento_para_filtrar.extend(["Débito Direto, Multibanco, Numerário/Payshop/CTT"])
    
    # Aplicar o filtro
    tf_processar = tf_processar[tf_processar['pagamento'].astype(str).str.strip().isin(pagamento_para_filtrar)]
    ti_processar = ti_processar[ti_processar['pagamento'].astype(str).str.strip().isin(pagamento_para_filtrar)]


#st.markdown("---")
# FIM Seletor de Modo de Visualização - NORMAL OU OPÇÃO HORÁRIA
# --- Construir Resumo dos Inputs para Exibição ---
cor_texto_resumo = "#333333"  # Um cinza escuro, bom para fundos claros

resumo_html_parts = [
    f"<div style='background-color: #f9f9f9; border: 1px solid #ddd; padding: 15px; border-radius: 6px; margin-bottom: 25px; color: {cor_texto_resumo};'>"
]
if is_diagram_mode:
    titulo_resumo = "Resumo da Simulação (Valores Originais do Ficheiro):"
else: # Modo Manual
    titulo_resumo = "Resumo da Simulação:"
resumo_html_parts.append(f"<h5 style='margin-top:0; color: {cor_texto_resumo};'>{titulo_resumo}</h5>")
resumo_html_parts.append("<ul style='list-style-type: none; padding-left: 0;'>")

# --- Adicionar detalhes dos filtros numa única linha ---
# 1. Tratar o caso especial do segmento "Ambos"
if selected_segmento_user == "Ambos":
    segmento_para_resumo = "Residencial e Empresarial"
else:
    segmento_para_resumo = selected_segmento_user

# 2. Construir a string combinada
# &nbsp; é o código HTML para um espaço, para dar um espaçamento agradável
linha_filtros = (
    f"<b>Segmento:</b> {segmento_para_resumo} &nbsp;&nbsp;|&nbsp;&nbsp; "
    f"<b>Faturação:</b> {selected_faturacao_user} &nbsp;&nbsp;|&nbsp;&nbsp; "
    f"<b>Pagamento:</b> {selected_pagamento_user}"
)

# 3. Adicionar a linha única ao resumo
resumo_html_parts.append(f"<li style='margin-bottom: 5px;'>{linha_filtros}</li>")
# --- FIM DO BLOCO ---

# 1. Potência contratada + Opção Horária e Ciclo
resumo_html_parts.append(f"<li style='margin-bottom: 5px;'><b>{potencia} kVA</b> em <b>{opcao_horaria}</b></li>")

# 2. Consumo dividido por opção
consumo_detalhe_str = ""
if opcao_horaria.lower() == "simples":
    consumo_detalhe_str = f"Simples: {consumo_simples:.0f} kWh"
elif opcao_horaria.lower().startswith("bi"):
    consumo_detalhe_str = f"Vazio: {consumo_vazio:.0f} kWh, Fora Vazio: {consumo_fora_vazio:.0f} kWh"
elif opcao_horaria.lower().startswith("tri"):
    consumo_detalhe_str = f"Vazio: {consumo_vazio:.0f} kWh, Cheias: {consumo_cheias:.0f} kWh, Ponta: {consumo_ponta:.0f} kWh"
resumo_html_parts.append(f"<li style='margin-bottom: 5px;'><b>Consumos ({consumo:.0f} kWh Total):</b> {consumo_detalhe_str}</li>")

# 3. Datas e Dias Faturados
# 'dias_default_calculado' e 'dias' já foram calculados
# 'dias_manual_input_val' é o valor do widget st.number_input para dias manuais
dias_manual_valor_do_input = st.session_state.get('dias_manual_input_key', dias_default_calculado)
usou_dias_manuais_efetivamente = False
if pd.notna(dias_manual_valor_do_input) and dias_manual_valor_do_input > 0 and \
   int(dias_manual_valor_do_input) != dias_default_calculado:
    usou_dias_manuais_efetivamente = True

if usou_dias_manuais_efetivamente:
    resumo_html_parts.append(f"<li style='margin-bottom: 5px;'><b>Período:</b> {dias} dias (definido manualmente)</li>")
else:
    resumo_html_parts.append(f"<li style='margin-bottom: 5px;'><b>Período:</b> De {data_inicio.strftime('%d/%m/%Y')} a {data_fim.strftime('%d/%m/%Y')} ({dias} dias)</li>")

# 4. Valores OMIE da opção escolhida, com a referência
omie_valores_str_parts = []
if opcao_horaria.lower() == "simples":
    val_s = st.session_state.get('omie_s_input_field', round(omie_medios_calculados.get('S',0), 2))
    omie_valores_str_parts.append(f"Simples: {val_s:.2f} €/MWh")
elif opcao_horaria.lower().startswith("bi"):
    val_v = st.session_state.get('omie_v_input_field', round(omie_medios_calculados.get('V',0), 2))
    val_f = st.session_state.get('omie_f_input_field', round(omie_medios_calculados.get('F',0), 2))
    omie_valores_str_parts.append(f"Vazio: {val_v:.2f} €/MWh")
    omie_valores_str_parts.append(f"Fora Vazio: {val_f:.2f} €/MWh")
elif opcao_horaria.lower().startswith("tri"):
    val_v = st.session_state.get('omie_v_input_field', round(omie_medios_calculados.get('V',0), 2))
    val_c = st.session_state.get('omie_c_input_field', round(omie_medios_calculados.get('C',0), 2))
    val_p = st.session_state.get('omie_p_input_field', round(omie_medios_calculados.get('P',0), 2))
    omie_valores_str_parts.append(f"Vazio: {val_v:.2f} €/MWh")
    omie_valores_str_parts.append(f"Cheias: {val_c:.2f} €/MWh")
    omie_valores_str_parts.append(f"Ponta: {val_p:.2f} €/MWh")

if omie_valores_str_parts: # Só mostra a secção OMIE se houver valores a exibir
    resumo_html_parts.append(f"<li style='margin-bottom: 5px;'><b>OMIE {nota_omie}:</b> {', '.join(omie_valores_str_parts)}</li>")

# 5. Perfil de consumo utilizado
perfil_consumo_calculado_str = calc.obter_perfil(consumo, dias, potencia) # Chamar a sua função
# Formatar para uma apresentação mais amigável
texto_perfil_apresentacao = perfil_consumo_calculado_str.replace("perfil_", "Perfil ").upper() # Ex: "Perfil A"
resumo_html_parts.append(f"<li style='margin-bottom: 5px;'><b>Perfil de Consumo:</b> {texto_perfil_apresentacao}</li>")

# 6. Tarifa Social (se ativa)
if tarifa_social:
    resumo_html_parts.append(f"<li style='margin-bottom: 5px; color: red;'><b>Benefício Aplicado:</b> Tarifa Social</li>")

# 7. Família Numerosa (se ativa)
if familia_numerosa:
    resumo_html_parts.append(f"<li style='margin-bottom: 5px; color: red;'><b>Benefício Aplicado:</b> Família Numerosa</li>")

resumo_html_parts.append("</ul>")
resumo_html_parts.append("</div>")
html_resumo_final = "".join(resumo_html_parts)

# Exibir o resumo
st.markdown(html_resumo_final, unsafe_allow_html=True)

# --- Resumo Dinâmico Pós-Autoconsumo ---
# Este bloco lê os dados guardados no session_state e pode ser colocado onde quiser na página.
if st.session_state.get("chk_autoconsumo_ativo", False) and \
   st.session_state.get("escolha_consumo_calculo") == "Consumo da Rede (após autoconsumo)" and \
   'consumos_finais_para_resumo' in st.session_state:

    # Ler os dados diretamente do "armazém" do session_state
    dados_consumo = st.session_state['consumos_finais_para_resumo']
    
    resumo_liquido_html_parts = [
        f"<div style='background-color: #e2efda; border: 1px solid #a9d08e; padding: 15px; border-radius: 6px; margin-bottom: 10px; color: #385723;'>",
        f"<h5 style='margin-top:0; color: #385723;'>Resumo da Simulação (Valores Após Autoconsumo):</h5>",
        "<ul style='list-style-type: none; padding-left: 0;'>",
    ]
    
    consumo_detalhe_liquido_str = ""
    if opcao_horaria.lower() == "simples":
        consumo_detalhe_liquido_str = f"Simples: {dados_consumo['simples']:.0f} kWh"
    elif opcao_horaria.lower().startswith("bi"):
        consumo_detalhe_liquido_str = f"Vazio: {dados_consumo['vazio']:.0f} kWh, Fora Vazio: {dados_consumo['fora_vazio']:.0f} kWh"
    elif opcao_horaria.lower().startswith("tri"):
        consumo_detalhe_liquido_str = f"Vazio: {dados_consumo['vazio']:.0f} kWh, Cheias: {dados_consumo['cheias']:.0f} kWh, Ponta: {dados_consumo['ponta']:.0f} kWh"
    
    resumo_liquido_html_parts.append(f"<li style='margin-bottom: 5px;'><b>Consumos da Rede ({dados_consumo['total']:.0f} kWh Total):</b> {consumo_detalhe_liquido_str}</li>")
    
    resumo_liquido_html_parts.append("</ul></div>")
    st.markdown("".join(resumo_liquido_html_parts), unsafe_allow_html=True)

# --- BLOCO PARA MOSTRAR A TABELA DE ANÁLISE SEPARADAMENTE ---
# Só mostra esta tabela se AMBAS as condições forem verdadeiras
if modo_de_comparacao_ativo and st.session_state.get('dados_completos_ficheiro') is not None:
    
    # --- LÓGICA PARA TÍTULO DINÂMICO ---
    titulo_tabela_analise = "Análise de Consumos e Médias OMIE (Valores Originais)"
    if st.session_state.get("chk_autoconsumo_ativo", False) and st.session_state.get("escolha_consumo_calculo") == "Consumo da Rede (após autoconsumo)":
        titulo_tabela_analise = "Análise de Consumos e Médias OMIE (Valores Após Autoconsumo)"
    
    st.subheader(titulo_tabela_analise)
    
    tabela_analise_html = criar_tabela_analise_completa_html(
        consumos_para_custos,
        omie_medios_calculados_para_todos_ciclos
    )
    st.markdown(tabela_analise_html, unsafe_allow_html=True)

resultados_list.clear() # Limpar a lista de resultados no início de cada execução

if modo_de_comparacao_ativo:
    # --- INÍCIO DO BLOCO PARA O MODO DE COMPARAÇÃO DE OPÇÕES HORÁRIAS ---
    st.subheader("📊 Tiago Felícia - Comparação de Custos entre Opções Tarifárias")
    st.markdown("➡️ [**Exportar Tabela de Comparação para Excel**](#exportar-excel-comparacao)")
    
    with st.spinner("A gerar a tabela de comparação... Este processo pode demorar um pouco."):

    #Vários JS, alguns repetidos
        #CORES PARA TARIFÁRIOS INDEXADOS:
        cor_fundo_indexado_media_css = "#FFE699"
        cor_texto_indexado_media_css = "black"
        cor_fundo_indexado_dinamico_css = "#4D79BC"  
        cor_texto_indexado_dinamico_css = "white"
        cor_fundo_indexado_diagrama_css = "#BDD7EE"  
        cor_texto_indexado_diagrama_css = "black"

        cell_style_nome_tarifario_js = JsCode(f"""
        function(params) {{
            // Estilo base aplicado a todas as células desta coluna
            let styleToApply = {{ 
                textAlign: 'center',
                borderRadius: '11px',  // O teu borderRadius desejado
                padding: '10px 10px'   // O teu padding desejado
                // Podes adicionar um backgroundColor default para células não especiais aqui, se quiseres
                // backgroundColor: '#f0f0f0' // Exemplo para tarifários fixos
            }};                                  

            if (params.data) {{
                const nomeExibir = params.data.NomeParaExibir;
                const tipoTarifario = params.data.Tipo;

                // VERIFICA SE O NOME COMEÇA COM "O Meu Tarifário"
                if (typeof nomeExibir === 'string' && nomeExibir.startsWith('O Meu Tarifário')) {{
                    styleToApply.backgroundColor = 'red';
                    styleToApply.color = 'white';
                    styleToApply.fontWeight = 'bold';
                }} else if (typeof nomeExibir === 'string' && nomeExibir.startsWith('Tarifário Personalizado')) {{
                    styleToApply.backgroundColor = '#92D050';
                    styleToApply.color = 'white';
                    styleToApply.fontWeight = 'bold';
                }} else if (tipoTarifario === 'Indexado Média') {{
                    styleToApply.backgroundColor = '{cor_fundo_indexado_media_css}';
                    styleToApply.color = '{cor_texto_indexado_media_css}';                
                }} else if (tipoTarifario === 'Indexado quarto-horário') {{
                    styleToApply.backgroundColor = '{cor_fundo_indexado_dinamico_css}';
                    styleToApply.color = '{cor_texto_indexado_dinamico_css}';
                }} else if (tipoTarifario === 'Indexado quarto-horário (Diagrama)') {{
                    styleToApply.backgroundColor = '{cor_fundo_indexado_diagrama_css}';
                    styleToApply.color = '{cor_texto_indexado_diagrama_css}';
                }} else if (tipoTarifario === 'Fixo') {{
                    styleToApply.backgroundColor = '#f0f0f0'; // Cor cinza claro
                    styleToApply.color = '#333333';    // Cor de texto escura
                }} else {{
                    // Para tarifários fixos ou outros tipos não explicitamente coloridos acima.
                    // Eles já terão o textAlign, borderRadius e padding do styleToApply.
                    // Se quiseres um fundo específico para eles diferente do default do styleToApply, define aqui.
                    // Ex: styleToApply.backgroundColor = '#e9ecef'; // Uma cor neutra para fixos
                }}
                return styleToApply;
            }}
            return styleToApply; 
        }}
        """)

        #tooltip Nome Tarifario
        tooltip_nome_tarifario_getter_js = JsCode("""
        function(params) {
            if (!params.data) { 
                return params.value || ''; 
            }

            const nomeExibir = params.data.NomeParaExibir || '';
            const notas = params.data.info_notas || ''; 

            let tooltipHtmlParts = [];

            if (nomeExibir) {
                tooltipHtmlParts.push("<strong>" + nomeExibir + "</strong>");
            }

            if (notas) {
                const notasHtml = notas.replace(/\\n/g, ' ').replace(/\n/g, ' ');
                // Usando aspas simples para atributos de estilo HTML para simplificar
                tooltipHtmlParts.push("<small style='display: block; margin-top: 5px;'><i>" + notasHtml + "</i></small>");
            }
        
            if (tooltipHtmlParts.length > 0) {
                // Se ambos nomeExibir e notas existem, queremos uma quebra de linha entre eles no tooltip.
                // Se join(''), eles ficam lado a lado. Se join('<br>'), ficam em linhas separadas.
                // Dado que a nota tem display:block, o join('') deve funcionar para colocá-los em "blocos" separados.
                return tooltipHtmlParts.join(''); // Para agora, vamos juntar diretamente.
                                                    // Se quiser uma quebra de linha explícita entre o nome e as notas,
                                                    // e ambos existirem, pode usar:
                                                    // return tooltipHtmlParts.join('<br style="margin-bottom:5px">');
            }
        
            return ''; 
        }
                """)
            
        # Custom_Tooltip
        custom_tooltip_component_js = JsCode("""
            class CustomTooltip {
                init(params) {
                    // params.value é a string que o seu tooltipValueGetter retorna
                    this.eGui = document.createElement('div');
                    // Para permitir HTML, definimos o innerHTML
                    // É importante que a string de params.value seja HTML seguro se vier de inputs do utilizador,
                    // mas no seu caso, está a construí-lo programaticamente.
                    this.eGui.innerHTML = params.value; 

                    // Aplicar algum estilo básico para o tooltip se desejar
                    this.eGui.style.backgroundColor = 'white'; // Ou outra cor de fundo
                    this.eGui.style.color = 'black';           // Cor do texto
                    this.eGui.style.border = '1px solid #ccc'; // Borda mais suave
                    this.eGui.style.padding = '10px';           // Mais padding
                    this.eGui.style.borderRadius = '11px';      // Cantos arredondados
                    this.eGui.style.boxShadow = '0 2px 5px rgba(0,0,0,0.15)'; // Sombra suave
                    this.eGui.style.maxWidth = '400px';        // Largura máxima
                    this.eGui.style.fontSize = '1.1em';        // Tamanho da fonte
                    this.eGui.style.fontFamily = 'Arial, sans-serif'; // Tipo de fonte                             
                    this.eGui.style.whiteSpace = 'normal';     // Para quebra de linha
                }

                getGui() {
                    return this.eGui;
                }
            }
        """)


        custom_css = {
            ".ag-header-cell-label": {
                "justify-content": "center !important",
                "text-align": "center !important",
                "font-size": "14px !important",
                "font-weight": "bold !important"
            },
            ".ag-center-header": {
                "justify-content": "center !important",
                "text-align": "center !important",
                "font-size": "14px !important"
            },
            ".ag-cell": {
                "font-size": "14px !important"
            },
            ".ag-center-cols-clip": {"justify-content": "center !important", "text-align": "center !important"}
        }

            # --- 1. DEFINIR O JsCode PARA LINK E TOOLTIP ---
        link_tooltip_renderer_js = JsCode("""
        class LinkTooltipRenderer {
            init(params) {
                this.eGui = document.createElement('div');
                let displayText = params.value; // Valor da célula (NomeParaExibir)
                let url = params.data.LinkAdesao; // Acede ao valor da coluna LinkAdesao da mesma linha

                if (url && typeof url === 'string' && url.toLowerCase().startsWith('http')) {
                    // HTML para o link clicável
                    // O atributo 'title' (tooltip) mostrará "Aderir/Saber mais: [URL]"
                    // O texto visível do link será o 'displayText' (NomeParaExibir)
                    this.eGui.innerHTML = `<a href="${url}" target="_blank" title="Aderir/Saber mais: ${url}" style="text-decoration: underline; color: inherit;">${displayText}</a>`;
                } else {
                    // Se não houver URL válido, apenas mostra o displayText com o próprio displayText como tooltip.
                    this.eGui.innerHTML = `<span title="${displayText}">${displayText}</span>`;
                }
            }
            getGui() { return this.eGui; }
        }
        """) # <--- FIM DA DEFINIÇÃO DE link_tooltip_renderer_js

        resultados_comparacao_list = [] # Usar uma lista dedicada para este modo

        # 1. Determinar se um ficheiro válido foi carregado
        ficheiro_foi_carregado = 'dados_completos_ficheiro' in st.session_state and st.session_state.dados_completos_ficheiro is not None

        # 2. Determinar colunas de destino com base na lógica
        opcoes_destino_db_nomes_comp, colunas_aggrid_custo_comp, coluna_ordenacao_aggrid_comp = \
            calc.determinar_opcoes_horarias_destino_e_ordenacao(
                opcao_horaria,
                potencia,
                opcoes_horarias_existentes,
                ficheiro_foi_carregado
            )
        
        if not opcoes_destino_db_nomes_comp:
            st.warning("Não há opções horárias válidas para comparação com os inputs selecionados.")
        else:
            # --- 3. Preparar os dicionários de consumo para cada coluna da tabela ---
            consumos_repartidos_finais_por_oh_comp = {}

            if ficheiro_foi_carregado:
                # CASO 1: FICHEIRO CARREGADO - Usar os consumos reais agregados para cada ciclo
                # Usamos 'consumos_para_custos'
                mapa_nomes_ciclos = {
                    "Simples": "Simples", "Bi-horário - Ciclo Diário": "BD",
                    "Bi-horário - Ciclo Semanal": "BS", "Tri-horário - Ciclo Diário": "TD",
                    "Tri-horário - Ciclo Semanal": "TS",
                    "Tri-horário > 20.7 kVA - Ciclo Diário": "TD",
                    "Tri-horário > 20.7 kVA - Ciclo Semanal": "TS"
                }
                for nome_longo, nome_curto in mapa_nomes_ciclos.items():
                    if nome_longo in opcoes_destino_db_nomes_comp:
                        if nome_curto == "Simples":
                            consumos_repartidos_finais_por_oh_comp[nome_longo] = {'S': consumos_para_custos.get('Simples', 0)}
                        else:
                            consumos_repartidos_finais_por_oh_comp[nome_longo] = consumos_para_custos.get(nome_curto, {})
            else:
                # CASO 2: INPUT MANUAL - Usar a função de estimativa como antes
                consumos_input_atuais_dict_comp = {}
                if opcao_horaria.lower() == "simples":
                    consumos_input_atuais_dict_comp['S'] = consumo_simples
                elif opcao_horaria.lower().startswith("bi-horário"):
                    consumos_input_atuais_dict_comp['V'] = consumo_vazio
                    consumos_input_atuais_dict_comp['F'] = consumo_fora_vazio
                elif opcao_horaria.lower().startswith("tri-horário"):
                    consumos_input_atuais_dict_comp['V'] = consumo_vazio
                    consumos_input_atuais_dict_comp['C'] = consumo_cheias
                    consumos_input_atuais_dict_comp['P'] = consumo_ponta

                consumos_repartidos_finais_por_oh_comp = calc.preparar_consumos_para_cada_opcao_destino(
                    opcao_horaria,
                    consumos_input_atuais_dict_comp,
                    opcoes_destino_db_nomes_comp
                )

            # Definir a variável que faltava, reunindo os valores OMIE dos inputs do utilizador.
            # Esta variável é necessária para a função `calcular_detalhes_custo_tarifario_indexado`.
            todos_omie_inputs_utilizador_comp_comparacao = {
                'S': st.session_state.get('omie_s_input_field', round(omie_medios_calculados.get('S', 0), 2)),
                'V': st.session_state.get('omie_v_input_field', round(omie_medios_calculados.get('V', 0), 2)),
                'F': st.session_state.get('omie_f_input_field', round(omie_medios_calculados.get('F', 0), 2)),
                'C': st.session_state.get('omie_c_input_field', round(omie_medios_calculados.get('C', 0), 2)),
                'P': st.session_state.get('omie_p_input_field', round(omie_medios_calculados.get('P', 0), 2))
            }

            # --- 4. Loop principal para calcular os custos e construir a tabela ---
            resultados_comparacao_list = []
            nomes_tarifarios_unicos_para_comparacao = []
            if not tf_processar.empty: # USAR tf_processar
                temp_fixos_unicos = tf_processar[['nome', 'comercializador', 'tipo', 'site_adesao', 'notas']].drop_duplicates(subset=['nome', 'comercializador'])
                for _, row_fixo in temp_fixos_unicos.iterrows():
                    nomes_tarifarios_unicos_para_comparacao.append(dict(row_fixo))

            if comparar_indexados and not ti_processar.empty: # USAR ti_processar
                temp_index_unicos = ti_processar[['nome', 'comercializador', 'tipo', 'site_adesao', 'notas', 'formula_calculo']].drop_duplicates(subset=['nome', 'comercializador'])
                for _, row_idx in temp_index_unicos.iterrows():
                    if not any(d['nome'] == row_idx['nome'] and d['comercializador'] == row_idx['comercializador'] for d in nomes_tarifarios_unicos_para_comparacao):
                        nomes_tarifarios_unicos_para_comparacao.append(dict(row_idx))

            if meu_tarifario_ativo and 'meu_tarifario_calculado' in st.session_state:
                omt_data = st.session_state['meu_tarifario_calculado']
                # Adicionar informação para identificar "O Meu Tarifário" e a sua opção horária original
                nomes_tarifarios_unicos_para_comparacao.append({
                    'nome': omt_data.get('NomeParaExibir', "O Meu Tarifário"),
                    'comercializador': omt_data.get('Comercializador', "-"),
                    'tipo': omt_data.get('Tipo', "Pessoal"),
                    'site_adesao': omt_data.get('LinkAdesao', '-'),
                    'notas': omt_data.get('info_notas', ''),
                    'is_meu_tarifario': True, # Flag para identificar "O Meu Tarifário"
                    'opcao_horaria_original_meu_tarifario': omt_data.get('opcao_horaria_calculada')
                })
            #st.subheader("DEBUG: DADOS DE MERCADO PARA COMPARAÇÃO") # Título para encontrar fácil
            #st.write("Dicionário de OMIEs Médios (todos os ciclos):")
            #st.json(omie_medios_calculados_para_todos_ciclos)

            #st.write("Dicionário de Perdas Médias (todos os ciclos):")
            #st.json(perdas_medias)

            #st.markdown("---") # Separador    
            for info_tar_base in nomes_tarifarios_unicos_para_comparacao:
                nome_t_comp = info_tar_base['nome']
                comerc_t_comp = info_tar_base['comercializador']
                tipo_t_comp = str(info_tar_base['tipo'])
                link_t_comp = info_tar_base.get('site_adesao', '-')
                notas_t_comp = info_tar_base.get('notas', '')
                formula_calculo_idx_comp = str(info_tar_base.get('formula_calculo', '')) # Para indexados

                nome_final_para_linha = nome_t_comp
                # Se a fórmula contiver 'BTN', é um quarto-horário que precisa do sufixo
                if 'BTN' in formula_calculo_idx_comp:
                    nome_final_para_linha = f"{nome_t_comp} - Perfil"

                linha_para_aggrid = {
                    'NomeParaExibir': nome_final_para_linha,
                    'Comercializador': comerc_t_comp,
                    'Tipo': tipo_t_comp,
                    'LinkAdesao': link_t_comp,
                    'info_notas': notas_t_comp
                }
                teve_pelo_menos_um_calculo_nesta_linha = False

                for oh_destino_db_nome in opcoes_destino_db_nomes_comp:
                    nome_coluna_aggrid_para_este_oh = f"Total {oh_destino_db_nome} (€)"
                    linha_para_aggrid[nome_coluna_aggrid_para_este_oh] = None # Inicializa a coluna de custo

                    consumos_para_calculo_nesta_oh = consumos_repartidos_finais_por_oh_comp.get(oh_destino_db_nome)
                    if not consumos_para_calculo_nesta_oh or sum(v for v in consumos_para_calculo_nesta_oh.values() if v is not None) == 0:
                        continue # Pula se não houver consumos para esta opção de destino
                    
                    dados_tarifario_especifico_para_calculo = None # DataFrame da linha específica do tarifário
                    
                    # Lógica para "O Meu Tarifário" na Tabela Comparativa
                    if info_tar_base.get('is_meu_tarifario'):
                        if 'meu_tarifario_calculado' in st.session_state:
                            if info_tar_base.get('opcao_horaria_original_meu_tarifario') == oh_destino_db_nome:
                                omt_data_calc = st.session_state['meu_tarifario_calculado']
                                linha_para_aggrid[nome_coluna_aggrid_para_este_oh] = omt_data_calc.get('Total (€)')
                                teve_pelo_menos_um_calculo_nesta_linha = True
                            # else: O custo para outras opções de destino para "O Meu Tarifário" não é calculado aqui.
                        continue # Vai para o próximo oh_destino_db_nome

                    # Lógica para tarifários do Excel
                    if tipo_t_comp == "Fixo":
                        df_match = tarifarios_fixos[
                            (tarifarios_fixos['nome'] == nome_t_comp) &
                            (tarifarios_fixos['comercializador'] == comerc_t_comp) &
                            (tarifarios_fixos['opcao_horaria_e_ciclo'] == oh_destino_db_nome) &
                            (tarifarios_fixos['potencia_kva'] == potencia)
                        ]
                        
                        if not df_match.empty:
                            dados_tarifario_especifico_para_calculo = df_match.iloc[0]
                    
                    elif tipo_t_comp.startswith("Indexado"):
                        df_match_idx = tarifarios_indexados[
                            (tarifarios_indexados['nome'] == nome_t_comp) &
                            (tarifarios_indexados['comercializador'] == comerc_t_comp) &
                            (tarifarios_indexados['opcao_horaria_e_ciclo'] == oh_destino_db_nome) &
                            (tarifarios_indexados['potencia_kva'] == potencia)
                        ]

                        if not df_match_idx.empty:
                            dados_tarifario_especifico_para_calculo = df_match_idx.iloc[0]

                    if dados_tarifario_especifico_para_calculo is not None:
                        resultado_celula = None
                        if tipo_t_comp == "Fixo":
                            resultado_celula = calc.calcular_detalhes_custo_tarifario_fixo(
                                dados_tarifario_especifico_para_calculo, oh_destino_db_nome,
                                consumos_para_calculo_nesta_oh, potencia, dias, tarifa_social, familia_numerosa,
                                valor_dgeg_user, valor_cav_user, incluir_quota_acp, desconto_continente,
                                CONSTANTES, dias_mes, mes, ano_atual, data_inicio, data_fim, FINANCIAMENTO_TSE_VAL, VALOR_QUOTA_ACP_MENSAL
                            )
                        elif tipo_t_comp.startswith("Indexado"):
                            resultado_celula = calc.calcular_detalhes_custo_tarifario_indexado(
                                dados_tarifario_especifico_para_calculo, oh_destino_db_nome, opcao_horaria,
                                consumos_para_calculo_nesta_oh, potencia, dias, tarifa_social, familia_numerosa,
                                valor_dgeg_user, valor_cav_user, CONSTANTES,
                                df_omie_ajustado,
                                perdas_medias,
                                todos_omie_inputs_utilizador_comp_comparacao,
                                omie_medios_calculados_para_todos_ciclos,
                                omie_medio_simples_real_kwh, # OMIE real simples para Luzigas
                                dias_mes, mes, ano_atual, data_inicio, data_fim, FINANCIAMENTO_TSE_VAL
                            )
                        
                        if resultado_celula and pd.notna(resultado_celula.get('Total (€)')):
                            linha_para_aggrid[nome_coluna_aggrid_para_este_oh] = resultado_celula['Total (€)']
                            # Guardar tooltips para cada célula da tabela comparativa
                            linha_para_aggrid[f'Tooltip_{nome_coluna_aggrid_para_este_oh}'] = resultado_celula.get('componentes_tooltip_custo_total_dict')
                            teve_pelo_menos_um_calculo_nesta_linha = True
                
                if teve_pelo_menos_um_calculo_nesta_linha:
                    resultados_comparacao_list.append(linha_para_aggrid)

        # ### Adicionar a linha do Tarifário Personalizado à tabela de comparação ###
        if st.session_state.get('dados_tarifario_personalizado', {}).get('ativo'):
            dados_pers = st.session_state['dados_tarifario_personalizado']
            
            linha_pers_comparacao = {'NomeParaExibir': "Tarifário Personalizado", 'Tipo': "Pessoal", 'Comercializador': "Personalizado"}
            
            # Loop pelas colunas de destino da tabela de comparação (Simples, Bi-Diário, etc.)
            for oh_destino in opcoes_destino_db_nomes_comp:
                nome_coluna = f"Total {oh_destino} (€)"
                precos_energia_pers = {}
                preco_potencia_pers = 0.0
                consumos_desta_coluna = consumos_repartidos_finais_por_oh_comp.get(oh_destino)
                
                if not consumos_desta_coluna:
                    continue

                # Escolher os preços personalizados corretos para esta coluna
                if "Simples" in oh_destino:
                    precos_energia_pers = {'S': dados_pers['precos_s']['energia']}
                    preco_potencia_pers = dados_pers['precos_s']['potencia']
                elif "Bi-horário" in oh_destino:
                    precos_energia_pers = {'V': dados_pers['precos_bi']['vazio'], 'F': dados_pers['precos_bi']['fora_vazio']}
                    preco_potencia_pers = dados_pers['precos_bi']['potencia']
                elif "Tri-horário" in oh_destino:
                    precos_energia_pers = {'V': dados_pers['precos_tri']['vazio'], 'C': dados_pers['precos_tri']['cheias'], 'P': dados_pers['precos_tri']['ponta']}
                    preco_potencia_pers = dados_pers['precos_tri']['potencia']

                # Calcular o custo para esta célula se houver preços definidos
                if preco_potencia_pers > 0 or any(p > 0 for p in precos_energia_pers.values()):
                    resultado_celula = calc.calcular_custo_personalizado(
                        precos_energia_pers, preco_potencia_pers, consumos_desta_coluna, dados_pers['flags'], CONSTANTES, FINANCIAMENTO_TSE_VAL,
                        dias=dias, potencia=potencia, tarifa_social=tarifa_social, familia_numerosa=familia_numerosa,
                        valor_dgeg_user=valor_dgeg_user, valor_cav_user=valor_cav_user, opcao_horaria_ref=oh_destino, 
                    )
                    linha_pers_comparacao[f"Total {oh_destino} (€)"] = round(resultado_celula['Total (€)'], 2)
                    linha_pers_comparacao[f"Tooltip_Total {oh_destino} (€)"] = resultado_celula.get('componentes_tooltip_custo_total_dict')


            resultados_comparacao_list.append(linha_pers_comparacao)

        # --- Adicionar Cálculos de Diagrama de Carga à Tabela de Comparação ---
        if ficheiro_foi_carregado and not df_consumos_filtrado.empty and comparar_indexados:
        
            # Filtra apenas os tarifários indexados que são quarto-horários (BTN) e correspondem à potência
            tarifarios_para_calculo_real = ti_processar[
                (ti_processar['formula_calculo'].str.contains('BTN', na=False)) &
                (ti_processar['potencia_kva'] == potencia)
            ].copy()

            # Agrupar por nome do tarifário para fazer um único cálculo por tarifário
            for nome_tarifario_agrupado, grupo in tarifarios_para_calculo_real.groupby('nome'):
            
                info_base_tarifario = grupo.iloc[0]
            
                # Criar uma nova linha para este resultado de "Quarto-horário - Diagrama"
                linha_aggrid_diagrama = {
                    'NomeParaExibir': f"{nome_tarifario_agrupado} - Diagrama",
                    'Comercializador': info_base_tarifario.get('comercializador'),
                    'Tipo': f"Indexado quarto-horário (Diagrama)", # Tipo explícito
                    'LinkAdesao': info_base_tarifario.get('site_adesao'),
                    'info_notas': info_base_tarifario.get('notas')
                }
            
                # Inicializar todas as colunas de custo com None
                for col_custo in colunas_aggrid_custo_comp:
                    linha_aggrid_diagrama[col_custo] = None
            
                # Iterar sobre as opções horárias DESTE tarifário (ex: um BTN pode ter entrada para Simples, BD e BS)
                for _, tarifario_real_especifico in grupo.iterrows():
                    oh_original_tarifario = tarifario_real_especifico['opcao_horaria_e_ciclo']
                    coluna_custo_correspondente = f"Total {oh_original_tarifario} (€)"
                
                    # Calcular o custo apenas se esta coluna de opção horária estiver na tabela de comparação
                    if coluna_custo_correspondente in colunas_aggrid_custo_comp:
                    
                        resultado_real_dict = calc.calcular_custo_completo_diagrama_carga(
                            tarifario_real_especifico, 
                            df_consumos_a_utilizar,
                            OMIE_PERDAS_CICLOS,
                            CONSTANTES,
                            dias, 
                            potencia, 
                            familia_numerosa, 
                            tarifa_social,
                            valor_dgeg_user, 
                            valor_cav_user, 
                            mes, 
                            ano_atual,
                            incluir_quota_acp, 
                            desconto_continente,
                            FINANCIAMENTO_TSE_VAL, 
                            VALOR_QUOTA_ACP_MENSAL
                        )

                        if resultado_real_dict:
                            linha_aggrid_diagrama[coluna_custo_correspondente] = resultado_real_dict.get('Total (€)')
                            # Adicionar dados para o tooltip, se necessário
                            tooltip_key = f"Tooltip_{coluna_custo_correspondente}"
                            linha_aggrid_diagrama[tooltip_key] = resultado_real_dict
            
                # Adicionar a linha completa à lista de resultados
                resultados_comparacao_list.append(linha_aggrid_diagrama)

        df_resultados_comparacao_aggrid = pd.DataFrame(resultados_comparacao_list)
        
    if not df_resultados_comparacao_aggrid.empty:
        if 'coluna_ordenacao_aggrid_comp' in locals() and \
            coluna_ordenacao_aggrid_comp and \
           coluna_ordenacao_aggrid_comp in df_resultados_comparacao_aggrid.columns:
            df_resultados_comparacao_aggrid = df_resultados_comparacao_aggrid.sort_values(
                by=coluna_ordenacao_aggrid_comp,
                ascending=True,
                na_position='last'
            ).reset_index(drop=True)
        
        # --- CONFIGURAÇÃO DO AGGRID PARA O MODO DE COMPARAÇÃO ---
        gb_comp = GridOptionsBuilder.from_dataframe(df_resultados_comparacao_aggrid)
        gb_comp.configure_default_column(
            sortable=True, resizable=True, editable=False, wrapText=True, autoHeight=True,
            wrapHeaderText=True, autoHeaderHeight=True
        )
            
        # Coluna NomeParaExibir com Link e Tooltip de Notas
        gb_comp.configure_column(
            field='NomeParaExibir', 
            headerName='Tarifário', 
            minWidth=200, flex=2.5, 
            filter='agTextColumnFilter', 
            cellRenderer=link_tooltip_renderer_js,
            cellStyle=cell_style_nome_tarifario_js,
            tooltipValueGetter=tooltip_nome_tarifario_getter_js, 
            tooltipComponent=custom_tooltip_component_js
        )
            
        # Ocultar colunas que agora estão integradas no NomeParaExibir ou não são necessárias
        if 'LinkAdesao' in df_resultados_comparacao_aggrid.columns:
             gb_comp.configure_column('LinkAdesao', hide=True)
        if 'info_notas' in df_resultados_comparacao_aggrid.columns:
             gb_comp.configure_column('info_notas', hide=True)
        if 'Comercializador' in df_resultados_comparacao_aggrid.columns:
            gb_comp.configure_column('Comercializador', hide=True)
        if 'Tipo' in df_resultados_comparacao_aggrid.columns:
            gb_comp.configure_column('Tipo', hide=True)

        # Colunas de Comercializador e Tipo (mais simples)
        if 'Comercializador' in df_resultados_comparacao_aggrid.columns:
            gb_comp.configure_column('Comercializador', headerName='Comercializador', minWidth=120, flex=1, cellStyle={'textAlign': 'center'})
        if 'Tipo' in df_resultados_comparacao_aggrid.columns:
            gb_comp.configure_column('Tipo', headerName='Tipo', minWidth=100, flex=1, cellStyle={'textAlign': 'center'})
        
        # Colunas de Custo (ex: "Total Simples (€)")
        if 'colunas_aggrid_custo_comp' in locals(): # Certifica-se de que a variável existe
            # Recriar min_max_data_comparativa_js e cell_style_cores_comparativa_js com base em df_resultados_comparacao_aggrid
            min_max_data_comparativa_js_local = {}
            for col_custo_comp_inner in colunas_aggrid_custo_comp:
                if col_custo_comp_inner in df_resultados_comparacao_aggrid:
                    series_comp_inner = pd.to_numeric(df_resultados_comparacao_aggrid[col_custo_comp_inner], errors='coerce').dropna()
                    if not series_comp_inner.empty:
                        min_max_data_comparativa_js_local[col_custo_comp_inner] = {'min': series_comp_inner.min(), 'max': series_comp_inner.max()}
                    else:
                        min_max_data_comparativa_js_local[col_custo_comp_inner] = {'min': 0, 'max': 0}
            min_max_data_comparativa_json_string_local = json.dumps(min_max_data_comparativa_js_local)

            cell_style_cores_comparativa_js_local = JsCode(f"""
            function(params) {{
                const minMaxConfig = {min_max_data_comparativa_json_string_local}; 
                let style = {{ textAlign: 'center', borderRadius: '11px', padding: '10px 10px' }};
                if (params.value == null || isNaN(parseFloat(params.value)) || !minMaxConfig[params.colDef.field]) return style;
                const min_val = minMaxConfig[params.colDef.field].min;
                const max_val = minMaxConfig[params.colDef.field].max;
                if (max_val === min_val) {{ style.backgroundColor = 'lightgrey'; return style; }}
                const normalized_value = Math.max(0, Math.min(1, (parseFloat(params.value) - min_val) / (max_val - min_val)));
                const cL={{r:90,g:138,b:198}},cM={{r:255,g:255,b:255}},cH={{r:247,g:150,b:70}}; let r,g,b;
                if(normalized_value < 0.5){{const t=normalized_value/0.5;r=Math.round(cL.r*(1-t)+cM.r*t);g=Math.round(cL.g*(1-t)+cM.g*t);b=Math.round(cL.b*(1-t)+cM.b*t);}}
                else{{const t=(normalized_value-0.5)/0.5;r=Math.round(cM.r*(1-t)+cH.r*t);g=Math.round(cM.g*(1-t)+cH.g*t);b=Math.round(cM.b*(1-t)+cH.b*t);}}
                style.backgroundColor=`rgb(${{r}},${{g}},${{b}})`;
                if((r*0.299+g*0.587+b*0.114)<140)style.color='white';else style.color='black';
                return style;
            }}
            """)
                
            # Tooltip para colunas de custo na tabela comparativa
            tooltip_custo_total_comparativa_js = JsCode("""
            function(params) {
                if (!params.data || params.value == null) { return String(params.value); }
                const colField = params.colDef.field; // Ex: "Total Simples (€)"
                const tooltipDataKey = "Tooltip_" + colField; 
                const tooltipData = params.data[tooltipDataKey];

                const formatCurrency = (num) => (typeof num === 'number' && !isNaN(num)) ? num.toFixed(2) : 'N/A';
                const formatUnitPrice = (num) => (typeof num === 'number' && !isNaN(num)) ? num.toFixed(4) : 'N/A'; // 4 casas para €/kWh ou €/dia

                let tooltipParts = [];
                    
                if (!tooltipData) { 
                    tooltipParts.push("<i>" + (params.data.NomeParaExibir || "Tarifário") + " - " + colField.replace("Total ", "").replace(" (€)", "") + "</i>");
                    tooltipParts.push("<b>Custo Total c/IVA: " + formatCurrency(parseFloat(params.value)) + " €</b>");
                    return tooltipParts.join("<br>");
                }

                tooltipParts.push("<i>" + (params.data.NomeParaExibir || "Tarifário") + " - " + colField.replace("Total ", "").replace(" (€)", "") + "</i>");
                    
                // Preços Unitários s/IVA
                tooltipParts.push("<b>Preços Unitários (s/IVA):</b>");
                let temPrecosUnitarios = false;
                if (colField.includes("Simples")) {
                    if (tooltipData.tt_preco_unit_energia_S_siva != null) {
                        tooltipParts.push("&nbsp;&nbsp;Energia Simples: " + formatUnitPrice(tooltipData.tt_preco_unit_energia_S_siva) + " €/kWh");
                        temPrecosUnitarios = true;
                    }
                } else if (colField.includes("Bi-horário")) {
                    if (tooltipData.tt_preco_unit_energia_V_siva != null) {
                        tooltipParts.push("&nbsp;&nbsp;Energia Vazio: " + formatUnitPrice(tooltipData.tt_preco_unit_energia_V_siva) + " €/kWh");
                        temPrecosUnitarios = true;
                    }
                    if (tooltipData.tt_preco_unit_energia_F_siva != null) {
                        tooltipParts.push("&nbsp;&nbsp;Energia Fora Vazio: " + formatUnitPrice(tooltipData.tt_preco_unit_energia_F_siva) + " €/kWh");
                        temPrecosUnitarios = true;
                    }
                } else if (colField.includes("Tri-horário")) {
                    if (tooltipData.tt_preco_unit_energia_V_siva != null) {
                        tooltipParts.push("&nbsp;&nbsp;Energia Vazio: " + formatUnitPrice(tooltipData.tt_preco_unit_energia_V_siva) + " €/kWh");
                        temPrecosUnitarios = true;
                    }
                    if (tooltipData.tt_preco_unit_energia_C_siva != null) {
                        tooltipParts.push("&nbsp;&nbsp;Energia Cheias: " + formatUnitPrice(tooltipData.tt_preco_unit_energia_C_siva) + " €/kWh");
                        temPrecosUnitarios = true;
                    }
                    if (tooltipData.tt_preco_unit_energia_P_siva != null) {
                        tooltipParts.push("&nbsp;&nbsp;Energia Ponta: " + formatUnitPrice(tooltipData.tt_preco_unit_energia_P_siva) + " €/kWh");
                        temPrecosUnitarios = true;
                    }
                }
                if (tooltipData.tt_preco_unit_potencia_siva != null) {
                    tooltipParts.push("&nbsp;&nbsp;Potência: " + formatUnitPrice(tooltipData.tt_preco_unit_potencia_siva) + " €/dia");
                    temPrecosUnitarios = true;
                }
                if(temPrecosUnitarios){
                     tooltipParts.push("------------------------------------");
                }

                // Decomposição original do tooltip
                tooltipParts.push("<b>Decomposição Custo Total:</b>");
                tooltipParts.push("------------------------------------");
                                                
                tooltipParts.push("Total Energia s/IVA: " + formatCurrency(tooltipData.tt_cte_energia_siva) + " €");
                tooltipParts.push("Total Potência s/IVA: " + formatCurrency(tooltipData.tt_cte_potencia_siva) + " €");
                if (tooltipData.tt_cte_iec_siva !== 0) { 
                    tooltipParts.push("IEC s/IVA: " + formatCurrency(tooltipData.tt_cte_iec_siva) + " €");
                }
                // ... (resto da decomposição original do tooltip, como estava antes) ...
                if (tooltipData.tt_cte_dgeg_siva !== 0) { tooltipParts.push("DGEG s/IVA: " + formatCurrency(tooltipData.tt_cte_dgeg_siva) + " €");}
                if (tooltipData.tt_cte_cav_siva !== 0) { tooltipParts.push("CAV s/IVA: " + formatCurrency(tooltipData.tt_cte_cav_siva) + " €");}
                tooltipParts.push("<b>Subtotal s/IVA: " + formatCurrency(tooltipData.tt_cte_total_siva) + " €</b>");
                tooltipParts.push("------------------------------------");
                if (tooltipData.tt_cte_valor_iva_6_total !== 0) { tooltipParts.push("Valor IVA (6%): " + formatCurrency(tooltipData.tt_cte_valor_iva_6_total) + " €");}
                if (tooltipData.tt_cte_valor_iva_23_total !== 0) { tooltipParts.push("Valor IVA (23%): " + formatCurrency(tooltipData.tt_cte_valor_iva_23_total) + " €");}
                tooltipParts.push("<b>Subtotal c/IVA: " + formatCurrency(tooltipData.tt_cte_subtotal_civa) + " €</b>");
                if (tooltipData.tt_cte_desc_finais_valor !== 0 || tooltipData.tt_cte_acres_finais_valor !== 0) { tooltipParts.push("------------------------------------");}
                if (tooltipData.tt_cte_desc_finais_valor !== 0) { tooltipParts.push("Outros Descontos: -" + formatCurrency(tooltipData.tt_cte_desc_finais_valor) + " €");}
                if (tooltipData.tt_cte_acres_finais_valor !== 0) { tooltipParts.push("Outros Acréscimos: +" + formatCurrency(tooltipData.tt_cte_acres_finais_valor) + " €");}
                if (tooltipData.tt_cte_desc_finais_valor !== 0 || tooltipData.tt_cte_acres_finais_valor !== 0) { tooltipParts.push("------------------------------------");}
                tooltipParts.push("<b>Custo Total c/IVA: " + formatCurrency(parseFloat(params.value)) + " €</b>");

                return tooltipParts.filter(part => part !== "").join("<br>");
            }
            """)

            for col_custo_nome_comp in colunas_aggrid_custo_comp:
                if col_custo_nome_comp in df_resultados_comparacao_aggrid.columns:
                    gb_comp.configure_column(
                        field=col_custo_nome_comp,
                        headerName=col_custo_nome_comp.replace("Total ", "").replace(" (€)", ""),
                        type=["numericColumn"],
                        valueFormatter=JsCode("function(params) { if(params.value == null) return '-'; return Number(params.value).toFixed(2); }"),
                        cellStyle=cell_style_cores_comparativa_js_local,
                        tooltipValueGetter=tooltip_custo_total_comparativa_js,
                        tooltipComponent=custom_tooltip_component_js,
                        minWidth=100, flex=1
                    )
            
        # Ocultar colunas de dados de tooltip que não sejam as de nome/comercializador/tipo ou os totais
        colunas_de_dados_tooltip_para_comparativa = [
            col for col in df_resultados_comparacao_aggrid.columns 
            if col.startswith("Tooltip_Total ")
        ]
        for col_ocultar_tooltip_comp in colunas_de_dados_tooltip_para_comparativa:
             gb_comp.configure_column(field=col_ocultar_tooltip_comp, hide=True)


        gb_comp.configure_grid_options(
            domLayout='autoHeight',
            suppressContextMenu=True,
            tooltipShowDelay=200, # Atraso para mostrar tooltip em ms
            tooltipMouseTrack=True # Tooltip segue o rato
        )
        gridOptions_comp = gb_comp.build()
        grid_response_comp = AgGrid(
            df_resultados_comparacao_aggrid,
            gridOptions=gridOptions_comp,
            custom_css=custom_css,
            allow_unsafe_jscode=True,
            fit_columns_on_grid_load=True,
            height = min( (len(df_resultados_comparacao_aggrid) + 1) * 40 + 5, 600),
            theme='alpine',
            key="aggrid_comparacao_opcoes_final",
            enable_enterprise_modules=True
        )
        st.markdown("<a id='exportar-excel-comparacao'></a>", unsafe_allow_html=True)

        st.markdown("---") # Separador
        with st.expander("📥 Exportar Tabela de Comparação para Excel"):
            if not df_resultados_comparacao_aggrid.empty:
                # Colunas visíveis na AgGrid comparativa por defeito
                default_cols_export_comp = ['NomeParaExibir']
                if 'colunas_aggrid_custo_comp' in locals():
                    default_cols_export_comp.extend(colunas_aggrid_custo_comp)

                # Todas as colunas disponíveis no DataFrame da tabela comparativa
                all_cols_comp_df = df_resultados_comparacao_aggrid.columns.tolist()
                
                # Outras colunas que podem ser úteis para exportar (originalmente ocultas na AgGrid)
                additional_useful_cols_comp = ['Comercializador', 'Tipo', 'LinkAdesao', 'info_notas']
                tooltip_data_cols_comp = [col for col in all_cols_comp_df if col.startswith("Tooltip_Total ")]
                
                # Construir lista de opções para o multiselect
                export_options_comp = default_cols_export_comp[:] # Começa com os defaults
                for col in additional_useful_cols_comp + tooltip_data_cols_comp:
                    if col in all_cols_comp_df and col not in export_options_comp:
                        export_options_comp.append(col)
                # Adicionar restantes colunas se houver alguma que não foi coberta
                for col in all_cols_comp_df:
                    if col not in export_options_comp:
                        export_options_comp.append(col)

                cols_to_export_comp_selected = st.multiselect(
                    "Selecione as colunas para exportar (Tabela Comparativa):",
                    options=export_options_comp,
                    default=default_cols_export_comp,
                    key="cols_export_excel_comp_selector"
                )

                # --- Função de interpolação de cores ---
                def gerar_estilo_completo_para_valor(valor, minimo, maximo):
                    estilo_css_final = 'text-align: center;' 
                    if pd.isna(valor): return estilo_css_final
                    try: val_float = float(valor)
                    except ValueError: return estilo_css_final
                    if maximo == minimo or minimo is None or maximo is None: return estilo_css_final
        
                    midpoint = (minimo + maximo) / 2
                    r_bg, g_bg, b_bg = 255,255,255 
                    verde_rgb, branco_rgb, vermelho_rgb = (90,138,198), (255,255,255), (247,150,70)

                    if val_float <= midpoint:
                        ratio = (val_float - minimo) / (midpoint - minimo) if midpoint != minimo else 0.0
                        r_bg = int(verde_rgb[0]*(1-ratio) + branco_rgb[0]*ratio)
                        g_bg = int(verde_rgb[1]*(1-ratio) + branco_rgb[1]*ratio)
                        b_bg = int(verde_rgb[2]*(1-ratio) + branco_rgb[2]*ratio)
                    else:
                        ratio = (val_float - midpoint) / (maximo - midpoint) if maximo != midpoint else 0.0
                        r_bg = int(branco_rgb[0]*(1-ratio) + vermelho_rgb[0]*ratio)
                        g_bg = int(branco_rgb[1]*(1-ratio) + vermelho_rgb[1]*ratio)
                        b_bg = int(branco_rgb[2]*(1-ratio) + vermelho_rgb[2]*ratio)
                
                    estilo_css_final += f' background-color: #{r_bg:02X}{g_bg:02X}{b_bg:02X};'
                    luminancia = (0.299 * r_bg + 0.587 * g_bg + 0.114 * b_bg)
                    cor_texto_css = '#000000' if luminancia > 140 else '#FFFFFF'
                    estilo_css_final += f' color: {cor_texto_css};'
                    return estilo_css_final

                                        # --- Função de estilo principal a ser aplicada ao DataFrame ---
                def estilo_geral_dataframe_para_exportar(df_a_aplicar_estilo, tipos_reais_para_estilo_serie, min_max_config_para_cores, nome_coluna_tarifario="Tarifário"):
                    df_com_estilos = pd.DataFrame('', index=df_a_aplicar_estilo.index, columns=df_a_aplicar_estilo.columns)
               
                    # Cores default (pode ajustar)
                    cor_fundo_indexado_media_css_local = "#FFE699"
                    cor_texto_indexado_media_css_local = "black"
                    cor_fundo_indexado_dinamico_css_local = "#4D79BC"  
                    cor_texto_indexado_dinamico_css_local = "white"
                    cor_fundo_indexado_diagrama_css_local = "#BDD7EE"
                    cor_texto_indexado_diagrama_css_local = "black"

                    for nome_coluna_df in df_a_aplicar_estilo.columns:
                        # Estilo para colunas de custo (Total (€) na detalhada, Total [Opção] (€) na comparativa)
                        if nome_coluna_df in min_max_config_para_cores:
                            try:
                                serie_valores_col = pd.to_numeric(df_a_aplicar_estilo[nome_coluna_df], errors='coerce')
                                min_valor_col = min_max_config_para_cores[nome_coluna_df]['min']
                                max_valor_col = min_max_config_para_cores[nome_coluna_df]['max']
                                df_com_estilos[nome_coluna_df] = serie_valores_col.apply(
                                    lambda valor_v: gerar_estilo_completo_para_valor(valor_v, min_valor_col, max_valor_col)
                                )
                            except Exception as e_estilo_custo:
                                print(f"Erro ao aplicar estilo de custo à coluna {nome_coluna_df}: {e_estilo_custo}")
                                df_com_estilos[nome_coluna_df] = 'text-align: center;' 

                        elif nome_coluna_df == nome_coluna_tarifario: # 'Tarifário' ou 'NomeParaExibir'
                            estilos_col_tarif_lista = []
                            for idx_linha_df, valor_nome_col_tarif in df_a_aplicar_estilo[nome_coluna_df].items():
                                tipo_tarif_str = tipos_reais_para_estilo_serie.get(idx_linha_df, '') if tipos_reais_para_estilo_serie is not None else ''
        
                                est_css_tarif = 'text-align: center; padding: 4px;' 
                                bg_cor_val, fonte_cor_val, fonte_peso_val = "#FFFFFF", "#000000", "normal" # Default Fixo/Outro (branco)

                                if isinstance(valor_nome_col_tarif, str) and valor_nome_col_tarif.startswith("O Meu Tarifário"):
                                    bg_cor_val, fonte_cor_val, fonte_peso_val = "#FF0000", "#FFFFFF", "bold"
                                elif isinstance(valor_nome_col_tarif, str) and valor_nome_col_tarif.startswith("Tarifário Personalizado"):
                                    bg_cor_val, fonte_cor_val, fonte_peso_val = "#92D050", "#FFFFFF", "bold"
                                elif tipo_tarif_str == 'Indexado Média':
                                    bg_cor_val, fonte_cor_val = cor_fundo_indexado_media_css_local, cor_texto_indexado_media_css_local
                                elif tipo_tarif_str == 'Indexado quarto-horário':
                                    bg_cor_val, fonte_cor_val = cor_fundo_indexado_dinamico_css_local, cor_texto_indexado_dinamico_css_local
                                elif tipo_tarif_str == 'Indexado quarto-horário (Diagrama)':
                                    bg_cor_val, fonte_cor_val = cor_fundo_indexado_diagrama_css_local, cor_texto_indexado_diagrama_css_local
                                elif tipo_tarif_str == 'Fixo': # Para dar um fundo um pouco diferente aos fixos
                                    bg_cor_val = "#F0F0F0" 
                
                                est_css_tarif += f' background-color: {bg_cor_val}; color: {fonte_cor_val}; font-weight: {fonte_peso_val};'
                                estilos_col_tarif_lista.append(est_css_tarif)
                            df_com_estilos[nome_coluna_df] = estilos_col_tarif_lista
                        else: # Outras colunas de texto ou sem estilização de cor baseada em valor
                            df_com_estilos[nome_coluna_df] = 'text-align: center;'
                    return df_com_estilos


                def exportar_excel_completo(df_para_exportar, styler_obj, resumo_html_para_excel, poupanca_texto_para_excel, identificador_cor_cabecalho, meu_tarifario_ativo_flag, personalizado_ativo_flag):
                    output_excel_buffer = io.BytesIO() 
                    with pd.ExcelWriter(output_excel_buffer, engine='openpyxl') as writer_excel:
                        sheet_name_excel = 'Tiago Felicia - Eletricidade'

                        # --- Escrever Resumo ---
                        dados_resumo_formatado = []
                        if resumo_html_para_excel:
                            soup_resumo = BeautifulSoup(resumo_html_para_excel, "html.parser")
                            
                            # --- LÓGICA ALTERADA PARA REARRANJAR O RESUMO ---
                            titulo_resumo = soup_resumo.find('h5')
                            if titulo_resumo:
                                dados_resumo_formatado.append([titulo_resumo.get_text(strip=True), None])

                            itens_lista_resumo = soup_resumo.find_all('li')
                            linha_filtros_texto = ""
                            linha_potencia_texto = ""
                            outras_linhas_resumo = []

                            for item in itens_lista_resumo:
                                # Usar .get_text() para obter o conteúdo limpo do item da lista
                                texto_item = item.get_text(separator=' ', strip=True)
                                
                                if "Segmento:" in texto_item:
                                    linha_filtros_texto = texto_item
                                elif "kVA" in texto_item:
                                    linha_potencia_texto = texto_item
                                else:
                                    # Processar as outras linhas normalmente
                                    parts = texto_item.split(':', 1)
                                    if len(parts) == 2:
                                        outras_linhas_resumo.append([parts[0].strip() + ":", parts[1].strip()])
                                    else:
                                        outras_linhas_resumo.append([texto_item, None])
                            
                            # Adicionar a linha combinada primeiro, na ordem que pediu
                            if linha_filtros_texto or linha_potencia_texto:
                                dados_resumo_formatado.append([linha_filtros_texto, linha_potencia_texto])
                            
                            # Adicionar o resto do resumo
                            dados_resumo_formatado.extend(outras_linhas_resumo)
                            # --- FIM DA LÓGICA ALTERADA ---
        
                        df_resumo_obj = pd.DataFrame(dados_resumo_formatado)

                        # 1. Deixe o Pandas criar/ativar a folha na primeira escrita
                        df_resumo_obj.to_excel(writer_excel, sheet_name=sheet_name_excel, index=False, header=False, startrow=0)

                        # 2. AGORA obtenha a referência à worksheet, que certamente existe
                        worksheet_excel = writer_excel.sheets[sheet_name_excel]


                        # Formatar Resumo (Negrito)
                        bold_font_obj = Font(bold=True) # Font já deve estar importado de openpyxl.styles
                        for i_resumo in range(len(df_resumo_obj)):
                            excel_row_idx_resumo = i_resumo + 1 # Linhas do Excel são 1-based
                            cell_resumo_rotulo = worksheet_excel.cell(row=excel_row_idx_resumo, column=1)
                            cell_resumo_rotulo.font = bold_font_obj
                            if df_resumo_obj.shape[1] > 1 and pd.notna(df_resumo_obj.iloc[i_resumo, 1]):
                                cell_resumo_valor = worksheet_excel.cell(row=excel_row_idx_resumo, column=2)
                                cell_resumo_valor.font = bold_font_obj

                        worksheet_excel.column_dimensions['A'].width = 35
                        worksheet_excel.column_dimensions['B'].width = 65

                        linha_atual_no_excel_escrita = len(df_resumo_obj) + 1

                        # --- Escrever Mensagem de Poupança ---
                        if poupanca_texto_para_excel: # Verifica se há texto para a mensagem de poupança
                            linha_atual_no_excel_escrita += 1 # Adiciona uma linha em branco

                            cor_p = st.session_state.get('poupanca_excel_cor', "000000") # Cor do session_state
                            negrito_p = st.session_state.get('poupanca_excel_negrito', False) # Negrito do session_state
            
                            poupanca_cell_escrita = worksheet_excel.cell(row=linha_atual_no_excel_escrita, column=1, value=poupanca_texto_para_excel)
                            poupanca_font_escrita = Font(bold=negrito_p, color=cor_p)
                            poupanca_cell_escrita.font = poupanca_font_escrita

                            # --- MODIFICAÇÃO PARA JUNTAR CÉLULAS ---
                            worksheet_excel.merge_cells(start_row=linha_atual_no_excel_escrita, start_column=1, end_row=linha_atual_no_excel_escrita, end_column=6)

                            # Aplicar alinhamento à célula fundida (a célula do canto superior esquerdo, poupanca_cell_escrita)
                            poupanca_cell_escrita.alignment = Alignment(wrap_text=True, horizontal='left', vertical='top')

                            linha_atual_no_excel_escrita += 1 # Avança para a próxima linha após a mensagem de poupança
        
                        linha_inicio_tab_dados_excel = linha_atual_no_excel_escrita + 3

                        # --- Adicionar linha de informação da simulação ---
                        # Adiciona uma linha em branco antes desta nova linha de informação
                        linha_info_simulacao_excel = linha_atual_no_excel_escrita + 1 

                        data_hoje_obj = datetime.date.today() # datetime já deve estar importado
                        data_hoje_formatada_str = data_hoje_obj.strftime('%d/%m/%Y')

                        espacador_info = "                                                                      " #: 70 espaços

                        texto_completo_info = (
                            f"          Simulação em {data_hoje_formatada_str}{espacador_info}"
                            f"https://www.tiagofelicia.pt{espacador_info}"
                            f"Tiago Felícia"
                        )

                        # Escrever o texto completo na primeira célula da área a ser fundida (Coluna A)
                        info_cell = worksheet_excel.cell(row=linha_info_simulacao_excel, column=1)
                        info_cell.value = texto_completo_info
            
                        # Aplicar negrito à célula
                        # Reutilizar bold_font_obj que já foi definido para o resumo, ou criar um novo se precisar de formatação diferente.
                        # Assumindo que bold_font_obj é Font(bold=True) e está no escopo.
                        # Se não, defina-o: from openpyxl.styles import Font; bold_font_obj = Font(bold=True)
                        if 'bold_font_obj' in locals() or 'bold_font_obj' in globals():
                             info_cell.font = bold_font_obj # Reutiliza o bold_font_obj do resumo
                        else:
                             info_cell.font = Font(bold=True) # Cria um novo se não existir

                        # Fundir as colunas A, B, C, e D para esta linha
                        worksheet_excel.merge_cells(start_row=linha_info_simulacao_excel, start_column=1, end_row=linha_info_simulacao_excel, end_column=6)

                        # Ajustar alinhamento da célula fundida (info_cell é a célula do topo-esquerda da área fundida)
                        # Alinhado à esquerda, centralizado verticalmente, com quebra de linha se necessário.
                        info_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True) 

                        # A linha de início para a tabela de dados principal virá depois desta linha de informação
                        # Adicionamos +1 para esta linha de informação e +1 para uma linha em branco antes da tabela
                        linha_inicio_tab_dados = linha_info_simulacao_excel + 2 
            
                        # --- Fim da adição da linha de informação ---

                        for row in worksheet_excel.iter_rows(min_row=1, max_row=worksheet_excel.max_row+100, min_col=1, max_col=20):
                            for cell in row:
                                cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

                        # O Styler escreverá na mesma folha 'sheet_name_excel'
                        styler_obj.to_excel(
                            writer_excel,
                            sheet_name=sheet_name_excel,
                            index=False,
                            startrow=linha_inicio_tab_dados_excel - 1, # startrow é 0-indexed
                            columns=df_para_exportar.columns.tolist()
                        )

                        # Determina cor do cabeçalho conforme a opção horária e ciclo
                        opcao_horaria_lower = str(opcao_horaria).lower() if 'opcao_horaria' in locals() else "simples"
                        cor_fundo = "A6A6A6"   # padrão Simples
                        cor_fonte = "000000"    # padrão preto

                        if isinstance(identificador_cor_cabecalho, str):
                            id_lower = identificador_cor_cabecalho.lower()
                            if id_lower == "simples": # Já coberto pelo default
                                pass
                            elif "bi-horário" in id_lower and "diário" in id_lower:
                                cor_fundo = "A9D08E"; cor_fonte = "000000"
                            elif "bi-horário" in id_lower and "semanal" in id_lower:
                                cor_fundo = "8EA9DB"; cor_fonte = "000000"
                            elif "tri-horário" in id_lower and "diário" in id_lower:
                                cor_fundo = "BF8F00"; cor_fonte = "FFFFFF"
                            elif "tri-horário" in id_lower and "semanal" in id_lower:
                                cor_fundo = "C65911"; cor_fonte = "FFFFFF"

                        # linha_inicio_tab_dados_excel já existe na função e corresponde ao header da tabela
                        for col_idx, _ in enumerate(df_para_exportar.columns):
                            celula = worksheet_excel.cell(row=linha_inicio_tab_dados_excel, column=col_idx + 1)
                            celula.fill = PatternFill(start_color=cor_fundo, end_color=cor_fundo, fill_type="solid")
                            celula.font = Font(color=cor_fonte, bold=True)

                        # ---- INÍCIO: ADICIONAR LEGENDA DE CORES APÓS A TABELA ----
                        # Calcular a linha de início para a legenda
                        # df_para_exportar é o DataFrame que foi escrito pelo styler (ex: df_export_final)
                        numero_linhas_dados_tabela_principal = len(df_para_exportar)
                        # linha_inicio_tab_dados_excel é a linha do cabeçalho da tabela principal (1-indexada)
                        ultima_linha_tabela_principal = linha_inicio_tab_dados_excel + numero_linhas_dados_tabela_principal
                    
                        linha_legenda_bloco_inicio = ultima_linha_tabela_principal + 2 # Deixa uma linha em branco após a tabela

                        # Título da Legenda
                        titulo_legenda_cell = worksheet_excel.cell(row=linha_legenda_bloco_inicio, column=1, value="Tipos de Tarifário:")
                        # Reutilizar bold_font_obj se definido ou criar um novo
                        if 'bold_font_obj' in locals() or 'bold_font_obj' in globals():
                            titulo_legenda_cell.font = bold_font_obj
                        else:
                            titulo_legenda_cell.font = Font(bold=True)
                        worksheet_excel.merge_cells(start_row=linha_legenda_bloco_inicio, start_column=1, end_row=linha_legenda_bloco_inicio, end_column=6)

                        linha_legenda_item_atual = linha_legenda_bloco_inicio + 1 # Primeira linha para item da legenda

                        titulo_legenda_cell.alignment = Alignment(horizontal='center', vertical='center')

                        itens_legenda_excel = []
                        # 1. Adicionar "O Meu Tarifário" se estiver ativo
                        if meu_tarifario_ativo_flag:
                            itens_legenda_excel.append(
                                {"cf": "FF0000", "ct": "FFFFFF", "b": True, "tA": "O Meu Tarifário", "tB": "Tarifário configurado pelo utilizador."}
                            )
                        # 2. Adicionar "Tarifário Personalizado" se estiver ativo
                        if personalizado_ativo_flag:
                            itens_legenda_excel.append(
                                {"cf": "92D050", "ct": "FFFFFF", "b": True, "tA": "Tarifário Personalizado", "tB": "Tarifário configurado pelo utilizador."}
                            )
                        # 3. Adicionar os tarifários base que aparecem sempre
                        itens_legenda_excel.extend([
                            {"cf": "FFE699", "ct": "000000", "b": False, "tA": "Indexado Média", "tB": "Preço de energia baseado na média OMIE do período."},
                            {"cf": "4D79BC", "ct": "FFFFFF", "b": False, "tA": "Indexado Quarto-horário - Perfil", "tB": "Preço de energia baseado nos valores OMIE horários/quarto-horários e perfil."},
                        ])
                        # 4. Adicionar condicionalmente a legenda do diagrama
                        if 'dados_completos_ficheiro' in st.session_state and st.session_state.dados_completos_ficheiro is not None:
                            itens_legenda_excel.append(
                                {"cf": "BDD7EE", "ct": "000000", "b": False, "tA": "Indexado Quarto-horário - Diagrama", "tB": "Preço de energia baseado nos valores OMIE quarto-horários e calculado com o ficheiro de consumo."}
                            )
                        # 5. Adicionar sempre o item 'Fixo' no final
                        itens_legenda_excel.append(
                            {"cf": "F0F0F0", "ct": "333333", "b": False, "tA": "Fixo", "tB": "Preços de energia constantes", "borda_cor": "CCCCCC"}
                        )

                        # Definir larguras das colunas para a legenda (pode ajustar conforme necessário)
                        worksheet_excel.column_dimensions[get_column_letter(1)].width = 30 # Coluna A para a amostra/nome
                        worksheet_excel.column_dimensions[get_column_letter(2)].width = 200 # Coluna B para a descrição (será junta)

                        for item in itens_legenda_excel:
                            celula_A_legenda = worksheet_excel.cell(row=linha_legenda_item_atual, column=1, value=item["tA"])
                            celula_A_legenda.fill = PatternFill(start_color=item["cf"], end_color=item["cf"], fill_type="solid")
                            celula_A_legenda.font = Font(color=item["ct"], bold=item["b"])
                            celula_A_legenda.alignment = Alignment(horizontal='center', vertical='center', indent=1)

                            if "borda_cor" in item:
                                cor_borda_hex = item["borda_cor"]
                                borda_legenda_obj = Border(
                                    top=Side(border_style="thin", color=cor_borda_hex),
                                    left=Side(border_style="thin", color=cor_borda_hex),
                                    right=Side(border_style="thin", color=cor_borda_hex),
                                    bottom=Side(border_style="thin", color=cor_borda_hex)
                                )
                                celula_A_legenda.border = borda_legenda_obj
                        
                            celula_B_legenda = worksheet_excel.cell(row=linha_legenda_item_atual, column=2, value=item["tB"])
                            celula_B_legenda.alignment = Alignment(vertical='center', wrap_text=True, horizontal='left')
                            # Fundir colunas B até D (ou Ajustar conforme a largura desejada para a descrição)
                            worksheet_excel.merge_cells(start_row=linha_legenda_item_atual, start_column=2,
                                                        end_row=linha_legenda_item_atual, end_column=6) 
                        
                            worksheet_excel.row_dimensions[linha_legenda_item_atual].height = 20 # Ajustar altura da linha da legenda
                            linha_legenda_item_atual += 1
                        # ---- FIM: ADICIONAR LEGENDA DE CORES ----

                        # Ajustar largura das colunas da tabela principal
                        for col_idx_iter, col_nome_iter_width in enumerate(df_para_exportar.columns):
                            col_letra_iter = get_column_letter(col_idx_iter + 1) # get_column_letter já deve estar importado
                            if "Tarifário" in col_nome_iter_width :
                                 worksheet_excel.column_dimensions[col_letra_iter].width = 95    
                            elif "Total Simples (€)" == col_nome_iter_width :
                                worksheet_excel.column_dimensions[col_letra_iter].width = 33
                            elif "Total Bi-horário - Ciclo Diário (€)" in col_nome_iter_width :
                                worksheet_excel.column_dimensions[col_letra_iter].width = 33
                            elif "Total Bi-horário - Ciclo Semanal (€)" in col_nome_iter_width :
                                 worksheet_excel.column_dimensions[col_letra_iter].width = 33    
                            elif "Total Tri-horário - Ciclo Diário (€)" in col_nome_iter_width :
                                worksheet_excel.column_dimensions[col_letra_iter].width = 33    
                            elif "Total Tri-horário - Ciclo Semanal (€)" in col_nome_iter_width :
                                 worksheet_excel.column_dimensions[col_letra_iter].width = 33    
                            else: 
                                worksheet_excel.column_dimensions[col_letra_iter].width = 25


                    output_excel_buffer.seek(0)
                    return output_excel_buffer
                        # --- Fim da definição de exportar_excel_completo ---

                limit_export_comp_selected = st.selectbox(
                    "Número de tarifários a exportar (Tabela Comparativa):",
                    options=["Todos"] + [f"Top {i}" for i in [10, 20, 30, 40, 50]],
                    index=0,
                    key="limit_export_excel_comp"
                )

                if st.button("Preparar Download Excel (Tabela Comparativa)", key="btn_prep_excel_comp_final"):
                    if not cols_to_export_comp_selected:
                        st.warning("Por favor, selecione pelo menos uma coluna para exportar.")
                    else:
                        with st.spinner("A gerar ficheiro Excel da Tabela Comparativa..."):
                            
                            df_data_from_grid_comp = pd.DataFrame()
                            if 'grid_response_comp' in locals() and grid_response_comp and grid_response_comp['data'] is not None:
                                df_data_from_grid_comp = pd.DataFrame(grid_response_comp['data'])
                            else:
                                st.warning("Não foi possível obter os dados da grelha. A exportar com base na tabela original.")
                                df_data_from_grid_comp = df_resultados_comparacao_aggrid.copy()

                            if not df_data_from_grid_comp.empty:
                                
                                tipos_reais_para_estilo_comp = df_data_from_grid_comp['Tipo'] if 'Tipo' in df_data_from_grid_comp else pd.Series(dtype=str)

                                min_max_config_excel = {}
                                colunas_custo_para_cor = [col for col in df_data_from_grid_comp.columns if col.startswith("Total ")]
                                for col_custo in colunas_custo_para_cor:
                                    serie = pd.to_numeric(df_data_from_grid_comp[col_custo], errors='coerce').dropna()
                                    if not serie.empty:
                                        min_max_config_excel[col_custo] = {'min': serie.min(), 'max': serie.max()}
                                    else:
                                        min_max_config_excel[col_custo] = {'min': 0, 'max': 0}

                                valid_export_cols_comp = [col for col in cols_to_export_comp_selected if col in df_data_from_grid_comp.columns]
                                if valid_export_cols_comp:
                                    df_export_comp_final = df_data_from_grid_comp[valid_export_cols_comp].copy()
                                else:
                                    st.warning("Nenhuma das colunas selecionadas para exportação está presente nos dados atuais da tabela comparativa.")
                                    df_export_comp_final = pd.DataFrame()
                                
                                if not df_export_comp_final.empty and limit_export_comp_selected != "Todos":
                                    try:
                                        num_to_export_comp = int(limit_export_comp_selected.split(" ")[1])
                                        df_export_comp_final = df_export_comp_final.head(num_to_export_comp)
                                        tipos_reais_para_estilo_comp = tipos_reais_para_estilo_comp.head(num_to_export_comp)
                                    except Exception as e:
                                        st.warning(f"Não foi possível aplicar o limite de tarifários à tabela comparativa: {e}")
                                
                                if not df_export_comp_final.empty:
                                    if 'NomeParaExibir' in df_export_comp_final.columns:
                                        df_export_comp_final.rename(columns={'NomeParaExibir': 'Tarifário'}, inplace=True)

                                    # --- MUDANÇA PRINCIPAL AQUI: Arredondar os dados diretamente no DataFrame ---
                                    for col in df_export_comp_final.columns:
                                        if "(€)" in col:
                                            # Garante que a coluna é numérica antes de arredondar
                                            df_export_comp_final[col] = pd.to_numeric(df_export_comp_final[col], errors='coerce').round(2)
                                    
                                    styler_comp_excel = df_export_comp_final.style.apply(
                                        lambda df: estilo_geral_dataframe_para_exportar(df, tipos_reais_para_estilo_comp, min_max_config_excel, "Tarifário"),
                                        axis=None
                                    )
                                    
                                    # O .format() já não é estritamente necessário para o arredondamento, mas podemos mantê-lo para formatação (ex: na_rep)
                                    styler_comp_excel = styler_comp_excel.format(formatter="{:.2f}", na_rep="-")
                                    
                                    styler_comp_excel = styler_comp_excel.set_table_styles([
                                        {'selector': 'th', 'props': [('background-color', '#404040'), ('color', 'white'), ('font-weight', 'bold'), ('text-align', 'center'), ('border', '1px solid black'), ('padding', '5px')]},
                                        {'selector': 'td', 'props': [('border', '1px solid #dddddd'), ('padding', '4px')]}
                                    ]).hide(axis="index")

                                    output_excel_comp_bytes = exportar_excel_completo(
                                        df_export_comp_final,
                                        styler_comp_excel,
                                        html_resumo_final,
                                        st.session_state.get('poupanca_excel_texto', ""),
                                        "Comparativa",
                                        meu_tarifario_ativo,
                                        personalizado_ativo
                                    )

                                    timestamp_comp_dl = int(time.time())
                                    filename_comp = f"Tiago_Felicia_Eletricidade_Comparacao_{timestamp_comp_dl}.xlsx"
                                    
                                    st.download_button(
                                        label=f"📥 Descarregar Excel ({filename_comp})",
                                        data=output_excel_comp_bytes.getvalue(),
                                        file_name=filename_comp,
                                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                        key=f"btn_dl_excel_comp_{timestamp_comp_dl}"
                                    )
                                    st.success(f"{filename_comp} pronto para download!")
                                elif df_export_comp_final.empty and not cols_to_export_comp_selected:
                                   pass
                                else:
                                    st.warning("Nenhum dado para exportar com os filtros e colunas selecionados para a tabela comparativa.")
                            else:
                                st.info("Tabela comparativa está vazia, nada para exportar.")

        # --- FIM DO EXPANDER DE EXPORTAÇÃO DA TABELA COMPARATIVA ---

else: # --- INÍCIO DO BLOCO PARA TABELA DETALHADA (Tiago Felícia - Tarifários de Eletricidade - Detalhado) ---
    #st.markdown("---")

    # --- Comparar Tarifários Fixos ---
    tarifarios_filtrados_fixos = tf_processar[
        (tf_processar['opcao_horaria_e_ciclo'] == opcao_horaria) &
        (tf_processar['potencia_kva'] == potencia)
    ].copy()

    is_billing_month = 28 <= dias <= 31

    with st.spinner("A calcular os custos para todos os tarifários... por favor, aguarde."):

        if not tarifarios_filtrados_fixos.empty:

            # Definir se é um mês de faturação completo UMA VEZ antes do loop
            is_billing_month = 28 <= dias <= 31

            for index, tarifario in tarifarios_filtrados_fixos.iterrows():
                # --- Get tariff specifics ---
                nome_tarifario = tarifario['nome']
                tipo_tarifario = tarifario['tipo']
                comercializador_tarifario = tarifario['comercializador']
                link_adesao_tf = tarifario.get('site_adesao')
                notas_tarifario_tf = tarifario.get('notas', '')
                segmento_tarifario = tarifario.get('segmento', '-')
                faturacao_tarifario = tarifario.get('faturacao', '-')
                pagamento_tarifario = tarifario.get('pagamento', '-')

                # --- Get Inputs and Flags ---
                preco_energia_input_tf = {}
                consumos_horarios_para_func_tf = {}
                
                # Obtém a referência ao dicionário de consumos corretos (brutos ou líquidos)
                consumos_para_este_calculo = consumos_para_custos

                if opcao_horaria.lower() == "simples":
                    # 1. Define o PREÇO a partir da linha do tarifário
                    preco_energia_input_tf['S'] = tarifario.get('preco_energia_simples')
                    # 2. Define o CONSUMO a partir dos dados já processados
                    consumos_horarios_para_func_tf = {'S': consumos_para_este_calculo.get('Simples', 0)}

                elif opcao_horaria.lower().startswith("bi"):
                    ciclo_a_usar = 'BD' if "Diário" in opcao_horaria else 'BS'
                    # 1. Define os PREÇOS
                    preco_energia_input_tf['V'] = tarifario.get('preco_energia_vazio_bi')
                    preco_energia_input_tf['F'] = tarifario.get('preco_energia_fora_vazio')
                    # 2. Define os CONSUMOS
                    consumos_horarios_para_func_tf = {
                        'V': consumos_para_este_calculo.get(ciclo_a_usar, {}).get('V', 0),
                        'F': consumos_para_este_calculo.get(ciclo_a_usar, {}).get('F', 0)
                    }

                elif opcao_horaria.lower().startswith("tri"):
                    ciclo_a_usar = 'TD' if "Diário" in opcao_horaria else 'TS'
                    # 1. Define os PREÇOS
                    preco_energia_input_tf['V'] = tarifario.get('preco_energia_vazio_tri')
                    preco_energia_input_tf['C'] = tarifario.get('preco_energia_cheias')
                    preco_energia_input_tf['P'] = tarifario.get('preco_energia_ponta')
                    # 2. Define os CONSUMOS
                    consumos_horarios_para_func_tf = {
                        'V': consumos_para_este_calculo.get(ciclo_a_usar, {}).get('V', 0),
                        'C': consumos_para_este_calculo.get(ciclo_a_usar, {}).get('C', 0),
                        'P': consumos_para_este_calculo.get(ciclo_a_usar, {}).get('P', 0)
                    }

                preco_potencia_input_tf = tarifario.get('preco_potencia_dia', 0.0)

                # Flags (com defaults sensatos)
                tar_incluida_energia_tf = tarifario.get('tar_incluida_energia', True)
                tar_incluida_potencia_tf = tarifario.get('tar_incluida_potencia', True)
                financiamento_tse_incluido_tf = tarifario.get('financiamento_tse_incluido', True) # Assumindo que fixos geralmente incluem

                # --- Passo 1: Identificar Componentes Base (Sem IVA, Sem TS) ---
                tar_energia_regulada_tf = {}
                for periodo in preco_energia_input_tf.keys():
                    tar_energia_regulada_tf[periodo] = calc.obter_tar_energia_periodo(opcao_horaria, periodo, potencia, CONSTANTES)

                tar_potencia_regulada_tf = calc.obter_tar_dia(potencia, CONSTANTES)

                preco_comercializador_energia_tf = {}
                for periodo, preco_in in preco_energia_input_tf.items():
                    preco_in_float = float(preco_in or 0.0)
                    if tar_incluida_energia_tf:
                        preco_comercializador_energia_tf[periodo] = preco_in_float - tar_energia_regulada_tf.get(periodo, 0.0)
                    else:
                        preco_comercializador_energia_tf[periodo] = preco_in_float

                preco_potencia_input_tf_float = float(preco_potencia_input_tf or 0.0)
                if tar_incluida_potencia_tf:
                    preco_comercializador_potencia_tf = preco_potencia_input_tf_float - tar_potencia_regulada_tf
                else:
                    preco_comercializador_potencia_tf = preco_potencia_input_tf_float
                preco_comercializador_potencia_tf = max(0.0, preco_comercializador_potencia_tf) # Limitar a 0

                financiamento_tse_a_adicionar_tf = FINANCIAMENTO_TSE_VAL if not financiamento_tse_incluido_tf else 0.0

                # --- Passo 2: Calcular Componentes TAR Finais (Com Desconto TS, Sem IVA) ---
                tar_energia_final_tf = {}
                tar_potencia_final_dia_tf = tar_potencia_regulada_tf

                if tarifa_social: # Flag global
                    desconto_ts_energia = calc.obter_constante('Desconto TS Energia', CONSTANTES)
                    desconto_ts_potencia_dia = calc.obter_constante(f'Desconto TS Potencia {potencia}', CONSTANTES)
                    for periodo, tar_reg in tar_energia_regulada_tf.items():
                        tar_energia_final_tf[periodo] = tar_reg - desconto_ts_energia
                    tar_potencia_final_dia_tf = max(0.0, tar_potencia_regulada_tf - desconto_ts_potencia_dia)
                else:
                    tar_energia_final_tf = tar_energia_regulada_tf.copy()

            # --- INÍCIO: CAMPOS PARA TOOLTIPS FIXOS ---
                # Para o tooltip do Preço Energia:
                componentes_tooltip_energia_dict_tf = {} # Dicionário para os componentes de energia deste tarifário

                # Flag global 'tarifa_social'
                ts_global_ativa = tarifa_social # Flag global de TS

                # Loop pelos períodos de energia (S, V, F, C, P) que existem para este tarifário
                for periodo_key_tf in preco_comercializador_energia_tf.keys():
                
                    comp_comerc_energia_base_tf = preco_comercializador_energia_tf.get(periodo_key_tf, 0.0)
                    tar_bruta_energia_periodo_tf = tar_energia_regulada_tf.get(periodo_key_tf, 0.0)
                
                    # Flag 'financiamento_tse_incluido_tf' lida do Excel para ESTE tarifário fixo
                    tse_declarado_incluido_excel_tf = financiamento_tse_incluido_tf 
                
                    tse_valor_nominal_const_tf = FINANCIAMENTO_TSE_VAL
                
                    ts_aplicada_energia_flag_para_tooltip_tf = ts_global_ativa
                    desconto_ts_energia_unitario_para_tooltip_tf = 0.0
                    if ts_global_ativa:
                        desconto_ts_energia_unitario_para_tooltip_tf = calc.obter_constante('Desconto TS Energia', CONSTANTES)

                    # Usar os nomes EXATOS que o JavaScript espera
                    componentes_tooltip_energia_dict_tf[f'tooltip_energia_{periodo_key_tf}_comerc_sem_tar'] = comp_comerc_energia_base_tf
                    componentes_tooltip_energia_dict_tf[f'tooltip_energia_{periodo_key_tf}_tar_bruta'] = tar_bruta_energia_periodo_tf
                    componentes_tooltip_energia_dict_tf[f'tooltip_energia_{periodo_key_tf}_tse_declarado_incluido'] = tse_declarado_incluido_excel_tf
                    componentes_tooltip_energia_dict_tf[f'tooltip_energia_{periodo_key_tf}_tse_valor_nominal'] = tse_valor_nominal_const_tf
                    componentes_tooltip_energia_dict_tf[f'tooltip_energia_{periodo_key_tf}_ts_aplicada_flag'] = ts_aplicada_energia_flag_para_tooltip_tf
                    componentes_tooltip_energia_dict_tf[f'tooltip_energia_{periodo_key_tf}_ts_desconto_valor'] = desconto_ts_energia_unitario_para_tooltip_tf
            
                desconto_ts_potencia_valor_aplicado = 0.0
                if tarifa_social: # Flag global
                    desconto_ts_potencia_dia_bruto = calc.obter_constante(f'Desconto TS Potencia {potencia}', CONSTANTES)
                    # O desconto efetivamente aplicado é o mínimo entre o desconto e a própria TAR
                    desconto_ts_potencia_valor_aplicado = min(tar_potencia_regulada_tf, desconto_ts_potencia_dia_bruto)

                # Para o tooltip do Preço Potência Fixos:
                componentes_tooltip_potencia_dict_tf = {
                    'tooltip_pot_comerc_sem_tar': preco_comercializador_potencia_tf,
                    'tooltip_pot_tar_bruta': tar_potencia_regulada_tf,
                    'tooltip_pot_ts_aplicada': ts_global_ativa,
                    'tooltip_pot_desconto_ts_valor': desconto_ts_potencia_valor_aplicado
                }
        
                # --- Passo 3: Calcular Preço Final Energia (€/kWh, Sem IVA) ---
                preco_energia_final_sem_iva_tf = {}
                for periodo in preco_comercializador_energia_tf.keys():
                    preco_energia_final_sem_iva_tf[periodo] = (
                        preco_comercializador_energia_tf[periodo]
                        + tar_energia_final_tf.get(periodo, 0.0)
                        + financiamento_tse_a_adicionar_tf
                    )

                # --- Passo 4: Calcular Componentes Finais Potência (€/dia, Sem IVA) ---
                preco_comercializador_potencia_final_sem_iva_tf = preco_comercializador_potencia_tf
                tar_potencia_final_dia_sem_iva_tf = tar_potencia_final_dia_tf

                # --- Passo 5: Calcular Custo Total Energia (Com IVA) ---
                custo_energia_tf_com_iva = calc.calcular_custo_energia_com_iva(
                    consumo,
                    preco_energia_final_sem_iva_tf.get('S') if opcao_horaria.lower() == "simples" else None,
                    {p: v for p, v in preco_energia_final_sem_iva_tf.items() if p != 'S'},
                    dias, potencia, opcao_horaria,
                    consumos_horarios_para_func_tf, # Já definido acima
                    familia_numerosa
                )

                # --- Passo 6: Calcular Custo Total Potência (Com IVA) ---
                custo_potencia_tf_com_iva = calc.calcular_custo_potencia_com_iva_final(
                    preco_comercializador_potencia_final_sem_iva_tf,
                    tar_potencia_final_dia_sem_iva_tf,
                    dias,
                    potencia
                )

                comercializador_tarifario_tf = tarifario['comercializador'] # Nome do comercializador deste tarifário

                # --- Passo 7: Calcular Taxas Adicionais ---
                consumo_total_para_taxas_tf = sum(consumos_horarios_para_func_tf.values())

                taxas_tf = calc.calcular_taxas_adicionais(
                    consumo_total_para_taxas_tf,
                    dias, tarifa_social,
                    valor_dgeg_user, valor_cav_user,
                    nome_comercializador_atual=comercializador_tarifario_tf,
                    aplica_taxa_fixa_mensal=is_billing_month
                )

                # --- Passo 8: Calcular Custo Total Final ---
                custo_total_antes_desc_fatura_tf = (
                custo_energia_tf_com_iva['custo_com_iva'] +
                custo_potencia_tf_com_iva['custo_com_iva'] +
                taxas_tf['custo_com_iva']
            )

                # Guardar o nome original do tarifário do Excel
                nome_tarifario_excel = str(tarifario['nome'])
                nome_a_exibir = nome_tarifario_excel # Começa com o nome original

                # A lógica 'e_mes_completo_selecionado' é substituída pela nossa variável 'is_billing_month'
                e_mes_completo_selecionado = is_billing_month

                # --- Aplicar desconto_fatura_mes (Com Limite e "s/ desc." visível) ---
                desconto_fatura_mensal_tf = float(tarifario.get('desconto_fatura_mes', 0.0) or 0.0)
                limite_meses_promo_tf = float(tarifario.get('desconto_meses_limite', 0.0) or 0.0)
                
                desconto_fatura_periodo_tf = 0.0

                if desconto_fatura_mensal_tf > 0:
                    limite_dias_promo = limite_meses_promo_tf * 30.0
                    
                    dias_efetivos = dias
                    txt_limite = ""
                    if limite_meses_promo_tf > 0:
                        dias_efetivos = min(dias, limite_dias_promo)
                        txt_limite = f" nos 1ºs {int(limite_meses_promo_tf)} meses"

                    if is_billing_month and (limite_meses_promo_tf == 0 or limite_meses_promo_tf >= 1):
                        desconto_fatura_periodo_tf = desconto_fatura_mensal_tf
                    else:
                        desconto_fatura_periodo_tf = (desconto_fatura_mensal_tf / 30.0) * dias_efetivos
                    
                    # --- ALTERAÇÃO AQUI: Capturar o custo ANTES de descontar ---
                    custo_sem_desconto_visual = custo_total_antes_desc_fatura_tf

                    nome_a_exibir += f" (INCLUI desc. {desconto_fatura_mensal_tf:.2f}€/mês{txt_limite}, s/ desc.={custo_sem_desconto_visual:.2f}€)"

                # Custo final após desconto
                custo_apos_desc_fatura_excel_tf = custo_total_antes_desc_fatura_tf - desconto_fatura_periodo_tf
                # --- FIM desconto_fatura_mes ---

                # Adicionar Quota ACP se aplicável
                custo_apos_acp_tf = custo_apos_desc_fatura_excel_tf
                quota_acp_periodo = 0.0
                # A flag incluir_quota_acp vem da checkbox geral
                # VALOR_QUOTA_ACP_MENSAL (constante global)
                if incluir_quota_acp and isinstance(nome_tarifario_excel, str) and nome_tarifario_excel.startswith("Goldenergy | ACP"):
                    if e_mes_completo_selecionado:
                        quota_acp_periodo = VALOR_QUOTA_ACP_MENSAL
                        custo_apos_acp_tf += quota_acp_periodo
                        nome_a_exibir += f" (INCLUI Quota ACP - {VALOR_QUOTA_ACP_MENSAL:.2f} €/mês)"
                    else:
                        quota_acp_periodo = (VALOR_QUOTA_ACP_MENSAL / 30.0) * dias if dias > 0 else 0
                        custo_apos_acp_tf += quota_acp_periodo
                        nome_a_exibir += f" (INCLUI Quota ACP - {VALOR_QUOTA_ACP_MENSAL:.2f} €/mês)"
                    # custo_apos_acp_tf já adiciona quota_acp_periodo

                # Inicializar o custo que será ajustado por este novo desconto MEO
                custo_antes_desconto_meo_tf = custo_apos_acp_tf # Ou custo_apos_desc_fatura_excel_tf se não houver ACP
                desconto_meo_aplicado_periodo = 0.0

                # --- LÓGICA PARA DESCONTO ESPECIAL MEO ---
                # Condições: Nome do tarifário e consumo
                nome_original_lower = str(nome_tarifario_excel).lower()
            
                consumo_mensal_equivalente = 0
                if dias > 0:
                    consumo_mensal_equivalente = (consumo / dias) * 30.0
            
                # Verifica se o nome contém a frase chave e se o consumo atinge o limite
                if "meo energia - tarifa fixa - clientes meo" in nome_original_lower and consumo_mensal_equivalente >= 216:
                    desconto_meo_mensal_base = 0.0
                    opcao_horaria_lower = str(opcao_horaria).lower()

                    if opcao_horaria_lower == "simples":
                        desconto_meo_mensal_base = 0
                    elif opcao_horaria_lower.startswith("bi"): # Cobre "bi-horário semanal" e "bi-horário diário"
                        desconto_meo_mensal_base = 0
                    elif opcao_horaria_lower.startswith("tri"): # Cobre "tri-horário semanal" e "tri-horário diário"
                        desconto_meo_mensal_base = 0
                
                    if desconto_meo_mensal_base > 0 and dias > 0:
                        desconto_meo_aplicado_periodo = (desconto_meo_mensal_base / 30.0) * dias
                        custo_antes_desconto_meo_tf -= desconto_meo_aplicado_periodo # Aplicar o desconto
                    
                        # Adicionar nota ao nome do tarifário
                        nome_a_exibir += f" (Desconto MEO Clientes {desconto_meo_aplicado_periodo:.2f}€ incl.)"
                # --- FIM DA LÓGICA DESCONTO ESPECIAL MEO ---

                # --- LÓGICA PARA DESCONTO CONTINENTE ---
                # A base para o desconto Continente deve ser o custo APÓS o desconto MEO
                custo_base_para_continente_tf = custo_antes_desconto_meo_tf
                custo_total_estimado_final_tf = custo_base_para_continente_tf
                valor_X_desconto_continente = 0.0

                if desconto_continente and isinstance(nome_tarifario_excel, str) and nome_tarifario_excel.startswith("Galp & Continente"):
            
                    # PASSO ADICIONAL: CALCULAR O CUSTO BRUTO (SEM TARIFA SOCIAL) APENAS PARA ESTE DESCONTO
            
                    # 1. Preço unitário bruto da energia (sem IVA e sem desconto TS)
                    preco_energia_bruto_sem_iva = {}
                    for p in preco_comercializador_energia_tf.keys():
                        preco_energia_bruto_sem_iva[p] = (
                            preco_comercializador_energia_tf.get(p, 0.0) + 
                            tar_energia_regulada_tf.get(p, 0.0) + # <--- USA A TAR BRUTA, sem desconto TS
                            financiamento_tse_a_adicionar_tf
                        )
            
                    # 2. Preço unitário bruto da potência (sem IVA e sem desconto TS)
                    # Requer as componentes brutas
                    preco_comercializador_potencia_bruto = preco_comercializador_potencia_tf 
                    tar_potencia_bruta = tar_potencia_regulada_tf # <--- USA A TAR BRUTA, sem desconto TS

                    # 3. Calcular o custo bruto COM IVA para a energia e potência
                    custo_energia_bruto_cIVA = calc.calcular_custo_energia_com_iva(
                        consumo,
                        preco_energia_bruto_sem_iva.get('S'),
                        {k: v for k, v in preco_energia_bruto_sem_iva.items() if k != 'S'},
                        dias, potencia, opcao_horaria, consumos_horarios_para_func_tf, familia_numerosa
                    )
                    custo_potencia_bruto_cIVA = calc.calcular_custo_potencia_com_iva_final(
                        preco_comercializador_potencia_bruto,
                        tar_potencia_bruta,
                        dias, potencia
                    )
                    # ### DESCONTO DE 10% ###
                    if nome_tarifario_excel.startswith("Galp & Continente (-10% DD)"):
                        valor_X_desconto_continente = (custo_energia_bruto_cIVA['custo_com_iva'] + custo_potencia_bruto_cIVA['custo_com_iva']) * 0.10
                        custo_total_estimado_final_tf = custo_base_para_continente_tf - valor_X_desconto_continente
                        nome_a_exibir += f" (INCLUI desc. Cont. de {valor_X_desconto_continente:.2f}€, s/ desc. Cont.={custo_base_para_continente_tf:.2f}€)"

                    # ### DESCONTO DE 7% ###
                    elif nome_tarifario_excel.startswith("Galp & Continente (-7% s/DD)"):
                        valor_X_desconto_continente = (custo_energia_bruto_cIVA['custo_com_iva'] + custo_potencia_bruto_cIVA['custo_com_iva']) * 0.07
                        custo_total_estimado_final_tf = custo_base_para_continente_tf - valor_X_desconto_continente
                        nome_a_exibir += f" (INCLUI desc. Cont. de {valor_X_desconto_continente:.2f}€, s/ desc. Cont.={custo_base_para_continente_tf:.2f}€)"

                # --- Passo 9: Preparar Resultados para Exibição ---
                valores_energia_exibir_tf = {} # Recalcular ou usar o já calculado 'preco_energia_final_sem_iva_tf'
                for p, v_energia_sem_iva in preco_energia_final_sem_iva_tf.items(): # Use os preços SEM IVA para exibição na tabela
                    periodo_nome = ""
                    if p == 'S': periodo_nome = "Simples"
                    elif p == 'V': periodo_nome = "Vazio"
                    elif p == 'F': periodo_nome = "Fora Vazio"
                    elif p == 'C': periodo_nome = "Cheias"
                    elif p == 'P': periodo_nome = "Ponta"
                    if periodo_nome:
                        valores_energia_exibir_tf[f'{periodo_nome} (€/kWh)'] = round(v_energia_sem_iva, 4)

                preco_potencia_total_final_sem_iva_tf = preco_comercializador_potencia_final_sem_iva_tf + tar_potencia_final_dia_sem_iva_tf

                # --- PASSO X: CALCULAR CUSTOS COM IVA E OBTER DECOMPOSIÇÃO PARA TOOLTIP ---

                # ENERGIA (Tarifários Fixos)
                preco_energia_simples_para_iva_tf = None
                precos_energia_horarios_para_iva_tf = {}
                if opcao_horaria.lower() == "simples":
                    preco_energia_simples_para_iva_tf = preco_energia_final_sem_iva_tf.get('S')
                else:
                    precos_energia_horarios_para_iva_tf = {
                        p: val for p, val in preco_energia_final_sem_iva_tf.items() if p != 'S'
                    }
                    
                decomposicao_custo_energia_tf = calc.calcular_custo_energia_com_iva(
                    consumo, # Consumo total global
                    preco_energia_simples_para_iva_tf,
                    precos_energia_horarios_para_iva_tf,
                    dias, potencia, opcao_horaria,
                    consumos_horarios_para_func_tf, # Dicionário de consumos por período para este tarifário
                    familia_numerosa
                )
                custo_energia_tf_com_iva = decomposicao_custo_energia_tf['custo_com_iva']
                tt_cte_energia_siva_tf = decomposicao_custo_energia_tf['custo_sem_iva']
                tt_cte_energia_iva_6_tf = decomposicao_custo_energia_tf['valor_iva_6']
                tt_cte_energia_iva_23_tf = decomposicao_custo_energia_tf['valor_iva_23']

                # POTÊNCIA (Tarifários Fixos)
                # preco_comercializador_potencia_final_sem_iva_tf e tar_potencia_final_dia_sem_iva_tf já incluem TS (se aplicável)
                decomposicao_custo_potencia_tf = calc.calcular_custo_potencia_com_iva_final(
                    preco_comercializador_potencia_final_sem_iva_tf, # Componente comercializador s/IVA, após TS (se TS afetasse isso)
                    tar_potencia_final_dia_sem_iva_tf,              # Componente TAR s/IVA, após TS
                    dias,
                    potencia
                )
                custo_potencia_tf_com_iva = decomposicao_custo_potencia_tf['custo_com_iva']
                tt_cte_potencia_siva_tf = decomposicao_custo_potencia_tf['custo_sem_iva']
                tt_cte_potencia_iva_6_tf = decomposicao_custo_potencia_tf['valor_iva_6']
                tt_cte_potencia_iva_23_tf = decomposicao_custo_potencia_tf['valor_iva_23']
                
                # TAXAS ADICIONAIS (Tarifários Fixos)
                consumo_total_para_taxas_tf = sum(consumos_horarios_para_func_tf.values())

                decomposicao_taxas_tf = calc.calcular_taxas_adicionais(
                    consumo_total_para_taxas_tf, dias, tarifa_social,
                    valor_dgeg_user, valor_cav_user,
                    nome_comercializador_atual=comercializador_tarifario_tf,
                    aplica_taxa_fixa_mensal=is_billing_month
                )
                taxas_tf_com_iva = decomposicao_taxas_tf['custo_com_iva']
                tt_cte_iec_siva_tf = decomposicao_taxas_tf['iec_sem_iva']
                tt_cte_dgeg_siva_tf = decomposicao_taxas_tf['dgeg_sem_iva']
                tt_cte_cav_siva_tf = decomposicao_taxas_tf['cav_sem_iva']
                tt_cte_taxas_iva_6_tf = decomposicao_taxas_tf['valor_iva_6']
                tt_cte_taxas_iva_23_tf = decomposicao_taxas_tf['valor_iva_23']

                # Custo Total antes de outros descontos específicos do tarifário fixo
                custo_total_antes_desc_especificos_tf = custo_energia_tf_com_iva + custo_potencia_tf_com_iva + taxas_tf_com_iva
                
                # Calcular totais para o tooltip do Custo Total Estimado
                tt_cte_total_siva_tf = tt_cte_energia_siva_tf + tt_cte_potencia_siva_tf + tt_cte_iec_siva_tf + tt_cte_dgeg_siva_tf + tt_cte_cav_siva_tf
                tt_cte_valor_iva_6_total_tf = tt_cte_energia_iva_6_tf + tt_cte_potencia_iva_6_tf + tt_cte_taxas_iva_6_tf
                tt_cte_valor_iva_23_total_tf = tt_cte_energia_iva_23_tf + tt_cte_potencia_iva_23_tf + tt_cte_taxas_iva_23_tf

                # Calcular Subtotal c/IVA (antes de descontos/acréscimos finais)
                tt_cte_subtotal_civa_tf = tt_cte_total_siva_tf + tt_cte_valor_iva_6_total_tf + tt_cte_valor_iva_23_total_tf
                
                tt_cte_desc_finais_valor_tf = 0.0
                if desconto_fatura_periodo_tf > 0: # Usa o valor proporcionalizado ou fixo já calculado
                    tt_cte_desc_finais_valor_tf += desconto_fatura_periodo_tf
                if 'desconto_meo_aplicado_periodo' in locals() and desconto_meo_aplicado_periodo > 0:
                    tt_cte_desc_finais_valor_tf += desconto_meo_aplicado_periodo
                if 'valor_X_desconto_continente' in locals() and valor_X_desconto_continente > 0:
                    tt_cte_desc_finais_valor_tf += valor_X_desconto_continente
                    
                tt_cte_acres_finais_valor_tf = 0.0
                if 'incluir_quota_acp' in locals() and incluir_quota_acp and 'quota_acp_periodo' in locals() and quota_acp_periodo > 0:
                    tt_cte_acres_finais_valor_tf += quota_acp_periodo

                # Adicionar os campos de tooltip ao resultado_fixo
                componentes_tooltip_custo_total_dict_tf = {
                    'tt_cte_energia_siva': tt_cte_energia_siva_tf,
                    'tt_cte_potencia_siva': tt_cte_potencia_siva_tf,
                    'tt_cte_iec_siva': tt_cte_iec_siva_tf,
                    'tt_cte_dgeg_siva': tt_cte_dgeg_siva_tf,
                    'tt_cte_cav_siva': tt_cte_cav_siva_tf,
                    'tt_cte_total_siva': tt_cte_total_siva_tf,
                    'tt_cte_valor_iva_6_total': tt_cte_valor_iva_6_total_tf,
                    'tt_cte_valor_iva_23_total': tt_cte_valor_iva_23_total_tf,
                    'tt_cte_subtotal_civa': tt_cte_subtotal_civa_tf,
                    'tt_cte_desc_finais_valor': tt_cte_desc_finais_valor_tf,
                    'tt_cte_acres_finais_valor': tt_cte_acres_finais_valor_tf
                }

                # Preparar o dicionário de resultado
                resultado_fixo = {
                    'NomeParaExibir': nome_a_exibir,
                    'LinkAdesao': link_adesao_tf,
                    'info_notas': notas_tarifario_tf,
                    'Tipo': tipo_tarifario,
                    'Segmento': segmento_tarifario,
                    'Faturação': faturacao_tarifario,
                    'Pagamento': pagamento_tarifario,
                    'Comercializador': comercializador_tarifario,
                    **valores_energia_exibir_tf,
                    'Potência (€/dia)': round(preco_potencia_total_final_sem_iva_tf, 4),
                    'Total (€)': round(custo_total_estimado_final_tf, 2),
                    # CAMPOS DO TOOLTIP DA POTÊNCIA FIXOS
                    **componentes_tooltip_potencia_dict_tf,
                    # CAMPOS DO TOOLTIP DA ENERGIA FIXOS
                    **componentes_tooltip_energia_dict_tf, 
                    # CAMPOS DO TOOLTIP DA CUSTO TOTAL FIXOS
                    **componentes_tooltip_custo_total_dict_tf, 
                    }
                resultados_list.append(resultado_fixo)

        # --- Fim do loop for tarifario_fixo ---

        # --- Comparar Tarifários Indexados ---
        if df_omie_ajustado.empty:
            st.warning("Não existem dados OMIE para o período selecionado. Tarifários indexados não podem ser calculados.")
        else:
            tarifarios_filtrados_indexados = ti_processar[
                (ti_processar['opcao_horaria_e_ciclo'] == opcao_horaria) &
                (ti_processar['potencia_kva'] == potencia)
            ].copy()

            # Usar a estrutura simplificada do if/else
            if not tarifarios_filtrados_indexados.empty:
                for index, tarifario_indexado in tarifarios_filtrados_indexados.iterrows():
                    # --- Get tariff specifics ---
                    nome_tarifario = tarifario_indexado['nome']
                    tipo_tarifario = tarifario_indexado['tipo']
                    comercializador_tarifario = tarifario_indexado['comercializador']
                    link_adesao_idx = tarifario_indexado.get('site_adesao')
                    notas_tarifario_idx = tarifario_indexado.get('notas', '') 
                    segmento_tarifario = tarifario_indexado.get('segmento', '-')
                    faturacao_tarifario = tarifario_indexado.get('faturacao', '-')
                    pagamento_tarifario = tarifario_indexado.get('pagamento', '-')
                    formula_energia = str(tarifario_indexado.get('formula_calculo', ''))
                    preco_potencia_dia = tarifario_indexado['preco_potencia_dia']

                    constantes = dict(zip(CONSTANTES["constante"], CONSTANTES["valor_unitário"]))

                    # Inicializar variáveis de preço
                    preco_energia_simples_indexado = None
                    preco_energia_vazio_indexado = None
                    preco_energia_fora_vazio_indexado = None
                    preco_energia_cheias_indexado = None
                    preco_energia_ponta_indexado = None

                    # --- CALCULAR PREÇO BASE INDEXADO (input energia) ---
                        # --- BLOCO 1: Cálculo para Indexados Quarto-Horários (BTN ou Luzboa "BTN SPOTDEF") ---
                        # Assume que 'BTN' em formula_energia ou o nome Luzboa identifica corretamente estes tarifários
                    if 'BTN' in formula_energia or nome_tarifario == "Luzboa | BTN SPOTDEF":

                        # --- Tratamento especial para Luzboa | BTN SPOTDEF ---
                        if nome_tarifario == "Luzboa | BTN SPOTDEF":
                            # [LÓGICA Luzboa | Mantida como estava na versão anterior que funcionava]
                            soma_luzboa_simples, count_luzboa_simples = 0.0, 0
                            soma_luzboa_vazio, count_luzboa_vazio = 0.0, 0
                            soma_luzboa_fv, count_luzboa_fv = 0.0, 0
                            soma_luzboa_cheias, count_luzboa_cheias = 0.0, 0
                            soma_luzboa_ponta, count_luzboa_ponta = 0.0, 0

                            coluna_ciclo_luzboa = None
                            if opcao_horaria.lower().startswith("bi"):
                                coluna_ciclo_luzboa = 'BD' if "Diário" in opcao_horaria else 'BS'
                            elif opcao_horaria.lower().startswith("tri"):
                                coluna_ciclo_luzboa = 'TD' if "Diário" in opcao_horaria else 'TS'

                            if coluna_ciclo_luzboa and coluna_ciclo_luzboa not in df_omie_ajustado.columns and not opcao_horaria.lower() == "simples":
                                st.warning(f"Coluna de ciclo '{coluna_ciclo_luzboa}' não encontrada para Luzboa. Energia será zero.")
                                if opcao_horaria.lower() == "simples": preco_energia_simples_indexado = 0.0
                                else: preco_energia_vazio_indexado, preco_energia_fora_vazio_indexado, preco_energia_cheias_indexado, preco_energia_ponta_indexado = 0.0, 0.0, 0.0, 0.0
                            else: # Calcular apenas se coluna de ciclo existe (ou se for simples)
                                for _, row_omie in df_omie_ajustado.iterrows():
                                    if not all(k in row_omie and pd.notna(row_omie[k]) for k in ['OMIE', 'Perdas']): continue
                                    omie_val = row_omie['OMIE'] / 1000; perdas_val = row_omie['Perdas']
                                    cgs_luzboa = constantes.get('Luzboa_CGS', 0.0); fa_luzboa = constantes.get('Luzboa_FA', 1.0); kp_luzboa = constantes.get('Luzboa_Kp', 0.0)
                                    valor_hora_luzboa = (omie_val + cgs_luzboa) * perdas_val * fa_luzboa + kp_luzboa
                                    if opcao_horaria.lower() == "simples": soma_luzboa_simples += valor_hora_luzboa; count_luzboa_simples += 1
                                    elif coluna_ciclo_luzboa and coluna_ciclo_luzboa in row_omie and pd.notna(row_omie[coluna_ciclo_luzboa]):
                                        ciclo_hora = row_omie[coluna_ciclo_luzboa]
                                        if opcao_horaria.lower().startswith("bi"):
                                            if ciclo_hora == 'V': soma_luzboa_vazio += valor_hora_luzboa; count_luzboa_vazio += 1
                                            elif ciclo_hora == 'F': soma_luzboa_fv += valor_hora_luzboa; count_luzboa_fv += 1
                                        elif opcao_horaria.lower().startswith("tri"):
                                            if ciclo_hora == 'V': soma_luzboa_vazio += valor_hora_luzboa; count_luzboa_vazio += 1
                                            elif ciclo_hora == 'C': soma_luzboa_cheias += valor_hora_luzboa; count_luzboa_cheias += 1
                                            elif ciclo_hora == 'P': soma_luzboa_ponta += valor_hora_luzboa; count_luzboa_ponta += 1
                                prec = 4
                                if opcao_horaria.lower() == "simples": preco_energia_simples_indexado = round(soma_luzboa_simples / count_luzboa_simples, 4) if count_luzboa_simples > 0 else 0.0
                                elif opcao_horaria.lower().startswith("bi"):
                                    preco_energia_vazio_indexado = round(soma_luzboa_vazio / count_luzboa_vazio, prec) if count_luzboa_vazio > 0 else 0.0
                                    preco_energia_fora_vazio_indexado = round(soma_luzboa_fv / count_luzboa_fv, prec) if count_luzboa_fv > 0 else 0.0
                                elif opcao_horaria.lower().startswith("tri"):
                                    preco_energia_vazio_indexado = round(soma_luzboa_vazio / count_luzboa_vazio, prec) if count_luzboa_vazio > 0 else 0.0
                                    preco_energia_cheias_indexado = round(soma_luzboa_cheias / count_luzboa_cheias, prec) if count_luzboa_cheias > 0 else 0.0
                                    preco_energia_ponta_indexado = round(soma_luzboa_ponta / count_luzboa_ponta, prec) if count_luzboa_ponta > 0 else 0.0
                            # --- FIM LÓGICA LUZBOA ---

                        else: # Outros Tarifários Quarto-Horários (Coopernico, Repsol, Galp, etc.)
                            # [LÓGICA PARA OUTROS BTN COM PERFIL - INCLUI AJUSTE REPSOL]
                            perfil_coluna = f"BTN_{calc.obter_perfil(consumo, dias, potencia).split('_')[1].upper()}"
                            # Verifica se coluna de perfil existe
                            if perfil_coluna not in df_omie_ajustado.columns:
                                st.warning(f"Coluna de perfil '{perfil_coluna}' não encontrada para '{nome_tarifario}'. Energia será zero.")
                                if opcao_horaria.lower() == "simples": preco_energia_simples_indexado = 0.0
                                else: preco_energia_vazio_indexado, preco_energia_fora_vazio_indexado, preco_energia_cheias_indexado, preco_energia_ponta_indexado = 0.0, 0.0, 0.0, 0.0
                            else: # Coluna de perfil existe, prosseguir com cálculos
                                soma_calculo_simples, soma_perfil_simples = 0.0, 0.0; soma_calculo_vazio, soma_perfil_vazio = 0.0, 0.0; soma_calculo_fv, soma_perfil_fv = 0.0, 0.0; soma_calculo_cheias, soma_perfil_cheias = 0.0, 0.0; soma_calculo_ponta, soma_perfil_ponta = 0.0, 0.0
                                coluna_ciclo = None
                                cycle_column_ok = True # Assumir que está OK por defeito

                                if not opcao_horaria.lower() == "simples":
                                    if opcao_horaria.lower().startswith("bi"): coluna_ciclo = 'BD' if "Diário" in opcao_horaria else 'BS'
                                    elif opcao_horaria.lower().startswith("tri"): coluna_ciclo = 'TD' if "Diário" in opcao_horaria else 'TS'
                                    
                                    if coluna_ciclo and coluna_ciclo not in df_omie_ajustado.columns:
                                        st.warning(f"Coluna de ciclo '{coluna_ciclo}' não encontrada para '{nome_tarifario}' com '{opcao_horaria}'. Preços específicos V/F/C/P podem ser zero.")
                                        cycle_column_ok = False
                                        # Definir preços específicos a zero, mas o simples ainda pode ser calculado
                                        preco_energia_vazio_indexado, preco_energia_fora_vazio_indexado, preco_energia_cheias_indexado, preco_energia_ponta_indexado = 0.0, 0.0, 0.0, 0.0

                                # Loop sobre os dados OMIE do período já filtrado (df_omie_ajustado)
                                for _, row_omie in df_omie_ajustado.iterrows():
                                    required_cols_check = ['OMIE', 'Perdas', perfil_coluna]
                                    if not all(k in row_omie and pd.notna(row_omie[k]) for k in required_cols_check): continue
                                    omie = row_omie['OMIE'] / 1000; perdas = row_omie['Perdas']; perfil = row_omie[perfil_coluna]
                                    if perfil <= 0: continue

                                    calculo_instantaneo_sem_perfil = 0.0
                                    # --- Fórmulas específicas BTN ---
                                    if nome_tarifario == "Coopérnico | Base": calculo_instantaneo_sem_perfil = (omie + constantes.get('Coop_CS_CR', 0.0) + constantes.get('Coop_K', 0.0)) * perdas
                                    elif nome_tarifario == "Coopérnico | GO": calculo_instantaneo_sem_perfil = (omie + constantes.get('Coop_CS_CR', 0.0) + constantes.get('Coop_K', 0.0)) * perdas + constantes.get('Coop_GO', 0.0)
                                    elif nome_tarifario == "Repsol | Leve Sem Mais": calculo_instantaneo_sem_perfil = (omie * perdas * constantes.get('Repsol_FA', 0.0) + constantes.get('Repsol_Q_Tarifa', 0.0))
                                    elif nome_tarifario == "Repsol | Leve PRO Sem Mais": calculo_instantaneo_sem_perfil = (omie * perdas * constantes.get('Repsol_FA', 0.0) + constantes.get('Repsol_Q_Tarifa_Pro', 0.0))
                                    elif nome_tarifario == "Galp | Plano Flexível / Dinâmico": calculo_instantaneo_sem_perfil = (omie + constantes.get('Galp_Ci', 0.0)) * perdas
                                    elif nome_tarifario == "Alfa Energia | ALFA POWER INDEX BTN": calculo_instantaneo_sem_perfil = ((omie + constantes.get('Alfa_CGS', 0.0)) * perdas + constantes.get('Alfa_K', 0.0))
                                    elif nome_tarifario == "Plenitude | Tendência": calculo_instantaneo_sem_perfil = (((omie + constantes.get('Plenitude_CGS', 0.0) + constantes.get('Plenitude_GDOs', 0.0))) * perdas + constantes.get('Plenitude_Fee', 0.0))
                                    elif nome_tarifario == "Meo Energia | Tarifa Variável": calculo_instantaneo_sem_perfil = (omie + constantes.get('Meo_K', 0.0)) * perdas
                                    elif nome_tarifario == "EDP | Eletricidade Indexada Horária": calculo_instantaneo_sem_perfil = (omie * perdas * constantes.get('EDP_H_K1', 1.0) + constantes.get('EDP_H_K2', 0.0))
                                    elif nome_tarifario == "EZU | Indexada": calculo_instantaneo_sem_perfil = (omie + constantes.get('EZU_K', 0.0) + constantes.get('EZU_CGS', 0.0)) * perdas
                                    elif nome_tarifario == "G9 | Smart Dynamic": calculo_instantaneo_sem_perfil = (omie * constantes.get('G9_FA', 0.0) * perdas + constantes.get('G9_CGS', 0.0) + constantes.get('G9_AC', 0.0))
                                    elif nome_tarifario == "G9 | Smart Dynamic (Empresarial)": calculo_instantaneo_sem_perfil = (omie * constantes.get('G9_FA', 0.0) * perdas + constantes.get('G9_CGS', 0.0) + constantes.get('G9_AC', 0.0))
                                    elif nome_tarifario == "Iberdrola | Simples Indexado Dinâmico": calculo_instantaneo_sem_perfil = (omie * perdas + constantes.get("Iberdrola_Dinamico_Q", 0.0) + constantes.get('Iberdrola_mFRR', 0.0))


                                    else: calculo_instantaneo_sem_perfil = omie * perdas # Fallback genérico
                                    # --- Fim Fórmulas ---

                                    # --- Acumular Somas ---
                                    # Acumula SEMPRE nas somas simples (gerais ponderadas pelo perfil)
                                    soma_calculo_simples += calculo_instantaneo_sem_perfil * perfil
                                    soma_perfil_simples += perfil

                                    # Acumula nas somas específicas do período SE aplicável e coluna de ciclo OK
                                    if cycle_column_ok and coluna_ciclo and coluna_ciclo in row_omie and pd.notna(row_omie[coluna_ciclo]):
                                        ciclo_hora = row_omie[coluna_ciclo]
                                        if opcao_horaria.lower().startswith("bi"):
                                            if ciclo_hora == 'V': soma_calculo_vazio += calculo_instantaneo_sem_perfil * perfil; soma_perfil_vazio += perfil
                                            elif ciclo_hora == 'F': soma_calculo_fv += calculo_instantaneo_sem_perfil * perfil; soma_perfil_fv += perfil
                                        elif opcao_horaria.lower().startswith("tri"):
                                            if ciclo_hora == 'V': soma_calculo_vazio += calculo_instantaneo_sem_perfil * perfil; soma_perfil_vazio += perfil
                                            elif ciclo_hora == 'C': soma_calculo_cheias += calculo_instantaneo_sem_perfil * perfil; soma_perfil_cheias += perfil
                                            elif ciclo_hora == 'P': soma_calculo_ponta += calculo_instantaneo_sem_perfil * perfil; soma_perfil_ponta += perfil
                                # --- Fim loop horas ---
                                    
                                prec = 4
                                # --- Cálculo de preços FINAIS para BTN ---
                                if nome_tarifario == "Repsol - Leve Sem Mais":
                                    # Repsol usa sempre o preço calculado como se fosse Simples
                                    preco_simples_repsol = round(soma_calculo_simples / soma_perfil_simples, prec) if soma_perfil_simples > 0 else 0.0
                                    preco_energia_simples_indexado = preco_simples_repsol
                                    preco_energia_vazio_indexado = preco_simples_repsol
                                    preco_energia_fora_vazio_indexado = preco_simples_repsol
                                    preco_energia_cheias_indexado = preco_simples_repsol
                                    preco_energia_ponta_indexado = preco_simples_repsol
                                elif nome_tarifario == "Repsol - Leve PRO Sem Mais":
                                    # Repsol usa sempre o preço calculado como se fosse Simples
                                    preco_simples_repsol_pro = round(soma_calculo_simples / soma_perfil_simples, prec) if soma_perfil_simples > 0 else 0.0
                                    preco_energia_simples_indexado = preco_simples_repsol_pro
                                    preco_energia_vazio_indexado = preco_simples_repsol_pro
                                    preco_energia_fora_vazio_indexado = preco_simples_repsol_pro
                                    preco_energia_cheias_indexado = preco_simples_repsol_pro
                                    preco_energia_ponta_indexado = preco_simples_repsol_pro                            
                                else:
                                    # Cálculo normal para os outros BTN
                                    if opcao_horaria.lower() == "simples":
                                        preco_energia_simples_indexado = round(soma_calculo_simples / soma_perfil_simples, prec) if soma_perfil_simples > 0 else 0.0
                                    elif opcao_horaria.lower().startswith("bi"):
                                        preco_energia_vazio_indexado = round(soma_calculo_vazio / soma_perfil_vazio, prec) if soma_perfil_vazio > 0 else 0.0
                                        preco_energia_fora_vazio_indexado = round(soma_calculo_fv / soma_perfil_fv, prec) if soma_perfil_fv > 0 else 0.0
                                    elif opcao_horaria.lower().startswith("tri"):
                                        preco_energia_vazio_indexado = round(soma_calculo_vazio / soma_perfil_vazio, prec) if soma_perfil_vazio > 0 else 0.0
                                        preco_energia_cheias_indexado = round(soma_calculo_cheias / soma_perfil_cheias, prec) if soma_perfil_cheias > 0 else 0.0
                                        preco_energia_ponta_indexado = round(soma_calculo_ponta / soma_perfil_ponta, prec) if soma_perfil_ponta > 0 else 0.0
                        # --- FIM LÓGICA OUTROS BTN ---

                    # --- BLOCO 2: Cálculo para Indexados Média ---
                    else: # Se não for Quarto-Horário (BTN ou Luzboa)
                        # --- INÍCIO LÓGICA MÉDIA CORRIGIDA ---
                        omie_medio_simples_input_kwh = None; omie_medio_vazio_kwh = None; omie_medio_fv_kwh = None; omie_medio_cheias_kwh = None; omie_medio_ponta_kwh = None
                        if opcao_horaria.lower() == "simples": omie_medio_simples_input_kwh = omie_para_tarifarios_media.get('S', 0.0) / 1000.0
                        elif opcao_horaria.lower().startswith("bi"): omie_medio_vazio_kwh = omie_para_tarifarios_media.get('V', 0.0) / 1000.0; omie_medio_fv_kwh = omie_para_tarifarios_media.get('F', 0.0) / 1000.0
                        elif opcao_horaria.lower().startswith("tri"): omie_medio_vazio_kwh = omie_para_tarifarios_media.get('V', 0.0) / 1000.0; omie_medio_cheias_kwh = omie_para_tarifarios_media.get('C', 0.0) / 1000.0; omie_medio_ponta_kwh = omie_para_tarifarios_media.get('P', 0.0) / 1000.0
                        prec = 4

                        if opcao_horaria.lower() == "simples":
                            perdas_a_usar = perdas_medias.get('Perdas_Anual_S', 1.0) # Usa Anual Simples
                            omie_a_usar = omie_medio_simples_input_kwh if omie_medio_simples_input_kwh is not None else 0.0
                            if nome_tarifario == "Iberdrola | Simples Indexado": preco_energia_simples_indexado = round(omie_a_usar * constantes.get('Iberdrola_Perdas', 1.0) + constantes.get("Iberdrola_Media_Q", 0.0) + constantes.get('Iberdrola_mFRR', 0.0), prec)
                            elif nome_tarifario == "Goldenergy | Tarifário Indexado 100%":
                                mes_num_calculo = list(dias_mes.keys()).index(mes) + 1; perdas_mensais_ge_map = {1: 1.29, 2: 1.18, 3: 1.18, 4: 1.15, 5: 1.11, 6: 1.10, 7: 1.15, 8: 1.13, 9: 1.10, 10: 1.10, 11: 1.16, 12: 1.25}; perdas_mensais_ge = perdas_mensais_ge_map.get(mes_num_calculo, 1.0)
                                preco_energia_simples_indexado = round(omie_a_usar * perdas_mensais_ge + constantes.get('GE_Q_Tarifa', 0.0) + constantes.get('GE_CG', 0.0), prec)
                            elif nome_tarifario == "Endesa | Tarifa Indexada": preco_energia_simples_indexado = round(omie_a_usar + constantes.get('Endesa_A_S', 0.0), prec)
                            elif nome_tarifario == "LUZiGÁS | Energy 8.8": preco_energia_simples_indexado = round((omie_a_usar + constantes.get('Luzigas_8_8_K', 0.0) + constantes.get('Luzigas_CGS', 0.0)) * perdas_a_usar, prec)
                            elif nome_tarifario == "LUZiGÁS | Super Lig Index": preco_energia_simples_indexado = round((omie_a_usar + constantes.get('Luzigas_K', 0.0) + constantes.get('Luzigas_CGS', 0.0)) * perdas_a_usar, prec)
                            elif nome_tarifario == "Ibelectra | Solução Família": preco_energia_simples_indexado = round((omie_a_usar + constantes.get('Ibelectra_CS', 0.0)) * constantes.get('Ibelectra_Perdas', 0.0) + constantes.get('Ibelectra_K', 0.0), prec)
                            elif nome_tarifario == "Ibelectra | Solução Amigo": preco_energia_simples_indexado = round((omie_a_usar + constantes.get('Ibelectra_CS', 0.0)) * constantes.get('Ibelectra_Perdas', 0.0) + constantes.get('Ibelectra_K_a', 0.0), prec)
                            elif nome_tarifario == "G9 | Smart Index": preco_energia_simples_indexado = round((omie_a_usar * constantes.get('G9_FA', 1.02)) * perdas_medias.get('Perdas_M_S', 1.16) + constantes.get('G9_CGS', 0.01) + constantes.get('G9_AC', 0.0055), prec)
                            elif nome_tarifario == "G9 | Smart Index (Empresarial)": preco_energia_simples_indexado = round((omie_a_usar * constantes.get('G9_FA', 1.02)) * perdas_medias.get('Perdas_M_S', 1.16) + constantes.get('G9_CGS', 0.01) + constantes.get('G9_AC', 0.0055), prec)
                            elif nome_tarifario == "EDP | Eletricidade Indexada Média": preco_energia_simples_indexado = round(omie_a_usar * constantes.get('EDP_M_Perdas', 1.0) * constantes.get('EDP_M_K1', 1.0) + constantes.get('EDP_M_K2', 0.0), prec)
                            else: st.warning(f"Fórmula não definida para tarifário médio Simples: {nome_tarifario}"); preco_energia_simples_indexado = omie_a_usar
                        elif opcao_horaria.lower().startswith("bi"):
                            ciclo_bi = 'BD' if "Diário" in opcao_horaria else 'BS'
                            perdas_v_anual = perdas_medias.get(f'Perdas_Anual_{ciclo_bi}_V', 1.0); perdas_f_anual = perdas_medias.get(f'Perdas_Anual_{ciclo_bi}_F', 1.0)
                            omie_v_a_usar = omie_medio_vazio_kwh if omie_medio_vazio_kwh is not None else 0.0; omie_f_a_usar = omie_medio_fv_kwh if omie_medio_fv_kwh is not None else 0.0
                            if nome_tarifario == "LUZiGÁS | Energy 8.8": k_luzigas = constantes.get('Luzigas_8_8_K', 0.0); cgs_luzigas = constantes.get('Luzigas_CGS', 0.0); calc_base = omie_medio_simples_real_kwh + k_luzigas + cgs_luzigas; preco_energia_vazio_indexado = round(calc_base * perdas_v_anual, prec); preco_energia_fora_vazio_indexado = round(calc_base * perdas_f_anual, prec)
                            elif nome_tarifario == "LUZiGÁS | Super Lig Index": k_luzigas = constantes.get('Luzigas_K', 0.0); cgs_luzigas = constantes.get('Luzigas_CGS', 0.0); calc_base = omie_medio_simples_real_kwh + k_luzigas + cgs_luzigas; preco_energia_vazio_indexado = round(calc_base * perdas_v_anual, prec); preco_energia_fora_vazio_indexado = round(calc_base * perdas_f_anual, prec)
                            elif nome_tarifario == "Endesa | Tarifa Indexada": preco_energia_vazio_indexado = round(omie_v_a_usar + constantes.get('Endesa_A_V', 0.0), prec); preco_energia_fora_vazio_indexado = round(omie_f_a_usar + constantes.get('Endesa_A_FV', 0.0), prec)
                            elif nome_tarifario == "Ibelectra | Solução Família": cs_ib = constantes.get('Ibelectra_CS', 0.0); k_ib = constantes.get('Ibelectra_K', 0.0); preco_energia_vazio_indexado = round((omie_v_a_usar + cs_ib) * constantes.get('Ibelectra_Perdas', 0.0) + k_ib, prec); preco_energia_fora_vazio_indexado = round((omie_f_a_usar + cs_ib) * constantes.get('Ibelectra_Perdas', 0.0) + k_ib, prec)                    
                            elif nome_tarifario == "Ibelectra | Solução Amigo": cs_ib = constantes.get('Ibelectra_CS', 0.0); k_ib = constantes.get('Ibelectra_K_a', 0.0); preco_energia_vazio_indexado = round((omie_v_a_usar + cs_ib) * constantes.get('Ibelectra_Perdas', 0.0) + k_ib, prec); preco_energia_fora_vazio_indexado = round((omie_f_a_usar + cs_ib) * constantes.get('Ibelectra_Perdas', 0.0) + k_ib, prec)                    
                            elif nome_tarifario == "G9 | Smart Index": preco_energia_vazio_indexado = round((omie_v_a_usar * constantes.get('G9_FA', 1.02) * perdas_medias.get(f'Perdas_M_{ciclo_bi}_V', 1.16)) + constantes.get('G9_CGS', 0.01) + constantes.get('G9_AC', 0.0055), prec); preco_energia_fora_vazio_indexado = round((omie_f_a_usar * constantes.get('G9_FA', 1.02) * perdas_medias.get(f'Perdas_M_{ciclo_bi}_F', 1.16)) + constantes.get('G9_CGS', 0.01) + constantes.get('G9_AC', 0.0055), prec)                    
                            elif nome_tarifario == "G9 | Smart Index (Empresarial)": preco_energia_vazio_indexado = round((omie_v_a_usar * constantes.get('G9_FA', 1.02) * perdas_medias.get(f'Perdas_M_{ciclo_bi}_V', 1.16)) + constantes.get('G9_CGS', 0.01) + constantes.get('G9_AC', 0.0055), prec); preco_energia_fora_vazio_indexado = round((omie_f_a_usar * constantes.get('G9_FA', 1.02) * perdas_medias.get(f'Perdas_M_{ciclo_bi}_F', 1.16)) + constantes.get('G9_CGS', 0.01) + constantes.get('G9_AC', 0.0055), prec)                    
                            elif nome_tarifario == "EDP | Eletricidade Indexada Média": perdas_const_edp = constantes.get('EDP_M_Perdas', 1.0); k1_edp = constantes.get('EDP_M_K1', 1.0); k2_edp = constantes.get('EDP_M_K2', 0.0); preco_energia_vazio_indexado = round(omie_v_a_usar * perdas_const_edp * k1_edp + k2_edp, prec); preco_energia_fora_vazio_indexado = round(omie_f_a_usar * perdas_const_edp * k1_edp + k2_edp, prec)
                            else: st.warning(f"Fórmula não definida para tarifário médio Bi-horário: {nome_tarifario}"); preco_energia_vazio_indexado = omie_v_a_usar; preco_energia_fora_vazio_indexado = omie_f_a_usar
                        elif opcao_horaria.lower().startswith("tri"):
                            ciclo_tri = 'TD' if "Diário" in opcao_horaria else 'TS'; perdas_v_anual = perdas_medias.get(f'Perdas_Anual_{ciclo_tri}_V', 1.0); perdas_c_anual = perdas_medias.get(f'Perdas_Anual_{ciclo_tri}_C', 1.0); perdas_p_anual = perdas_medias.get(f'Perdas_Anual_{ciclo_tri}_P', 1.0)
                            omie_v_a_usar = omie_medio_vazio_kwh if omie_medio_vazio_kwh is not None else 0.0; omie_c_a_usar = omie_medio_cheias_kwh if omie_medio_cheias_kwh is not None else 0.0; omie_p_a_usar = omie_medio_ponta_kwh if omie_medio_ponta_kwh is not None else 0.0
                            if nome_tarifario == "LUZiGÁS | Energy 8.8": k_luzigas = constantes.get('Luzigas_8_8_K', 0.0); cgs_luzigas = constantes.get('Luzigas_CGS', 0.0); calc_base = omie_medio_simples_real_kwh + k_luzigas + cgs_luzigas; preco_energia_vazio_indexado = round(calc_base * perdas_v_anual, prec); preco_energia_cheias_indexado = round(calc_base * perdas_c_anual, prec); preco_energia_ponta_indexado = round(calc_base * perdas_p_anual, prec)
                            elif nome_tarifario == "LUZiGÁS | Super Lig Index": k_luzigas = constantes.get('Luzigas_K', 0.0); cgs_luzigas = constantes.get('Luzigas_CGS', 0.0); calc_base = omie_medio_simples_real_kwh + k_luzigas + cgs_luzigas; preco_energia_vazio_indexado = round(calc_base * perdas_v_anual, prec); preco_energia_cheias_indexado = round(calc_base * perdas_c_anual, prec); preco_energia_ponta_indexado = round(calc_base * perdas_p_anual, prec)
                            elif nome_tarifario == "Ibelectra | Solução Família": cs_ib = constantes.get('Ibelectra_CS', 0.0); k_ib = constantes.get('Ibelectra_K', 0.0); preco_energia_vazio_indexado = round((omie_v_a_usar + cs_ib) * constantes.get('Ibelectra_Perdas', 0.0) + k_ib, prec); preco_energia_cheias_indexado = round((omie_c_a_usar + cs_ib) * constantes.get('Ibelectra_Perdas', 0.0) + k_ib, prec); preco_energia_ponta_indexado = round((omie_p_a_usar + cs_ib) * constantes.get('Ibelectra_Perdas', 0.0)+ k_ib, prec)
                            elif nome_tarifario == "Ibelectra | Solução Amigo": cs_ib = constantes.get('Ibelectra_CS', 0.0); k_ib = constantes.get('Ibelectra_K_a', 0.0); preco_energia_vazio_indexado = round((omie_v_a_usar + cs_ib) * constantes.get('Ibelectra_Perdas', 0.0) + k_ib, prec); preco_energia_cheias_indexado = round((omie_c_a_usar + cs_ib) * constantes.get('Ibelectra_Perdas', 0.0) + k_ib, prec); preco_energia_ponta_indexado = round((omie_p_a_usar + cs_ib) * constantes.get('Ibelectra_Perdas', 0.0) + k_ib, prec)
                            elif nome_tarifario == "G9 | Smart Index": preco_energia_vazio_indexado = round((omie_v_a_usar * constantes.get('G9_FA', 1.02) * perdas_medias.get(f'Perdas_M_{ciclo_tri}_V', 1.16)) + constantes.get('G9_CGS', 0.01) + constantes.get('G9_AC', 0.0055), prec); preco_energia_cheias_indexado = round((omie_c_a_usar * constantes.get('G9_FA', 1.02) * perdas_medias.get(f'Perdas_M_{ciclo_tri}_C', 1.16)) + constantes.get('G9_CGS', 0.01) + constantes.get('G9_AC', 0.0055), prec); preco_energia_ponta_indexado = round((omie_p_a_usar * constantes.get('G9_FA', 1.02) * perdas_medias.get(f'Perdas_M_{ciclo_tri}_P', 1.16)) + constantes.get('G9_CGS', 0.01) + constantes.get('G9_AC', 0.0055), prec) 
                            elif nome_tarifario == "G9 | Smart Index (Empresarial)": preco_energia_vazio_indexado = round((omie_v_a_usar * constantes.get('G9_FA', 1.02) * perdas_medias.get(f'Perdas_M_{ciclo_tri}_V', 1.16)) + constantes.get('G9_CGS', 0.01) + constantes.get('G9_AC', 0.0055), prec); preco_energia_cheias_indexado = round((omie_c_a_usar * constantes.get('G9_FA', 1.02) * perdas_medias.get(f'Perdas_M_{ciclo_tri}_C', 1.16)) + constantes.get('G9_CGS', 0.01) + constantes.get('G9_AC', 0.0055), prec); preco_energia_ponta_indexado = round((omie_p_a_usar * constantes.get('G9_FA', 1.02) * perdas_medias.get(f'Perdas_M_{ciclo_tri}_P', 1.16)) + constantes.get('G9_CGS', 0.01) + constantes.get('G9_AC', 0.0055), prec) 
                            elif nome_tarifario == "EDP | Eletricidade Indexada Média": perdas_const_edp = constantes.get('EDP_M_Perdas', 1.0); k1_edp = constantes.get('EDP_M_K1', 1.0); k2_edp = constantes.get('EDP_M_K2', 0.0); preco_energia_vazio_indexado = round(omie_v_a_usar * perdas_const_edp * k1_edp + k2_edp, prec); preco_energia_cheias_indexado = round(omie_c_a_usar * perdas_const_edp * k1_edp + k2_edp, prec); preco_energia_ponta_indexado = round(omie_p_a_usar * perdas_const_edp * k1_edp + k2_edp, prec)
                            else: st.warning(f"Fórmula não definida para tarifário médio Tri-horário: {nome_tarifario}"); preco_energia_vazio_indexado = omie_v_a_usar; preco_energia_cheias_indexado = omie_c_a_usar; preco_energia_ponta_indexado = omie_p_a_usar
                        # --- FIM LÓGICA MÉDIA ---

                    # SE FOR QUARTO-HORÁRIO E HOUVER FICHEIRO, CALCULA O CUSTO REAL
                    if 'BTN' in formula_energia and "Luzboa | BTN SPOTDEF" not in nome_tarifario and st.session_state.get('dados_completos_ficheiro') is not None:
                
                        df_consumos_reais = df_consumos_a_utilizar
                
                        resultado_real = calc.calcular_custo_completo_diagrama_carga(
                            tarifario_indexado, 
                            df_consumos_reais,
                            OMIE_PERDAS_CICLOS,
                            CONSTANTES,
                            dias, potencia, familia_numerosa, tarifa_social,
                            valor_dgeg_user, valor_cav_user, mes, ano_atual,
                            incluir_quota_acp,
                            desconto_continente,
                            FINANCIAMENTO_TSE_VAL, 
                            VALOR_QUOTA_ACP_MENSAL
                        )
                
                        if resultado_real:
                            resultados_list.append(resultado_real)

                    # --- Fim do bloco de cálculo base indexado ---

                    # Criar dict de input
                    preco_energia_input_idx = {}
                    consumos_horarios_para_func_idx = {}
                    
                    # Obtém a referência ao dicionário de consumos corretos (brutos ou líquidos)
                    consumos_para_este_calculo = consumos_para_custos

                    if opcao_horaria.lower() == "simples":
                        # 1. Define o PREÇO a partir dos valores calculados para indexados
                        preco_energia_input_idx['S'] = preco_energia_simples_indexado
                        # 2. Define o CONSUMO a partir dos dados já processados
                        consumos_horarios_para_func_idx = {'S': consumos_para_este_calculo.get('Simples', 0)}

                    elif opcao_horaria.lower().startswith("bi"):
                        ciclo_a_usar = 'BD' if "Diário" in opcao_horaria else 'BS'
                        # 1. Define os PREÇOS
                        preco_energia_input_idx['V'] = preco_energia_vazio_indexado
                        preco_energia_input_idx['F'] = preco_energia_fora_vazio_indexado
                        # 2. Define os CONSUMOS
                        consumos_horarios_para_func_idx = {
                            'V': consumos_para_este_calculo.get(ciclo_a_usar, {}).get('V', 0),
                            'F': consumos_para_este_calculo.get(ciclo_a_usar, {}).get('F', 0)
                        }

                    elif opcao_horaria.lower().startswith("tri"):
                        ciclo_a_usar = 'TD' if "Diário" in opcao_horaria else 'TS'
                        # 1. Define os PREÇOS
                        preco_energia_input_idx['V'] = preco_energia_vazio_indexado
                        preco_energia_input_idx['C'] = preco_energia_cheias_indexado
                        preco_energia_input_idx['P'] = preco_energia_ponta_indexado
                        # 2. Define os CONSUMOS
                        consumos_horarios_para_func_idx = {
                            'V': consumos_para_este_calculo.get(ciclo_a_usar, {}).get('V', 0),
                            'C': consumos_para_este_calculo.get(ciclo_a_usar, {}).get('C', 0),
                            'P': consumos_para_este_calculo.get(ciclo_a_usar, {}).get('P', 0)
                        }

                    preco_potencia_input_idx = tarifario_indexado.get('preco_potencia_dia', 0.0)

                    # Flags (verificar defaults adequados para indexados)
                    tar_incluida_energia_idx = tarifario_indexado.get('tar_incluida_energia', False)
                    tar_incluida_potencia_idx = tarifario_indexado.get('tar_incluida_potencia', True)
                    financiamento_tse_incluido_idx = tarifario_indexado.get('financiamento_tse_incluido', False)

                    # --- Passo 1: Identificar Componentes Base (Sem IVA, Sem TS) ---
                    tar_energia_regulada_idx = {}
                    for periodo in preco_energia_input_idx.keys():
                        tar_energia_regulada_idx[periodo] = calc.obter_tar_energia_periodo(opcao_horaria, periodo, potencia, CONSTANTES)

                    tar_potencia_regulada_idx = calc.obter_tar_dia(potencia, CONSTANTES)

                    preco_comercializador_energia_idx = {}
                    for periodo, preco_in in preco_energia_input_idx.items():
                        preco_in_float = float(preco_in or 0.0)
                        if tar_incluida_energia_idx:
                            preco_comercializador_energia_idx[periodo] = preco_in_float - tar_energia_regulada_idx.get(periodo, 0.0)
                        else:
                            preco_comercializador_energia_idx[periodo] = preco_in_float

                    preco_potencia_input_idx_float = float(preco_potencia_input_idx or 0.0)
                    if tar_incluida_potencia_idx:
                        preco_comercializador_potencia_idx = preco_potencia_input_idx_float - tar_potencia_regulada_idx
                    else:
                        preco_comercializador_potencia_idx = preco_potencia_input_idx_float

                    financiamento_tse_a_adicionar_idx = FINANCIAMENTO_TSE_VAL if not financiamento_tse_incluido_idx else 0.0

                    # --- Passo 2: Calcular Componentes TAR Finais (Com Desconto TS, Sem IVA) ---
                    tar_energia_final_idx = {}
                    tar_potencia_final_dia_idx = tar_potencia_regulada_idx

                    if tarifa_social: # Flag global
                        desconto_ts_energia = calc.obter_constante('Desconto TS Energia', CONSTANTES)
                        desconto_ts_potencia_dia = calc.obter_constante(f'Desconto TS Potencia {potencia}', CONSTANTES)
                        for periodo, tar_reg in tar_energia_regulada_idx.items():
                            tar_energia_final_idx[periodo] = tar_reg - desconto_ts_energia
                        tar_potencia_final_dia_idx = max(0.0, tar_potencia_regulada_idx - desconto_ts_potencia_dia)
                    else:
                        tar_energia_final_idx = tar_energia_regulada_idx.copy()

                    desconto_ts_potencia_valor_aplicado = 0.0
                    if tarifa_social: # Flag global
                        desconto_ts_potencia_dia_bruto = calc.obter_constante(f'Desconto TS Potencia {potencia}', CONSTANTES)
                        # O desconto efetivamente aplicado é o mínimo entre o desconto e a própria TAR
                        desconto_ts_potencia_valor_aplicado = min(tar_potencia_regulada_idx, desconto_ts_potencia_dia_bruto)

                    # --- Passo 3: Calcular Preço Final Energia (€/kWh, Sem IVA) ---
                    preco_energia_final_sem_iva_idx = {}
                    for periodo in preco_comercializador_energia_idx.keys():
                        preco_energia_final_sem_iva_idx[periodo] = (
                            preco_comercializador_energia_idx[periodo]
                            + tar_energia_final_idx.get(periodo, 0.0)
                            + financiamento_tse_a_adicionar_idx
                        )

                    # --- Passo 4: Calcular Componentes Finais Potência (€/dia, Sem IVA) ---
                    preco_comercializador_potencia_final_sem_iva_idx = preco_comercializador_potencia_idx
                    tar_potencia_final_dia_sem_iva_idx = tar_potencia_final_dia_idx

                    # --- Passo 5: Calcular Custo Total Energia (Com IVA) ---
                    custo_energia_idx_com_iva = calc.calcular_custo_energia_com_iva(
                        consumo,
                        preco_energia_final_sem_iva_idx.get('S') if opcao_horaria.lower() == "simples" else None,
                        {p: v for p, v in preco_energia_final_sem_iva_idx.items() if p != 'S'},
                        dias, potencia, opcao_horaria,
                        consumos_horarios_para_func_idx,
                        familia_numerosa
                    )

                    # --- Passo 6: Calcular Custo Total Potência (Com IVA) ---
                    custo_potencia_idx_com_iva = calc.calcular_custo_potencia_com_iva_final(
                        preco_comercializador_potencia_final_sem_iva_idx,
                        tar_potencia_final_dia_sem_iva_idx,
                        dias,
                        potencia
                    )

                    comercializador_tarifario_idx = tarifario_indexado['comercializador'] # Nome do comercializador

                    # --- Passo 7: Calcular Taxas Adicionais ---
                    consumo_total_para_taxas_idx = sum(consumos_horarios_para_func_idx.values())

                    taxas_idx = calc.calcular_taxas_adicionais(
                        consumo_total_para_taxas_idx,
                        dias, tarifa_social,
                        valor_dgeg_user, valor_cav_user,
                        nome_comercializador_atual=comercializador_tarifario_idx,
                        aplica_taxa_fixa_mensal=is_billing_month
                    )

                    # --- Passo 8: Calcular Custo Total Final ---
                    custo_total_antes_desc_fatura_idx = custo_energia_idx_com_iva['custo_com_iva'] + custo_potencia_idx_com_iva['custo_com_iva'] + taxas_idx['custo_com_iva']

                    # A lógica 'e_mes_completo_selecionado' é substituída pela nossa variável 'is_billing_month'
                    e_mes_completo_selecionado = is_billing_month

# --- Aplicar desconto_fatura_mes (Indexados - Com Limite e "s/ desc.") ---
                    desconto_fatura_mensal_idx = float(tarifario_indexado.get('desconto_fatura_mes', 0.0) or 0.0)
                    limite_meses_promo_idx = float(tarifario_indexado.get('desconto_meses_limite', 0.0) or 0.0)
                    
                    desconto_fatura_periodo_idx = 0.0

                    if desconto_fatura_mensal_idx > 0:
                        limite_dias_promo = limite_meses_promo_idx * 30.0
                        
                        dias_efetivos = dias
                        txt_limite = ""
                        if limite_meses_promo_idx > 0:
                            dias_efetivos = min(dias, limite_dias_promo)
                            txt_limite = f" nos 1ºs {int(limite_meses_promo_idx)} meses"

                        if is_billing_month and (limite_meses_promo_idx == 0 or limite_meses_promo_idx >= 1):
                            desconto_fatura_periodo_idx = desconto_fatura_mensal_idx
                        else:
                            desconto_fatura_periodo_idx = (desconto_fatura_mensal_idx / 30.0) * dias_efetivos

                        # --- ALTERAÇÃO AQUI: Capturar o custo ANTES de descontar ---
                        custo_sem_desconto_visual = custo_total_antes_desc_fatura_idx

                        nome_tarifario += f" (INCLUI desc. {desconto_fatura_mensal_idx:.2f}€/mês{txt_limite}, s/ desc.={custo_sem_desconto_visual:.2f}€)"

                    custo_total_estimado_idx = custo_total_antes_desc_fatura_idx - desconto_fatura_periodo_idx
                    # --- FIM Aplicar desconto_fatura_mes ---

                    # --- INÍCIO: CAMPOS PARA TOOLTIPS DE ENERGIA (INDEXADOS) ---
                    componentes_tooltip_energia_dict_idx = {}
                    ts_global_ativa_idx = tarifa_social # Flag global de TS

                    # Loop pelos períodos de energia (S, V, F, C, P) que existem para este tarifário indexado
                    # Certifique-se que preco_comercializador_energia_idx.keys() tem os períodos corretos (S ou V,F ou V,C,P)
                    for periodo_key_idx in preco_comercializador_energia_idx.keys():
                        comp_comerc_energia_base_idx = preco_comercializador_energia_idx.get(periodo_key_idx, 0.0)
                        tar_bruta_energia_periodo_idx = tar_energia_regulada_idx.get(periodo_key_idx, 0.0)

                        # Flag 'financiamento_tse_incluido_idx' lida do Excel para ESTE tarifário
                        tse_declarado_incluido_excel_idx = financiamento_tse_incluido_idx
                        tse_valor_nominal_const_idx = FINANCIAMENTO_TSE_VAL

                        ts_aplicada_energia_flag_para_tooltip_idx = ts_global_ativa_idx
                        desconto_ts_energia_unitario_para_tooltip_idx = 0.0
                        if ts_global_ativa_idx:
                            desconto_ts_energia_unitario_para_tooltip_idx = calc.obter_constante('Desconto TS Energia', CONSTANTES)

                        componentes_tooltip_energia_dict_idx[f'tooltip_energia_{periodo_key_idx}_comerc_sem_tar'] = comp_comerc_energia_base_idx
                        componentes_tooltip_energia_dict_idx[f'tooltip_energia_{periodo_key_idx}_tar_bruta'] = tar_bruta_energia_periodo_idx
                        componentes_tooltip_energia_dict_idx[f'tooltip_energia_{periodo_key_idx}_tse_declarado_incluido'] = tse_declarado_incluido_excel_idx
                        componentes_tooltip_energia_dict_idx[f'tooltip_energia_{periodo_key_idx}_tse_valor_nominal'] = tse_valor_nominal_const_idx
                        componentes_tooltip_energia_dict_idx[f'tooltip_energia_{periodo_key_idx}_ts_aplicada_flag'] = ts_aplicada_energia_flag_para_tooltip_idx
                        componentes_tooltip_energia_dict_idx[f'tooltip_energia_{periodo_key_idx}_ts_desconto_valor'] = desconto_ts_energia_unitario_para_tooltip_idx
                    # --- FIM: CAMPOS PARA TOOLTIPS DE ENERGIA (INDEXADOS) ---

                    desconto_ts_potencia_valor_aplicado_idx = 0.0
                    if ts_global_ativa_idx:
                        desconto_ts_potencia_dia_bruto_idx = calc.obter_constante(f'Desconto TS Potencia {potencia}', CONSTANTES)
                        # tar_potencia_regulada_idx é a TAR bruta para este tarifário indexado
                        desconto_ts_potencia_valor_aplicado_idx = min(tar_potencia_regulada_idx, desconto_ts_potencia_dia_bruto_idx)

                    # Para o tooltip do Preço Potência Indexados:
                    componentes_tooltip_potencia_dict_idx = {
                        'tooltip_pot_comerc_sem_tar': preco_comercializador_potencia_idx,
                        'tooltip_pot_tar_bruta': tar_potencia_regulada_idx,
                        'tooltip_pot_ts_aplicada': ts_global_ativa,
                        'tooltip_pot_desconto_ts_valor': desconto_ts_potencia_valor_aplicado
                    }

                    # --- PASSO X: CALCULAR CUSTOS COM IVA E OBTER DECOMPOSIÇÃO PARA TOOLTIP ---

                    # ENERGIA (Tarifários Indexados)
                    preco_energia_simples_para_iva_idx = None
                    precos_energia_horarios_para_iva_idx = {}
                    if opcao_horaria.lower() == "simples":
                        preco_energia_simples_para_iva_idx = preco_energia_final_sem_iva_idx.get('S')
                    else:
                        precos_energia_horarios_para_iva_idx = {
                            p: val for p, val in preco_energia_final_sem_iva_idx.items() if p != 'S'
                        }

                    decomposicao_custo_energia_idx = calc.calcular_custo_energia_com_iva(
                        consumo, # Consumo total global
                        preco_energia_simples_para_iva_idx,
                        precos_energia_horarios_para_iva_idx,
                        dias, potencia, opcao_horaria,
                        consumos_horarios_para_func_idx, # Dicionário de consumos por período para este tarifário
                        familia_numerosa
                    )
                    custo_energia_idx_com_iva = decomposicao_custo_energia_idx['custo_com_iva']
                    tt_cte_energia_siva_idx = decomposicao_custo_energia_idx['custo_sem_iva']
                    tt_cte_energia_iva_6_idx = decomposicao_custo_energia_idx['valor_iva_6']
                    tt_cte_energia_iva_23_idx = decomposicao_custo_energia_idx['valor_iva_23']

                    # POTÊNCIA (Tarifários Indexados)
                    decomposicao_custo_potencia_idx = calc.calcular_custo_potencia_com_iva_final(
                        preco_comercializador_potencia_final_sem_iva_idx,
                        tar_potencia_final_dia_sem_iva_idx, # Esta já tem TS se aplicável
                        dias,
                        potencia
                    )
                    custo_potencia_idx_com_iva = decomposicao_custo_potencia_idx['custo_com_iva']
                    tt_cte_potencia_siva_idx = decomposicao_custo_potencia_idx['custo_sem_iva']
                    tt_cte_potencia_iva_6_idx = decomposicao_custo_potencia_idx['valor_iva_6']
                    tt_cte_potencia_iva_23_idx = decomposicao_custo_potencia_idx['valor_iva_23']
                    
                    # TAXAS ADICIONAIS (Tarifários Indexados)
                    consumo_total_para_taxas_idx = sum(consumos_horarios_para_func_idx.values())

                    decomposicao_taxas_idx = calc.calcular_taxas_adicionais(
                        consumo_total_para_taxas_idx, dias, tarifa_social,
                        valor_dgeg_user, valor_cav_user,
                        nome_comercializador_atual=comercializador_tarifario_idx, # Passa o comercializador
                        aplica_taxa_fixa_mensal=is_billing_month 
                    )
                    taxas_idx_com_iva = decomposicao_taxas_idx['custo_com_iva']
                    tt_cte_iec_siva_idx = decomposicao_taxas_idx['iec_sem_iva']
                    tt_cte_dgeg_siva_idx = decomposicao_taxas_idx['dgeg_sem_iva']
                    tt_cte_cav_siva_idx = decomposicao_taxas_idx['cav_sem_iva']
                    tt_cte_taxas_iva_6_idx = decomposicao_taxas_idx['valor_iva_6']
                    tt_cte_taxas_iva_23_idx = decomposicao_taxas_idx['valor_iva_23']

                    # Custo Total antes de outros descontos específicos do tarifário indexado
                    custo_total_antes_desc_especificos_idx = custo_energia_idx_com_iva + custo_potencia_idx_com_iva + taxas_idx_com_iva

                    # Calcular totais para o tooltip do Custo Total Estimado
                    tt_cte_total_siva_idx = tt_cte_energia_siva_idx + tt_cte_potencia_siva_idx + tt_cte_iec_siva_idx + tt_cte_dgeg_siva_idx + tt_cte_cav_siva_idx
                    tt_cte_valor_iva_6_total_idx = tt_cte_energia_iva_6_idx + tt_cte_potencia_iva_6_idx + tt_cte_taxas_iva_6_idx
                    tt_cte_valor_iva_23_total_idx = tt_cte_energia_iva_23_idx + tt_cte_potencia_iva_23_idx + tt_cte_taxas_iva_23_idx

                    # Calcular Subtotal c/IVA (antes de descontos/acréscimos finais)
                    tt_cte_subtotal_civa_idx = tt_cte_total_siva_idx + tt_cte_valor_iva_6_total_idx + tt_cte_valor_iva_23_total_idx
                
                    # Consolidar Outros Descontos e Acréscimos Finais
                    # Para indexados, geralmente é só o desconto_fatura_periodo_idx
                    tt_cte_desc_finais_valor_idx = 0.0
                    if 'desconto_fatura_periodo_idx' in locals() and desconto_fatura_periodo_idx > 0:
                        tt_cte_desc_finais_valor_idx = desconto_fatura_periodo_idx
                    
                    tt_cte_acres_finais_valor_idx = 0.0 # Tipicamente não há para indexados, a menos que adicione

                    # Adicionar os campos de tooltip ao resultado_indexado
                    componentes_tooltip_custo_total_dict_idx = {
                        'tt_cte_energia_siva': tt_cte_energia_siva_idx,
                        'tt_cte_potencia_siva': tt_cte_potencia_siva_idx,
                        'tt_cte_iec_siva': tt_cte_iec_siva_idx,
                        'tt_cte_dgeg_siva': tt_cte_dgeg_siva_idx,
                        'tt_cte_cav_siva': tt_cte_cav_siva_idx,
                        'tt_cte_total_siva': tt_cte_total_siva_idx,
                        'tt_cte_valor_iva_6_total': tt_cte_valor_iva_6_total_idx,
                        'tt_cte_valor_iva_23_total': tt_cte_valor_iva_23_total_idx,
                        'tt_cte_subtotal_civa': tt_cte_subtotal_civa_idx,
                        'tt_cte_desc_finais_valor': tt_cte_desc_finais_valor_idx,
                        'tt_cte_acres_finais_valor': tt_cte_acres_finais_valor_idx

                        }
                        # resultados_list.append(resultado_indexado)

                    # --- Passo 9: Preparar Resultados para Exibição ---
                    valores_energia_exibir_idx = {}
                    for p, v in preco_energia_final_sem_iva_idx.items(): # Preços finais s/IVA
                        periodo_nome = ""
                        if p == 'S': periodo_nome = "Simples"
                        elif p == 'V': periodo_nome = "Vazio"
                        elif p == 'F': periodo_nome = "Fora Vazio"
                        elif p == 'C': periodo_nome = "Cheias"
                        elif p == 'P': periodo_nome = "Ponta"
                        if periodo_nome:
                            valores_energia_exibir_idx[f'{periodo_nome} (€/kWh)'] = round(v, 4)

                    preco_potencia_total_final_sem_iva_idx = preco_comercializador_potencia_final_sem_iva_idx + tar_potencia_final_dia_sem_iva_idx

                    if pd.notna(custo_total_estimado_idx):
                        resultado_indexado = {
                            'NomeParaExibir': f"{nome_tarifario} - Perfil" if 'BTN' in formula_energia else nome_tarifario,
                            'LinkAdesao': link_adesao_idx,
                            'info_notas': notas_tarifario_idx,
                            'Tipo': tipo_tarifario,
                            'Segmento': segmento_tarifario,
                            'Faturação': faturacao_tarifario,
                            'Pagamento': pagamento_tarifario,
                            'Comercializador': comercializador_tarifario,
                            **valores_energia_exibir_idx,
                            'Potência (€/dia)': round(preco_potencia_total_final_sem_iva_idx, 4), # Preço final s/IVA
                            'Total (€)': round(custo_total_estimado_idx, 2),
                            # CAMPOS DO TOOLTIP DA POTÊNCIA INDEXADOS
                            **componentes_tooltip_potencia_dict_idx,
                            # CAMPOS DO TOOLTIP DA ENERGIA INDEXADOS
                            **componentes_tooltip_energia_dict_idx, 
                            # CAMPOS DO TOOLTIP DA CUSTO TOTAL FIXOS
                            **componentes_tooltip_custo_total_dict_idx, 
                            }
                        resultados_list.append(resultado_indexado)
        # --- Fim do loop for tarifario_indexado ---
        # ### Bloco para adicionar o Tarifário Personalizado à Tabela Detalhada ###
        if st.session_state.get('dados_tarifario_personalizado', {}).get('ativo'):
            dados_pers = st.session_state['dados_tarifario_personalizado']
            precos_energia_a_usar = {}
            preco_potencia_a_usar = 0.0
            consumos_a_usar = {}
            
            # Escolher os preços e consumos corretos com base na OPÇÃO HORÁRIA PRINCIPAL
            if opcao_horaria.lower() == "simples":
                precos_energia_a_usar = {'S': dados_pers['precos_s']['energia']}
                preco_potencia_a_usar = dados_pers['precos_s']['potencia']
                consumos_a_usar = {'S': consumos_para_custos.get('Simples', 0)}
            elif opcao_horaria.lower().startswith("bi"):
                ciclo_a_usar = 'BD' if "diário" in opcao_horaria.lower() else 'BS'
                precos_energia_a_usar = {'V': dados_pers['precos_bi']['vazio'], 'F': dados_pers['precos_bi']['fora_vazio']}
                preco_potencia_a_usar = dados_pers['precos_bi']['potencia']
                consumos_a_usar = {
                    'V': consumos_para_custos.get(ciclo_a_usar, {}).get('V', 0),
                    'F': consumos_para_custos.get(ciclo_a_usar, {}).get('F', 0)
                }
            elif opcao_horaria.lower().startswith("tri"):
                ciclo_a_usar = 'TD' if "diário" in opcao_horaria.lower() else 'TS'
                precos_energia_a_usar = {'V': dados_pers['precos_tri']['vazio'], 'C': dados_pers['precos_tri']['cheias'], 'P': dados_pers['precos_tri']['ponta']}
                preco_potencia_a_usar = dados_pers['precos_tri']['potencia']
                consumos_a_usar = {
                    'V': consumos_para_custos.get(ciclo_a_usar, {}).get('V', 0),
                    'C': consumos_para_custos.get(ciclo_a_usar, {}).get('C', 0),
                    'P': consumos_para_custos.get(ciclo_a_usar, {}).get('P', 0)
                }

            # Só calcula se houver algum preço definido para a estrutura atual
            if preco_potencia_a_usar > 0 or any(p > 0 for p in precos_energia_a_usar.values()):
                resultado_pers = calc.calcular_custo_personalizado(
                    precos_energia_a_usar, preco_potencia_a_usar, consumos_a_usar, dados_pers['flags'],
                    CONSTANTES,
                    FINANCIAMENTO_TSE_VAL,
                    dias=dias, potencia=potencia, tarifa_social=tarifa_social, familia_numerosa=familia_numerosa,
                    valor_dgeg_user=valor_dgeg_user, valor_cav_user=valor_cav_user, opcao_horaria_ref=opcao_horaria
                )
                
                linha_pers_detalhada = {
                    'NomeParaExibir': "Tarifário Personalizado",
                    'Tipo': "Pessoal", 'Comercializador': "Personalizado",
                    **{f"{p.replace('S', 'Simples').replace('V', 'Vazio').replace('F', 'Fora Vazio').replace('C', 'Cheias').replace('P', 'Ponta')} (€/kWh)": v for p, v in resultado_pers['PrecosFinaisSemIVA'].items()},
                    'Potência (€/dia)': resultado_pers['PrecoPotenciaFinalSemIVA'],
                    'Total (€)': round(resultado_pers['Total (€)'], 2),
                    # Desempacotar os dados de tooltip na linha
                    **resultado_pers.get('componentes_tooltip_energia_dict', {}),
                    **resultado_pers.get('componentes_tooltip_potencia_dict', {}),
                    **resultado_pers.get('componentes_tooltip_custo_total_dict', {})
                }
                resultados_list.append(linha_pers_detalhada)

    # --- Processamento final e exibição da tabela de resultados ---
    st.subheader("💰 Tiago Felícia - Tarifários de Eletricidade - Detalhado")


    vista_simplificada = st.checkbox(
        "📱 Ativar vista simplificada (ideal em ecrãs menores)",
        value=True,
        key="chk_vista_simplificada"
        )

    st.write("**Total** com todos os componentes, taxas e impostos. **Valores unitários** de **Energia e Potência** sem IVA.")
    st.write("**O nome do tarifário tem link para mais informações/adesão sobre o mesmo.**")

    st.markdown("➡️ [**Exportar Tabela Detalhada para Excel**](#exportar-excel-detalhada)")

    # Verifica se "O Meu Tarifário" deve ser incluído
    final_results_list = resultados_list.copy() # Começa com os tarifários fixos e/ou indexados
    if meu_tarifario_ativo and 'meu_tarifario_calculado' in st.session_state:
        dados_meu_tarifario_guardado = st.session_state['meu_tarifario_calculado']
        # 'opcao_horaria' aqui é o valor atual do selectbox principal "Opção Horária e Ciclo"
        if dados_meu_tarifario_guardado.get('opcao_horaria_calculada') == opcao_horaria:
            final_results_list.append(dados_meu_tarifario_guardado)

    df_resultados = pd.DataFrame(final_results_list)

    try:
        # Inicializar/resetar variáveis do session_state para a mensagem do Excel
        # Isto garante que não usamos dados de uma execução anterior se as condições atuais não gerarem uma nova mensagem.
        st.session_state.poupanca_excel_texto = ""
        st.session_state.poupanca_excel_cor = "000000"  # Preto por defeito (formato RRGGBB)
        st.session_state.poupanca_excel_negrito = False
        st.session_state.poupanca_excel_disponivel = False # Flag para indicar se há mensagem para o Excel

        if meu_tarifario_ativo and not df_resultados.empty: # df_resultados é o DataFrame da UI
            meu_tarifario_linha = df_resultados[df_resultados['NomeParaExibir'].str.contains("O Meu Tarifário", case=False, na=False)]

            if not meu_tarifario_linha.empty:
                custo_meu_tarifario = meu_tarifario_linha['Total (€)'].iloc[0]
                nome_meu_tarifario_ui = meu_tarifario_linha['NomeParaExibir'].iloc[0]

                if pd.notna(custo_meu_tarifario):
                    outros_tarifarios_ui_df = df_resultados[
                        ~df_resultados['NomeParaExibir'].str.contains("O Meu Tarifário", case=False, na=False)
                    ]
                    custos_outros_validos_ui = outros_tarifarios_ui_df['Total (€)'].dropna()
                    
                    mensagem_poupanca_html_ui = "" # Para a UI

                    if not custos_outros_validos_ui.empty:
                        custo_minimo_outros_ui = custos_outros_validos_ui.min()
                        linha_mais_barata_outros_ui = outros_tarifarios_ui_df.loc[custos_outros_validos_ui.idxmin()]
                        nome_tarifario_mais_barato_outros_ui = linha_mais_barata_outros_ui['NomeParaExibir']

                        if custo_meu_tarifario > custo_minimo_outros_ui:
                            poupanca_abs_ui = custo_meu_tarifario - custo_minimo_outros_ui
                            poupanca_rel_ui = (poupanca_abs_ui / custo_meu_tarifario) * 100 if custo_meu_tarifario != 0 else 0
                            
                            mensagem_poupanca_html_ui = (
                                f"<span style='color:red; font-weight:bold;'>Poupança entre '{nome_meu_tarifario_ui}' ({custo_meu_tarifario:.2f} €) e o mais económico da lista, "
                                f"{nome_tarifario_mais_barato_outros_ui} ({custo_minimo_outros_ui:.2f} €): </span>"
                                f"<span style='color:red; font-weight:bold;'>{poupanca_abs_ui:.2f} €</span> "
                                f"<span style='color:red; font-weight:bold;'>({poupanca_rel_ui:.2f} %).</span>"
                            )
                            # Guardar para Excel
                            st.session_state.poupanca_excel_texto = (
                                f"Poupança entre '{nome_meu_tarifario_ui}' ({custo_meu_tarifario:.2f} €) e o mais económico da lista, "
                                f"{nome_tarifario_mais_barato_outros_ui} ({custo_minimo_outros_ui:.2f} €): "
                                f"{poupanca_abs_ui:.2f} € ({poupanca_rel_ui:.2f} %)."
                            )
                            st.session_state.poupanca_excel_cor = "FF0000" # Vermelho
                            st.session_state.poupanca_excel_negrito = True
                            st.session_state.poupanca_excel_disponivel = True
                        
                        elif custo_meu_tarifario <= custo_minimo_outros_ui:
                            mensagem_poupanca_html_ui = f"<span style='color:green; font-weight:bold;'>Parabéns! O seu tarifário ('{nome_meu_tarifario_ui}' - {custo_meu_tarifario:.2f}€) já é o mais económico ou está entre os mais económicos da lista!</span>"
                            st.session_state.poupanca_excel_texto = f"Parabéns! O seu tarifário ('{nome_meu_tarifario_ui}' - {custo_meu_tarifario:.2f}€) já é o mais económico ou está entre os mais económicos da lista!"
                            st.session_state.poupanca_excel_cor = "008000" # Verde
                            st.session_state.poupanca_excel_negrito = True
                            st.session_state.poupanca_excel_disponivel = True
                    
                    elif len(df_resultados) == 1: # Só "O Meu Tarifário" com custo válido e mais nenhum
                        mensagem_poupanca_html_ui = f"<span style='color:green; font-weight:bold;'>'{nome_meu_tarifario_ui}' ({custo_meu_tarifario:.2f}€) é o único tarifário na lista.</span>"
                        st.session_state.poupanca_excel_texto = f"'{nome_meu_tarifario_ui}' ({custo_meu_tarifario:.2f}€) é o único tarifário na lista."
                        st.session_state.poupanca_excel_cor = "000000" # Preto
                        st.session_state.poupanca_excel_negrito = True # Ou False, conforme preferir
                        st.session_state.poupanca_excel_disponivel = True
                    else: # Meu tarifário tem custo, mas não há outros para comparar
                        mensagem_poupanca_html_ui = f"<span style='color:black; font-weight:normal;'>Não há outros tarifários com custos válidos para comparar com '{nome_meu_tarifario_ui}' ({custo_meu_tarifario:.2f}€).</span>"
                        st.session_state.poupanca_excel_texto = f"Não há outros tarifários com custos válidos para comparar com '{nome_meu_tarifario_ui}' ({custo_meu_tarifario:.2f}€)."
                        st.session_state.poupanca_excel_cor = "000000" # Preto
                        st.session_state.poupanca_excel_negrito = False
                        st.session_state.poupanca_excel_disponivel = True
                    
                    if mensagem_poupanca_html_ui:
                        st.markdown(mensagem_poupanca_html_ui, unsafe_allow_html=True)

                else: # Custo do Meu Tarifário é NaN ou não é válido
                    st.info("Custo do 'Meu Tarifário' não pôde ser calculado. Não é possível determinar poupança.")
                    st.session_state.poupanca_excel_texto = "Custo do 'Meu Tarifário' não pôde ser calculado. Não é possível determinar poupança."
                    st.session_state.poupanca_excel_disponivel = True # Há uma mensagem informativa para o Excel

            # else: Se "Meu Tarifário" não foi encontrado, as variáveis de session_state ficam com os valores de inicialização (mensagem vazia, flag False)
        
        elif meu_tarifario_ativo: # Meu tarifário está ativo mas df_resultados está vazio ou não contém o meu tarifário
            st.info("Ative e calcule 'O Meu Tarifário' ou verifique os resultados para ver a poupança na interface.")
            # Para o Excel, podemos também querer indicar isto
            st.session_state.poupanca_excel_texto = "Informação de poupança não disponível (verifique 'O Meu Tarifário' ou os resultados)."
            st.session_state.poupanca_excel_disponivel = True

    except Exception as e_poupanca: # Renomeado para e_poupanca_ui para evitar conflitos se houver outro try-except
        st.error(f"Erro ao processar a informação de poupança para UI: {e_poupanca}")
        st.session_state.poupanca_excel_texto = "Erro ao calcular a informação de poupança."
        st.session_state.poupanca_excel_disponivel = True # Indica que houve um problema, pode ser útil no Excel
    # --- FIM DO BLOCO PARA EXIBIR POUPANÇA ---

    #ATENÇÃO, PODE CAUSAR PROBLEMAS
    #st.empty()
    #import time
    #time.sleep(0.2) # Geralmente uma má ideia em apps Streamlit

    if not df_resultados.empty:
        colunas_visiveis_presentes = []
        
        if vista_simplificada:
            colunas_base = ['NomeParaExibir', 'Total (€)']
            coluna_potencia = 'Potência (€/dia)'
            colunas_energia_a_mostrar = []
            
            opcao_lower = opcao_horaria.lower()
            if opcao_lower == "simples": colunas_energia_a_mostrar = ['Simples (€/kWh)']
            elif opcao_lower.startswith("bi"): colunas_energia_a_mostrar = ['Vazio (€/kWh)', 'Fora Vazio (€/kWh)']
            elif opcao_lower.startswith("tri"): colunas_energia_a_mostrar = ['Vazio (€/kWh)', 'Cheias (€/kWh)', 'Ponta (€/kWh)']
            
            colunas_visiveis_presentes = colunas_base + colunas_energia_a_mostrar
            if coluna_potencia in df_resultados.columns:
                colunas_visiveis_presentes.append(coluna_potencia)
                
        else: # Vista detalhada
            colunas_base = ['NomeParaExibir', 'LinkAdesao', 'Total (€)']
            coluna_potencia = 'Potência (€/dia)'
            colunas_energia_a_mostrar = []
            
            opcao_lower = opcao_horaria.lower()
            if opcao_lower == "simples": colunas_energia_a_mostrar = ['Simples (€/kWh)']
            elif opcao_lower.startswith("bi"): colunas_energia_a_mostrar = ['Vazio (€/kWh)', 'Fora Vazio (€/kWh)']
            elif opcao_lower.startswith("tri"): colunas_energia_a_mostrar = ['Vazio (€/kWh)', 'Cheias (€/kWh)', 'Ponta (€/kWh)']

            colunas_visiveis_presentes = colunas_base + colunas_energia_a_mostrar
            if coluna_potencia in df_resultados.columns:
                colunas_visiveis_presentes.append(coluna_potencia)
            colunas_visiveis_presentes.extend(['Tipo', 'Comercializador', 'Segmento', 'Faturação', 'Pagamento'])

        # Garantir que apenas colunas que realmente existem no DataFrame são usadas
        colunas_visiveis_presentes = [col for col in colunas_visiveis_presentes if col in df_resultados.columns]

        # --- Definir colunas necessárias para os dados dos tooltips ---
        colunas_dados_tooltip = [
            'tooltip_pot_comerc_sem_tar', 'tooltip_pot_tar_bruta', 'tooltip_pot_ts_aplicada', 'tooltip_pot_desconto_ts_valor',
            # Energia Simples (S)
            'tooltip_energia_S_comerc_sem_tar', 'tooltip_energia_S_tar_bruta', 
            'tooltip_energia_S_tse_declarado_incluido', 'tooltip_energia_S_tse_valor_nominal',
            'tooltip_energia_S_ts_aplicada_flag', 'tooltip_energia_S_ts_desconto_valor',
        
            # Energia Vazio (V) - Adicione se tiver opção Bi ou Tri
            'tooltip_energia_V_comerc_sem_tar', 'tooltip_energia_V_tar_bruta', 
            'tooltip_energia_V_tse_declarado_incluido', 'tooltip_energia_V_tse_valor_nominal',
            'tooltip_energia_V_ts_aplicada_flag', 'tooltip_energia_V_ts_desconto_valor',

            # Energia Fora Vazio (F) - Adicione se tiver opção Bi
            'tooltip_energia_F_comerc_sem_tar', 'tooltip_energia_F_tar_bruta', 
            'tooltip_energia_F_tse_declarado_incluido', 'tooltip_energia_F_tse_valor_nominal',
            'tooltip_energia_F_ts_aplicada_flag', 'tooltip_energia_F_ts_desconto_valor',

            # Energia Cheias (C) - Adicione se tiver opção Tri
            'tooltip_energia_C_comerc_sem_tar', 'tooltip_energia_C_tar_bruta', 
            'tooltip_energia_C_tse_declarado_incluido', 'tooltip_energia_C_tse_valor_nominal',
            'tooltip_energia_C_ts_aplicada_flag', 'tooltip_energia_C_ts_desconto_valor',

            # Energia Ponta (P) - Adicione se tiver opção Tri
            'tooltip_energia_P_comerc_sem_tar', 'tooltip_energia_P_tar_bruta', 
            'tooltip_energia_P_tse_declarado_incluido', 'tooltip_energia_P_tse_valor_nominal',
            'tooltip_energia_P_ts_aplicada_flag', 'tooltip_energia_P_ts_desconto_valor',

            # Para Custo Total
            'tt_cte_energia_siva', 'tt_cte_potencia_siva', 'tt_cte_iec_siva',
            'tt_cte_dgeg_siva', 'tt_cte_cav_siva', 'tt_cte_total_siva',
            'tt_cte_valor_iva_6_total', 'tt_cte_valor_iva_23_total',
            'tt_cte_subtotal_civa','tt_cte_desc_finais_valor','tt_cte_acres_finais_valor'
        ]

        # Colunas que DEVEM estar presentes nos dados do AgGrid para lógica JS, mesmo que ocultas visualmente
        colunas_essenciais_para_js = ['Tipo', 'NomeParaExibir', 'LinkAdesao', 'info_notas'] # Adicionar outras se necessário
        colunas_essenciais_para_js.extend(colunas_dados_tooltip) # as de tooltip já estão aqui

        # Unir colunas visíveis e essenciais para JS, removendo duplicados e mantendo a ordem das visíveis primeiro
        colunas_para_aggrid_final = list(dict.fromkeys(colunas_visiveis_presentes + colunas_essenciais_para_js))
        
        # Filtrar para garantir que todas as colunas em colunas_para_aggrid_final existem em df_resultados
        colunas_para_aggrid_final = [col for col in colunas_para_aggrid_final if col in df_resultados.columns]


        # Verifica se as colunas essenciais 'NomeParaExibir' e 'LinkAdesao' existem
        # Se não existirem, o AgGrid pode não funcionar como esperado para os links.
        if not all(col in df_resultados.columns for col in ['NomeParaExibir', 'LinkAdesao']):
            st.error("Erro: O DataFrame de resultados não contém as colunas 'NomeParaExibir' e/ou 'LinkAdesao' necessárias para o AgGrid. Verifique a construção da lista de resultados.")

        else:
            df_resultados_para_aggrid = df_resultados[colunas_para_aggrid_final].copy()

            if 'Total (€)' in df_resultados_para_aggrid.columns:
                df_resultados_para_aggrid = df_resultados_para_aggrid.sort_values(by='Total (€)')
            df_resultados_para_aggrid = df_resultados_para_aggrid.reset_index(drop=True)

            # ---- INÍCIO DA CONFIGURAÇÃO DO AGGRID ----
            gb = GridOptionsBuilder.from_dataframe(df_resultados_para_aggrid)

            # --- Configurações Padrão para Colunas ---
            gb.configure_default_column(
                sortable=True,
                resizable=True,
                editable=False,
                wrapText=True,
                autoHeight=True,
                wrapHeaderText=True,    # Permite quebra de linha no TEXTO DO CABEÇALHO
                autoHeaderHeight=True   # Ajusta a ALTURA DO CABEÇALHO para o texto quebrado
            )

            # --- 1. DEFINIR O JsCode PARA LINK E TOOLTIP ---
            link_tooltip_renderer_js = JsCode("""
            class LinkTooltipRenderer {
                init(params) {
                    this.eGui = document.createElement('div');
                    let displayText = params.value; // Valor da célula (NomeParaExibir)
                    let url = params.data.LinkAdesao; // Acede ao valor da coluna LinkAdesao da mesma linha

                    if (url && typeof url === 'string' && url.toLowerCase().startsWith('http')) {
                        // HTML para o link clicável
                        // O atributo 'title' (tooltip) mostrará "Aderir/Saber mais: [URL]"
                        // O texto visível do link será o 'displayText' (NomeParaExibir)
                        this.eGui.innerHTML = `<a href="${url}" target="_blank" title="Aderir/Saber mais: ${url}" style="text-decoration: underline; color: inherit;">${displayText}</a>`;
                    } else {
                        // Se não houver URL válido, apenas mostra o displayText com o próprio displayText como tooltip.
                        this.eGui.innerHTML = `<span title="${displayText}">${displayText}</span>`;
                    }
                }
                getGui() { return this.eGui; }
            }
            """) # <--- FIM DA DEFINIÇÃO DE link_tooltip_renderer_js
            
            #CORES PARA TARIFÁRIOS INDEXADOS:
            cor_fundo_indexado_media_css = "#FFE699"
            cor_texto_indexado_media_css = "black"
            cor_fundo_indexado_dinamico_css = "#4D79BC"  
            cor_texto_indexado_dinamico_css = "white"
            cor_fundo_indexado_diagrama_css = "#BDD7EE"  
            cor_texto_indexado_diagrama_css = "black"

            cell_style_nome_tarifario_js = JsCode(f"""
            function(params) {{
                // Estilo base aplicado a todas as células desta coluna
                let styleToApply = {{ 
                    textAlign: 'center',
                    borderRadius: '11px',  // O teu borderRadius desejado
                    padding: '10px 10px'   // O teu padding desejado
                    // Podes adicionar um backgroundColor default para células não especiais aqui, se quiseres
                    // backgroundColor: '#f0f0f0' // Exemplo para tarifários fixos
                }};                                  

                if (params.data) {{
                    const nomeExibir = params.data.NomeParaExibir;
                    const tipoTarifario = params.data.Tipo;

                    // VERIFICA SE O NOME COMEÇA COM "O Meu Tarifário"
                    if (typeof nomeExibir === 'string' && nomeExibir.startsWith('O Meu Tarifário')) {{
                        styleToApply.backgroundColor = 'red';
                        styleToApply.color = 'white';
                        styleToApply.fontWeight = 'bold';
                    }} else if (typeof nomeExibir === 'string' && nomeExibir.startsWith('Tarifário Personalizado')) {{
                        styleToApply.backgroundColor = '#92D050';
                        styleToApply.color = 'white';
                        styleToApply.fontWeight = 'bold';                                              
                    }} else if (tipoTarifario === 'Indexado Média') {{
                        styleToApply.backgroundColor = '{cor_fundo_indexado_media_css}';
                        styleToApply.color = '{cor_texto_indexado_media_css}';
                    }} else if (tipoTarifario === 'Indexado quarto-horário') {{
                        styleToApply.backgroundColor = '{cor_fundo_indexado_dinamico_css}';
                        styleToApply.color = '{cor_texto_indexado_dinamico_css}';
                    }} else if (tipoTarifario === 'Indexado quarto-horário (Diagrama)') {{
                        styleToApply.backgroundColor = '#BDD7EE';
                        styleToApply.color = '#000000';
                    }} else if (tipoTarifario === 'Fixo') {{
                        styleToApply.backgroundColor = '#f0f0f0'; // Cor cinza claro
                        styleToApply.color = '#333333';    // Cor de texto escura
                    }} else {{
                        // Para tarifários fixos ou outros tipos não explicitamente coloridos acima.
                        // Eles já terão o textAlign, borderRadius e padding do styleToApply.
                        // Se quiseres um fundo específico para eles diferente do default do styleToApply, define aqui.
                        // Ex: styleToApply.backgroundColor = '#e9ecef'; // Uma cor neutra para fixos
                    }}
                    return styleToApply;
                }}
                return styleToApply; 
            }}
            """)

            #tooltip Nome Tarifario
            tooltip_nome_tarifario_getter_js = JsCode("""
            function(params) {
                if (!params.data) { 
                    return params.value || ''; 
                }

                const nomeExibir = params.data.NomeParaExibir || '';
                const notas = params.data.info_notas || ''; 

                let tooltipHtmlParts = [];

                if (nomeExibir) {
                    tooltipHtmlParts.push("<strong>" + nomeExibir + "</strong>");
                }

                if (notas) {
                    const notasHtml = notas.replace(/\\n/g, ' ').replace(/\n/g, ' ');
                    // Usando aspas simples para atributos de estilo HTML para simplificar
                    tooltipHtmlParts.push("<small style='display: block; margin-top: 5px;'><i>" + notasHtml + "</i></small>");
                }
        
                if (tooltipHtmlParts.length > 0) {
                    // Se ambos nomeExibir e notas existem, queremos uma quebra de linha entre eles no tooltip.
                    // Se join(''), eles ficam lado a lado. Se join('<br>'), ficam em linhas separadas.
                    // Dado que a nota tem display:block, o join('') deve funcionar para colocá-los em "blocos" separados.
                    return tooltipHtmlParts.join(''); // Para agora, vamos juntar diretamente.
                                                    // Se quiser uma quebra de linha explícita entre o nome e as notas,
                                                    // e ambos existirem, pode usar:
                                                    // return tooltipHtmlParts.join('<br style="margin-bottom:5px">');
                }
        
                return ''; 
            }
                    """)
            
            # Custom_Tooltip
            custom_tooltip_component_js = JsCode("""
                class CustomTooltip {
                    init(params) {
                        // params.value é a string que o seu tooltipValueGetter retorna
                        this.eGui = document.createElement('div');
                        // Para permitir HTML, definimos o innerHTML
                        // É importante que a string de params.value seja HTML seguro se vier de inputs do utilizador,
                        // mas no seu caso, está a construí-lo programaticamente.
                        this.eGui.innerHTML = params.value; 

                        // Aplicar algum estilo básico para o tooltip se desejar
                        this.eGui.style.backgroundColor = 'white'; // Ou outra cor de fundo
                        this.eGui.style.color = 'black';           // Cor do texto
                        this.eGui.style.border = '1px solid #ccc'; // Borda mais suave
                        this.eGui.style.padding = '10px';           // Mais padding
                        this.eGui.style.borderRadius = '11px';      // Cantos arredondados
                        this.eGui.style.boxShadow = '0 2px 5px rgba(0,0,0,0.15)'; // Sombra suave
                        this.eGui.style.maxWidth = '400px';        // Largura máxima
                        this.eGui.style.fontSize = '1.1em';        // Tamanho da fonte
                        this.eGui.style.fontFamily = 'Arial, sans-serif'; // Tipo de fonte                             
                        this.eGui.style.whiteSpace = 'normal';     // Para quebra de linha
                    }

                    getGui() {
                        return this.eGui;
                    }
                }
            """)

            # --- Configuração Coluna Tarifário com Link e Tooltip ---
            gb.configure_column(field='NomeParaExibir', headerName='Tarifário', cellRenderer=link_tooltip_renderer_js, width=450, minWidth=100, flex=2, filter='agTextColumnFilter', tooltipValueGetter=tooltip_nome_tarifario_getter_js, tooltipComponent=custom_tooltip_component_js,
        cellStyle=cell_style_nome_tarifario_js)
            if 'LinkAdesao' in df_resultados_para_aggrid.columns:
                gb.configure_column(field='LinkAdesao', hide=True) # Desativar filtro explicitamente
                        
            # --- 2. Formatação Condicional de Cores ---
            cols_para_cor = [
                col for col in df_resultados_para_aggrid.columns
                if '(€/kWh)' in col or '(€/dia)' in col or 'Total (€)' == col
            ]
            min_max_data_for_js = {}

            for col_name in cols_para_cor:
                if col_name in df_resultados_para_aggrid:
                    series = pd.to_numeric(df_resultados_para_aggrid[col_name], errors='coerce').dropna()
                    if not series.empty:
                        min_max_data_for_js[col_name] = {'min': series.min(), 'max': series.max()}
                    else:
                        min_max_data_for_js[col_name] = {'min': 0, 'max': 0}

            min_max_data_json_string = json.dumps(min_max_data_for_js)

        # Função get_color para JavaScript (para cor nas colunas de valor)
            cell_style_cores_js = JsCode(f"""
            function(params) {{
                const colName = params.colDef.field;
                const value = parseFloat(params.value);
                const minMaxConfig = {min_max_data_json_string}; //

                let style = {{ 
                    textAlign: 'center',
                    borderRadius: '11px',
                    padding: '10px 10px'
                }};

                if (isNaN(value) || !minMaxConfig[colName]) {{
                    return style; // Sem cor para NaN ou se não houver config min/max
                }}

                const min_val = minMaxConfig[colName].min;
                const max_val = minMaxConfig[colName].max;

                if (max_val === min_val) {{
                    style.backgroundColor = 'lightgrey'; // Ou 'transparent'
                    return style;
                }}

                const normalized_value = Math.max(0, Math.min(1, (value - min_val) / (max_val - min_val)));
                // Cores alvo do Excel
                const colorLow = {{ r: 90, g: 138, b: 198 }};  // Azul #5A8AC6
                const colorMid = {{ r: 255, g: 255, b: 255 }}; // Branco #FFFFFF
                const colorHigh = {{ r: 247, g: 150, b: 70 }}; // Laranja #F79646

                let r, g, b;

                if (normalized_value < 0.5) {{
                    // Interpolar entre colorLow (Azul) e colorMid (Branco)
                    // t vai de 0 (no min) a 1 (no meio)
                    const t = normalized_value / 0.5; 
                    r = Math.round(colorLow.r * (1 - t) + colorMid.r * t);
                    g = Math.round(colorLow.g * (1 - t) + colorMid.g * t);
                    b = Math.round(colorLow.b * (1 - t) + colorMid.b * t);
                }} else {{
                    // Interpolar entre colorMid (Branco) e colorHigh (Laranja)
                    // t vai de 0 (no meio) a 1 (no max)
                    const t = (normalized_value - 0.5) / 0.5;
                    r = Math.round(colorMid.r * (1 - t) + colorHigh.r * t);
                    g = Math.round(colorMid.g * (1 - t) + colorHigh.g * t);
                    b = Math.round(colorMid.b * (1 - t) + colorHigh.b * t);
                }}
                
                style.backgroundColor = `rgb(${'{r}'},${'{g}'},${'{b}'})`;
            
                // Lógica de contraste para a cor do texto (preto/branco)
                // Esta heurística calcula a luminância percebida.
                // Fundos mais escuros (<140-150) recebem texto branco, fundos mais claros recebem texto preto.
                // Pode ajustar o limiar 140 se necessário.
                if ((r * 0.299 + g * 0.587 + b * 0.114) < 140) {{ 
                    style.color = 'white';
                }} else {{
                    style.color = 'black';
                }}

                return style;
            }}
            """)
            # --- FIM DA DEFINIÇÃO DE cell_style_cores_js

            #Tooltip Preço Energia
            tooltip_preco_energia_js = JsCode("""
            function(params) {
                if (!params.data) { 
                    // console.error("Tooltip Energia: params.data está AUSENTE para a célula com valor:", params.value, "e coluna:", params.colDef.field);
                    // Decidi retornar apenas o valor da célula se não houver dados, em vez de uma string de erro no tooltip.
                    return String(params.value); 
                }
                                                    
                const colField = params.colDef.field;
                let periodoKey = "";
                let nomePeriodoCompletoParaTitulo = "Energia"; // Um default caso algo falhe

                // Determinar a chave do período e o nome completo para o título
                if (colField.includes("Simples")) {
                    periodoKey = "S";
                    nomePeriodoCompletoParaTitulo = "Simples";
                } else if (colField.includes("Fora Vazio")) { // Importante verificar "Fora Vazio" antes de "Vazio"
                    periodoKey = "F";
                    nomePeriodoCompletoParaTitulo = "Fora Vazio";
                } else if (colField.includes("Vazio")) {
                    periodoKey = "V";
                    nomePeriodoCompletoParaTitulo = "Vazio";
                } else if (colField.includes("Cheias")) {
                    periodoKey = "C";
                    nomePeriodoCompletoParaTitulo = "Cheias";
                } else if (colField.includes("Ponta")) {
                    periodoKey = "P";
                    nomePeriodoCompletoParaTitulo = "Ponta";
                }

                if (!periodoKey) {
                    // console.error("Tooltip Energia: Não foi possível identificar o período a partir de colField:", colField);
                    return String(params.value); // Retorna o valor da célula se o período não for identificado
                }
        
                // Nomes exatos dos campos como definidos em Python
                const field_comerc = 'tooltip_energia_' + periodoKey + '_comerc_sem_tar';
                const field_tar_bruta = 'tooltip_energia_' + periodoKey + '_tar_bruta';
                const field_tse_declarado = 'tooltip_energia_' + periodoKey + '_tse_declarado_incluido';
                const field_tse_nominal = 'tooltip_energia_' + periodoKey + '_tse_valor_nominal';
                const field_ts_aplicada = 'tooltip_energia_' + periodoKey + '_ts_aplicada_flag';
                const field_ts_desconto = 'tooltip_energia_' + periodoKey + '_ts_desconto_valor';

                // Verificar se os campos de dados necessários para o tooltip existem
                if (typeof params.data[field_comerc] === 'undefined' || 
                    typeof params.data[field_tar_bruta] === 'undefined' ||
                    typeof params.data[field_tse_declarado] === 'undefined' ||
                    typeof params.data[field_tse_nominal] === 'undefined' ||
                    typeof params.data[field_ts_aplicada] === 'undefined' ||
                    typeof params.data[field_ts_desconto] === 'undefined') {
            
                    // console.warn("Tooltip Energia (" + periodoKey + "): Um ou mais campos de dados para o tooltip estão UNDEFINED. Coluna:", colField);
                    return "Info decomposição indisponível."; // Mensagem mais clara se os dados não estiverem lá
                }
                                            
                const comercializador = parseFloat(params.data[field_comerc] || 0);
                const tarBruta = parseFloat(params.data[field_tar_bruta] || 0);
                const tseDeclaradoIncluido = params.data[field_tse_declarado]; // Booleano
                const tseValorNominal = parseFloat(params.data[field_tse_nominal] || 0);
                const tsAplicadaEnergia = params.data[field_ts_aplicada]; // Booleano
                const tsDescontoValorEnergia = parseFloat(params.data[field_ts_desconto] || 0);

                const formatPrice = (num, decimalPlaces) => {
                    if (typeof num === 'number' && !isNaN(num)) {
                        return num.toFixed(decimalPlaces);
                    }
                    // console.warn("formatPrice (Energia): Tentativa de formatar valor não numérico:", num);
                    return 'N/A';
                };
                
                // MODIFICADO: Construir o título dinamicamente
                let tituloTooltip = "<b>Decomposição Preço " + nomePeriodoCompletoParaTitulo + " (s/IVA):</b>";
                let tooltipParts = [tituloTooltip];

                tooltipParts.push("Comercializador (s/TAR): " + formatPrice(comercializador, 4) + " €/kWh");
                tooltipParts.push("TAR (Tarifa Acesso Redes): " + formatPrice(tarBruta, 4) + " €/kWh");

                if (tseDeclaradoIncluido === true) {
                    tooltipParts.push("<i>(Financiamento TSE incluído no preço)</i>");
                } else if (tseDeclaradoIncluido === false && tseValorNominal > 0) { // Mostrar apenas se houver valor
                    tooltipParts.push("Financiamento TSE: " + formatPrice(tseValorNominal, 7) + " €/kWh");
                } else if (tseDeclaradoIncluido !== true && tseDeclaradoIncluido !== false) { // Se não for booleano
                    // console.warn("Tooltip Energia ("+periodoKey+"): Flag 'tseDeclaradoIncluido' tem valor inesperado:", tseDeclaradoIncluido);
                    tooltipParts.push("<i>(Info Fin. TSE indisponível)</i>");
                }
        
                if (tsAplicadaEnergia === true && tsDescontoValorEnergia > 0) {
                    tooltipParts.push("Desconto Tarifa Social: -" + formatPrice(tsDescontoValorEnergia, 4) + " €/kWh");
                }
        
                tooltipParts.push("----------------------------------------------------"); // Separador
                tooltipParts.push("<b>Custo Final : " + formatPrice(parseFloat(params.value), 4) + " €/kWh</b>");
        
                return tooltipParts.join("<br>");
            }
            """)

            # Configuração Coluna 'Preço Energia Simples (€/kWh)'
            col_energia_s_nome = 'Simples (€/kWh)'
            if col_energia_s_nome in df_resultados_para_aggrid.columns:
                casas_decimais_energia = 4
        
                js_value_formatter_energia = JsCode(f"""
                    function(params) {{
                        // ... (sua lógica de valueFormatter para Energia)
                        if (params.value == null || typeof params.value === 'undefined' || String(params.value).trim() === '') {{
                            return '';
                        }}
                        let num = Number(params.value);
                        if (isNaN(num)) {{ return ''; }}
                        try {{
                            return num.toFixed({casas_decimais_energia});
                        }} catch (e) {{
                            console.error("Erro valueFormatter Energia:", e, params.value);
                            return String(params.value);
                        }}
                    }}
                """)

                gb.configure_column(
                    field=col_energia_s_nome,
                    headerName=col_energia_s_nome,
                    type=["numericColumn"],
                    filter=False,
                    valueFormatter=js_value_formatter_energia,
                    cellStyle=cell_style_cores_js,
                    tooltipValueGetter=tooltip_preco_energia_js,
                    tooltipComponent=custom_tooltip_component_js,
                    minWidth=60, # Ajustar conforme necessário
                    flex=1
                )

            # Configuração Coluna 'Preço Energia Vazio (€/kWh)'
            col_energia_v_nome = 'Vazio (€/kWh)'
            if col_energia_v_nome in df_resultados_para_aggrid.columns:
                casas_decimais_energia = 4
        
                js_value_formatter_energia = JsCode(f"""
                    function(params) {{
                        // ... (sua lógica de valueFormatter para Energia)
                        if (params.value == null || typeof params.value === 'undefined' || String(params.value).trim() === '') {{
                            return '';
                        }}
                        let num = Number(params.value);
                        if (isNaN(num)) {{ return ''; }}
                        try {{
                            return num.toFixed({casas_decimais_energia});
                        }} catch (e) {{
                            console.error("Erro valueFormatter Energia:", e, params.value);
                            return String(params.value);
                        }}
                    }}
                """)

                gb.configure_column(
                    field=col_energia_v_nome,
                    headerName=col_energia_v_nome,
                    type=["numericColumn"],
                    filter=False,
                    valueFormatter=js_value_formatter_energia,
                    cellStyle=cell_style_cores_js,
                    tooltipValueGetter=tooltip_preco_energia_js,
                    tooltipComponent=custom_tooltip_component_js,
                    minWidth=60, # Ajustar conforme necessário
                    flex=1
                )

            # Configuração Coluna 'Preço Energia Fora Vazio (€/kWh)'
            col_energia_f_nome = 'Fora Vazio (€/kWh)'
            if col_energia_f_nome in df_resultados_para_aggrid.columns:
                casas_decimais_energia = 4 # Preço energia geralmente tem mais casas decimais
        
                js_value_formatter_energia = JsCode(f"""
                    function(params) {{
                        // ... (sua lógica de valueFormatter para Energia)
                        if (params.value == null || typeof params.value === 'undefined' || String(params.value).trim() === '') {{
                            return '';
                        }}
                        let num = Number(params.value);
                        if (isNaN(num)) {{ return ''; }}
                        try {{
                            return num.toFixed({casas_decimais_energia});
                        }} catch (e) {{
                            console.error("Erro valueFormatter Energia:", e, params.value);
                            return String(params.value);
                        }}
                    }}
                """)

                gb.configure_column(
                    field=col_energia_f_nome,
                    headerName=col_energia_f_nome,
                    type=["numericColumn"],
                    filter=False,
                    valueFormatter=js_value_formatter_energia,
                    cellStyle=cell_style_cores_js,
                    tooltipValueGetter=tooltip_preco_energia_js,
                    tooltipComponent=custom_tooltip_component_js,
                    minWidth=60, # Ajustar conforme necessário
                    flex=1
                )

            # Configuração Coluna 'Preço Energia Cheias (€/kWh)'
            col_energia_c_nome = 'Cheias (€/kWh)'
            if col_energia_c_nome in df_resultados_para_aggrid.columns:
                casas_decimais_energia = 4 # Preço energia geralmente tem mais casas decimais
        
                js_value_formatter_energia = JsCode(f"""
                    function(params) {{
                        // ... (sua lógica de valueFormatter para Energia)
                        if (params.value == null || typeof params.value === 'undefined' || String(params.value).trim() === '') {{
                            return '';
                        }}
                        let num = Number(params.value);
                        if (isNaN(num)) {{ return ''; }}
                        try {{
                            return num.toFixed({casas_decimais_energia});
                        }} catch (e) {{
                            console.error("Erro valueFormatter Energia:", e, params.value);
                            return String(params.value);
                        }}
                    }}
                """)

                gb.configure_column(
                    field=col_energia_c_nome,
                    headerName=col_energia_c_nome,
                    type=["numericColumn"],
                    filter=False,
                    valueFormatter=js_value_formatter_energia,
                    cellStyle=cell_style_cores_js,
                    tooltipValueGetter=tooltip_preco_energia_js,
                    tooltipComponent=custom_tooltip_component_js,
                    minWidth=60, # Ajustar conforme necessário
                    flex=1
                )

            # Configuração Coluna 'Preço Energia Ponta (€/kWh)'
            col_energia_p_nome = 'Ponta (€/kWh)'
            if col_energia_p_nome in df_resultados_para_aggrid.columns:
                casas_decimais_energia = 4
        
                js_value_formatter_energia = JsCode(f"""
                    function(params) {{
                        // ... (sua lógica de valueFormatter para Energia)
                        if (params.value == null || typeof params.value === 'undefined' || String(params.value).trim() === '') {{
                            return '';
                        }}
                        let num = Number(params.value);
                        if (isNaN(num)) {{ return ''; }}
                        try {{
                            return num.toFixed({casas_decimais_energia});
                        }} catch (e) {{
                            console.error("Erro valueFormatter Energia:", e, params.value);
                            return String(params.value);
                        }}
                    }}
                """)

                gb.configure_column(
                    field=col_energia_p_nome,
                    headerName=col_energia_p_nome,
                    type=["numericColumn"],
                    filter=False,
                    valueFormatter=js_value_formatter_energia,
                    cellStyle=cell_style_cores_js,
                    tooltipValueGetter=tooltip_preco_energia_js,
                    tooltipComponent=custom_tooltip_component_js,
                    minWidth=60, # Ajustar conforme necessário
                    flex=1
                )

            #Tooltip Preço Potencia
            tooltip_preco_potencia_js = JsCode("""
            function(params) {
                // params.value é o valor exibido na célula (Potência (€/dia) final sem IVA)
                // params.data contém todos os dados da linha
                if (!params.data) {
                    // Se não houver dados da linha, retorna apenas o valor da célula como tooltip
                    return String(params.value); 
                }

                // console.log para depurar os valores que chegam:
                console.log("Tooltip Potência Dados:", 
                    params.data.tooltip_pot_comerc_sem_tar, 
                    params.data.tooltip_pot_tar_bruta, 
                    params.data.tooltip_pot_ts_aplicada, 
                    params.data.tooltip_pot_desconto_ts_valor,
                    params.value // Valor da célula
                );                                   


                // Aceder aos campos que adicionou em Python
                // Use (params.data.NOME_CAMPO || 0) para tratar casos onde o campo pode ser nulo/undefined
                const comercializador = parseFloat(params.data.tooltip_pot_comerc_sem_tar || 0);
                const tarBruta = parseFloat(params.data.tooltip_pot_tar_bruta || 0);
                const tsAplicada = params.data.tooltip_pot_ts_aplicada; 
                const descontoTSValor = parseFloat(params.data.tooltip_pot_desconto_ts_valor || 0);

                // Função helper para formatar números com 4 casas decimais
                const formatPrice = (num) => {
                    if (typeof num === 'number' && !isNaN(num)) {
                        return num.toFixed(4);
                    }
                    return 'N/A'; // Ou algum outro placeholder
                };

                let tooltipParts = [];
                tooltipParts.push("<b>Decomposição Potência (s/IVA):</b>");
                tooltipParts.push("Comercializador (s/TAR): " + formatPrice(comercializador) + " €/dia");
                tooltipParts.push("TAR (Tarifa Acesso Redes): " + formatPrice(tarBruta) + " €/dia");

                if (tsAplicada === true && descontoTSValor > 0) { // Garantir que tsAplicada é explicitamente true
                    tooltipParts.push("Desconto Tarifa Social: -" + formatPrice(descontoTSValor) + " €/dia");
                }
        
                tooltipParts.push("----------------------------------------------------");
                tooltipParts.push("<b>Custo Final : " + formatPrice(parseFloat(params.value)) + " €/dia</b>");

                var finalTooltipHtml = tooltipParts.join("<br>");
                // Log para ver o HTML final
                // console.log("Tooltip HTML Final para Potência:", finalTooltipHtml);
                return finalTooltipHtml;
            }
            """)

            # Exemplo para a coluna 'Preço Potência (€/dia)'
            col_potencia_nome = 'Potência (€/dia)'
            if col_potencia_nome in df_resultados_para_aggrid.columns and col_potencia_nome in cols_para_cor:
                casas_decimais_pot = 4
        
                js_value_formatter_potencia = JsCode(f"""
                    function(params) {{
                        // ... (sua lógica de valueFormatter para potência)
                        if (params.value == null || typeof params.value === 'undefined' || String(params.value).trim() === '') {{
                            return '';
                        }}
                        let num = Number(params.value);
                        if (isNaN(num)) {{ return ''; }}
                        try {{
                            return num.toFixed({casas_decimais_pot});
                        }} catch (e) {{
                            console.error("Erro valueFormatter Potência:", e, params.value);
                            return String(params.value);
                        }}
                    }}
                """)

                #Configuração Coluna Potência
                gb.configure_column(
                    field=col_potencia_nome,
                    headerName=col_potencia_nome,
                    type=["numericColumn"],
                    filter=False,
                    valueFormatter=js_value_formatter_potencia,
                    cellStyle=cell_style_cores_js,
                    tooltipValueGetter=tooltip_preco_potencia_js,
                    tooltipComponent=custom_tooltip_component_js,
                    # tooltipComponentParams={'color': '#aabbcc'}
                )

            #Tooltip Custo Total
            tooltip_custo_total_js = JsCode("""
            function(params) {
                if (!params.data) { return String(params.value); }

                const formatCurrency = (num) => {
                    if (typeof num === 'number' && !isNaN(num)) {
                        return num.toFixed(2); // 2 casas decimais para custos
                    }
                    return 'N/A';
                };

                const nomeTarifario = params.data.NomeParaExibir || "Tarifário"; // Fallback se não existir

                // Linha de título do tooltip atualizada para incluir o nome do tarifário
                let tooltipParts = [
                    "<i>" + nomeTarifario + "</i>", // Nome do tarifário em negrito na primeira linha
                    "<b>Decomposição Custo Total:</b>" // Título em itálico na segunda linha
                    // Pode adicionar uma linha em branco se quiser mais espaçamento: ""
                ];
                tooltipParts.push("------------------------------------");
                                            
                const energia_siva = parseFloat(params.data.tt_cte_energia_siva || 0);
                const potencia_siva = parseFloat(params.data.tt_cte_potencia_siva || 0);
                const iec_siva = parseFloat(params.data.tt_cte_iec_siva || 0);
                const dgeg_siva = parseFloat(params.data.tt_cte_dgeg_siva || 0);
                const cav_siva = parseFloat(params.data.tt_cte_cav_siva || 0);
                const total_siva = parseFloat(params.data.tt_cte_total_siva || 0);
                const iva_6 = parseFloat(params.data.tt_cte_valor_iva_6_total || 0);
                const iva_23 = parseFloat(params.data.tt_cte_valor_iva_23_total || 0);
                const subtotal_civa_antes_desc_acr = parseFloat(params.data.tt_cte_subtotal_civa || 0);
                const desc_finais_valor = parseFloat(params.data.tt_cte_desc_finais_valor || 0);
                const acres_finais_valor = parseFloat(params.data.tt_cte_acres_finais_valor || 0);

                const custo_total_celula = parseFloat(params.value);

                tooltipParts.push("Total Energia s/IVA: " + formatCurrency(energia_siva) + " €");
                tooltipParts.push("Total Potência s/IVA: " + formatCurrency(potencia_siva) + " €");
                // Condição para mostrar IEC: se o valor for maior que zero OU se a tarifa social não estiver ativa.
                // Precisa adicionar 'tarifa_social' (a flag booleana global) aos dados da linha se quiser esta condição precisa.
                // Por agora, vou simplificar para mostrar se iec_siva > 0.
                if (iec_siva !== 0) { 
                    tooltipParts.push("IEC s/IVA: " + formatCurrency(iec_siva) + " €");
                }
                if (dgeg_siva !== 0) {
                    tooltipParts.push("DGEG s/IVA: " + formatCurrency(dgeg_siva) + " €");
                }
                if (cav_siva !== 0) {
                    tooltipParts.push("CAV s/IVA: " + formatCurrency(cav_siva) + " €");
                }
                tooltipParts.push("<b>Subtotal s/IVA: " + formatCurrency(total_siva) + " €</b>");
                tooltipParts.push("------------------------------------");
                if (iva_6 !== 0) {
                    tooltipParts.push("Valor IVA (6%): " + formatCurrency(iva_6) + " €");
                }
                if (iva_23 !== 0) {
                    tooltipParts.push("Valor IVA (23%): " + formatCurrency(iva_23) + " €");
                }
                tooltipParts.push("<b>Subtotal c/IVA: " + formatCurrency(subtotal_civa_antes_desc_acr) + " €</b>");
                tooltipParts.push("------------------------------------");
        
                // Mostrar descontos e acréscimos apenas se existirem
                if (desc_finais_valor !== 0) {
                    tooltipParts.push("Outros Descontos: -" + formatCurrency(desc_finais_valor) + " €");
                }
                if (acres_finais_valor !== 0) {
                    tooltipParts.push("Outros Acréscimos: +" + formatCurrency(acres_finais_valor) + " €");
                }
        
                if (desc_finais_valor !== 0 || acres_finais_valor !== 0) {
                    tooltipParts.push("------------------------------------");
                }
                tooltipParts.push("<b>Custo Total c/IVA: " + formatCurrency(custo_total_celula) + " €</b>");

                return tooltipParts.join("<br>");
            }
            """)

            # Configuração Coluna 'Custo Total (€)'
            col_custo_total_nome = 'Total (€)'
            if col_custo_total_nome in df_resultados_para_aggrid.columns:
                casas_decimais_total = 2
        
                js_value_formatter_energia = JsCode(f"""
                    function(params) {{
                        // ... (sua lógica de valueFormatter para Energia)
                        if (params.value == null || typeof params.value === 'undefined' || String(params.value).trim() === '') {{
                            return '';
                        }}
                        let num = Number(params.value);
                        if (isNaN(num)) {{ return ''; }}
                        try {{
                            return num.toFixed({casas_decimais_total});
                        }} catch (e) {{
                            console.error("Erro valueFormatter Energia:", e, params.value);
                            return String(params.value);
                        }}
                    }}
                """)

                gb.configure_column(
                    field=col_custo_total_nome,
                    headerName=col_custo_total_nome,
                    type=["numericColumn"],
                    filter=False,
                    valueFormatter=js_value_formatter_energia,
                    cellStyle=cell_style_cores_js,
                    tooltipValueGetter=tooltip_custo_total_js, 
                    tooltipComponent=custom_tooltip_component_js, 
                    minWidth=80, # Ajustar conforme necessário
                    flex=1
                )

            # --- Configuração de Colunas com Set Filter ---
            set_filter_params = {
                'buttons': ['apply', 'reset'],
                'excelMode': 'mac',
                'suppressMiniFilter': False, # Garante que a caixa de pesquisa dentro do Set Filter aparece
            }

            is_visible = col_name in colunas_visiveis_presentes

            text_columns_with_set_filter = ['Tipo', 'Segmento', 'Faturação', 'Pagamento']
            for col_name in text_columns_with_set_filter:
                if col_name in df_resultados_para_aggrid.columns:
                    min_w = 100 if col_name == 'Tipo' else 120
                    fx = 0.5 if col_name == 'Tipo' else 0.75
                    gb.configure_column(
                        col_name, 
                        headerName=col_name,
                        minWidth=min_w,
                        flex=fx,
                        filter='agSetColumnFilter',
                        filterParams=set_filter_params,
                        cellStyle={
                            'textAlign': 'center', 
                            'borderRadius': '11px',
                            'padding': '10px 10px',
                            'backgroundColor': '#f0f0f0' 
                        },
                        hide=(not is_visible) 

                    )
            is_visible_comerc = 'Comercializador' in colunas_visiveis_presentes

            # --- Configuração Coluna Comercializador (Text Filter) ---
            if 'Comercializador' in df_resultados_para_aggrid.columns:
                gb.configure_column(
                    "Comercializador",
                    headerName="Comercializador",
                    minWidth=50,
                    flex=1,
                    filter='agTextColumnFilter',
                    cellStyle={
                        'textAlign': 'center', 
                        'borderRadius': '11px', 
                        'padding': '10px 10px', 
                        'backgroundColor': '#f0f0f0'
                    },
                    hide=(not is_visible_comerc)
                )

            # Configurar outras colunas (Tipo, Comercializador e colunas de dados)
            for col_nome_num in df_resultados_para_aggrid.columns:
                if col_nome_num in cols_para_cor:
                    casas_decimais = 4 if '€/kWh' in col_nome_num or '€/dia' in col_nome_num else 2
            
                    # Definir o JsCode para valueFormatter DENTRO do loop
                    # para que capture o valor correto de 'casas_decimais' para esta coluna específica.
                    js_value_formatter_para_coluna = JsCode(f"""
                        function(params) {{
                            // Verificar se o valor é nulo ou não é um número
                            if (params.value == null || typeof params.value === 'undefined' || String(params.value).trim() === '') {{
                                return ''; // Retornar string vazia para valores nulos, indefinidos ou vazios
                            }}
                    
                            let num = Number(params.value); // Tentar converter para número
                    
                            if (isNaN(num)) {{
                                // Se após a tentativa de conversão ainda for NaN, retornar string vazia
                                // ou pode optar por retornar o valor original se fizer sentido: return params.value;
                                return ''; 
                            }}
                    
                            // Se for um número válido, formatá-lo com o número correto de casas decimais
                            try {{
                                return num.toFixed({casas_decimais});
                            }} catch (e) {{
                                console.error("Erro ao formatar o número no valueFormatter:", e, 
                                                "Valor original:", params.value, 
                                                "Casas decimais:", {casas_decimais});
                                return String(params.value); // Fallback para string do valor original em caso de erro inesperado
                            }}
                        }}
                    """)

                    gb.configure_column(
                        field=col_nome_num,
                        headerName=col_nome_num,
                        type=["numericColumn", "numberColumnFilter"], 
                        valueFormatter=js_value_formatter_para_coluna,
                        cellStyle=cell_style_cores_js,
                        minWidth=50,
                        flex=1
                    )

            # --- 3. Formatação para "O Meu Tarifário" ---
            get_row_style_meu_tarifario_js = JsCode("""
            function(params) {
                // Verifica se params.data existe e se NomeParaExibir é uma string que COMEÇA COM "O Meu Tarifário"
                if (params.data && typeof params.data.NomeParaExibir === 'string' && params.data.NomeParaExibir.startsWith('O Meu Tarifário')) {
                    return { fontWeight: 'bold' }; // Aplica negrito a toda a linha
                }
                return null; // Sem estilo especial para outras linhas
            }
            """)

            # Ocultar 'Tipo' e 'LinkAdesao' na vista simplificada se não estiverem em colunas_visiveis_presentes
            if vista_simplificada:
                if 'Tipo' not in colunas_visiveis_presentes and 'Tipo' in df_resultados_para_aggrid.columns:
                    gb.configure_column(field='Tipo', hide=True)
                if 'LinkAdesao' not in colunas_visiveis_presentes and 'LinkAdesao' in df_resultados_para_aggrid.columns:
                    gb.configure_column(field='LinkAdesao', hide=True)
                # Oculte outras colunas que estão nos dados mas não são visíveis na vista simplificada
                colunas_desktop_a_ocultar_na_vista_movel = ['Segmento', 'Faturação', 'Pagamento', 'Comercializador']
                for col_ocultar in colunas_desktop_a_ocultar_na_vista_movel:
                    if col_ocultar not in colunas_visiveis_presentes and col_ocultar in df_resultados_para_aggrid.columns:
                        gb.configure_column(field=col_ocultar, hide=True)


            # Ocultar colunas de dados de tooltip
            colunas_de_dados_tooltip_a_ocultar = [
                'info_notas', 
                'tooltip_pot_comerc_sem_tar', 'tooltip_pot_tar_bruta', 'tooltip_pot_ts_aplicada', 'tooltip_pot_desconto_ts_valor',
                # Energia Simples (S)
                'tooltip_energia_S_comerc_sem_tar', 'tooltip_energia_S_tar_bruta', 
                'tooltip_energia_S_tse_declarado_incluido', 'tooltip_energia_S_tse_valor_nominal',
                'tooltip_energia_S_ts_aplicada_flag', 'tooltip_energia_S_ts_desconto_valor',
        
                # Energia Vazio (V) - Adicione se tiver opção Bi ou Tri
                'tooltip_energia_V_comerc_sem_tar', 'tooltip_energia_V_tar_bruta', 
                'tooltip_energia_V_tse_declarado_incluido', 'tooltip_energia_V_tse_valor_nominal',
                'tooltip_energia_V_ts_aplicada_flag', 'tooltip_energia_V_ts_desconto_valor',

                # Energia Fora Vazio (F) - Adicione se tiver opção Bi
                'tooltip_energia_F_comerc_sem_tar', 'tooltip_energia_F_tar_bruta', 
                'tooltip_energia_F_tse_declarado_incluido', 'tooltip_energia_F_tse_valor_nominal',
                'tooltip_energia_F_ts_aplicada_flag', 'tooltip_energia_F_ts_desconto_valor',

                # Energia Cheias (C) - Adicione se tiver opção Tri
                'tooltip_energia_C_comerc_sem_tar', 'tooltip_energia_C_tar_bruta', 
                'tooltip_energia_C_tse_declarado_incluido', 'tooltip_energia_C_tse_valor_nominal',
                'tooltip_energia_C_ts_aplicada_flag', 'tooltip_energia_C_ts_desconto_valor',

                # Energia Ponta (P) - Adicione se tiver opção Tri
                'tooltip_energia_P_comerc_sem_tar', 'tooltip_energia_P_tar_bruta', 
                'tooltip_energia_P_tse_declarado_incluido', 'tooltip_energia_P_tse_valor_nominal',
                'tooltip_energia_P_ts_aplicada_flag', 'tooltip_energia_P_ts_desconto_valor',

                # Para Custo Total
                'tt_cte_energia_siva', 'tt_cte_potencia_siva', 'tt_cte_iec_siva',
                'tt_cte_dgeg_siva', 'tt_cte_cav_siva', 'tt_cte_total_siva',
                'tt_cte_valor_iva_6_total', 'tt_cte_valor_iva_23_total',
                'tt_cte_subtotal_civa','tt_cte_desc_finais_valor','tt_cte_acres_finais_valor'
            ]

            for col_para_ocultar in colunas_de_dados_tooltip_a_ocultar:
                if col_para_ocultar in df_resultados_para_aggrid.columns:
                    gb.configure_column(field=col_para_ocultar, hide=True)

            gb.configure_grid_options(
                domLayout='autoHeight', # Para altura automática
                getRowStyle=get_row_style_meu_tarifario_js,
                suppressContextMenu=True  # Adicionado para desativar o menu de contexto
            )

            gb.configure_default_column(headerClass='center-header')
            gridOptions = gb.build()
            custom_css = {
                ".ag-header-cell-label": {
                    "justify-content": "center !important",
                    "text-align": "center !important",
                    "font-size": "14px !important",      # <-- aumenta o header
                    "font-weight": "bold !important"
                },
                ".ag-center-header": {
                    "justify-content": "center !important",
                    "text-align": "center !important",
                    "font-size": "14px !important"       # <-- reforço para headers
                },
                ".ag-cell": {
                    "font-size": "14px !important"       # <-- aumenta valores das células
                },
                ".ag-center-cols-clip": {"justify-content": "center !important", "text-align": "center !important"}
            }

            # --- Construir a chave dinâmica para a AgGrid ---
            # Inclua todos os inputs cujo estado deve levar a um "reset" da AgGrid
            # (filtros manuais, ordenação manual na grelha são resetados quando a key muda)
            key_parts_para_aggrid = [
                str(potencia), str(opcao_horaria), str(mes),
                str(data_inicio), str(data_fim), str(dias),
                # Consumos
                str(consumo_simples), str(consumo_vazio), str(consumo_fora_vazio),
                str(consumo_cheias), str(consumo_ponta),
                # Opções adicionais (garanta que estas variáveis estão definidas no escopo)
                str(tarifa_social), str(familia_numerosa), str(comparar_indexados),
                str(valor_dgeg_user), str(valor_cav_user),
                str(incluir_quota_acp), str(desconto_continente),
                # Inputs OMIE manuais (usar os valores do session_state que alimentam os number_input)
                str(st.session_state.get('omie_s_input_field')),
            str(st.session_state.get('omie_v_input_field')),
                str(st.session_state.get('omie_f_input_field')),
                str(st.session_state.get('omie_c_input_field')),
                str(st.session_state.get('omie_p_input_field')),
                # Estado e dados do "Meu Tarifário"
                str(meu_tarifario_ativo)
            ]
            if meu_tarifario_ativo and 'meu_tarifario_calculado' in st.session_state:
            # Adicionar uma representação do "Meu Tarifário" à chave,
                key_parts_para_aggrid.append(str(st.session_state['meu_tarifario_calculado'].get('Total (€)', '')))

            # Juntar todas as partes para formar a chave única
            aggrid_dinamica_key = "aggrid_principal_" + "_".join(key_parts_para_aggrid)


            # Exibir a Grelha
            grid_response = AgGrid(
                df_resultados_para_aggrid,
                gridOptions=gridOptions,
                custom_css=custom_css,
                # update_mode diz ao Streamlit para atualizar os dados quando os filtros ou a ordenação mudam na AgGrid
                update_mode=GridUpdateMode.FILTERING_CHANGED | GridUpdateMode.SORTING_CHANGED | GridUpdateMode.SELECTION_CHANGED,
                allow_unsafe_jscode=True,
                fit_columns_on_grid_load=True,
                theme='alpine', 
                key=aggrid_dinamica_key, # Uma key para a instância interativa
                enable_enterprise_modules=True,
                # reload_data=True # Considere usar se os dados de entrada (df_resultados_para_aggrid) puderem mudar dinamicamente por outras interações
            )
            # ---- FIM DA CONFIGURAÇÃO DO AGGRID ----

        st.markdown("<a id='exportar-excel-detalhada'></a>", unsafe_allow_html=True)
        #st.markdown("---")
        with st.expander("📥 Exportar Tabela Detalhada para Excel"):
            colunas_dados_tooltip_a_ocultar = [
                'info_notas', 'LinkAdesao',
                'tooltip_pot_comerc_sem_tar', 'tooltip_pot_tar_bruta', # ... e todas as outras ...
                'tt_cte_subtotal_civa','tt_cte_desc_finais_valor','tt_cte_acres_finais_valor'
            ]

            if 'df_resultados_para_aggrid' in locals() and \
            isinstance(df_resultados_para_aggrid, pd.DataFrame) and \
            not df_resultados_para_aggrid.empty and \
            'colunas_visiveis_presentes' in locals() and \
            isinstance(colunas_visiveis_presentes, list) and \
            'colunas_dados_tooltip_a_ocultar' in locals() and \
            isinstance(colunas_dados_tooltip_a_ocultar, list):

            # Colunas disponíveis para seleção:
            # Começamos com todas as colunas que estão no DataFrame que alimenta o AgGrid.
                todas_as_colunas_no_df_aggrid = df_resultados_para_aggrid.columns.tolist()
            
            # Organizar as opções para o multiselect:
            # 1. Colunas visíveis primeiro
            # 2. Depois, colunas de tooltip (que não estão já nas visíveis)
            # 3. Depois, outras colunas (se houver e fizer sentido oferecer)
            
                opcoes_export_excel = []
            # Adicionar colunas visíveis primeiro, mantendo a sua ordem
                for col_vis in colunas_visiveis_presentes:
                    if col_vis in todas_as_colunas_no_df_aggrid and col_vis not in opcoes_export_excel:
                        opcoes_export_excel.append(col_vis)
            
            # Adicionar colunas de dados de tooltip que não estão já nas visíveis
            # e que existem no df_resultados_para_aggrid
                for col_tooltip in colunas_dados_tooltip_a_ocultar: # Esta lista contém os nomes das colunas de tooltip
                    if col_tooltip in todas_as_colunas_no_df_aggrid and col_tooltip not in opcoes_export_excel:
                        opcoes_export_excel.append(col_tooltip)
            
            # Adicionar quaisquer outras colunas restantes do df_resultados_para_aggrid se desejado
            # (excluindo as que já foram adicionadas)
            # Se 'colunas_para_aggrid_final' foi usado para criar df_resultados_para_aggrid,
            # ele pode já ser uma boa base, mas vamos usar todas_as_colunas_no_df_aggrid para garantir
                for col_restante in todas_as_colunas_no_df_aggrid:
                    if col_restante not in opcoes_export_excel:
                        opcoes_export_excel.append(col_restante)

            # Colunas pré-selecionadas: apenas as que estão atualmente visíveis no AgGrid
                default_cols_excel = [col for col in colunas_visiveis_presentes if col in opcoes_export_excel]
                
                if not default_cols_excel and 'NomeParaExibir' in opcoes_export_excel:
                    default_cols_excel.append('NomeParaExibir')
                if not default_cols_excel and 'Total (€)' in opcoes_export_excel:
                    default_cols_excel.append('Total (€)')


                colunas_para_exportar_excel_selecionadas = st.multiselect(
                    "Selecione as colunas para exportar para Excel:",
                    options=opcoes_export_excel, 
                    default=default_cols_excel,
                    key="cols_export_excel_selector_dados_com_tooltips"
                )
                
                def exportar_excel_completo(df_para_exportar, styler_obj, resumo_html_para_excel, poupanca_texto_para_excel, identificador_cor_cabecalho, meu_tarifario_ativo_flag, personalizado_ativo_flag):
                    output_excel_buffer = io.BytesIO() 
                    with pd.ExcelWriter(output_excel_buffer, engine='openpyxl') as writer_excel:
                        sheet_name_excel = 'Tiago Felicia - Eletricidade'

                        # --- Escrever Resumo ---
                        dados_resumo_formatado = []
                        if resumo_html_para_excel:
                            soup_resumo = BeautifulSoup(resumo_html_para_excel, "html.parser")
                            
                            titulo_resumo = soup_resumo.find('h5')
                            if titulo_resumo:
                                dados_resumo_formatado.append([titulo_resumo.get_text(strip=True), None])

                            itens_lista_resumo = soup_resumo.find_all('li')
                            linha_filtros_texto = ""
                            linha_potencia_texto = ""
                            outras_linhas_resumo = []

                            for item in itens_lista_resumo:
                                # Usar .get_text() para obter o conteúdo limpo do item da lista
                                texto_item = item.get_text(separator=' ', strip=True)
                                    
                                if "Segmento:" in texto_item:
                                    linha_filtros_texto = texto_item
                                elif "kVA" in texto_item:
                                    linha_potencia_texto = texto_item
                                else:
                                    # Processar as outras linhas normalmente
                                    parts = texto_item.split(':', 1)
                                    if len(parts) == 2:
                                        outras_linhas_resumo.append([parts[0].strip() + ":", parts[1].strip()])
                                    else:
                                        outras_linhas_resumo.append([texto_item, None])
                                
                            # Adicionar a linha combinada primeiro, na ordem que pediu
                            if linha_filtros_texto or linha_potencia_texto:
                                dados_resumo_formatado.append([linha_filtros_texto, linha_potencia_texto])
                            
                            # Adicionar o resto do resumo
                            dados_resumo_formatado.extend(outras_linhas_resumo)
                            # --- FIM DA LÓGICA ALTERADA ---
            
                        df_resumo_obj = pd.DataFrame(dados_resumo_formatado)

                        # 1. Deixe o Pandas criar/ativar a folha na primeira escrita
                        df_resumo_obj.to_excel(writer_excel, sheet_name=sheet_name_excel, index=False, header=False, startrow=0)

                        # 2. AGORA obtenha a referência à worksheet, que certamente existe
                        worksheet_excel = writer_excel.sheets[sheet_name_excel]

                        # Formatar Resumo (Negrito)
                        bold_font_obj = Font(bold=True) # Font já deve estar importado de openpyxl.styles
                        for i_resumo in range(len(df_resumo_obj)):
                            excel_row_idx_resumo = i_resumo + 1 # Linhas do Excel são 1-based
                            cell_resumo_rotulo = worksheet_excel.cell(row=excel_row_idx_resumo, column=1)
                            cell_resumo_rotulo.font = bold_font_obj
                            if df_resumo_obj.shape[1] > 1 and pd.notna(df_resumo_obj.iloc[i_resumo, 1]):
                                cell_resumo_valor = worksheet_excel.cell(row=excel_row_idx_resumo, column=2)
                                cell_resumo_valor.font = bold_font_obj
            
                        worksheet_excel.column_dimensions['A'].width = 35
                        worksheet_excel.column_dimensions['B'].width = 65

                        linha_atual_no_excel_escrita = len(df_resumo_obj) + 1

                        # --- Escrever Mensagem de Poupança ---
                        if poupanca_texto_para_excel: # Verifica se há texto para a mensagem de poupança
                            linha_atual_no_excel_escrita += 1 # Adiciona uma linha em branco
                
                            cor_p = st.session_state.get('poupanca_excel_cor', "000000") # Cor do session_state
                            negrito_p = st.session_state.get('poupanca_excel_negrito', False) # Negrito do session_state
                
                            poupanca_cell_escrita = worksheet_excel.cell(row=linha_atual_no_excel_escrita, column=1, value=poupanca_texto_para_excel)
                            poupanca_font_escrita = Font(bold=negrito_p, color=cor_p)
                            poupanca_cell_escrita.font = poupanca_font_escrita

                            # --- JUNTAR CÉLULAS ---
                            worksheet_excel.merge_cells(start_row=linha_atual_no_excel_escrita, start_column=1, end_row=linha_atual_no_excel_escrita, end_column=4)

                            # Aplicar alinhamento à célula fundida (a célula do canto superior esquerdo, poupanca_cell_escrita)
                            poupanca_cell_escrita.alignment = Alignment(wrap_text=True, horizontal='left', vertical='top')
                    
                            linha_atual_no_excel_escrita += 1 # Avança para a próxima linha após a mensagem de poupança
            
                        linha_inicio_tab_dados_excel = linha_atual_no_excel_escrita + 3

                        # --- Adicionar linha de informação da simulação ---
                        # Adiciona uma linha em branco antes desta nova linha de informação
                        linha_info_simulacao_excel = linha_atual_no_excel_escrita + 1 

                        data_hoje_obj = datetime.date.today() # datetime já deve estar importado
                        data_hoje_formatada_str = data_hoje_obj.strftime('%d/%m/%Y')
                
                        espacador_info = "                                                                      " # Exemplo: 70 espaços

                        texto_completo_info = (
                            f"          Simulação em {data_hoje_formatada_str}{espacador_info}"
                            f"https://www.tiagofelicia.pt{espacador_info}"
                            f"Tiago Felícia"
                        )

                        # Escrever o texto completo na primeira célula da área a ser fundida (Coluna A)
                        info_cell = worksheet_excel.cell(row=linha_info_simulacao_excel, column=1)
                        info_cell.value = texto_completo_info
                
                        # Aplicar negrito à célula
                        # Reutilizar bold_font_obj que já foi definido para o resumo, ou criar um novo se precisar de formatação diferente.
                        # Assumindo que bold_font_obj é Font(bold=True) e está no escopo.
                        # Se não, defina-o: from openpyxl.styles import Font; bold_font_obj = Font(bold=True)
                        if 'bold_font_obj' in locals() or 'bold_font_obj' in globals():
                            info_cell.font = bold_font_obj # Reutiliza o bold_font_obj do resumo
                        else:
                            info_cell.font = Font(bold=True) # Cria um novo se não existir

                        # Fundir as colunas A, B, C, e D para esta linha
                        worksheet_excel.merge_cells(start_row=linha_info_simulacao_excel, start_column=1, end_row=linha_info_simulacao_excel, end_column=4)
                
                        # Ajustar alinhamento da célula fundida (info_cell é a célula do topo-esquerda da área fundida)
                        # Alinhado à esquerda, centralizado verticalmente, com quebra de linha se necessário.
                        info_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True) 

                        # A linha de início para a tabela de dados principal virá depois desta linha de informação
                        # Adicionamos +1 para esta linha de informação e +1 para uma linha em branco antes da tabela
                        linha_inicio_tab_dados = linha_info_simulacao_excel + 2 
                
                        # --- Fim da adição da linha de informação ---

                        for row in worksheet_excel.iter_rows(min_row=1, max_row=worksheet_excel.max_row+100, min_col=1, max_col=20):
                            for cell in row:
                                cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

                        # O Styler escreverá na mesma folha 'sheet_name_excel'
                        styler_obj.to_excel(
                            writer_excel,
                            sheet_name=sheet_name_excel,
                            index=False,
                            startrow=linha_inicio_tab_dados_excel - 1, # startrow é 0-indexed
                            columns=df_para_exportar.columns.tolist()
                        )

                        # Determina cor do cabeçalho conforme a opção horária e ciclo
                        opcao_horaria_lower = str(opcao_horaria).lower() if 'opcao_horaria' in locals() else "simples"
                        cor_fundo = "A6A6A6"   # padrão Simples
                        cor_fonte = "000000"    # padrão preto

                        if isinstance(identificador_cor_cabecalho, str):
                            id_lower = identificador_cor_cabecalho.lower()
                            if id_lower == "simples": # Já coberto pelo default
                                pass
                            elif "bi-horário" in id_lower and "diário" in id_lower:
                                cor_fundo = "A9D08E"; cor_fonte = "000000"
                            elif "bi-horário" in id_lower and "semanal" in id_lower:
                                cor_fundo = "8EA9DB"; cor_fonte = "000000"
                            elif "tri-horário" in id_lower and "diário" in id_lower:
                                cor_fundo = "BF8F00"; cor_fonte = "FFFFFF"
                            elif "tri-horário" in id_lower and "semanal" in id_lower:
                                cor_fundo = "C65911"; cor_fonte = "FFFFFF"

                        # linha_inicio_tab_dados_excel já existe na função e corresponde ao header da tabela
                        for col_idx, _ in enumerate(df_para_exportar.columns):
                            celula = worksheet_excel.cell(row=linha_inicio_tab_dados_excel, column=col_idx + 1)
                            celula.fill = PatternFill(start_color=cor_fundo, end_color=cor_fundo, fill_type="solid")
                            celula.font = Font(color=cor_fonte, bold=True)

                        # ---- INÍCIO: ADICIONAR LEGENDA DE CORES APÓS A TABELA ----
                        # Calcular a linha de início para a legenda
                        # df_para_exportar é o DataFrame que foi escrito pelo styler (ex: df_export_final)
                        numero_linhas_dados_tabela_principal = len(df_para_exportar)
                        # linha_inicio_tab_dados_excel é a linha do cabeçalho da tabela principal (1-indexada)
                        ultima_linha_tabela_principal = linha_inicio_tab_dados_excel + numero_linhas_dados_tabela_principal
                        
                        linha_legenda_bloco_inicio = ultima_linha_tabela_principal + 2 # Deixa uma linha em branco após a tabela

                        # Título da Legenda
                        titulo_legenda_cell = worksheet_excel.cell(row=linha_legenda_bloco_inicio, column=1, value="Tipos de Tarifário:")
                        # Reutilizar bold_font_obj se definido ou criar um novo
                        if 'bold_font_obj' in locals() or 'bold_font_obj' in globals():
                            titulo_legenda_cell.font = bold_font_obj
                        else:
                            titulo_legenda_cell.font = Font(bold=True)
                        worksheet_excel.merge_cells(start_row=linha_legenda_bloco_inicio, start_column=1, end_row=linha_legenda_bloco_inicio, end_column=4) # Fundir para o título
                        
                        linha_legenda_item_atual = linha_legenda_bloco_inicio + 1 # Primeira linha para item da legenda

                        titulo_legenda_cell.alignment = Alignment(horizontal='center', vertical='center')

                        itens_legenda_excel = []
                        # 1. Adicionar "O Meu Tarifário" se estiver ativo
                        if meu_tarifario_ativo_flag:
                            itens_legenda_excel.append(
                                {"cf": "FF0000", "ct": "FFFFFF", "b": True, "tA": "O Meu Tarifário", "tB": "Tarifário configurado pelo utilizador."}
                            )
                        # 2. Adicionar "Tarifário Personalizado" se estiver ativo
                        if personalizado_ativo_flag:
                            itens_legenda_excel.append(
                                {"cf": "92D050", "ct": "FFFFFF", "b": True, "tA": "Tarifário Personalizado", "tB": "Tarifário configurado pelo utilizador."}
                            )
                        # 3. Adicionar os tarifários base que aparecem sempre
                        itens_legenda_excel.extend([
                            {"cf": "FFE699", "ct": "000000", "b": False, "tA": "Indexado Média", "tB": "Preço de energia baseado na média OMIE do período."},
                            {"cf": "4D79BC", "ct": "FFFFFF", "b": False, "tA": "Indexado Quarto-horário - Perfil", "tB": "Preço de energia baseado nos valores OMIE horários/quarto-horários e perfil."},
                        ])
                        # 4. Adicionar condicionalmente a legenda do diagrama
                        if 'dados_completos_ficheiro' in st.session_state and st.session_state.dados_completos_ficheiro is not None:
                            itens_legenda_excel.append(
                                {"cf": "BDD7EE", "ct": "000000", "b": False, "tA": "Indexado Quarto-horário - Diagrama", "tB": "Preço de energia baseado nos valores OMIE quarto-horários e calculado com o ficheiro de consumo."}
                            )
                        # 5. Adicionar sempre o item 'Fixo' no final
                        itens_legenda_excel.append(
                            {"cf": "F0F0F0", "ct": "333333", "b": False, "tA": "Fixo", "tB": "Preços de energia constantes", "borda_cor": "CCCCCC"}
                        )
                        
                        # Definir larguras das colunas para a legenda (pode ajustar conforme necessário)
                        worksheet_excel.column_dimensions[get_column_letter(1)].width = 30 # Coluna A para a amostra/nome
                        worksheet_excel.column_dimensions[get_column_letter(2)].width = 70 # Coluna B para a descrição (será junta)

                        for item in itens_legenda_excel:
                            celula_A_legenda = worksheet_excel.cell(row=linha_legenda_item_atual, column=1, value=item["tA"])
                            celula_A_legenda.fill = PatternFill(start_color=item["cf"], end_color=item["cf"], fill_type="solid")
                            celula_A_legenda.font = Font(color=item["ct"], bold=item["b"])
                            celula_A_legenda.alignment = Alignment(horizontal='center', vertical='center', indent=1)

                            if "borda_cor" in item:
                                cor_borda_hex = item["borda_cor"]
                                borda_legenda_obj = Border(
                                    top=Side(border_style="thin", color=cor_borda_hex),
                                    left=Side(border_style="thin", color=cor_borda_hex),
                                    right=Side(border_style="thin", color=cor_borda_hex),
                                    bottom=Side(border_style="thin", color=cor_borda_hex)
                                )
                                celula_A_legenda.border = borda_legenda_obj
                            
                            celula_B_legenda = worksheet_excel.cell(row=linha_legenda_item_atual, column=2, value=item["tB"])
                            celula_B_legenda.alignment = Alignment(vertical='center', wrap_text=True, horizontal='left')
                            # Fundir colunas B até D (ou Ajustar conforme a largura desejada para a descrição)
                            worksheet_excel.merge_cells(start_row=linha_legenda_item_atual, start_column=2,
                                                        end_row=linha_legenda_item_atual, end_column=4) 
                            
                            worksheet_excel.row_dimensions[linha_legenda_item_atual].height = 20 # Ajustar altura da linha da legenda
                            linha_legenda_item_atual += 1
                        # ---- FIM: ADICIONAR LEGENDA DE CORES ----

                        # Ajustar largura das colunas da tabela principal
                        for col_idx_iter, col_nome_iter_width in enumerate(df_para_exportar.columns):
                            col_letra_iter = get_column_letter(col_idx_iter + 1) # get_column_letter já deve estar importado
                            if "Tarifário" in col_nome_iter_width :
                                worksheet_excel.column_dimensions[col_letra_iter].width = 95    
                            elif "Total (€)" == col_nome_iter_width :
                                worksheet_excel.column_dimensions[col_letra_iter].width = 25
                            elif "(€/kWh)" in col_nome_iter_width or "(€/dia)" in col_nome_iter_width:
                                worksheet_excel.column_dimensions[col_letra_iter].width = 25
                            elif "Comercializador" in col_nome_iter_width :
                                worksheet_excel.column_dimensions[col_letra_iter].width = 30    
                            elif "Faturação" in col_nome_iter_width :
                                worksheet_excel.column_dimensions[col_letra_iter].width = 33    
                            elif "Pagamento" in col_nome_iter_width :
                                worksheet_excel.column_dimensions[col_letra_iter].width = 50    
                            else: 
                                worksheet_excel.column_dimensions[col_letra_iter].width = 25


                    output_excel_buffer.seek(0)
                    return output_excel_buffer
                        # --- Fim da definição de exportar_excel_completo ---

                # --- Início do Bloco Numero Tarifários exportados ---
                opcoes_limite_export = ["Todos"] + [f"Top {i}" for i in [10, 20, 30, 40, 50]]
                limite_export_selecionado = st.selectbox(
                    "Número de tarifários a exportar (ordenados pelo 'Total (€)' atual da tabela):",
                    options=opcoes_limite_export,
                    index=0, # "Todos" como padrão
                    key="limite_tarifarios_export_excel"
                )
                # --- Fim do Bloco Numero Tarifários exportados ---

                # --- Bloco Preparar Download ---
                if st.button("Preparar Download do Ficheiro Excel (Dados Selecionados)", key="btn_prep_excel_download_dados_com_tooltips_corrigido"):
                    if not colunas_para_exportar_excel_selecionadas:
                        st.warning("Por favor, selecione pelo menos uma coluna para exportar.")
                    else:
                        # O código de geração do Excel e o st.download_button VÊM AQUI DENTRO
                        with st.spinner("A gerar ficheiro Excel..."):
                            #df_export_final = df_resultados_para_aggrid[colunas_para_exportar_excel_selecionadas].copy()
                            if grid_response and grid_response['data'] is not None: # Verifica se grid_response e os dados existem
                            # grid_response['data'] contém os dados filtrados e ordenados da AgGrid como uma lista de dicionários
                                df_dados_filtrados_da_grid = pd.DataFrame(grid_response['data'])


                            if df_dados_filtrados_da_grid.empty and not df_resultados_para_aggrid.empty:
                                # Isto pode acontecer se os filtros resultarem numa tabela vazia
                                st.warning("Os filtros aplicados resultaram numa tabela vazia. A exportar um ficheiro vazio ou com cabeçalhos apenas.")
                                # Decide o que fazer: exportar ficheiro vazio ou parar.
                                # Para exportar ficheiro vazio com cabeçalhos:
                                df_export_final = pd.DataFrame(columns=colunas_para_exportar_excel_selecionadas)

                            elif not df_dados_filtrados_da_grid.empty:
                                # Assegurar que apenas as colunas selecionadas pelo utilizador para exportação são usadas,
                                # e que estas colunas existem no df_dados_filtrados_da_grid.
                                colunas_export_validas_no_filtrado = [
                                    col for col in colunas_para_exportar_excel_selecionadas 
                                    if col in df_dados_filtrados_da_grid.columns
                                ]
                                if not colunas_export_validas_no_filtrado:
                                    st.warning("Nenhuma das colunas selecionadas para exportação está presente nos dados filtrados atuais da tabela.")
                        
                                df_export_final = df_dados_filtrados_da_grid[colunas_export_validas_no_filtrado].copy()
                            else: # Se grid_response['data'] for None ou vazio e df_resultados_para_aggrid também era vazio
                                st.warning("Não há dados na tabela para exportar.")

                            # Aplicar limite de tarifários, se não for "Todos"
                            if limite_export_selecionado != "Todos":
                                try:
                                    # Extrai o número do "Top N"
                                    num_a_exportar = int(limite_export_selecionado.split(" ")[1])
                                    
                                    # df_export_final já deve estar na ordem da AgGrid (que é ordenada por 'Total (€)' por defeito)
                                    # Se precisar garantir a ordenação por 'Total (€)' ascendentemente aqui:
                                    # if 'Total (€)' in df_export_final.columns:
                                    #     df_export_final = df_export_final.sort_values(by='Total (€)', ascending=True)
                                    
                                    if len(df_export_final) > num_a_exportar:
                                        df_export_final = df_export_final.head(num_a_exportar)
                                        st.info(f"A exportar os {num_a_exportar} primeiros tarifários da tabela atual.")
                                except Exception as e_limite_export:
                                    st.warning(f"Não foi possível aplicar o limite de tarifários: {e_limite_export}")


                            nome_coluna_tarifario_excel = None
                            if 'NomeParaExibir' in df_export_final.columns:
                                df_export_final.rename(columns={'NomeParaExibir': 'Tarifário'}, inplace=True)
                                nome_coluna_tarifario_excel = 'Tarifário'
                            elif 'Tarifário' in df_export_final.columns:
                                nome_coluna_tarifario_excel = 'Tarifário'

                            # --- Obter a coluna 'Tipo' do DataFrame original para usar na estilização ---
                            # Isto garante que temos os tipos mesmo que a coluna 'Tipo' não seja exportada.
                            # Assumimos que df_export_final mantém o índice de df_resultados_para_aggrid.
                            tipos_reais_para_estilo = None
                            if 'Tipo' in df_dados_filtrados_da_grid.columns: # Usar df_dados_filtrados_da_grid
                                try:
                                    # df_export_final agora tem um novo índice (0, 1, 2...).
                                    # Precisamos de alinhar com base no índice de df_dados_filtrados_da_grid que corresponde
                                    # às linhas de df_export_final. Se df_export_final é apenas uma seleção de colunas
                                    # de df_dados_filtrados_da_grid, o índice direto deve funcionar.
                                    tipos_reais_para_estilo = df_dados_filtrados_da_grid.loc[df_export_final.index, 'Tipo']
                                except KeyError:
                                    tipos_reais_para_estilo = pd.Series(index=df_export_final.index, dtype=str)
                            else:
                                tipos_reais_para_estilo = pd.Series(index=df_export_final.index, dtype=str)

                            # --- Função de interpolação de cores ---
                            def gerar_estilo_completo_para_valor(valor, minimo, maximo):
                                estilo_css_final = 'text-align: center;' 
                                if pd.isna(valor): return estilo_css_final
                                try: val_float = float(valor)
                                except ValueError: return estilo_css_final
                                if maximo == minimo or minimo is None or maximo is None: return estilo_css_final
                    
                                midpoint = (minimo + maximo) / 2
                                r_bg, g_bg, b_bg = 255,255,255 
                                verde_rgb, branco_rgb, vermelho_rgb = (90,138,198), (255,255,255), (247,150,70)

                                if val_float <= midpoint:
                                    ratio = (val_float - minimo) / (midpoint - minimo) if midpoint != minimo else 0.0
                                    r_bg = int(verde_rgb[0]*(1-ratio) + branco_rgb[0]*ratio)
                                    g_bg = int(verde_rgb[1]*(1-ratio) + branco_rgb[1]*ratio)
                                    b_bg = int(verde_rgb[2]*(1-ratio) + branco_rgb[2]*ratio)
                                else:
                                    ratio = (val_float - midpoint) / (maximo - midpoint) if maximo != midpoint else 0.0
                                    r_bg = int(branco_rgb[0]*(1-ratio) + vermelho_rgb[0]*ratio)
                                    g_bg = int(branco_rgb[1]*(1-ratio) + vermelho_rgb[1]*ratio)
                                    b_bg = int(branco_rgb[2]*(1-ratio) + vermelho_rgb[2]*ratio)
                    
                                estilo_css_final += f' background-color: #{r_bg:02X}{g_bg:02X}{b_bg:02X};'
                                luminancia = (0.299 * r_bg + 0.587 * g_bg + 0.114 * b_bg)
                                cor_texto_css = '#000000' if luminancia > 140 else '#FFFFFF'
                                estilo_css_final += f' color: {cor_texto_css};'
                                return estilo_css_final


                            # --- Função de estilo principal a ser aplicada ao DataFrame ---
                            def estilo_geral_dataframe_para_exportar(df_a_aplicar_estilo, tipos_reais_para_estilo_serie, min_max_config_para_cores, nome_coluna_tarifario="Tarifário"):
                                df_com_estilos = pd.DataFrame('', index=df_a_aplicar_estilo.index, columns=df_a_aplicar_estilo.columns)
                
                                # Cores default (pode ajustar)
                                cor_fundo_indexado_media_css_local = "#FFE699"
                                cor_texto_indexado_media_css_local = "black"
                                cor_fundo_indexado_dinamico_css_local = "#4D79BC"  
                                cor_texto_indexado_dinamico_css_local = "white"
                                cor_fundo_indexado_diagrama_css_local = "#BDD7EE"
                                cor_texto_indexado_diagrama_css_local = "black"

                                for nome_coluna_df in df_a_aplicar_estilo.columns:
                                    # Estilo para colunas de custo (Total (€) na detalhada, Total [Opção] (€) na comparativa)
                                    if nome_coluna_df in min_max_config_para_cores: # min_max_config_para_cores é o min_max_data_X_json_string convertido para dict
                                        try:
                                            serie_valores_col = pd.to_numeric(df_a_aplicar_estilo[nome_coluna_df], errors='coerce')
                                            min_valor_col = min_max_config_para_cores[nome_coluna_df]['min']
                                            max_valor_col = min_max_config_para_cores[nome_coluna_df]['max']
                                            df_com_estilos[nome_coluna_df] = serie_valores_col.apply(
                                                lambda valor_v: gerar_estilo_completo_para_valor(valor_v, min_valor_col, max_valor_col)
                                            )
                                        except Exception as e_estilo_custo:
                                            print(f"Erro ao aplicar estilo de custo à coluna {nome_coluna_df}: {e_estilo_custo}")
                                            df_com_estilos[nome_coluna_df] = 'text-align: center;' 
            
                                    elif nome_coluna_df == nome_coluna_tarifario: # 'Tarifário' ou 'NomeParaExibir'
                                        estilos_col_tarif_lista = []
                                        for idx_linha_df, valor_nome_col_tarif in df_a_aplicar_estilo[nome_coluna_df].items():
                                            tipo_tarif_str = tipos_reais_para_estilo_serie.get(idx_linha_df, '') if tipos_reais_para_estilo_serie is not None else ''
                    
                                            est_css_tarif = 'text-align: center; padding: 4px;' 
                                            bg_cor_val, fonte_cor_val, fonte_peso_val = "#FFFFFF", "#000000", "normal" # Default Fixo/Outro (branco)

                                            if isinstance(valor_nome_col_tarif, str) and valor_nome_col_tarif.startswith("O Meu Tarifário"):
                                                bg_cor_val, fonte_cor_val, fonte_peso_val = "#FF0000", "#FFFFFF", "bold"
                                            elif isinstance(valor_nome_col_tarif, str) and valor_nome_col_tarif.startswith("Tarifário Personalizado"):
                                                bg_cor_val, fonte_cor_val, fonte_peso_val = "#92D050", "#FFFFFF", "bold"
                                            elif tipo_tarif_str == 'Indexado Média':
                                                bg_cor_val, fonte_cor_val = cor_fundo_indexado_media_css_local, cor_texto_indexado_media_css_local
                                            elif tipo_tarif_str == 'Indexado quarto-horário':
                                                bg_cor_val, fonte_cor_val = cor_fundo_indexado_dinamico_css_local, cor_texto_indexado_dinamico_css_local
                                            elif tipo_tarif_str == 'Indexado quarto-horário (Diagrama)':
                                                bg_cor_val, fonte_cor_val = cor_fundo_indexado_diagrama_css_local, cor_texto_indexado_diagrama_css_local
                                            elif tipo_tarif_str == 'Fixo': # Para dar um fundo um pouco diferente aos fixos
                                                bg_cor_val = "#F0F0F0" 
                    
                                            est_css_tarif += f' background-color: {bg_cor_val}; color: {fonte_cor_val}; font-weight: {fonte_peso_val};'
                                            estilos_col_tarif_lista.append(est_css_tarif)
                                        df_com_estilos[nome_coluna_df] = estilos_col_tarif_lista
                                    else: # Outras colunas de texto ou sem estilização de cor baseada em valor
                                        df_com_estilos[nome_coluna_df] = 'text-align: center;'
                                return df_com_estilos

                            # 1. Aplicar a função de estilo principal que retorna strings CSS
                            # Para a tabela detalhada, min_max_data_for_js é o dicionário com min/max para as colunas de preço e Total.
                            styler_excel = df_export_final.style.apply(
                                lambda df: estilo_geral_dataframe_para_exportar(df, tipos_reais_para_estilo, min_max_data_for_js, "Tarifário"), # Passa min_max_data_for_js
                                axis=None
                            )

                            # 2. Aplicar formatação de número (casas decimais)
                            for coluna_formatar in df_export_final.columns:
                                if '(€/kWh)' in coluna_formatar or '(€/dia)' in coluna_formatar:
                                    styler_excel = styler_excel.format(formatter="{:.4f}", subset=[coluna_formatar], na_rep="-")
                                elif 'Total (€)' in coluna_formatar:
                                    styler_excel = styler_excel.format(formatter="{:.2f}", subset=[coluna_formatar], na_rep="-")

                
                            # 3. Aplicar estilos de tabela gerais (cabeçalhos, bordas para todas as células td)
                            styler_excel = styler_excel.set_table_styles([
                                {'selector': 'th', 'props': [
                                    ('background-color', '#404040'), ('color', 'white'),
                                    ('font-weight', 'bold'), ('text-align', 'center'),
                                    ('border', '1px solid black'), ('padding', '5px')]},
                                {'selector': 'td', 'props': [ 
                                    ('border', '1px solid #dddddd'), ('padding', '4px')
                                ]}
                            ]).hide(axis="index")
                
                            # Obter o resumo_html e a mensagem de poupança
                            # Certifique-se que html_resumo_final está definido e acessível neste escopo
                            resumo_html_para_excel_func = html_resumo_final if 'html_resumo_final' in locals() else "Resumo não disponível."
                            poupanca_texto_para_excel_func = st.session_state.get('poupanca_excel_texto', "")

                            output_excel_bytes = exportar_excel_completo(
                                df_export_final,
                                styler_excel,
                                html_resumo_final,
                                st.session_state.get('poupanca_excel_texto', ""),
                                opcao_horaria,
                                meu_tarifario_ativo,
                                personalizado_ativo
                            )

                            timestamp_final_dl = int(time.time()) # import time no início do script
                            nome_ficheiro_final_dl = f"Tiago_Felicia_Eletricidade_detalhe_{timestamp_final_dl}.xlsx"
                
                            st.download_button(
                                label=f"📥 Descarregar Excel ({nome_ficheiro_final_dl})",
                                data=output_excel_bytes.getvalue(), # output_excel_bytes é o BytesIO retornado por exportar_excel_completo
                                file_name=nome_ficheiro_final_dl,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key=f"btn_dl_excel_completo_{timestamp_final_dl}" 
                            )
                            st.success(f"{nome_ficheiro_final_dl} pronto para download!")

        # Inicio Secção "Pódio da Poupança"

        st.subheader("🏆 O Seu Pódio da Poupança")
        st.markdown("Estas são as 3 opções mais económicas para si, com base nos seus consumos atuais.")

        # Garantir que o DataFrame está ordenado e o índice está correto
        df_resultados_ordenado = df_resultados.sort_values(by="Total (€)").reset_index(drop=True)
        top3 = df_resultados_ordenado.head(3)

        if len(top3) >= 3:
            # --- Lógica de Referência para a Poupança ---
            custo_referencia = None
            nome_referencia = ""
        
            if meu_tarifario_ativo and 'meu_tarifario_calculado' in st.session_state:
                meu_tar_resultado = st.session_state['meu_tarifario_calculado']
                if 'Total (€)' in meu_tar_resultado and pd.notna(meu_tar_resultado['Total (€)']):
                    custo_referencia = meu_tar_resultado['Total (€)']
                    nome_referencia = meu_tar_resultado['NomeParaExibir']

            if custo_referencia is None:
                if not df_resultados_ordenado.empty:
                    pior_tarifario = df_resultados_ordenado.iloc[-1]
                    custo_referencia = pior_tarifario['Total (€)']
                    nome_referencia = pior_tarifario['NomeParaExibir']
        
            if custo_referencia is not None:
                st.caption(f"A comparação é feita em relação ao seu ponto de referência (o Meu tarifário, ou se este não existir, o mais caro da tabela): **'{nome_referencia}' ({custo_referencia:.2f} €)**.")
        
            # --- Apresentação do Pódio ---
            col2, col1, col3 = st.columns([1, 1.2, 1])

            def apresentar_item_podio(coluna, dados_podio, emoji):
                with coluna:
                    st.markdown(f"<p style='text-align: center; font-size: 24px;'>{emoji}</p>", unsafe_allow_html=True)
                    with st.container(border=True):
                        st.markdown(f"<p style='text-align: center; font-weight: bold;'>{dados_podio['NomeParaExibir']}</p>", unsafe_allow_html=True)
                        st.metric("Custo Estimado", f"{dados_podio['Total (€)']:.2f} €")
                    
                        if custo_referencia is not None:
                            diferenca = dados_podio['Total (€)'] - custo_referencia
                        
                            if diferenca < 0:
                                # É mais barato que a referência -> Poupança
                                st.metric("Poupança", f"{abs(diferenca):.2f} €/mês", delta_color="off")
                            elif diferenca > 0:
                                # É mais caro que a referência -> Custo Adicional
                                st.metric("Custo Adicional", f"{diferenca:.2f} €/mês", delta=f"{diferenca:.2f} €", delta_color="inverse")
                            else:
                                st.metric("Custo", "Igual à referência", delta_color="off")
                            
                        if pd.notna(dados_podio['LinkAdesao']) and 'http' in str(dados_podio['LinkAdesao']):
                            st.link_button("Saber Mais", dados_podio['LinkAdesao'], use_container_width=True)

            # 🥇 1º Lugar (coluna do meio)
            apresentar_item_podio(col1, top3.iloc[0], "🥇 1º lugar")
        
            # 🥈 2º Lugar (coluna da esquerda)
            apresentar_item_podio(col2, top3.iloc[1], "🥈2º lugar")
        
            # 🥉 3º Lugar (coluna da direita)
            apresentar_item_podio(col3, top3.iloc[2], "🥉3º lugar")

        st.markdown("---") # Separador antes da tabela detalhada
            # FIM Secção "Pódio da Poupança"

    # ##################################################################
    # --- INÍCIO: SECÇÃO ANÁLISE DE POUPANÇA COM AUTOCONSUMO ---
    # ##################################################################
    if st.session_state.get("chk_autoconsumo_ativo", False) and is_diagram_mode:
        with st.expander("📊 Analisar Poupança Detalhada com Autoconsumo", expanded=False):
            
            st.info("Clique no botão para calcular e comparar o custo de cada tarifário (disponível para a sua Opção Horária selecionada) com e sem a produção dos seus painéis solares fotovoltaicos.")

            if st.button("Calcular Tabela de Poupança", key="btn_calcular_poupanca_solar", use_container_width=True):
                with st.spinner("A calcular a poupança para os tarifários aplicáveis..."):
                                    
                    # --- PASSO 1: FILTRAR OS TARIFÁRIOS DE PERFIL (FIXOS E INDEXADOS MÉDIA/PERFIL) ---
                    tarifarios_fixos_filtrados = tf_processar[
                        (tf_processar['opcao_horaria_e_ciclo'] == opcao_horaria) &
                        (tf_processar['potencia_kva'] == potencia)
                    ]
                    tarifarios_indexados_perfil_filtrados = ti_processar[
                        (ti_processar['opcao_horaria_e_ciclo'] == opcao_horaria) &
                        (ti_processar['potencia_kva'] == potencia)
                    ]
                    tarifarios_de_perfil_para_analise = pd.concat([tarifarios_fixos_filtrados, tarifarios_indexados_perfil_filtrados])

                    lista_poupanca = []

                    # --- PASSO 2: PREPARAR OS DICIONÁRIOS DE CONSUMO (BRUTO E LÍQUIDO) ---
                    oh_lower_global = opcao_horaria.lower()
                    consumos_brutos_repartidos = {}
                    consumos_liquidos_repartidos = {}
                    if oh_lower_global == "simples":
                        consumos_brutos_repartidos = {'S': consumos_agregados_brutos.get('Simples', 0)}
                        consumos_liquidos_repartidos = {'S': consumos_para_custos.get('Simples', 0)}
                    elif oh_lower_global.startswith("bi"):
                        ciclo_a_usar = 'BD' if 'diário' in oh_lower_global else 'BS'
                        consumos_brutos_repartidos = consumos_agregados_brutos.get(ciclo_a_usar, {})
                        consumos_liquidos_repartidos = consumos_para_custos.get(ciclo_a_usar, {})
                    elif oh_lower_global.startswith("tri"):
                        ciclo_a_usar = 'TD' if 'diário' in oh_lower_global else 'TS'
                        consumos_brutos_repartidos = consumos_agregados_brutos.get(ciclo_a_usar, {})
                        consumos_liquidos_repartidos = consumos_para_custos.get(ciclo_a_usar, {})

                    # --- PASSO 3: ITERAR E CALCULAR TARIFÁRIOS DE PERFIL ---
                    for _, tarifario_linha in tarifarios_de_perfil_para_analise.iterrows():
                        custo_sem_pv, custo_com_pv = None, None
                        tipo_de_tarifario = str(tarifario_linha.get('tipo', '')).strip()
                        res_bruto, res_liquido = None, None

                        if tipo_de_tarifario == 'Fixo':
                            res_bruto = calc.calcular_detalhes_custo_tarifario_fixo(tarifario_linha, opcao_horaria, consumos_brutos_repartidos, potencia, dias, tarifa_social, familia_numerosa, valor_dgeg_user, valor_cav_user, incluir_quota_acp, desconto_continente, CONSTANTES, dias_mes, mes, ano_atual, data_inicio, data_fim, FINANCIAMENTO_TSE_VAL, VALOR_QUOTA_ACP_MENSAL)
                            res_liquido = calc.calcular_detalhes_custo_tarifario_fixo(tarifario_linha, opcao_horaria, consumos_liquidos_repartidos, potencia, dias, tarifa_social, familia_numerosa, valor_dgeg_user, valor_cav_user, incluir_quota_acp, desconto_continente, CONSTANTES, dias_mes, mes, ano_atual, data_inicio, data_fim, FINANCIAMENTO_TSE_VAL, VALOR_QUOTA_ACP_MENSAL)
                        elif tipo_de_tarifario.startswith('Indexado'):
                            res_bruto = calc.calcular_detalhes_custo_tarifario_indexado(tarifario_linha, opcao_horaria, opcao_horaria, consumos_brutos_repartidos, potencia, dias, tarifa_social, familia_numerosa, valor_dgeg_user, valor_cav_user, CONSTANTES, df_omie_ajustado, perdas_medias, todos_omie_inputs_utilizador_comp, omie_medios_calculados_para_todos_ciclos, omie_medio_simples_real_kwh, dias_mes, mes, ano_atual, data_inicio, data_fim, FINANCIAMENTO_TSE_VAL)
                            res_liquido = calc.calcular_detalhes_custo_tarifario_indexado(tarifario_linha, opcao_horaria, opcao_horaria, consumos_liquidos_repartidos, potencia, dias, tarifa_social, familia_numerosa, valor_dgeg_user, valor_cav_user, CONSTANTES, df_omie_ajustado, perdas_medias, todos_omie_inputs_utilizador_comp, omie_medios_calculados_para_todos_ciclos, omie_medio_simples_real_kwh, dias_mes, mes, ano_atual, data_inicio, data_fim, FINANCIAMENTO_TSE_VAL)
                        
                        if res_bruto: custo_sem_pv = res_bruto.get('Total (€)')
                        if res_liquido: custo_com_pv = res_liquido.get('Total (€)')

                        if custo_sem_pv is not None and custo_com_pv is not None:
                            poupanca_eur = custo_sem_pv - custo_com_pv
                            poupanca_perc = (poupanca_eur / custo_sem_pv * 100) if custo_sem_pv > 0 else 0
                            nome_original = tarifario_linha['nome']
                            formula_calculo = str(tarifario_linha.get('formula_calculo', ''))

                            nome_para_exibir = f"{nome_original} - Perfil" if 'BTN' in formula_calculo else nome_original

                            lista_poupanca.append({
                                "Tarifário": nome_para_exibir,
                                "Tipo": tipo_de_tarifario,
                                "Custo Sem PV (€)": custo_sem_pv,
                                "Custo Com PV (€)": custo_com_pv,
                                "Poupança (€)": poupanca_eur,
                                "Poupança (%)": poupanca_perc
                            })
                    # --- PASSO 4: LOOP DEDICADO PARA TARIFÁRIOS DE DIAGRAMA ---
                    tarifarios_diagrama_filtrados = ti_processar[
                        (ti_processar['opcao_horaria_e_ciclo'] == opcao_horaria) &
                        (ti_processar['potencia_kva'] == potencia) &
                        (ti_processar['formula_calculo'].str.contains('BTN', na=False))
                    ]
                    for _, tarifario_linha in tarifarios_diagrama_filtrados.iterrows():
                        custo_sem_pv, custo_com_pv = None, None
                        res_bruto_diag = calc.calcular_custo_completo_diagrama_carga(tarifario_linha, df_consumos_bruto_filtrado, OMIE_PERDAS_CICLOS, CONSTANTES, dias, potencia, familia_numerosa, tarifa_social, valor_dgeg_user, valor_cav_user, mes, ano_atual, incluir_quota_acp, desconto_continente, FINANCIAMENTO_TSE_VAL, VALOR_QUOTA_ACP_MENSAL)
                        if res_bruto_diag: custo_sem_pv = res_bruto_diag.get('Total (€)')
                        res_liquido_diag = calc.calcular_custo_completo_diagrama_carga(tarifario_linha, df_consumos_a_utilizar, OMIE_PERDAS_CICLOS, CONSTANTES, dias, potencia, familia_numerosa, tarifa_social, valor_dgeg_user, valor_cav_user, mes, ano_atual, incluir_quota_acp, desconto_continente, FINANCIAMENTO_TSE_VAL, VALOR_QUOTA_ACP_MENSAL)
                        if res_liquido_diag: custo_com_pv = res_liquido_diag.get('Total (€)')

                        if custo_sem_pv is not None and custo_com_pv is not None:
                            poupanca_eur = custo_sem_pv - custo_com_pv
                            poupanca_perc = (poupanca_eur / custo_sem_pv * 100) if custo_sem_pv > 0 else 0
                            tipo_de_tarifario = str(tarifario_linha.get('tipo', '')).strip()
                            lista_poupanca.append({
                                "Tarifário": f"{tarifario_linha['nome']} - Diagrama",
                                "Tipo": f"{tipo_de_tarifario} (Diagrama)",
                                "Custo Sem PV (€)": custo_sem_pv,
                                "Custo Com PV (€)": custo_com_pv,
                                "Poupança (€)": poupanca_eur,
                                "Poupança (%)": poupanca_perc
                            })
                    # --- PASSO 5: CÁLCULO PARA "O MEU TARIFÁRIO" ---
                    if meu_tarifario_ativo:
                        
                        # CÁLCULO 1: SEM PV (USA DADOS BRUTOS)
                        resultado_bruto = calc.calcular_detalhes_custo_meu_tarifario(
                            st.session_state, opcao_horaria, consumos_brutos_repartidos, potencia, dias, tarifa_social, familia_numerosa,
                            valor_dgeg_user, valor_cav_user, CONSTANTES, FINANCIAMENTO_TSE_VAL
                        )
                        # CÁLCULO 2: COM PV (USA DADOS LÍQUIDOS)
                        resultado_liquido = calc.calcular_detalhes_custo_meu_tarifario(
                            st.session_state, opcao_horaria, consumos_liquidos_repartidos, potencia, dias, tarifa_social, familia_numerosa,
                            valor_dgeg_user, valor_cav_user, CONSTANTES, FINANCIAMENTO_TSE_VAL
                        )

                        custo_sem_pv = resultado_bruto.get('Total (€)') if resultado_bruto else None
                        custo_com_pv = resultado_liquido.get('Total (€)') if resultado_liquido else None

                        if custo_sem_pv is not None and custo_com_pv is not None:
                            poupanca_eur = custo_sem_pv - custo_com_pv
                            poupanca_perc = (poupanca_eur / custo_sem_pv * 100) if custo_sem_pv > 0 else 0
                            lista_poupanca.append({
                                "Tarifário": resultado_liquido.get('NomeParaExibir', "O Meu Tarifário"),
                                "Tipo": "Pessoal",
                                "Custo Sem PV (€)": custo_sem_pv,
                                "Custo Com PV (€)": custo_com_pv,
                                "Poupança (€)": poupanca_eur,
                                "Poupança (%)": poupanca_perc
                            })

                    # --- PASSO 6: CÁLCULO PARA "TARIFÁRIO PERSONALIZADO" ---
                    if personalizado_ativo and st.session_state.get('dados_tarifario_personalizado', {}).get('ativo'):
                        dados_pers = st.session_state['dados_tarifario_personalizado']
                        precos_energia_pers = {}
                        preco_potencia_pers = 0.0

                        if opcao_horaria.lower() == "simples":
                            precos_energia_pers = {'S': dados_pers['precos_s']['energia']}
                            preco_potencia_pers = dados_pers['precos_s']['potencia']
                        elif opcao_horaria.lower().startswith("bi"):
                            precos_energia_pers = {'V': dados_pers['precos_bi']['vazio'], 'F': dados_pers['precos_bi']['fora_vazio']}
                            preco_potencia_pers = dados_pers['precos_bi']['potencia']
                        elif opcao_horaria.lower().startswith("tri"):
                            precos_energia_pers = {'V': dados_pers['precos_tri']['vazio'], 'C': dados_pers['precos_tri']['cheias'], 'P': dados_pers['precos_tri']['ponta']}
                            preco_potencia_pers = dados_pers['precos_tri']['potencia']

                        if preco_potencia_pers > 0 or any(p > 0 for p in precos_energia_pers.values()):
                            resultado_bruto = calc.calcular_custo_personalizado(precos_energia_pers, preco_potencia_pers, consumos_brutos_repartidos, dados_pers['flags'], CONSTANTES, FINANCIAMENTO_TSE_VAL, dias=dias, potencia=potencia, tarifa_social=tarifa_social, familia_numerosa=familia_numerosa, valor_dgeg_user=valor_dgeg_user, valor_cav_user=valor_cav_user, opcao_horaria_ref=opcao_horaria)
                            resultado_liquido = calc.calcular_custo_personalizado(precos_energia_pers, preco_potencia_pers, consumos_liquidos_repartidos, dados_pers['flags'], CONSTANTES, FINANCIAMENTO_TSE_VAL, dias=dias, potencia=potencia, tarifa_social=tarifa_social, familia_numerosa=familia_numerosa, valor_dgeg_user=valor_dgeg_user, valor_cav_user=valor_cav_user, opcao_horaria_ref=opcao_horaria)
                            
                            custo_sem_pv = resultado_bruto.get('Total (€)')
                            custo_com_pv = resultado_liquido.get('Total (€)')

                            if custo_sem_pv is not None and custo_com_pv is not None:
                                poupanca_eur = custo_sem_pv - custo_com_pv
                                poupanca_perc = (poupanca_eur / custo_sem_pv * 100) if custo_sem_pv > 0 else 0
                                lista_poupanca.append({
                                    "Tarifário": "Tarifário Personalizado",
                                    "Tipo": "Pessoal",
                                    "Custo Sem PV (€)": custo_sem_pv,
                                    "Custo Com PV (€)": custo_com_pv,
                                    "Poupança (€)": poupanca_eur,
                                    "Poupança (%)": poupanca_perc
                                })
                    
                    # --- PASSO 7: RENDERIZAR A TABELA ---
                    if lista_poupanca:
                        df_poupanca = pd.DataFrame(lista_poupanca).sort_values(by="Custo Com PV (€)", ascending=True).reset_index(drop=True)
                        
                        # --- Configuração do AgGrid para a nova tabela ---
                        gb_poupanca = GridOptionsBuilder.from_dataframe(df_poupanca)
                        gb_poupanca.configure_default_column(sortable=True, resizable=True, wrapHeaderText=True, autoHeaderHeight=True)
                        gb_poupanca.configure_column("Tarifário", minWidth=300, flex=2)
                        # --- Definir as cores e o estilo para a coluna 'Tarifário' ---
                        cor_fundo_indexado_media_css = "#FFE699"
                        cor_fundo_indexado_perfil_css = "#4D79BC"
                        cor_fundo_indexado_diagrama_css = "#BDD7EE"
                        cor_fundo_fixo_css = "#f0f0f0"

                        cell_style_poupanca_tarifario_js = JsCode(f"""
                        function(params) {{
                            let style = {{ textAlign: 'left' }};
                            if (params.data) {{
                                const tipo = params.data.Tipo;
                                const nome = params.data.Tarifário;

                                if (tipo === 'Pessoal') {{
                                    if (nome && nome.startsWith('O Meu Tarifário')) {{
                                        style.backgroundColor = 'red';
                                        style.color = 'white';
                                        style.fontWeight = 'bold';
                                    }} else {{
                                        style.backgroundColor = '#92D050'; // Verde para o Tarifário Personalizado
                                        style.color = 'white';
                                        style.fontWeight = 'bold';
                                    }}
                                }}
                                else if (tipo.includes('Diagrama')) {{
                                    style.backgroundColor = '{cor_fundo_indexado_diagrama_css}';
                                    style.color = 'black';
                                }} else if (tipo === 'Indexado quarto-horário') {{
                                    style.backgroundColor = '{cor_fundo_indexado_perfil_css}';
                                    style.color = 'white';
                                }} else if (tipo === 'Indexado Média') {{
                                    style.backgroundColor = '{cor_fundo_indexado_media_css}';
                                    style.color = 'black';
                                }} else if (tipo === 'Fixo') {{
                                    style.backgroundColor = '{cor_fundo_fixo_css}';
                                    style.color = 'black';
                                }}
                            }}
                            return style;
                        }}
                        """)
                        gb_poupanca.configure_column("Tarifário", minWidth=350, flex=2, cellStyle=cell_style_poupanca_tarifario_js)

                        # Adicione esta linha para ocultar a coluna 'Tipo', que só usamos para a cor
                        gb_poupanca.configure_column("Tipo", hide=True)

                        formatter_eur_js = JsCode("""
                            function(params) {
                                if (params.value === null || params.value === undefined || isNaN(params.value)) { return ''; }
                                return '€ ' + Number(params.value).toFixed(2);
                            }""")
                        formatter_perc_js = JsCode("""
                            function(params) {
                                if (params.value === null || params.value === undefined || isNaN(params.value)) { return ''; }
                                return Number(params.value).toFixed(2) + ' %';
                            }""")
                        
                        gb_poupanca.configure_column("Custo Sem PV (€)", type=["numericColumn"], valueFormatter=formatter_eur_js, minWidth=150, flex=1, cellStyle={'textAlign': 'center'})
                        gb_poupanca.configure_column("Custo Com PV (€)", type=["numericColumn"], valueFormatter=formatter_eur_js, minWidth=150, flex=1, cellStyle={'textAlign': 'center'})
                        
                        cell_style_poupanca_eur_js = JsCode("""
                            function(params) {
                                if (params.value > 0) return { 'backgroundColor': '#C6E0B4', 'color': 'black', 'fontWeight': 'bold', 'textAlign': 'center' };
                                if (params.value < 0) return { 'backgroundColor': '#F8CBAD', 'color': 'black', 'fontWeight': 'bold', 'textAlign': 'center' };
                                return { 'textAlign': 'center' };
                            }""")
                        cell_style_poupanca_perc_js = JsCode("""
                            function(params) {
                                if (params.value > 0) return { 'backgroundColor': '#DDEBF7', 'color': 'black', 'fontWeight': 'bold', 'textAlign': 'center' };
                                if (params.value < 0) return { 'backgroundColor': '#F8CBAD', 'color': 'black', 'fontWeight': 'bold', 'textAlign': 'center' };
                                return { 'textAlign': 'center' };
                            }""")
                        
                        gb_poupanca.configure_column("Poupança (€)", type=["numericColumn"], valueFormatter=formatter_eur_js, cellStyle=cell_style_poupanca_eur_js, minWidth=130, flex=1)
                        gb_poupanca.configure_column("Poupança (%)", type=["numericColumn"], valueFormatter=formatter_perc_js, cellStyle=cell_style_poupanca_perc_js, minWidth=130, flex=1)

                        gridOptions_poupanca = gb_poupanca.build()
                        AgGrid(df_poupanca, gridOptions=gridOptions_poupanca, fit_columns_on_grid_load=True, theme='alpine', allow_unsafe_jscode=True, key="aggrid_poupanca_solar", enable_enterprise_modules=True)

                        # --- BOTÃO PARA EXPORTAR DADOS DA TABELA DE POUPANÇA ---
                        if 'df_poupanca' in locals() and not df_poupanca.empty:
                            
                            # PASSO 1: Preparar o DataFrame para exportação (arredondar para 2 casas decimais)
                            df_para_exportar = df_poupanca.copy()
                            colunas_para_arredondar = [
                                "Custo Sem PV (€)", "Custo Com PV (€)", "Poupança (€)", "Poupança (%)"
                            ]
                            for col in colunas_para_arredondar:
                                if col in df_para_exportar.columns:
                                    # Garante que a coluna é numérica antes de arredondar
                                    df_para_exportar[col] = pd.to_numeric(df_para_exportar[col], errors='coerce').round(2)
                            
                            # PASSO 2: Converter o DataFrame para um ficheiro Excel em memória
                            output = io.BytesIO()
                            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                                df_para_exportar.to_excel(writer, index=False, sheet_name='Poupanca_Autoconsumo')
                            
                            processed_data = output.getvalue()
                            
                            # PASSO 3: Criar o botão de download atualizado para Excel
                            st.download_button(
                                label="📥 Descarregar Tabela de Poupança (Excel)",
                                data=processed_data,
                                file_name='analise_poupanca_autoconsumo.xlsx',
                                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', # <-- MIME type para Excel
                                key='btn_download_poupanca_xlsx',
                                use_container_width=True
                            )
                    else:
                        st.warning("Não foi possível calcular os dados de poupança para os tarifários selecionados.")

    # ##################################################################
    # --- FIM: SECÇÃO ANÁLISE DE POUPANÇA COM AUTOCONSUMO ---
    # ##################################################################

    if is_diagram_mode:

        if 'df_resultados' in locals() and not df_resultados.empty:
            df_para_analise = df_resultados.copy()

            # Filtrar apenas os resultados que são baseados no diagrama
            tarifarios_diagrama = df_para_analise[df_para_analise['Tipo'].str.contains("Diagrama", na=False)]
            tarifarios_perfil = df_para_analise[df_para_analise['NomeParaExibir'].str.contains("Perfil", na=False)]

            if not tarifarios_diagrama.empty and not tarifarios_perfil.empty:

                with st.expander("🔍 Análise de Desvios (Perfil Real vs. Perfil Padrão ERSE)", expanded=False):
            
                    st.markdown("""
                    Esta análise compara o custo da sua fatura usando o seu **perfil de consumo real** (extraído do ficheiro) 
                    contra o custo que teria se o seu consumo seguisse o **perfil padrão definido pela ERSE**.
                    Isto ajuda a perceber se o seu padrão de consumo é, por si só, mais económico que a média.
                    """)

                    analise_list = []

                    for _, linha_diagrama in tarifarios_diagrama.iterrows():
                        nome_base_diagrama = extrair_nome_base_tarifario(linha_diagrama['NomeParaExibir'])
                        for _, linha_perfil in tarifarios_perfil.iterrows():
                            nome_base_perfil = extrair_nome_base_tarifario(linha_perfil['NomeParaExibir'])
                        
                            if nome_base_diagrama == nome_base_perfil:
                                custo_real = linha_diagrama['Total (€)']
                                custo_perfil_erse = linha_perfil['Total (€)']
                            
                                if pd.notna(custo_real) and pd.notna(custo_perfil_erse):
                                    diferenca = custo_real - custo_perfil_erse
                                    analise_list.append({
                                        "Tarifário": nome_base_diagrama,
                                        "Custo com o seu Perfil Real (€)": custo_real,
                                        "Custo com Perfil Padrão ERSE (€)": custo_perfil_erse,
                                        "Diferença (€)": diferenca
                                    })
                                    break

                    if analise_list:
                        df_analise = pd.DataFrame(analise_list).sort_values(by="Custo com Perfil Padrão ERSE (€)")
                        
                        # --- INÍCIO DA CONFIGURAÇÃO DA AGGRID ---
                        gb_analise = GridOptionsBuilder.from_dataframe(df_analise)
                        
                        gb_analise.configure_default_column(
                            sortable=True, resizable=True, wrapHeaderText=True, autoHeaderHeight=True
                        )

                        gb_analise.configure_grid_options(domLayout='autoHeight')

                        formatter_eur_js = JsCode("""
                            function(params) {
                                if (params.value === null || params.value === undefined || isNaN(params.value)) { return ''; }
                                return Number(params.value).toFixed(2) + ' €';
                            }""")

                        formatter_diff_js = JsCode("""
                            function(params) {
                                if (params.value === null || params.value === undefined || isNaN(params.value)) { return ''; }
                                let value = Number(params.value);
                                let prefix = value > 0 ? '+' : '';
                                return prefix + value.toFixed(2) + ' €';
                            }""")
                        
                        cell_style_diferenca_js = JsCode("""
                            function(params) {
                                let style = { textAlign: 'center', color: 'white', fontWeight: 'bold' };
                                if (params.value <= 0) {
                                    style.backgroundColor = '#28a745'; // Verde
                                } else {
                                    style.backgroundColor = '#dc3545'; // Vermelho
                                }
                                return style;
                            }""")

                        gb_analise.configure_column("Tarifário", headerName="Tarifário", minWidth=300, flex=2, cellStyle={'textAlign': 'center'})
                        gb_analise.configure_column("Custo com o seu Perfil Real (€)", headerName="Custo Perfil Real (€)", type=["numericColumn"], valueFormatter=formatter_eur_js, minWidth=150, flex=1, cellStyle={'textAlign': 'center'})
                        gb_analise.configure_column("Custo com Perfil Padrão ERSE (€)", headerName="Custo Perfil Padrão (€)", type=["numericColumn"], valueFormatter=formatter_eur_js, minWidth=150, flex=1, cellStyle={'textAlign': 'center'})
                        gb_analise.configure_column("Diferença (€)", headerName="Diferença (€)", type=["numericColumn"], valueFormatter=formatter_diff_js, minWidth=130, flex=1, cellStyle=cell_style_diferenca_js)

                        gridOptions_analise = gb_analise.build()
                        AgGrid(
                            df_analise,
                            gridOptions=gridOptions_analise,
                            custom_css=custom_css, 
                            fit_columns_on_grid_load=True,
                            theme='alpine',
                            allow_unsafe_jscode=True,
                            key="aggrid_analise_desvios",
                            enable_enterprise_modules=True
                        )
                        # --- FIM DA CONFIGURAÇÃO DA AGGRID ---

                        st.caption("Valores negativos na 'Diferença' (verde) significam que o seu perfil é MAIS ECONÓMICO que o padrão.")
                    else:
                        st.info("Não foi possível encontrar pares de tarifários (Diagrama e Perfil) para comparar com os filtros atuais.")


    # --- FIM DA SECÇÃO ---
    # ##################################################################

    # Esta secção inteira só será apresentada se NÃO estivermos em modo de diagrama.
    if not is_diagram_mode:
        st.subheader("🔗 Partilhar Simulação")

        if st.query_params:
            base_url = "https://tiagofelicia.streamlit.app/"
            # Filtra parâmetros, mantendo "0" apenas para ACP e Continente
            params_filtrados = {}
            for k, v in st.query_params.items():
                if v:  # ignora None ou vazio
                    if v == "0" and k not in ("acp", "cont", "m_te", "m_tp", "m_tse", "c_s", "c_v", "c_fv", "c_c", "c_p"):
                        continue  # continua a filtrar zeros de outros parâmetros
                    params_filtrados[k] = v

            if params_filtrados:
                query_string = "&".join([f"{k}={v}" for k, v in params_filtrados.items()])
                shareable_link = f"{base_url}?{query_string}"

                # --- Componente HTML/JS para o campo de texto e botão de copiar ---
                html_componente_copiar = f"""
                <div style="display: flex; align-items: center; gap: 8px; font-family: sans-serif;">
                    <input 
                        type="text" 
                        id="shareable-link-input" 
                        value="{shareable_link}" 
                        readonly 
                        style="width: 100%; padding: 8px; border-radius: 6px; border: 1px solid #ccc; font-size: 14px;"
                    >
                    <button 
                        id="copy-button" 
                        onclick="copyLinkToClipboard()"
                        style="
                            padding: 8px 12px; 
                            border-radius: 6px; 
                            border: 1px solid #ccc;
                            background-color: #f0f2f6; 
                            cursor: pointer;
                            font-size: 14px;
                            white-space: nowrap;
                        "
                    >
                        📋 Copiar Link
                    </button>
                </div>

                <script>
                function copyLinkToClipboard() {{
                    // 1. Obter o elemento do input
                    const linkInput = document.getElementById("shareable-link-input");
                    
                    // 2. Selecionar o texto
                    linkInput.select();
                    linkInput.setSelectionRange(0, 99999); // Necessário para telemóveis

                    // 3. Copiar para a área de transferência
                    navigator.clipboard.writeText(linkInput.value).then(() => {{
                        // 4. Dar feedback ao utilizador
                        const copyButton = document.getElementById("copy-button");
                        copyButton.innerText = "Copiado!";
                        // Voltar ao texto original após 2 segundos
                        setTimeout(() => {{
                            copyButton.innerHTML = "&#128203; Copiar Link"; // &#128203; é o emoji da prancheta
                        }}, 2000);
                    }}).catch(err => {{
                        console.error('Falha ao copiar o link: ', err);
                        const copyButton = document.getElementById("copy-button");
                        copyButton.innerText = "Erro!";
                    }});
                }}
                </script>
                """
                st.components.v1.html(html_componente_copiar, height=55)

            else:
                st.info("Altere um dos parâmetros para gerar um link de partilha.")
        else:
            st.info("Altere um dos parâmetros (Potência, Opção ou Consumos) para gerar um link de partilha.")

    # ##################################################################
    # FIM DO BLOCO
    # ##################################################################

# Legenda das Colunas da Tabela Tarifários de Eletricidade
st.markdown("---")
st.subheader("📖 Legenda das Colunas da Tabela Tarifários de Eletricidade")
st.caption("""
* **Tarifário**: Nome identificativo do tarifário. Pode incluir notas sobre descontos de fatura específicos.
* **Tipo**: Indica se o tarifário é:
    * `Fixo`: Preços de energia e potência são constantes.
    * `Indexado Média`: Preço da energia baseado na média do OMIE para os períodos horários.
    * `Indexado quarto-horário`: Preço da energia baseado nos valores OMIE horários/quarto-horários e no perfil de consumo. Também conhecidos como "Dinâmicos".
            
    * `Pessoal`: O seu tarifário, conforme introduzido.
* **Comercializador**: Empresa que oferece o tarifário.
* **[...] (€/kWh)**: Custo unitário da energia para o período indicado (Simples, Vazio, Fora Vazio, Cheias, Ponta), **sem IVA**.
    * Para "O Meu Tarifário", este valor já reflete quaisquer descontos percentuais de energia e o desconto da Tarifa Social que tenhas configurado.
    * Para os outros tarifários, é o preço base sem IVA, já considerando o desconto da Tarifa Social se ativa.
* **Potência (€/dia)**: Custo unitário diário da potência contratada e Termo Fixo **sem IVA**.
    * Para "O Meu Tarifário", este valor já reflete quaisquer descontos percentuais de potência e o desconto da Tarifa Social que tenhas configurado.
    * Para os outros tarifários, é o preço base sem IVA, já considerando o desconto da Tarifa Social se ativa.
* **Total (€)**: Valor do custo final estimado da fatura para o período simulado. Este custo inclui:
    * Custo da energia consumida (com IVA aplicado conforme as regras).
    * Custo da potência contratada (com IVA aplicado conforme as regras).
    * Taxas adicionais: IEC (Imposto Especial de Consumo, isento com Tarifa Social), DGEG (Taxa de Exploração da Direção-Geral de Energia e Geologia) e CAV (Contribuição Audiovisual).
    * Quaisquer descontos de fatura em euros (para "O Meu Tarifário" ou especificados nos tarifários).
""")

# Usar f-strings para construir o HTML da legenda
st.subheader("🎨 Legenda de Cores por Tipo de Tarifário")

# Definições de cores
cor_fundo_meu_tarifario_legenda = "red"
cor_texto_meu_tarifario_legenda = "white"
cor_fundo_tarifario_personalizado_legenda = "#92D050"
cor_texto_tarifario_personalizado_legenda = "white"
cor_fundo_indexado_media_css = "#FFE699"
cor_texto_indexado_media_css = "black"
cor_fundo_indexado_dinamico_css = "#4D79BC"
cor_texto_indexado_dinamico_css = "white"
cor_fundo_indexado_diagrama_css = "#BDD7EE"
cor_texto_indexado_diagrama_css = "black"
cor_fundo_fixo_legenda = "#f0f0f0" # Cor ligeiramente cinza para Fixo
cor_texto_fixo_legenda = "#333333"
borda_fixo_legenda = "#CCCCCC"     # Borda para o quadrado branco ser visível

# Determinar se um ficheiro está carregado
ficheiro_foi_carregado = 'dados_completos_ficheiro' in st.session_state and st.session_state.dados_completos_ficheiro is not None

# Construir a string HTML da legenda passo a passo
html_items = ""

# Item: O Meu Tarifário
if meu_tarifario_ativo:
    html_items += f"""<div style="display: flex; align-items: center; margin-bottom: 5px;">
        <div style="width: 18px; height: 18px; background-color: {cor_fundo_meu_tarifario_legenda}; border: 1px solid #ccc; border-radius: 4px; margin-right: 8px;"></div>
        <span style="background-color: {cor_fundo_meu_tarifario_legenda}; color: {cor_texto_meu_tarifario_legenda}; padding: 2px 6px; border-radius: 4px; font-weight: bold;">O Meu Tarifário</span>
        <span style="margin-left: 8px;">- Tarifário configurado pelo utilizador.</span>
    </div>"""

# Item: Tarifário Personalizado
if personalizado_ativo:
    html_items += f"""<div style="display: flex; align-items: center; margin-bottom: 5px;">
        <div style="width: 18px; height: 18px; background-color: {cor_fundo_tarifario_personalizado_legenda}; border: 1px solid #ccc; border-radius: 4px; margin-right: 8px;"></div>
        <span style="background-color: {cor_fundo_tarifario_personalizado_legenda}; color: {cor_texto_tarifario_personalizado_legenda}; padding: 2px 6px; border-radius: 4px; font-weight: bold;">Tarifário Personalizado</span>
        <span style="margin-left: 8px;">- Tarifário configurado pelo utilizador.</span>
    </div>"""

# Item: Indexado Média
html_items += f"""<div style="display: flex; align-items: center; margin-bottom: 5px;">
    <div style="width: 18px; height: 18px; background-color: {cor_fundo_indexado_media_css}; border: 1px solid #ccc; border-radius: 4px; margin-right: 8px;"></div>
    <span style="background-color: {cor_fundo_indexado_media_css}; color: {cor_texto_indexado_media_css}; padding: 2px 6px; border-radius: 4px;">Indexado Média</span>
    <span style="margin-left: 8px;">- Preço de energia baseado na média OMIE do período definido.</span>
</div>"""

# Item: Indexado Quarto-horário - Perfil
html_items += f"""<div style="display: flex; align-items: center; margin-bottom: 5px;">
    <div style="width: 18px; height: 18px; background-color: {cor_fundo_indexado_dinamico_css}; border: 1px solid #ccc; border-radius: 4px; margin-right: 8px;"></div>
    <span style="background-color: {cor_fundo_indexado_dinamico_css}; color: {cor_texto_indexado_dinamico_css}; padding: 2px 6px; border-radius: 4px;">Indexado Quarto-horário - Perfil</span>
    <span style="margin-left: 8px;">- Preço de energia baseado nos valores OMIE horários/quarto-horários e perfil de consumo ERSE.</span>
</div>"""

# Item Condicional: Indexado Quarto-horário - Diagrama
if ficheiro_foi_carregado:
    html_items += f"""<div style="display: flex; align-items: center; margin-bottom: 5px;">
        <div style="width: 18px; height: 18px; background-color: {cor_fundo_indexado_diagrama_css}; border: 1px solid #ccc; border-radius: 4px; margin-right: 8px;"></div>
        <span style="background-color: {cor_fundo_indexado_diagrama_css}; color: {cor_texto_indexado_diagrama_css}; padding: 2px 6px; border-radius: 4px;">Indexado Quarto-horário - Diagrama</span>
        <span style="margin-left: 8px;">- Preço de energia baseado nos valores OMIE horários/quarto-horários calculado com base no ficheiro de consumo da E-Redes.</span>
    </div>"""

# Item: Fixo
html_items += f"""<div style="display: flex; align-items: center; margin-bottom: 5px;">
    <div style="width: 18px; height: 18px; background-color: {cor_fundo_fixo_legenda}; border: 1px solid {borda_fixo_legenda}; border-radius: 4px; margin-right: 8px;"></div>
    <span style="background-color: {cor_fundo_fixo_legenda}; color: {cor_texto_fixo_legenda}; padding: 2px 6px; border-radius: 4px;">Tarifário Fixo</span>
    <span style="margin-left: 8px;">- Preços de energia constantes.</span>
</div>"""

# Criar o HTML final e renderizar
legenda_html_completa = f"""<div style="font-size: 14px;">{html_items}</div>"""
st.markdown(legenda_html_completa, unsafe_allow_html=True)

# --- DATAS DE REFERÊNCIA ---
st.markdown("---") # Adiciona um separador visual
st.subheader("📅 Datas de Referência dos Valores de Mercado no simulador")

# 1. Processar e exibir Data_Valores_OMIE
# data_valores_omie_dt já foi processada no início do script.
data_omie_formatada_str = "Não disponível"
if data_valores_omie_dt and isinstance(data_valores_omie_dt, datetime.date):
    try:
        data_omie_formatada_str = data_valores_omie_dt.strftime('%d/%m/%Y')
    except ValueError: # No caso de uma data inválida que passou pelo isinstance
        data_omie_formatada_str = f"Data inválida ({data_valores_omie_dt})"
elif data_valores_omie_dt: # Se existe mas não é um objeto date (pode indicar erro no processamento inicial)
    data_omie_formatada_str = f"Valor não reconhecido como data ({data_valores_omie_dt})"

st.markdown(f"**Valores OMIE (SPOT) até** {data_omie_formatada_str}")

# 2. Processar e exibir Data_Valores_OMIP
data_omip_formatada_str = "Não disponível"
constante_omip_df_row = CONSTANTES[CONSTANTES['constante'] == 'Data_Valores_OMIP']

if not constante_omip_df_row.empty:
    valor_bruto_omip = constante_omip_df_row['valor_unitário'].iloc[0]
    if pd.notna(valor_bruto_omip):
        data_omip_dt_temp = None # Variável temporária para a data OMIP
        try:
            # Tenta converter para pd.Timestamp, que é mais flexível, e depois para objeto date
            if isinstance(valor_bruto_omip, (datetime.datetime, pd.Timestamp)):
                data_omip_dt_temp = valor_bruto_omip.date()
            else:
                timestamp_convertido_omip = pd.to_datetime(valor_bruto_omip, errors='coerce')
                if pd.notna(timestamp_convertido_omip):
                    data_omip_dt_temp = timestamp_convertido_omip.date()
            
            if data_omip_dt_temp and isinstance(data_omip_dt_temp, datetime.date):
                data_omip_formatada_str = data_omip_dt_temp.strftime('%d/%m/%Y')
            elif valor_bruto_omip: # Se a conversão falhou mas havia um valor
                data_omip_formatada_str = f"Valor não reconhecido como data ({valor_bruto_omip})"
            # Se valor_bruto_omip for pd.NaT ou a conversão falhar completamente, mantém "Não disponível"
        except Exception: # Captura outros erros de conversão
            if valor_bruto_omip:
                data_omip_formatada_str = f"Erro ao processar valor ({valor_bruto_omip})"
    # Se valor_bruto_omip for NaN, data_omip_formatada_str permanece "Não disponível"
# Se a constante 'Data_Valores_OMIP' não for encontrada, data_omip_formatada_str permanece "Não disponível"

st.markdown(f"**Valores OMIP (Futuros) atualizados em** {data_omip_formatada_str}")
# --- FIM DA SECÇÃO ---

# --- INÍCIO DA SECÇÃO DE APOIO ---
st.markdown("---") # Adiciona um separador visual antes da secção de apoio
st.subheader("💖 Apoie este Projeto")

st.markdown(
    "Se quiser apoiar a manutenção do site e o desenvolvimento contínuo deste simulador, "
    "pode fazê-lo através de uma das seguintes formas:"
)

# Link para BuyMeACoffee
st.markdown(
    "☕ [**Compre-me um café em BuyMeACoffee**](https://buymeacoffee.com/tiagofelicia)"
)

st.markdown("ou através do botão PayPal:")

# Código HTML para o botão do PayPal
paypal_button_html = """
<div style="text-align: left; margin-top: 10px; margin-bottom: 15px;">
    <form action="https://www.paypal.com/donate" method="post" target="_blank" style="display: inline-block;">
    <input type="hidden" name="hosted_button_id" value="W6KZHVL53VFJC">
    <input type="image" src="https://www.paypalobjects.com/pt_PT/PT/i/btn/btn_donate_SM.gif" border="0" name="submit" title="PayPal - The safer, easier way to pay online!" alt="Faça donativos com o botão PayPal">
    <img alt="" border="0" src="https://www.paypal.com/pt_PT/i/scr/pixel.gif" width="1" height="1">
    </form>
</div>
"""
st.markdown(paypal_button_html, unsafe_allow_html=True)
# --- FIM DA SECÇÃO DE APOIO ---

st.markdown("---")
# Título para as redes sociais
st.subheader("Redes sociais, onde poderão seguir o projeto:")

# URLs das redes sociais
url_x = "https://x.com/tiagofelicia"
url_facebook = "https://www.facebook.com/profile.php?id=61555007360529"
url_instagram = "https://www.instagram.com/tiago_felicia/"
url_youtube = "https://youtube.com/@tiagofelicia"
url_tiktok = "https://www.tiktok.com/@Tiago_Felicia"

icon_url_x = "https://upload.wikimedia.org/wikipedia/commons/thumb/c/cc/X_icon.svg/120px-X_icon.svg.png?20250519203220"
icon_url_facebook = "https://upload.wikimedia.org/wikipedia/commons/thumb/b/b9/2023_Facebook_icon.svg/120px-2023_Facebook_icon.svg.png"
icon_url_instagram = "https://upload.wikimedia.org/wikipedia/commons/a/a5/Instagram_icon.png"
icon_url_youtube = "https://upload.wikimedia.org/wikipedia/commons/thumb/f/fd/YouTube_full-color_icon_%282024%29.svg/120px-YouTube_full-color_icon_%282024%29.svg.png"
icon_url_tiktok = "https://upload.wikimedia.org/wikipedia/commons/a/a6/Tiktok_icon.svg"


svg_icon_style_dark_mode_friendly = "filter: invert(0.8) sepia(0) saturate(1) hue-rotate(0deg) brightness(1.5) contrast(0.8);"

col_social1, col_social2, col_social3, col_social4, col_social5 = st.columns(5)

with col_social1:
    st.markdown(
        f"""
        <a href="{url_x}" target="_blank" style="text-decoration: none; color: inherit; display: flex; flex-direction: column; align-items: center; text-align: center;">
            <img src="{icon_url_x}" width="40" alt="X" style="margin-bottom: 8px; object-fit: contain;">
            X
        </a>
        """,
        unsafe_allow_html=True
    )

with col_social2:
    st.markdown(
        f"""
        <a href="{url_facebook}" target="_blank" style="text-decoration: none; color: inherit; display: flex; flex-direction: column; align-items: center; text-align: center;">
            <img src="{icon_url_facebook}" width="40" alt="Facebook" style="margin-bottom: 8px; object-fit: contain;">
            Facebook
        </a>
        """,
        unsafe_allow_html=True
    )

with col_social3:
    st.markdown(
        f"""
        <a href="{url_instagram}" target="_blank" style="text-decoration: none; color: inherit; display: flex; flex-direction: column; align-items: center; text-align: center;">
            <img src="{icon_url_instagram}" width="40" alt="Instagram" style="margin-bottom: 8px; object-fit: contain;">
            Instagram
        </a>
        """,
        unsafe_allow_html=True
    )

with col_social4:
    st.markdown(
        f"""
        <a href="{url_youtube}" target="_blank" style="text-decoration: none; color: inherit; display: flex; flex-direction: column; align-items: center; text-align: center;">
            <img src="{icon_url_youtube}" width="40" alt="YouTube" style="margin-bottom: 8px; object-fit: contain;">
            YouTube
        </a>
        """,
        unsafe_allow_html=True
    )

with col_social5:
    st.markdown(
        f"""
        <a href="{url_tiktok}" target="_blank" style="text-decoration: none; color: inherit; display: flex; flex-direction: column; align-items: center; text-align: center;">
            <img src="{icon_url_tiktok}" width="40" alt="Tiktok" style="margin-bottom: 8px; object-fit: contain;">
            Tiktok
        </a>
        """,
        unsafe_allow_html=True
    )
st.markdown("<br>", unsafe_allow_html=True) # Adiciona um espaço vertical

# Texto de Copyright
ano_copyright = 2026
nome_autor = "Tiago Felícia"
texto_copyright_html = f"© {ano_copyright} Todos os direitos reservados | {nome_autor}"

st.markdown(
    f"<div style='text-align: center; font-size: 0.9em; color: grey;'>{texto_copyright_html}</div>",
    unsafe_allow_html=True
)
